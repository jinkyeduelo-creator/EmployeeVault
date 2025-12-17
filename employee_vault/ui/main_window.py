"""
Main Window
Main application window with sidebar navigation
"""

import os
import sys
import json
import time
import shutil
import socket
import logging
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, Any

from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *

from employee_vault.config import *
from employee_vault.database import DB
from employee_vault.utils import remove_background
from employee_vault.ui.pages.dashboard import EnhancedDashboardPage
from employee_vault.ui.pages.employees import EmployeesPage
from employee_vault.ui.dialogs.employee_form import EmployeeForm
from employee_vault.ui.dialogs.user_management import UserManagementDialog
from employee_vault.ui.dialogs.bulk_operations import BulkOperationsDialog
from employee_vault.ui.dialogs.id_card import IDCardGeneratorBackenderator
from employee_vault.ui.dialogs.print_dialogs import BatchPrintDialog, PrintSystemDialog


# Phase 3.1: DatabaseInitWorker - Async initial data loading
class DatabaseInitWorker(QThread):
    """Background thread for initial data loading to prevent startup freeze"""
    data_ready = Signal(dict)  # employees, db_mtime

    def __init__(self, db):
        super().__init__()
        self.db = db

    def run(self):
        """Load all initial data in background"""
        try:
            employees = self.db.all_employees()
            db_mtime = db_latest_mtime(DB_FILE)
            self.data_ready.emit({'employees': employees, 'db_mtime': db_mtime})
        except Exception as e:
            logging.error(f"Database init worker error: {e}")
            # Emit empty data on error
            self.data_ready.emit({'employees': [], 'db_mtime': None})
from employee_vault.ui.dialogs.letter_generation import LetterGenerationDialog
from employee_vault.ui.dialogs.store_management import StoreManagementDialog
from employee_vault.ui.dialogs.permissions import PermissionEditorDialog
from employee_vault.ui.dialogs.session_monitor import SessionMonitorDialog
from employee_vault.ui.dialogs.login import LoginDialog
from employee_vault.ui.widgets import *
from employee_vault.ui.modern_ui_helper import show_success_toast, show_error_toast, show_warning_toast, show_info_toast
from employee_vault.ui.widgets import disable_cursor_changes
from employee_vault.ui.widgets import ModernAnimatedButton, PulseButton
from employee_vault.ui.widgets import AnimatedGradientBackground
from employee_vault.ui.widgets import NotificationCenter, NotificationBell, FloatingNotificationPanel
from employee_vault.ui.widgets import get_thumbnail_cache
from employee_vault.ui.ios_button_styles import apply_ios_style
from employee_vault.ui.widgets.page_transitions import PageTransitionManager

# v4.5.0: Animation system imports
try:
    from employee_vault.animation_manager import get_animation_manager
    from employee_vault.ui.particle_effects import create_celebration_effect
    ANIMATIONS_AVAILABLE = True
except ImportError:
    ANIMATIONS_AVAILABLE = False
    get_animation_manager = None
    create_celebration_effect = None


# ============================================================================
# BACKGROUND WORKER FOR DATABASE CHECKPOINT (v5.3)
# ============================================================================
class CheckpointWorker(QThread):
    """Background thread worker for database checkpoint operations.
    Prevents UI freezes by running checkpoint in separate thread."""
    finished = Signal(bool)  # success
    error = Signal(str)  # error message
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self._running = False
    
    def run(self):
        if self._running:
            return  # Prevent overlapping checkpoint operations
        
        self._running = True
        try:
            if self.db:
                self.db.checkpoint_database()
                self.finished.emit(True)
        except Exception as e:
            logging.warning(f"Background checkpoint failed: {e}")
            self.error.emit(str(e))
            self.finished.emit(False)
        finally:
            self._running = False

class MainWindow(QMainWindow):
    def __init__(self, db, username, user_row, icon=None):
        super().__init__(); self.setWindowTitle(f"{APP_TITLE} — {user_row['name']}")
        
        # v3.3: Modern window styling - themed title bar
        self.setAttribute(Qt.WA_TranslucentBackground, False)  # Keep solid background
        
        # Apply modern styling to the window
        current_theme = load_theme_preference()
        theme_colors = MODERN_THEMES.get(current_theme, MODERN_THEMES["default"])
        
        # Set window background to match theme
        self.setStyleSheet(f"""
            QMainWindow {{
                background-color: {theme_colors['background']};
                border: 2px solid {theme_colors['primary']};
                border-radius: 12px;
            }}
        """)
        
        if icon:
            self.setWindowIcon(icon)
        # Set fixed window size
        self.setMinimumSize(1200, 600)
        self.resize(1200, 600)

        # Centralized animation manager (theme-aware durations/easings)
        self.anim_manager = get_animation_manager() if ANIMATIONS_AVAILABLE else None
        # Center the window properly
        screen = QApplication.primaryScreen().availableGeometry()
        self.move(
            (screen.width() - self.width()) // 2,
            (screen.height() - self.height()) // 2
        )

        self.db=db; self.current_user=username; self.user_row=user_row

        # Phase 3.1: Don't load employees during init - defer until after window shows
        self.employees = []  # Empty initially
        self._db_mtime = None

        # v2.0: Load user permissions
        self.user_permissions = self.db.get_user_permissions(self.current_user)

        # v2.0: Create session tracking
        try:
            computer_name = socket.gethostname()
            self.db.create_session(self.current_user, "", computer_name)
        except Exception as e:
            logging.warning(f"Could not create session: {e}")

        # v2.0: Setup session keepalive timer
        self.session_timer = QTimer(self)
        self.session_timer.timeout.connect(lambda: self.db.update_session_activity(self.current_user))
        self.session_timer.start(60000)  # Update every minute

        # v5.2: Periodic checkpoint timer for multi-PC sync
        # Forces WAL checkpoint every 30 seconds to ensure changes are visible across PCs
        self.checkpoint_timer = QTimer(self)
        self.checkpoint_timer.timeout.connect(self._periodic_checkpoint)
        self.checkpoint_timer.start(30000)  # Checkpoint every 30 seconds

        # WEEK 2 FEATURE #1: Idle timeout (auto-lock after inactivity)
        self.idle_timeout_minutes = 30  # Configurable: 30 minutes default
        self.last_activity_time = datetime.now()
        self.idle_timer = QTimer(self)
        self.idle_timer.timeout.connect(self._check_idle_timeout)
        self.idle_timer.start(60000)  # Check every minute

        # Install event filter to track user activity
        self.installEventFilter(self)

        # Log login action
        self.db.log_action(username=self.current_user, action="LOGIN", details=f"User logged in")

        # Page container with smooth transitions
        self.stack = PageTransitionManager()
        self.stack.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.dashboard = EnhancedDashboardPage()

        # Accessibility: Screen reader support for main pages
        self.dashboard.setAccessibleName("Dashboard")
        self.dashboard.setAccessibleDescription("Main dashboard showing employee statistics and quick access cards")

        # Connect dashboard signals for interactive features
        self.dashboard.navigate_to_employees.connect(self._handle_dashboard_navigate)
        self.dashboard.open_employee_detail.connect(self._handle_dashboard_open_employee)
        self.dashboard.open_employee_edit.connect(self._handle_dashboard_edit_employee)

        # Create the form first
        self.emp_form = EmployeeForm(self.db, self.current_user, self._on_form_saved)

        # Create a scroll area and put the form inside it
        self.form_scroll_area = QScrollArea()
        self.form_scroll_area.setWidgetResizable(True)
        self.form_scroll_area.setWidget(self.emp_form)

        self.employees_page=EmployeesPage(self._view_employee, self._edit_employee, self._delete_selected)
        self.employees_page.advanced_search_btn.clicked.connect(self._show_enhanced_search)
        # v5.1: Connect refresh button signal
        self.employees_page.refresh_requested.connect(self._refresh_all)

        # Accessibility: Screen reader support for employees page
        self.employees_page.setAccessibleName("Employees List")
        self.employees_page.setAccessibleDescription("View and manage all employees with search and filter options")

        # Phase 3.1: Don't load data yet - will load after window shows
        # self.dashboard.refresh(self.employees); self.employees_page.set_data(self.employees)

        # Sidebar
        # Create the sidebar frame and store a reference on the instance so
        # toggle_sidebar() can animate its width.  We set both a
        # fixed width and maximum width initially to the expanded size.
        sidebar = QFrame()
        self.sidebar = sidebar
        # Expanded and collapsed widths per reference dashboard
        # 220px expanded, 70px collapsed for modern, clean look (Option D)
        self.sidebar_expanded_width = 220
        self.sidebar_collapsed_width = 70
        self.is_sidebar_collapsed = True  # Start collapsed
        self.sidebar_hover_expanded = False  # Track if sidebar is temporarily expanded on hover
        self.sidebar.setFixedWidth(self.sidebar_collapsed_width)  # Start collapsed
        self.sidebar.setMaximumWidth(self.sidebar_expanded_width)

        # Install event filter for hover-expand functionality
        self.sidebar.installEventFilter(self)

        # iOS-style frosted glass sidebar (no top border-radius for header integration)
        sidebar.setStyleSheet(f"""
            QFrame {{
                background: rgba(30, 30, 40, 0.95);
                border-right: 1px solid rgba(255, 255, 255, 0.06);
                border-radius: 0px;
            }}
        """)

        sidebar_main_layout = QVBoxLayout(sidebar)
        sidebar_main_layout.setContentsMargins(0, 16, 0, 16)
        sidebar_main_layout.setSpacing(0)

        # Store theme colors for header styling
        self.theme_colors = theme_colors

        # Create scroll area for sections
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        # Widget to hold all collapsible sections
        sections_widget = QWidget()
        s = QVBoxLayout(sections_widget)
        s.setContentsMargins(0, 0, 0, 0)
        s.setSpacing(0)

        # --- Sidebar section tracking ---
        # Initialise the lists to hold the collapsible sections and their
        # corresponding icons.  These will be used by the sidebar
        # collapse/expand logic to adjust the button labels and tooltips.
        self.sidebar_sections = []
        self.section_icons = {}

        # ═══════════════════════════════════════════════════════════
        # Main Section - Core employee management features
        # ═══════════════════════════════════════════════════════════
        main_section = CollapsibleSection("Main", start_collapsed=True, color="#5b8ec9", icon="🏠", main_window=self, theme_colors=theme_colors)
        main_section.add_button("  📊 Dashboard", lambda: self._show_page(self.dashboard))
        main_section.add_button("  👥 Employees", lambda: self._show_page(self.employees_page))
        main_section.add_button("  ➕ Add New Employee", lambda: self._show_add_new())
        main_section.add_button("  📊 Reports", self._show_reports)
        s.addWidget(main_section)
        main_section.full_title = "Main"
        self.sidebar_sections.append(main_section)
        self.section_icons[main_section] = "🏠"

        # ═══════════════════════════════════════════════════════════
        # Documents Section - ID cards and letters
        # ═══════════════════════════════════════════════════════════
        documents_section = CollapsibleSection("Documents", start_collapsed=True, color="#5a9a6a", icon="📄", main_window=self, theme_colors=theme_colors)

        # Letter generation (available to all users with permission)
        if self.user_permissions.get('letters', True):
            documents_section.add_button("  📝 Generate Letter", self._show_letter_generation)

        # ID Card Generator (admin only)
        if self.user_permissions.get('user_management'):
            documents_section.add_button("  🆔 ID Card Generator", self._show_id_card_generator)
            documents_section.add_button("  🖨️ Print System", self._show_print_system)

        s.addWidget(documents_section)
        documents_section.full_title = "Documents"
        self.sidebar_sections.append(documents_section)
        self.section_icons[documents_section] = "📄"

        # ═══════════════════════════════════════════════════════════
        # Data Management Section - Backup, export, bulk operations
        # ═══════════════════════════════════════════════════════════
        data_section = CollapsibleSection("Data Management", start_collapsed=True, color="#c9943a", icon="💾", main_window=self, theme_colors=theme_colors)

        # Backup (available to all)
        data_section.add_button("  💾 Backup Now", self._backup_data)
        data_section.add_button("  📁 Export Data", self._export_data)
        data_section.add_button("  📦 Bulk Operations", self._show_bulk_operations)

        # Admin-only features
        if self.user_permissions.get('user_management'):
            data_section.add_button("  ⏰ Scheduled Backup", self._show_scheduled_backup)
            data_section.add_button("  📦 Archive Manager", self._show_archive_manager)

        s.addWidget(data_section)
        data_section.full_title = "Data Management"
        self.sidebar_sections.append(data_section)
        self.section_icons[data_section] = "💾"

        # ═══════════════════════════════════════════════════════════
        # Administration Section - Admin-only system management
        # ═══════════════════════════════════════════════════════════
        if self.user_permissions.get('user_management'):
            admin_section = CollapsibleSection("Administration", start_collapsed=True, color="#c95a5a", icon="🔐", main_window=self, theme_colors=theme_colors)
            admin_section.add_button("  👥 User Management", self._show_user_management)
            admin_section.add_button("  👥 Active Sessions", self._show_session_monitor)
            admin_section.add_button("  🔄 Swap Employee IDs", self._show_emp_id_swap)
            admin_section.add_button("  🌐 Network Config", self._show_network_config)
            admin_section.add_button("  📜 Audit Log", self._show_audit_log)
            admin_section.add_button("  🛑 Force Close All", self._show_force_close_dialog)
            s.addWidget(admin_section)
            admin_section.full_title = "Administration"
            self.sidebar_sections.append(admin_section)
            self.section_icons[admin_section] = "🔐"

        # ═══════════════════════════════════════════════════════════
        # Settings Section - Appearance and help
        # ═══════════════════════════════════════════════════════════
        settings_section = CollapsibleSection("Settings", start_collapsed=True, color="#8a6aaa", icon="⚙️", main_window=self, theme_colors=theme_colors)
        settings_section.add_button("  🎨 Change Theme", self._show_theme_selector)
        settings_section.add_button("  ℹ️ About", self._show_about)
        s.addWidget(settings_section)
        settings_section.full_title = "Settings"
        self.sidebar_sections.append(settings_section)
        self.section_icons[settings_section] = "⚙"

        # Add stretch to push content up but allow scrolling
        s.addStretch(1)

        # Set the sections widget to scroll area
        scroll_area.setWidget(sections_widget)
        sidebar_main_layout.addWidget(scroll_area)

        # Phase 1.5: Pre-cache sidebar stylesheets to avoid regeneration during animation
        self._init_sidebar_styles()

        # Add expanding spacer at bottom
        sidebar_main_layout.addStretch(1)

        # Create content area with animated background
        content_area = QWidget()
        content_inner_layout = QVBoxLayout(content_area)
        content_inner_layout.setContentsMargins(0, 0, 0, 0)
        content_inner_layout.setSpacing(0)

        # v5.0: Animated gradient background (performance-aware)
        self._setup_animated_background(content_area, theme_colors)

        content_inner_layout.addWidget(self.stack, 1)

        # Store reference and install event filter for background resize
        self.content_area = content_area
        content_area.installEventFilter(self)

        # ═══════════════════════════════════════════════════════════════
        # OPTION D LAYOUT: Top Header Bar + Content (Sidebar + Main)
        # ═══════════════════════════════════════════════════════════════
        container = QWidget()
        main_layout = QVBoxLayout(container)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # ─────────────────────────────────────────────────────────────────
        # TOP HEADER BAR - Theme-aware with hamburger, centered logo, user profile
        # ─────────────────────────────────────────────────────────────────
        self.top_header = QFrame()
        self.top_header.setFixedHeight(56)
        self.top_header.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba({self._hex_to_rgb(theme_colors['primary'])}, 0.85),
                    stop:1 rgba({self._hex_to_rgb(theme_colors['primary_dark'])}, 0.95));
                border-bottom: 1px solid rgba(255, 255, 255, 0.08);
            }}
        """)

        top_header_layout = QHBoxLayout(self.top_header)
        top_header_layout.setContentsMargins(16, 0, 16, 0)
        top_header_layout.setSpacing(12)

        # LEFT: Hamburger button - three horizontal bars (auto-centered via layout)
        self.hamburger_btn = QPushButton()
        self.hamburger_btn.setFixedSize(40, 40)

        # Use layout inside button to auto-center the icon
        btn_layout = QVBoxLayout(self.hamburger_btn)
        btn_layout.setContentsMargins(0, 0, 0, 0)
        btn_layout.setAlignment(Qt.AlignCenter)

        # 3 bars of 2px + 2 gaps of 5px = 16px height
        hamburger_widget = QWidget()
        hamburger_widget.setFixedSize(20, 16)
        hamburger_widget.setAttribute(Qt.WA_TranslucentBackground)
        hamburger_widget.setStyleSheet("background: transparent;")
        btn_layout.addWidget(hamburger_widget, 0, Qt.AlignCenter)

        hamburger_layout = QVBoxLayout(hamburger_widget)
        hamburger_layout.setContentsMargins(0, 0, 0, 0)
        hamburger_layout.setSpacing(5)
        for _ in range(3):
            bar = QFrame()
            bar.setFixedSize(20, 2)
            bar.setStyleSheet("background-color: rgba(255, 255, 255, 0.9); border-radius: 1px; border: none;")
            hamburger_layout.addWidget(bar)
        self.hamburger_bars = hamburger_widget.findChildren(QFrame)
        self.hamburger_btn.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
            }
            QPushButton:hover {
                background-color: transparent;
            }
            QPushButton:pressed {
                background-color: transparent;
            }
        """)
        self.hamburger_btn.setToolTip("Toggle Sidebar")
        self.hamburger_btn.clicked.connect(self._toggle_sidebar)
        top_header_layout.addWidget(self.hamburger_btn, 0, Qt.AlignVCenter)

        # Left stretch for true centering
        top_header_layout.addStretch(1)

        # CENTER: Logo + Company Name - visually centered (offset for right-side widgets)
        center_widget = QWidget()
        center_widget.setStyleSheet("background-color: transparent;")
        center_layout = QHBoxLayout(center_widget)
        # Add left margin to offset for the heavier right side (user dropdown + bell)
        center_layout.setContentsMargins(80, 0, 0, 0)
        center_layout.setSpacing(10)

        # Logo
        self.logo_label = QLabel()
        self.logo_label.setObjectName("headerLogo")
        self.logo_label.setStyleSheet("background-color: transparent;")
        logo_path = self._resolve_header_logo_path()
        try:
            logo_pixmap = QPixmap(logo_path) if logo_path else QPixmap()
            if not logo_pixmap.isNull():
                logo_pixmap = logo_pixmap.scaled(36, 36, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.logo_label.setPixmap(logo_pixmap)
                self.logo_label.setToolTip("Cuddly International Corporation")
            else:
                self.logo_label.setText("🏢")
                self.logo_label.setStyleSheet("font-size: 24px; background-color: transparent;")
        except Exception as e:
            logging.warning(f"Could not load logo: {e}")
            self.logo_label.setText("🏢")
            self.logo_label.setStyleSheet("font-size: 24px; background-color: transparent;")

        # v5.4: Add subtle glow effect to logo + hover animation
        logo_glow = QGraphicsDropShadowEffect()
        logo_glow.setBlurRadius(12)
        logo_glow.setColor(QColor(255, 255, 255, 80))
        logo_glow.setOffset(0, 0)
        self.logo_label.setGraphicsEffect(logo_glow)
        self.logo_label.installEventFilter(self)
        self._logo_glow_effect = logo_glow
        self._logo_glow_animation = QPropertyAnimation(logo_glow, b"blurRadius", self)
        self._logo_glow_animation.setDuration(self.anim_manager.get_theme_duration("hover") if self.anim_manager else 180)
        center_layout.addWidget(self.logo_label, 0, Qt.AlignVCenter)

        # Company name - no background
        self.company_label = QLabel("CUDDLY INTERNATIONAL CORPORATION")
        self.company_label.setStyleSheet("""
            color: white;
            font-size: 16px;
            font-weight: 600;
            background-color: transparent;
            border: none;
        """)
        # v5.4: Animated pulsing glow
        self.company_glow = QGraphicsDropShadowEffect()
        self.company_glow.setBlurRadius(12)
        self.company_glow.setColor(QColor(255, 255, 255, 120))
        self.company_glow.setOffset(0, 0)
        self.company_label.setGraphicsEffect(self.company_glow)

        self._glow_animation = QPropertyAnimation(self.company_glow, b"blurRadius")
        self._glow_animation.setDuration(2000)
        self._glow_animation.setStartValue(8)
        self._glow_animation.setEndValue(16)
        self._glow_animation.setEasingCurve(QEasingCurve.InOutSine)
        self._glow_animation.setLoopCount(-1)
        self._glow_animation.start()
        center_layout.addWidget(self.company_label, 0, Qt.AlignVCenter)

        top_header_layout.addWidget(center_widget, 0, Qt.AlignCenter)

        # Right stretch for true centering
        top_header_layout.addStretch(1)

        # RIGHT: User Profile Dropdown + Notification Bell
        user_widget = QWidget()
        user_widget.setStyleSheet("background-color: transparent;")
        user_layout = QHBoxLayout(user_widget)
        user_layout.setContentsMargins(0, 0, 0, 0)
        user_layout.setSpacing(10)

        # User Avatar
        self.user_photo_label = QLabel()
        self.user_photo_label.setFixedSize(36, 36)
        self.user_photo_label.setAlignment(Qt.AlignCenter)
        self.user_photo_label.setStyleSheet(f"""
            QLabel {{
                font-size: 18px;
                background: rgba({self._hex_to_rgb(theme_colors['primary'])}, 0.3);
                border: 2px solid rgba(255, 255, 255, 0.2);
                border-radius: 18px;
            }}
            QLabel:hover {{
                border: 2px solid rgba(255, 255, 255, 0.4);
            }}
        """)
        self._load_user_photo()
        user_layout.addWidget(self.user_photo_label, 0, Qt.AlignVCenter)

        # User Dropdown Button
        self.user_dropdown_btn = QPushButton(f"{user_row['name']} ▼")
        self.user_dropdown_btn.setStyleSheet(f"""
            QPushButton {{
                color: rgba(255, 255, 255, 0.85);
                font-size: 13px;
                background-color: transparent;
                border: none;
                padding: 8px 4px;
            }}
            QPushButton:hover {{ color: white; }}
            QPushButton::menu-indicator {{ image: none; }}
        """)

        # User Menu
        user_menu = QMenu(self)
        user_menu.setStyleSheet(f"""
            QMenu {{
                background: {theme_colors['surface']};
                border: 1px solid rgba(255, 255, 255, 0.1);
                border-radius: 12px;
                padding: 8px;
            }}
            QMenu::item {{
                color: {theme_colors['text_primary']};
                padding: 10px 24px;
                border-radius: 6px;
                margin: 2px 4px;
            }}
            QMenu::item:selected {{
                background: rgba({self._hex_to_rgb(theme_colors['primary'])}, 0.3);
            }}
            QMenu::separator {{
                height: 1px;
                background: rgba(255, 255, 255, 0.08);
                margin: 6px 12px;
            }}
        """)
        user_menu.addAction("👤  Profile Settings")
        user_menu.addAction("🔑  Change Password")
        user_menu.addSeparator()

        # Logout action
        logout_action = user_menu.addAction("🚪  Logout")
        logout_action.triggered.connect(self._logout)

        self.user_dropdown_btn.setMenu(user_menu)
        user_layout.addWidget(self.user_dropdown_btn, 0, Qt.AlignVCenter)

        top_header_layout.addWidget(user_widget, 0, Qt.AlignVCenter)

        # Notification bell at far right - no hover background
        self.notification_bell = NotificationBell()
        self.notification_bell.setToolTip("Notifications - Click to view alerts")
        self.notification_bell.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
                font-size: 18px;
            }
            QPushButton:hover {
                background-color: transparent;
            }
            QPushButton:pressed {
                background-color: transparent;
            }
        """)
        self.notification_bell.clicked.connect(self._toggle_notification_panel)
        top_header_layout.addWidget(self.notification_bell, 0, Qt.AlignVCenter)

        # Add header to main layout
        main_layout.addWidget(self.top_header)

        # ─────────────────────────────────────────────────────────────────
        # CONTENT AREA: Sidebar + Main Content
        # ─────────────────────────────────────────────────────────────────
        content_widget = QWidget()
        content_h_layout = QHBoxLayout(content_widget)
        content_h_layout.setContentsMargins(0, 0, 0, 0)
        content_h_layout.setSpacing(0)
        content_h_layout.addWidget(sidebar)
        content_h_layout.addWidget(content_area, 1)

        main_layout.addWidget(content_widget, 1)
        self.setCentralWidget(container)
        self.stack.addWidget(self.dashboard)
        self.stack.addWidget(self.employees_page)
        self.stack.addWidget(self.form_scroll_area) # <-- Add the SCROLL AREA here
        self._show_page(self.dashboard)

        # Tray icon
        pm=QPixmap(64,64); pm.fill(Qt.transparent); p=QPainter(pm); p.setRenderHint(QPainter.Antialiasing,True); p.setBrush(QColor("#3d6cff")); p.setPen(Qt.NoPen); p.drawEllipse(0,0,64,64); p.end()
        self.tray=QSystemTrayIcon(QIcon(pm), self); self.tray.setToolTip("Cuddly Employees Information"); self.tray.setVisible(True)

        # Timers - v2.0 FIXED: 3 seconds for better multi-user sync
        self.timer=QTimer(self); self.timer.setInterval(3000); self.timer.timeout.connect(self._auto_refresh); self.timer.start()
        self.contract_timer=QTimer(self); self.contract_timer.setInterval(30*60*1000); self.contract_timer.timeout.connect(self._notify_contracts); self.contract_timer.start()

        # QUICK WIN #3: Setup keyboard shortcuts
        self._setup_keyboard_shortcuts()

        # v2.0 FIXED: Apply user permissions after UI is built
        self._apply_user_permissions()
        
        # v5.3: Initialize notification center for contract expiry alerts
        self._setup_notification_center()

        # Initialize sidebar in collapsed state
        self._initialize_sidebar_collapsed()

        # Phase 3.1: Load data asynchronously after window is shown (100ms delay)
        QTimer.singleShot(100, self._load_initial_data)

    def _load_initial_data(self):
        """Phase 3.1: Load employee data in background after window is visible"""
        # Start background worker
        self.data_worker = DatabaseInitWorker(self.db)
        self.data_worker.data_ready.connect(self._on_data_loaded)
        self.data_worker.start()

    def _on_data_loaded(self, data):
        """Phase 3.1: Handler when data is loaded from background thread"""
        self.employees = data['employees']
        self._db_mtime = data['db_mtime']

        # Now populate UI with loaded data
        self.dashboard.refresh(self.employees)
        self.employees_page.set_data(self.employees)

        print(f"[PERF] Initial data loaded: {len(self.employees)} employees")

    def _initialize_sidebar_collapsed(self):
        """Initialize sidebar sections to collapsed/icon-only mode on startup"""
        if not hasattr(self, 'sidebar_sections') or not self.is_sidebar_collapsed:
            return

        for section in self.sidebar_sections:
            icon_char = self.section_icons.get(section, "")
            # Icon-only mode with centered styling
            section.toggle_button.setText(icon_char)
            section.toggle_button.setToolTip(f"{section.full_title} - Click for menu")
            section.toggle_button.setStyleSheet(f"""
                QPushButton {{
                    text-align: center;
                    padding: 14px 8px;
                    font-size: 20px;
                    color: rgba(255, 255, 255, 0.7);
                    background: transparent;
                    border: none;
                    border-radius: 10px;
                    margin: 2px 6px;
                }}
                QPushButton:hover {{
                    color: white;
                    background: rgba(255, 255, 255, 0.1);
                }}
            """)
            # Collapse content area
            section.is_collapsed = True
            section.content_area.setMaximumHeight(0)
            section.content_area.setVisible(False)

    def _setup_animated_background(self, content_area, theme_colors):
        """Setup animated gradient background with performance settings"""
        # Load settings
        settings_file = os.path.join(os.path.dirname(DB_FILE), "settings.json")
        enabled = True
        mode = "wave"
        fps = 30
        
        try:
            if os.path.exists(settings_file):
                with open(settings_file, 'r') as f:
                    settings = json.load(f)
                    enabled = settings.get("enable_animated_background", True)
                    mode = settings.get("animated_background_mode", "wave")
                    fps = settings.get("animated_background_fps", 30)
        except Exception as e:
            logging.warning(f"Could not load animated background settings: {e}")
        
        # Create animated background widget
        self.animated_background = AnimatedGradientBackground(
            content_area,
            theme_colors=theme_colors,
            mode=mode,
            fps=fps
        )
        self.animated_background.setGeometry(content_area.rect())
        self.animated_background.lower()  # Send to back
        
        if enabled:
            self.animated_background.start()
        else:
            self.animated_background.set_enabled(False)
        
        # Store reference for toggle functionality
        self._animated_bg_enabled = enabled
        
    def toggle_animated_background(self, enabled=None):
        """Toggle animated background on/off"""
        if enabled is None:
            enabled = not self._animated_bg_enabled
            
        self._animated_bg_enabled = enabled
        
        if hasattr(self, 'animated_background'):
            self.animated_background.set_enabled(enabled)
            
        # Save to settings
        settings_file = os.path.join(os.path.dirname(DB_FILE), "settings.json")
        try:
            settings = {}
            if os.path.exists(settings_file):
                with open(settings_file, 'r') as f:
                    settings = json.load(f)
            settings["enable_animated_background"] = enabled
            with open(settings_file, 'w') as f:
                json.dump(settings, f, indent=2)
        except Exception as e:
            logging.warning(f"Could not save animated background setting: {e}")
            
        return enabled

    def _hex_to_rgb(self, hex_color):
        """Convert hex color to RGB string for rgba()"""
        hex_color = hex_color.lstrip('#')
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        return f"{r}, {g}, {b}"

    def _setup_keyboard_shortcuts(self):
        """QUICK WIN #3: Setup keyboard shortcuts for power users"""
        from PySide6.QtGui import QShortcut, QKeySequence

        # Ctrl+F - Focus search on employees page
        self.shortcut_search = QShortcut(QKeySequence("Ctrl+F"), self)
        self.shortcut_search.activated.connect(self._focus_search)

        # F5 - Refresh current page
        self.shortcut_refresh = QShortcut(QKeySequence("F5"), self)
        self.shortcut_refresh.activated.connect(self._refresh_all)

        # Ctrl+N - New employee
        self.shortcut_new = QShortcut(QKeySequence("Ctrl+N"), self)
        self.shortcut_new.activated.connect(self._show_add_new)

        # Ctrl+E - Edit selected employee
        self.shortcut_edit = QShortcut(QKeySequence("Ctrl+E"), self)
        self.shortcut_edit.activated.connect(self._edit_selected_employee)

        # Ctrl+B - Backup database
        self.shortcut_backup = QShortcut(QKeySequence("Ctrl+B"), self)
        self.shortcut_backup.activated.connect(self._backup_data)

        # Ctrl+P - Print/Export menu
        self.shortcut_print = QShortcut(QKeySequence("Ctrl+P"), self)
        self.shortcut_print.activated.connect(self._show_print_menu)

        # Ctrl+R - Refresh data
        self.shortcut_refresh_r = QShortcut(QKeySequence("Ctrl+R"), self)
        self.shortcut_refresh_r.activated.connect(self._refresh_all)
        
        # Ctrl+D - Go to Dashboard
        self.shortcut_dashboard = QShortcut(QKeySequence("Ctrl+D"), self)
        self.shortcut_dashboard.activated.connect(lambda: self._show_page(self.dashboard))

    def _setup_notification_center(self):
        """v5.3: Setup notification center for contract expiry alerts"""
        # Create notification center (hidden by default)
        self.notification_center = NotificationCenter(self.db, self)
        
        # Create floating panel
        self.notification_panel = FloatingNotificationPanel(self.notification_center, self)
        self.notification_panel.hide()
        
        # Connect notification center signals
        self.notification_center.notification_clicked.connect(self._handle_notification_click)
        
        # Start monitoring
        self.notification_center.start_monitoring()
        
        # Sync badge count
        QTimer.singleShot(5000, self._update_notification_badge)
        
        # v5.4: Start force close monitor (checks every 10 seconds if admin requested shutdown)
        self.force_close_timer = QTimer(self)
        self.force_close_timer.timeout.connect(self._check_force_close)
        self.force_close_timer.start(10000)  # Check every 10 seconds
        
    def _toggle_notification_panel(self):
        """Toggle the notification panel visibility"""
        if self.notification_panel.isVisible():
            self.notification_panel.hide()
        else:
            # Position panel below the bell
            bell_pos = self.notification_bell.mapToGlobal(QPoint(0, self.notification_bell.height()))
            self.notification_panel.show_at(bell_pos)
            
    def _update_notification_badge(self):
        """Update the notification bell badge count"""
        count = len(self.notification_center.notifications)
        self.notification_bell.set_count(count)
        
    def _handle_notification_click(self, notification: dict):
        """Handle when a notification is clicked"""
        # Hide the panel
        self.notification_panel.hide()
        
        # Navigate based on notification type
        ntype = notification.get('type', '')
        data = notification.get('data', {})
        
        if ntype in ['contract_expired', 'critical', 'warning', 'info', 'contract_expiring']:
            # Show employees page with expiring contracts filter
            self._show_page(self.employees_page)
            
            # Show toast with details
            if data.get('employees'):
                emp_list = data['employees']
                show_info_toast(self, f"Viewing {len(emp_list)} employee(s) with expiring contracts")
            else:
                show_info_toast(self, "Viewing contract expiry alerts")

    def _check_force_close(self):
        """v5.4: Check if admin has requested force close for all users"""
        try:
            force_close_data = self.db.check_force_close()
            if force_close_data:
                # Don't close the admin who triggered it
                requested_by = force_close_data.get('requested_by', '')
                if requested_by == self.current_user and self.user_permissions.get('user_management'):
                    return  # Admin who triggered it stays open
                
                # Stop the timer to prevent multiple dialogs
                self.force_close_timer.stop()
                
                # Show message to user
                message = force_close_data.get('message', 'Application is being updated. Please restart.')
                from PySide6.QtWidgets import QMessageBox
                msg = QMessageBox(self)
                msg.setWindowTitle("🛑 Application Closing")
                msg.setText("The administrator has requested all users to close the application.")
                msg.setInformativeText(f"\n{message}\n\nAny unsaved work will be auto-saved as a draft.\nThe application will close in 10 seconds.")
                msg.setIcon(QMessageBox.Icon.Warning)
                msg.setStandardButtons(QMessageBox.StandardButton.Ok)
                
                # Auto-close after 10 seconds
                def auto_close():
                    msg.accept()
                    self._force_quit()
                
                QTimer.singleShot(10000, auto_close)
                msg.exec()
                
                # If user clicked OK before timeout, close now
                self._force_quit()
                
        except Exception as e:
            logging.warning(f"Error checking force close: {e}")
    
    def _force_quit(self):
        """Force quit the application with auto-save of any unsaved work"""
        try:
            # Auto-save any draft in progress (employee form)
            if hasattr(self, 'emp_form'):
                try:
                    # Force immediate draft save if there's unsaved content
                    if hasattr(self.emp_form, '_save_draft'):
                        self.emp_form._save_draft(force=True)
                        logging.info("Draft auto-saved before force close")
                except Exception as e:
                    logging.warning(f"Could not auto-save draft: {e}")
            
            # Clean up session
            if hasattr(self, 'session_timer'):
                self.session_timer.stop()
            if hasattr(self, 'force_close_timer'):
                self.force_close_timer.stop()
            if hasattr(self, 'notification_center'):
                self.notification_center.stop_monitoring()
            
            # Remove session from database
            self.db.remove_session(self.current_user)
            
            # Close application
            QApplication.instance().quit()
        except Exception as e:
            logging.error(f"Error during force quit: {e}")
            QApplication.instance().quit()

    def _focus_search(self):
        """Focus search box on employees page"""
        if self.stack.currentWidget() == self.employees_page:
            self.employees_page.search.setFocus()
            self.employees_page.search.selectAll()
        else:
            # Switch to employees page and focus search
            self._show_page(self.employees_page)
            QTimer.singleShot(100, lambda: self.employees_page.search.setFocus())

    def _edit_selected_employee(self):
        """Edit currently selected employee"""
        if self.stack.currentWidget() == self.employees_page:
            selection = self.employees_page.table.selectionModel().selectedRows()
            if selection:
                idx = self.employees_page.proxy.mapToSource(selection[0])
                emp = self.employees_page.model.employees[idx.row()]
                self._edit_employee(emp)

    def _show_print_menu(self):
        """Show print/export options menu"""
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: rgba(45, 45, 48, 0.95);
                border: 1px solid rgba(255, 255, 255, 0.1);
                border-radius: 8px;
                padding: 8px;
            }
            QMenu::item {
                padding: 8px 32px 8px 16px;
                border-radius: 4px;
            }
            QMenu::item:selected {
                background-color: rgba(33, 150, 243, 0.3);
            }
        """)

        # FEATURE: Recent Actions submenu
        recent_menu = QMenu("🕐 Recent Actions", self)
        recent_menu.setStyleSheet(menu.styleSheet())

        # Get last 5 modified employees
        recent_employees = sorted(self.employees, key=lambda e: e.get('modified', ''), reverse=True)[:5]

        if recent_employees and recent_employees[0].get('modified'):
            for emp in recent_employees:
                name = emp.get('name', 'Unknown')
                emp_id = emp.get('emp_id', '')
                modified = emp.get('modified', '')[:16]  # Just date and time
                action_text = f"{name} ({emp_id}) - {modified}"

                # Create action with employee data
                action = recent_menu.addAction(action_text)
                action.setData(emp)  # Store employee data
                action.triggered.connect(lambda checked=False, e=emp: self._edit_employee(e))
        else:
            no_recent = recent_menu.addAction("No recent activity")
            no_recent.setEnabled(False)

        menu.addMenu(recent_menu)
        menu.addSeparator()

        menu.addAction("🖨️ Print System", self._show_print_system)
        menu.addAction("🆔 ID Card Generator", self._show_id_card_generator)
        menu.addSeparator()
        menu.addAction("📥 Import from Excel/CSV", self._show_import_dialog)
        menu.addAction("📦 Batch Photo Upload", self._show_batch_photo_upload)
        menu.addAction("🔧 Quick Fix Actions", self._show_quick_fix_actions)
        menu.addSeparator()
        menu.addAction("📊 Generate Reports", self._show_reports)
        menu.addAction("📄 Export Data", self._export_data)

        # Show menu at cursor position
        menu.exec(QCursor.pos())

    def _show_page(self, w: QWidget):
        """Show a page with smooth transition, checking for unsaved changes if navigating away from form"""
        # Check if currently on employee form and navigating away
        if self.stack.currentWidget() == self.form_scroll_area and w != self.form_scroll_area:
            if not self.emp_form._check_unsaved_changes():
                return  # User cancelled navigation
        # Get widget index and use animated transition
        index = self.stack.indexOf(w)
        if index >= 0:
            self.stack.animated_set_current_index(index)
    def _show_add_new(self): self.emp_form.edit_employee(None); self._show_page(self.form_scroll_area)

    def _view_employee(self, emp):
        # v4.4.1: Animated dialog for employee details
        from employee_vault.ui.widgets import SmoothAnimatedDialog
        dlg=SmoothAnimatedDialog(self, animation_style="slide"); dlg.setWindowTitle(f"Employee Details — {emp.get('name','?')}"); dlg.resize(720,600)

        # Main dialog layout
        main_layout = QVBoxLayout(dlg)

        # Create scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        # Content widget inside scroll area
        content_widget = QWidget()
        v = QVBoxLayout(content_widget)

        # Header with photo - v5.2: Use new folder structure first, then legacy
        header=QHBoxLayout()
        emp_id = emp.get('emp_id', '')
        photos = get_employee_photos(emp_id)
        photo = photos[0] if photos else None
        # Fallback to legacy location
        if not photo:
            legacy_path = os.path.join(PHOTOS_DIR, f"{emp_id}.png")
            if os.path.exists(legacy_path):
                photo = legacy_path
        
        img=QLabel()
        if photo and os.path.exists(photo):
            img.setPixmap(QPixmap(photo).scaled(60, 60, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            img.setText("📷")
        header.addWidget(img); info=QVBoxLayout(); status="🟢 Active" if not emp.get("resign_date") else "🔴 Resigned"; info.addWidget(QLabel(f"<b>{emp.get('name','?')}</b> — {status}")); info.addWidget(QLabel(f"{emp.get('emp_id','?')} • {emp.get('position','N/A')}")); header.addLayout(info,1); v.addLayout(header)

        # Employee fields
        fields=[("👤 Full Name", emp.get('name','?')),
                ("🆔 Employee ID", emp.get('emp_id','?')),
                ("🔢 SSS #", emp.get('sss_number','—') or "—"),
                ("💳 TIN", emp.get('tin_number','—') or "—"),
                ("🏦 Pag-IBIG #", emp.get('pagibig_number','—') or "—"),
                ("🏥 PhilHealth #", emp.get('philhealth_number','—') or "—"),
                ("📧 Email", emp.get('email','N/A')),
                ("📱 Phone", emp.get('phone','N/A')),
                ("🏢 Department", emp.get('department','N/A')),
                ("🏢 Agency", emp.get('agency','—') or "—"),
                ("💼 Position", emp.get('position','N/A')),
                ("📅 Hire Date", emp.get('hire_date','N/A')),
                ("📅 Resign Date", emp.get('resign_date','—') or "—"),
                ("💰 Salary/Day", f"{emp.get('salary',0):,.2f}")]

        # Add contract information if available
        if emp.get('contract_start_date') or emp.get('contract_expiry'):
            fields.append(("📄 Contract Start", emp.get('contract_start_date','—') or "—"))
            if emp.get('contract_months'):
                months = emp.get('contract_months')
                if months >= 12:
                    years = months // 12
                    remaining = months % 12
                    duration = f"{years}y {remaining}m" if remaining else f"{years} year(s)"
                else:
                    duration = f"{months} month(s)"
                fields.append(("⏱️ Contract Duration", duration))
            fields.append(("📄 Contract Expiry", emp.get('contract_expiry','—') or "—"))

        for lab,val in fields:
            r=QHBoxLayout(); l=QLabel(f"<b>{lab}</b>"); l.setFixedWidth(170); r.addWidget(l); r.addWidget(QLabel(str(val)),1); v.addLayout(r)

        # Emergency contact section
        if emp.get('emergency_contact_name') or emp.get('emergency_contact_phone'):
            v.addWidget(QLabel("<b style='color:#ff9966;'>🚨 Emergency Contact</b>"))
            emergency_fields = []
            if emp.get('emergency_contact_name'):
                emergency_fields.append(("👤 Name", emp.get('emergency_contact_name')))
            if emp.get('emergency_contact_phone'):
                emergency_fields.append(("📱 Phone", emp.get('emergency_contact_phone')))
            for lab, val in emergency_fields:
                r=QHBoxLayout(); l=QLabel(f"<b>{lab}</b>"); l.setFixedWidth(170); r.addWidget(l); r.addWidget(QLabel(str(val)),1); v.addLayout(r)

        # Contract status
        d=contract_days_left(emp)
        if d is not None:
            if d<0: v.addWidget(QLabel(f"<b style='color:#ff6b6b'>Contract expired {-d} day(s) ago</b>"))
            elif d==0: v.addWidget(QLabel("<b style='color:#ff6b6b'>Contract expires today</b>"))
            elif d<=ALERT_DAYS: v.addWidget(QLabel(f"<b style='color:#ffcc66'>Contract expires in {d} day(s)</b>"))
            else: v.addWidget(QLabel(f"<b style='color:#9ad17a'>Contract valid ({d} day(s) left)</b>"))

        # Notes section
        if emp.get('notes'):
            v.addWidget(QLabel("<b>📝 Notes</b>"))
            notes_label = QLabel(emp.get('notes', ''))
            notes_label.setWordWrap(True)
            notes_label.setStyleSheet("padding: 10px; background: #1c1c1c; border-radius: 8px;")
            v.addWidget(notes_label)

        # Attached files - IMPROVED
        v.addWidget(QLabel("<b>📎 Attached Files</b>"))
        folder=os.path.join(FILES_DIR, emp.get('emp_id',''))
        if os.path.exists(folder):
            files=os.listdir(folder)
            if files:
                files_container = QWidget()
                files_container.setStyleSheet("background: #1c1c1c; border-radius: 8px; padding: 10px;")
                files_layout = QVBoxLayout(files_container)
                files_layout.setContentsMargins(5, 5, 5, 5)
                files_layout.setSpacing(5)

                for f in files:
                    pth=os.path.join(folder,f)
                    btn=ModernAnimatedButton(f"📄 {f}")
                    btn.setFlat(True)
                    btn.setStyleSheet("""
                        QPushButton {
                            text-align: left;
                            padding: 8px;
                            background: #2a2a2a;
                            border-radius: 4px;
                        }
                        QPushButton:hover {
                            background: #3a3a3a;
                        }
                    """)
                    from PySide6.QtCore import QUrl
                    btn.clicked.connect(lambda _=False, pp=pth: QDesktopServices.openUrl(QUrl.fromLocalFile(pp)))
                    files_layout.addWidget(btn)

                v.addWidget(files_container)
            else:
                v.addWidget(QLabel("<i style='color:#888;'>No files attached</i>"))
        else:
            v.addWidget(QLabel("<i style='color:#888;'>No files attached</i>"))

        # Add stretch to push content to top
        v.addStretch()

        # Set content widget to scroll area
        scroll.setWidget(content_widget)

        # Add scroll area to main layout
        main_layout.addWidget(scroll)

        # Close button outside scroll area (always visible)
        button_layout = QHBoxLayout()

        # Add Edit button - allows user to quickly edit this employee
        edit_btn = ModernAnimatedButton("✏️ Edit")
        apply_ios_style(edit_btn, 'blue')
        def open_edit_and_close():
            dlg.close()
            self._edit_employee(emp)
        edit_btn.clicked.connect(open_edit_and_close)
        button_layout.addWidget(edit_btn)

        # Add View History button for admins
        if self.user_row.get('role') == 'admin':
            history_btn = ModernAnimatedButton("📜 View History")
            apply_ios_style(history_btn, 'purple')
            history_btn.clicked.connect(lambda: self._show_employee_history(emp))
            button_layout.addWidget(history_btn)

        button_layout.addStretch()

        close_btn = ModernAnimatedButton("Close (ESC)")
        apply_ios_style(close_btn, 'gray')
        close_btn.clicked.connect(dlg.close)
        button_layout.addWidget(close_btn)

        main_layout.addLayout(button_layout)

        dlg.exec()

    def _show_employee_history(self, emp):
        """Show audit history for a specific employee"""
        # v4.4.1: Animated dialog for employee history
        from employee_vault.ui.widgets import AnimatedDialogBase
        dlg = AnimatedDialogBase(self, animation_style="fade")
        dlg.setWindowTitle(f"📜 History - {emp.get('name', '?')}")
        dlg.resize(900, 500)

        layout = QVBoxLayout(dlg)
        layout.addWidget(QLabel(f"<h2>📜 Activity History for {emp.get('name', '?')}</h2>"))
        layout.addWidget(QLabel(f"Employee ID: <b>{emp.get('emp_id', '?')}</b>"))

        # Table for history
        from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QHeaderView
        table = QTableWidget()
        disable_cursor_changes(table)  # Remove hand cursor
        table.setColumnCount(7)  # Added row number column
        table.setHorizontalHeaderLabels(["#", "Timestamp", "User", "Action", "Old Value", "New Value", "Details"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        table.horizontalHeader().setStretchLastSection(True)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)

        # Load history
        history = self.db.get_employee_history(emp.get('emp_id'))

        table.setRowCount(len(history))
        for i, log in enumerate(history):
            # Row number (1-indexed)
            row_num_item = QTableWidgetItem(str(i + 1))
            row_num_item.setTextAlignment(Qt.AlignCenter)
            row_num_item.setForeground(QColor('#888888'))
            table.setItem(i, 0, row_num_item)
            
            table.setItem(i, 1, QTableWidgetItem(log.get('timestamp', '')))
            table.setItem(i, 2, QTableWidgetItem(log.get('username', '')))

            # Color code actions
            action_item = QTableWidgetItem(log.get('action', ''))
            action_text = log.get('action', '')
            if action_text == 'INSERT' or action_text == 'ADDED':
                action_item.setForeground(QColor('#2196F3'))  # Blue
            elif action_text == 'UPDATE' or action_text == 'EDITED':
                action_item.setForeground(QColor('#FF9800'))  # Orange
            elif action_text == 'DELETE' or action_text == 'DELETED':
                action_item.setForeground(QColor('#F44336'))  # Red
            table.setItem(i, 3, action_item)

            table.setItem(i, 4, QTableWidgetItem(log.get('old_value', '') or ''))
            table.setItem(i, 5, QTableWidgetItem(log.get('new_value', '') or ''))

            details_text = log.get('details', '') or ''
            details_item = QTableWidgetItem(details_text[:100] + ('...' if len(details_text) > 100 else ''))
            details_item.setToolTip(details_text)
            table.setItem(i, 6, details_item)

        table.resizeColumnsToContents()
        layout.addWidget(table)

        # Stats
        layout.addWidget(QLabel(f"Total changes: <b>{len(history)}</b>"))

        # Close button
        close_btn = ModernAnimatedButton("Close")
        close_btn.clicked.connect(dlg.close)
        layout.addWidget(close_btn)

        dlg.exec()

    def _edit_employee(self, emp): self.emp_form.edit_employee(emp); self._show_page(self.form_scroll_area)

    def _handle_dashboard_navigate(self, filter_type: str):
        """Handle KPI card click - navigate to employees with filter"""
        # Switch to employees page
        self._show_page(self.employees_page)

        # Apply filter based on card type
        if filter_type == "total":
            # Show all employees - clear any filters
            self.employees_page.set_data(self.employees)
        elif filter_type == "active":
            # Filter to active employees only
            active_employees = [e for e in self.employees if not e.get('resign_date')]
            self.employees_page.set_data(active_employees)
        elif filter_type == "expiring":
            # Filter to employees with contracts expiring within 30 days
            from datetime import datetime, timedelta
            today = datetime.now().date()
            expiring_employees = []
            for emp in self.employees:
                if emp.get('contract_expiry'):
                    try:
                        expiry_date = datetime.strptime(emp['contract_expiry'], "%m-%d-%Y").date()
                        days = (expiry_date - today).days
                        if 0 <= days <= 30:
                            expiring_employees.append(emp)
                    except ValueError:
                        # Invalid date format, skip this employee
                        pass
            self.employees_page.set_data(expiring_employees)

    def _handle_dashboard_open_employee(self, emp_id: str):
        """Handle recent activity item click - open employee detail"""
        # Find employee by ID
        employee = next((e for e in self.employees if e.get('emp_id') == emp_id), None)
        if employee:
            self._view_employee(employee)
        else:
            from employee_vault.ui.modern_ui_helper import show_error_toast
            show_error_toast(self, f"Employee {emp_id} not found")

    def _handle_dashboard_edit_employee(self, emp_id: str):
        """Handle missing photos item click - open employee edit dialog"""
        # Find employee by ID
        employee = next((e for e in self.employees if e.get('emp_id') == emp_id), None)
        if employee:
            self._edit_employee(employee)
        else:
            from employee_vault.ui.modern_ui_helper import show_error_toast
            show_error_toast(self, f"Employee {emp_id} not found")

    def _delete_selected(self, rows):
        # Only admins can delete employees
        if not self.user_permissions.get('delete_employee'):
            show_warning_toast(
                self, "Only administrators can delete employees.\n\n"
                "Please contact your administrator if you need to delete employee records."
            )
            return

        if not rows:
            return

        preview = "<br>".join(f"{r['emp_id']} — {r['name']}" for r in rows[:10])
        if len(rows) > 10:
            preview += "<br>…"

        # Ask for archive reason
        # Build the prompt outside of the f-string to avoid backslash
        # confusion within the f-string expression.  Newlines are
        # escaped here so that Python does not interpret them inside an
        # f-string expression.  Using join() would also work but this
        # approach is clearer.
        prompt_lines = [
            f"Archive {len(rows)} employee(s)?",
            "",
            preview.replace('<br>', '\n'),
            "",
            "These employees will be moved to the archive and can be restored later.",
            "",
            "Reason for archiving (optional):"
        ]
        prompt_text = "\n".join(prompt_lines)
        reason, ok = QInputDialog.getText(self, "Archive Reason", prompt_text)

        if not ok:
            return

        # Archive employees instead of deleting
        for row in rows:
            emp_id = row["emp_id"]
            self.db.archive_employee(emp_id, self.current_user, reason or "No reason specified")

            # Move file folders to archive
            src_folder = os.path.join(FILES_DIR, emp_id)
            if os.path.isdir(src_folder):
                archive_folder = os.path.join(FILES_DIR, "_archived", emp_id)
                os.makedirs(os.path.dirname(archive_folder), exist_ok=True)
                try:
                    shutil.move(src_folder, archive_folder)
                except (OSError, shutil.Error):
                    # File may already be moved or locked
                    pass

            # Move photos to archive
            src_photo = os.path.join(PHOTOS_DIR, f"{emp_id}.png")
            if os.path.exists(src_photo):
                archive_photo = os.path.join(PHOTOS_DIR, "_archived", f"{emp_id}.png")
                os.makedirs(os.path.dirname(archive_photo), exist_ok=True)
                try:
                    shutil.move(src_photo, archive_photo)
                except (OSError, shutil.Error):
                    # Photo may already be moved or locked
                    pass

        show_success_toast(
            self, f"{len(rows)} employee(s) have been archived.\n\n"
            "You can restore them from the Archive Manager."
        )

        self._refresh_all()

    def _on_form_saved(self, cancel_only=False, switch_page=True):
        if not cancel_only:
            self._refresh_all()
            # v4.5.0: Show celebration effect on employee save
            if ANIMATIONS_AVAILABLE and create_celebration_effect:
                try:
                    create_celebration_effect(self, "success")
                except Exception as e:
                    logging.debug(f"Animation effect skipped: {e}")
            # After save, go to employees page
            if switch_page: self._show_page(self.employees_page)
        else:
            # On cancel/ESC, go to Dashboard instead of employees page
            if switch_page: self._show_page(self.dashboard)

    def _refresh_all(self):
        """Refresh all data - checkpoint WAL first to get changes from all PCs"""
        try:
            # Step 1: Checkpoint to merge any pending WAL changes from all users
            self.db.checkpoint_database()
            logging.info("Checkpoint completed before refresh")
        except Exception as e:
            logging.warning(f"Checkpoint before refresh failed: {e}")
        
        # Step 2: Reload fresh data from the merged database
        self.employees = self.db.all_employees()
        self.dashboard.refresh(self.employees)
        self.employees_page.set_data(self.employees)
        logging.info(f"Data refreshed: {len(self.employees)} employees loaded")

    def _show_enhanced_search(self):
        """Show enhanced search dialog"""
        # v4.4.1: Animated dialog for enhanced search
        from employee_vault.ui.widgets import AnimatedDialogBase
        dialog = AnimatedDialogBase(self, animation_style="fade")
        dialog.setWindowTitle("🔍 Enhanced Search")
        dialog.setMinimumWidth(500)
        layout = QFormLayout(dialog)

        # Search criteria
        name_input = NeumorphicGradientLineEdit("Employee name")
        name_input.setMinimumHeight(70)
        dept_combo = NeumorphicGradientComboBox("Select Department")
        dept_combo.setMinimumHeight(70)
        dept_combo.addItems(["", "Office", "Store - Sm Novaliches", "Store - Sm San Fernando"])
        position_input = NeumorphicGradientLineEdit("Position/Role")
        position_input.setMinimumHeight(70)
        status_combo = NeumorphicGradientComboBox("Select Status")
        status_combo.setMinimumHeight(70)
        status_combo.addItems(["All", "Active", "Resigned"])

        layout.addRow("Name:", name_input)
        layout.addRow("Department:", dept_combo)
        layout.addRow("Position:", position_input)
        layout.addRow("Status:", status_combo)

        # Buttons
        btn_layout = QHBoxLayout()
        search_btn = ModernAnimatedButton("Search")
        clear_btn = ModernAnimatedButton("Clear")
        close_btn = ModernAnimatedButton("Close")

        def do_search():
            # Filter employees based on criteria
            results = self.employees.copy()

            if name_input.line_edit.text():
                results = [e for e in results if name_input.line_edit.text().lower() in e['name'].lower()]
            if dept_combo.combo_box.currentText():
                results = [e for e in results if e.get('department') == dept_combo.combo_box.currentText()]
            if position_input.line_edit.text():
                results = [e for e in results if position_input.line_edit.text().lower() in e.get('position', '').lower()]
            if status_combo.combo_box.currentText() == "Active":
                results = [e for e in results if not e.get('resign_date')]
            elif status_combo.combo_box.currentText() == "Resigned":
                results = [e for e in results if e.get('resign_date')]

            # Update table
            self.employees_page.refresh(results)
            dialog.accept()
            self._show_page(self.employees_page)

        search_btn.clicked.connect(do_search)
        clear_btn.clicked.connect(lambda: [name_input.clear(), position_input.clear()])
        close_btn.clicked.connect(dialog.reject)

        btn_layout.addWidget(search_btn)
        btn_layout.addWidget(clear_btn)
        btn_layout.addWidget(close_btn)
        layout.addRow(btn_layout)

        dialog.exec()


    def _show_id_card_generator(self):
        """Open ID card generator"""
        dlg = IDCardGeneratorBackenderator(self, self.db)
        dlg.exec()

    def _show_batch_photo_upload(self):
        """Open batch photo upload from folder or ZIP"""
        from employee_vault.ui.dialogs.batch_photo_upload import BatchPhotoUploadDialog
        dlg = BatchPhotoUploadDialog(self.db, self)
        dlg.exec()
        # Refresh to show updated photos
        self._refresh_all()

    def _show_import_dialog(self):
        """Open Excel/CSV import dialog"""
        from employee_vault.ui.dialogs.excel_import import ExcelImportDialog
        dlg = ExcelImportDialog(self.db, self)
        if dlg.exec() == QDialog.Accepted:
            # Refresh to show imported employees
            self._refresh_all()
            show_success_toast(self, "Employees imported successfully!")

    def _show_quick_fix_actions(self):
        """Open quick fix actions dialog"""
        from employee_vault.ui.dialogs.quick_fix_actions import QuickFixActionsDialog
        dlg = QuickFixActionsDialog(self.db, self)
        dlg.exec()
        # Refresh in case any bulk updates were made
        self._refresh_all()

    def _show_print_system(self):
        """Open print system dialog"""
        dlg = PrintSystemDialog(self, self.db, self.employees)
        dlg.exec()

    def _show_bulk_operations(self):
        """Open bulk operations dialog"""
        dlg = BulkOperationsDialog(self, self.db, self.employees)
        if dlg.exec() == QDialog.Accepted:
            self._refresh_all()


    def _export_data(self):
        """Export employee data to JSON, Excel, or PDF"""
        formats = "PDF Report (*.pdf);;Excel (*.xlsx);;JSON (*.json)"
        path, selected_filter = QFileDialog.getSaveFileName(
            self, "Export Employee Data", "employees_export.pdf", formats
        )
        if not path:
            return

        try:
            if path.endswith('.pdf'):
                self._export_to_pdf(path)
            elif path.endswith('.xlsx'):
                self._export_to_excel(path)
            else:
                # JSON export
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(self.employees, f, indent=2)
            show_success_toast(self, f"Data exported successfully to:\n{path}")
        except Exception as e:
            logging.error(f"Export error: {e}")
            show_error_toast(self, f"Failed to export data:\n{str(e)}")

    def _export_to_excel(self, filepath):
        """Export employee data to Excel with formatting and progress dialog"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            show_warning_toast(
                self, "Excel export requires 'openpyxl' library.\n\n"
                "Install it with:\npip install openpyxl"
            )
            return

        # Create progress dialog
        progress = QProgressDialog("Exporting to Excel...", "Cancel", 0, len(self.employees) + 2, self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setWindowTitle("Exporting")
        progress.setMinimumDuration(0)
        progress.setValue(0)

        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        # Headers
        headers = ["Employee ID", "Name", "SSS #", "Email", "Phone", "Department",
                  "Position", "Agency", "Hire Date", "Resign Date", "Salary/Day",
                  "Contract Start", "Contract Duration", "Contract Expiry", "Status",
                  "Emergency Contact", "Emergency Phone"]

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        progress.setValue(1)
        progress.setLabelText("Writing employee data...")
        QApplication.processEvents()

        # Data rows
        for row_num, emp in enumerate(self.employees, 2):
            if progress.wasCanceled():
                return
            
            progress.setValue(row_num)
            QApplication.processEvents()
            
            status = "Active" if not emp.get('resign_date') else "Resigned"

            # Calculate contract duration
            contract_duration = ""
            if emp.get('contract_months'):
                months = emp.get('contract_months')
                if months >= 12:
                    years = months // 12
                    remaining = months % 12
                    contract_duration = f"{years}y {remaining}m" if remaining else f"{years} year(s)"
                else:
                    contract_duration = f"{months} month(s)"

            row_data = [
                emp.get('emp_id', ''),
                emp.get('name', ''),
                emp.get('sss_number', ''),
                emp.get('email', ''),
                emp.get('phone', ''),
                emp.get('department', ''),
                emp.get('position', ''),
                emp.get('agency', ''),
                emp.get('hire_date', ''),
                emp.get('resign_date', ''),
                emp.get('salary', 0),
                emp.get('contract_start_date', ''),
                contract_duration,
                emp.get('contract_expiry', ''),
                status,
                emp.get('emergency_contact_name', ''),
                emp.get('emergency_contact_phone', '')
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                if col_num == 11:  # Salary column
                    cell.number_format = '#,##0.00'
                # Color-code status
                if col_num == 15:  # Status column
                    if status == "Active":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        progress.setLabelText("Saving file...")
        progress.setValue(len(self.employees) + 1)
        QApplication.processEvents()

        wb.save(filepath)
        
        progress.setValue(len(self.employees) + 2)
        logging.info(f"Exported {len(self.employees)} employees to Excel: {filepath}")

    def _export_to_pdf(self, filepath):
        """Export employee data to professional PDF report with progress dialog"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            from datetime import datetime
        except ImportError:
            show_warning_toast(
                self, "PDF export requires 'reportlab' library.\n\n"
                "Install it with:\npip install reportlab"
            )
            return

        # Create progress dialog
        progress = QProgressDialog("Generating PDF report...", "Cancel", 0, 100, self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setWindowTitle("Exporting PDF")
        progress.setMinimumDuration(0)
        progress.setValue(0)

        # Create PDF with landscape orientation for better table fit
        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                               rightMargin=0.5*inch, leftMargin=0.5*inch,
                               topMargin=0.75*inch, bottomMargin=0.5*inch)

        progress.setValue(10)
        progress.setLabelText("Setting up document styles...")
        QApplication.processEvents()

        # Container for PDF elements
        elements = []
        styles = getSampleStyleSheet()

        # Custom title style
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2196F3'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        # Subtitle style
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.gray,
            spaceAfter=20,
            alignment=TA_CENTER
        )

        # Add title
        title = Paragraph("Employee Directory Report", title_style)
        elements.append(title)

        # Add generation info
        now = datetime.now().strftime("%B %d, %Y at %I:%M %p")
        subtitle = Paragraph(f"Generated on {now} | Total Employees: {len(self.employees)}", subtitle_style)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.2*inch))

        # Prepare table data
        data = [['Emp ID', 'Name', 'Department', 'Position', 'Hire Date', 'Status', 'Contact']]

        progress.setValue(30)
        progress.setLabelText("Processing employee data...")
        QApplication.processEvents()

        for idx, emp in enumerate(self.employees):
            if progress.wasCanceled():
                return
            
            # Update progress every 10 employees
            if idx % 10 == 0:
                progress.setValue(30 + int((idx / len(self.employees)) * 40))
                QApplication.processEvents()
            
            status = "Active" if not emp.get('resign_date') else f"Resigned\n{emp.get('resign_date', '')}"

            row = [
                emp.get('emp_id', '')[:10],
                emp.get('name', '')[:25],
                emp.get('department', '')[:15],
                emp.get('position', '')[:20],
                emp.get('hire_date', ''),
                status,
                emp.get('phone', '')[:15] or emp.get('email', '')[:20]
            ]
            data.append(row)

        # Create table with proper column widths
        col_widths = [0.9*inch, 2*inch, 1.3*inch, 1.5*inch, 1*inch, 1.2*inch, 1.5*inch]
        table = Table(data, colWidths=col_widths, repeatRows=1)

        # Table styling
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),

            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2196F3')),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),

            # Status column color coding
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        # Add color coding for resigned employees
        for i, emp in enumerate(self.employees, 1):
            if emp.get('resign_date'):
                table.setStyle(TableStyle([
                    ('TEXTCOLOR', (5, i), (5, i), colors.HexColor('#F44336')),
                    ('FONTNAME', (5, i), (5, i), 'Helvetica-Bold'),
                ]))

        elements.append(table)

        progress.setValue(80)
        progress.setLabelText("Adding summary...")
        QApplication.processEvents()

        # Add footer with stats
        elements.append(Spacer(1, 0.3*inch))

        active_count = sum(1 for e in self.employees if not e.get('resign_date'))
        resigned_count = len(self.employees) - active_count

        footer_text = f"Summary: {active_count} Active Employees | {resigned_count} Resigned"
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        footer = Paragraph(footer_text, footer_style)
        elements.append(footer)

        progress.setValue(90)
        progress.setLabelText("Saving PDF file...")
        QApplication.processEvents()

        # Build PDF
        doc.build(elements)
        
        progress.setValue(100)
        logging.info(f"Exported {len(self.employees)} employees to PDF: {filepath}")

    def _backup_data(self):
        """Manual backup with timestamp"""
        folder = QFileDialog.getExistingDirectory(self, "Choose Backup Folder")
        if not folder:
            return

        try:
            # Create timestamped backup folder
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_folder = os.path.join(folder, f"employee_vault_backup_{timestamp}")
            os.makedirs(backup_folder, exist_ok=True)

            # Copy database files
            files_copied = 0
            for fn in [DB_FILE, DB_FILE + "-wal", DB_FILE + "-shm"]:
                if os.path.exists(fn):
                    dest = os.path.join(backup_folder, os.path.basename(fn))
                    shutil.copy2(fn, dest)
                    files_copied += 1

            # v5.2: Backup employee files (new structure includes photos in subfolders)
            if os.path.exists(FILES_DIR):
                shutil.copytree(FILES_DIR, os.path.join(backup_folder, "employee_files"))
            
            # Also backup legacy photos folder if it exists (for backwards compatibility)
            if os.path.exists(PHOTOS_DIR):
                shutil.copytree(PHOTOS_DIR, os.path.join(backup_folder, "employee_photos"))

            # Create backup info file
            info = {
                "backup_date": datetime.now().isoformat(),
                "total_employees": len(self.employees),
                "database_files": files_copied,
                "backed_up_by": self.current_user,
                "structure_version": "5.2"  # New folder structure version
            }
            with open(os.path.join(backup_folder, "backup_info.json"), "w") as f:
                json.dump(info, f, indent=2)

            logging.info(f"Backup created: {backup_folder}")
            show_success_toast(
                self, f"Backup created successfully!\n\n"
                f"Location: {backup_folder}\n"
                f"Files backed up: {files_copied} database files + attachments"
            )
        except Exception as e:
            logging.error(f"Backup error: {e}")
            show_error_toast(self, f"Failed to create backup:\n{str(e)}")

    def _setup_auto_backup(self):
        """Setup automated daily backup"""
        # This would be called from settings in a full implementation
        pass

    def _show_network_config(self):
        """Show enhanced network configuration"""
        from employee_vault.ui.dialogs.network_config import NetworkConfigDialog
        dialog = NetworkConfigDialog(self)
        dialog.exec()

    def _test_network_paths(self, db_path, files_path, photos_path):
        """Test if network paths are accessible"""
        results = []

        # Test database
        if os.path.exists(db_path) and os.access(db_path, os.R_OK | os.W_OK):
            results.append("✅ Database path is accessible")
        else:
            results.append("❌ Database path is not accessible")

        # Test files directory
        if os.path.exists(files_path) and os.access(files_path, os.R_OK | os.W_OK):
            results.append("✅ Files directory is accessible")
        else:
            results.append("❌ Files directory is not accessible")

        # Test photos directory
        if os.path.exists(photos_path) and os.access(photos_path, os.R_OK | os.W_OK):
            results.append("✅ Photos directory is accessible")
        else:
            results.append("❌ Photos directory is not accessible")

        show_success_toast(self, "\n".join(results))


    def _show_scheduled_backup(self):
        """Show scheduled backup configuration dialog"""
        # v4.4.1: Animated dialog for scheduled backup
        from employee_vault.ui.widgets import AnimatedDialogBase
        dlg = AnimatedDialogBase(self, animation_style="fade")
        dlg.setWindowTitle("⏰ Scheduled Backup")
        dlg.resize(650, 600)

        wheel_guard = WheelGuard(dlg)

        try:
            config_str = self.db.get_setting('backup_config', '{}')
            config = json.loads(config_str)
        except Exception:
            config = {}  # Use defaults if loading fails

        # This is the MAIN layout for the dialog (non-scrolling)
        layout = QVBoxLayout(dlg)
        layout.addWidget(QLabel("<h2>⏰ Scheduled Backup Configuration</h2>"))
        layout.addWidget(QLabel("Automatically backup your data on a schedule"))

        # --- Scroll Area Setup (as before) ---
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_content_widget)


        # --- Enable/Disable ---
        enable_box = QGroupBox("Status")
        enable_layout = QVBoxLayout(enable_box)

        enable_check = QCheckBox("Enable Scheduled Backups")
        # Load saved value
        enable_check.setChecked(config.get('enabled', False))
        enable_layout.addWidget(enable_check)

        status_label = QLabel("Status: ❌ Disabled")
        status_label.setStyleSheet("color: #ff6b6b; padding: 5px;")
        enable_layout.addWidget(status_label)

        def toggle_status():
            if enable_check.isChecked():
                status_label.setText("Status: ✅ Enabled")
                status_label.setStyleSheet("color: #9ad17a; padding: 5px;")
            else:
                status_label.setText("Status: ❌ Disabled")
                status_label.setStyleSheet("color: #ff6b6b; padding: 5px;")

        enable_check.stateChanged.connect(toggle_status)
        toggle_status() # Set initial text
        scroll_layout.addWidget(enable_box)

        # --- Schedule settings ---
        schedule_box = QGroupBox("Schedule")
        schedule_layout = QFormLayout(schedule_box)

        frequency_combo = NeumorphicGradientComboBox("Select Frequency")
        frequency_combo.setMinimumHeight(70)
        frequency_combo.combo_box.setFocusPolicy(Qt.ClickFocus)
        frequency_combo.combo_box.installEventFilter(wheel_guard)
        frequency_combo.addItems(["Daily", "Weekly", "Monthly"])
        # Load saved value
        frequency_combo.combo_box.setCurrentText(config.get('frequency', 'Daily'))
        schedule_layout.addRow("Frequency:", frequency_combo)

        # Time selection with dropdowns
        time_layout = QHBoxLayout()
        hour_combo = NeumorphicGradientComboBox("Hour")
        hour_combo.setMinimumHeight(70)
        hour_combo.combo_box.setFocusPolicy(Qt.ClickFocus)
        hour_combo.combo_box.installEventFilter(wheel_guard)
        hour_combo.addItems([f"{h:02d}" for h in range(24)])
        # Load saved value
        hour_combo.combo_box.setCurrentText(config.get('hour', '18'))

        minute_combo = NeumorphicGradientComboBox("Minute")
        minute_combo.setMinimumHeight(70)
        minute_combo.combo_box.setFocusPolicy(Qt.ClickFocus)
        minute_combo.combo_box.installEventFilter(wheel_guard)
        minute_combo.addItems([f"{m:02d}" for m in range(0, 60, 5)])
        # Load saved value
        minute_combo.combo_box.setCurrentText(config.get('minute', '00'))

        time_layout.addWidget(QLabel("Hour:"))
        time_layout.addWidget(hour_combo)
        time_layout.addWidget(QLabel("Minute:"))
        time_layout.addWidget(minute_combo)
        time_layout.addStretch()
        schedule_layout.addRow("Backup Time:", time_layout)

        day_combo = NeumorphicGradientComboBox("Select Day")
        day_combo.setMinimumHeight(70)
        day_combo.combo_box.setFocusPolicy(Qt.ClickFocus)
        day_combo.combo_box.installEventFilter(wheel_guard)
        day_combo.addItems(["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"])
        # Load saved value
        day_combo.combo_box.setCurrentText(config.get('day_of_week', 'Monday'))

        schedule_layout.addRow("Day of Week:", day_combo)

        def update_day_combo():
            day_combo.setEnabled(frequency_combo.currentText() == "Weekly")

        frequency_combo.currentTextChanged.connect(update_day_combo)
        update_day_combo() # Set initial state

        scroll_layout.addWidget(schedule_box)

        # --- Backup location ---
        location_box = QGroupBox("Backup Location")
        location_layout = QVBoxLayout(location_box)

        location_path = QLineEdit()
        # Load saved value
        location_path.setText(config.get('location', ''))
        location_path.setPlaceholderText("\\\\SERVER\\Backups\\EmployeeVault")
        # Phase 3: iOS frosted glass styling
        location_browse = ModernAnimatedButton("Browse...")
        apply_ios_style(location_browse, 'blue')

        def browse_backup_location():
            folder = QFileDialog.getExistingDirectory(dlg, "Select Backup Location")
            if folder:
                # Normalize path to use forward slashes for consistency
                location_path.setText(folder.replace("\\", "/"))

        location_browse.clicked.connect(browse_backup_location)

        loc_layout = QHBoxLayout()
        loc_layout.addWidget(location_path)
        loc_layout.addWidget(location_browse)
        location_layout.addLayout(loc_layout)

        scroll_layout.addWidget(location_box)

        # --- Retention policy ---
        retention_box = QGroupBox("Retention Policy")
        retention_layout = QFormLayout(retention_box)

        keep_backups = QLineEdit()
        # Load saved value
        keep_backups.setText(config.get('keep_n', '30'))
        keep_backups.setPlaceholderText("Number of backups to keep")
        retention_layout.addRow("Keep Last N Backups:", keep_backups)

        auto_delete = QCheckBox("Auto-delete old backups")
        # Load saved value
        auto_delete.setChecked(config.get('auto_delete', True))
        retention_layout.addRow("", auto_delete)

        scroll_layout.addWidget(retention_box)

        # --- Notifications ---
        notif_box = QGroupBox("Notifications")
        notif_layout = QVBoxLayout(notif_box)

        notify_success = QCheckBox("Notify on successful backup")
        # Load saved value
        notify_success.setChecked(config.get('notify_success', True))
        notif_layout.addWidget(notify_success)

        notify_fail = QCheckBox("Notify on backup failure")
        # Load saved value
        notify_fail.setChecked(config.get('notify_fail', True))
        notif_layout.addWidget(notify_fail)

        scroll_layout.addWidget(notif_box)

        # --- End of scrolling content ---
        scroll_layout.addStretch(1)
        scroll.setWidget(scroll_content_widget)
        layout.addWidget(scroll)


        # --- NEW: Real Save Function ---
        def _save_settings():
            # Create a dictionary with all the current values
            new_config = {
                'enabled': enable_check.isChecked(),
                'frequency': frequency_combo.currentText(),
                'hour': hour_combo.currentText(),
                'minute': minute_combo.currentText(),
                'day_of_week': day_combo.currentText(),
                'location': location_path.text(),
                'keep_n': keep_backups.text(),
                'auto_delete': auto_delete.isChecked(),
                'notify_success': notify_success.isChecked(),
                'notify_fail': notify_fail.isChecked(),
            }
            try:
                # Save the dictionary to the database as a JSON string
                self.db.set_setting('backup_config', json.dumps(new_config))
                # Log this action
                self.db.log_action(self.current_user, "BACKUP_CONFIG_CHANGED", details="Scheduled backup settings updated.")
                show_success_toast(dlg, "Scheduled backup configuration saved successfully!")
                dlg.close()
            except Exception as e:
                show_error_toast(
                    dlg, f"Unable to save settings to database:\n{e}\n\n"
                    "Possible causes:\n"
                    "• Database is locked by another user\n"
                    "• Insufficient permissions\n"
                    "• Database file is read-only\n\n"
                    "Your settings changes have not been saved.\n"
                    "Please try again or contact your administrator."
                )

        # --- NEW: Real Test Backup Function ---
        def _run_test_backup():
            location = location_path.text()
            if not location:
                show_warning_toast(
                    dlg, "Please set a backup location before running a test backup.\n\n"
                    "Click the 'Browse...' button to select a folder where\n"
                    "backup files will be saved."
                )
                return

            if not os.path.exists(location):
                try:
                    os.makedirs(location, exist_ok=True)
                except Exception as e:
                    show_error_toast(
                        dlg, f"Unable to create backup directory:\n{e}\n\n"
                        "Possible causes:\n"
                        "• Path contains invalid characters\n"
                        "• Insufficient permissions\n"
                        "• Network drive is disconnected\n"
                        "• Drive is read-only or out of space\n\n"
                        "Please choose a different backup location."
                    )
                    return

            # Create a timestamped folder for this specific test
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_folder = os.path.join(location, f"TEST_BACKUP_{timestamp}")

            try:
                os.makedirs(backup_folder, exist_ok=True)

                # Fix: Use SQLite backup API instead of file copying to avoid WinError 33 (locked files)
                # This properly handles the database even when it's in use
                backup_path = self.db.backup_database(backup_folder)

                if backup_path:
                    show_success_toast(dlg, f"Test backup created successfully!\n\n"
                        f"Backup file:\n{backup_path}\n\n"
                        f"The database was safely backed up using SQLite's backup API,\n"
                        f"which avoids file locking issues.")

                    # Log this action
                    self.db.log_action(self.current_user, "BACKUP_TEST_RUN", details=f"Test backup created at {backup_path}")
                else:
                    show_warning_toast(dlg, "Backup operation completed but no backup file was created.\n"
                        "Check the logs for details.")

            except Exception as e:
                show_error_toast(dlg, f"An error occurred while creating backup folder:\n{e}")

        # --- Test backup button (Fixed at bottom) - Phase 3: iOS frosted glass ---
        test_btn = ModernAnimatedButton("🧪 Test Backup Now")
        apply_ios_style(test_btn, 'orange')
        # --- UPDATED: Connect to the new function ---
        test_btn.clicked.connect(_run_test_backup)
        layout.addWidget(test_btn)

        # --- Buttons (Fixed at bottom) - Phase 3: iOS frosted glass ---
        button_box = QHBoxLayout()
        save_btn = PulseButton("💾 Save Autobackup Settings")
        apply_ios_style(save_btn, 'green')
        save_btn.start_pulse()
        # --- UPDATED: Connect to the new function ---
        save_btn.clicked.connect(_save_settings)

        cancel_btn = ModernAnimatedButton("Cancel")
        apply_ios_style(cancel_btn, 'gray')
        cancel_btn.clicked.connect(dlg.close)

        button_box.addWidget(save_btn)
        button_box.addWidget(cancel_btn)
        layout.addLayout(button_box)

        dlg.exec()

    def _show_archive_manager(self):
        """Show Archive Manager dialog (Priority #2 - Delete Protection)"""
        # v4.4.1: Animated dialog for archive manager
        from employee_vault.ui.widgets import SmoothAnimatedDialog
        dlg = SmoothAnimatedDialog(self, animation_style="slide")
        dlg.setWindowTitle("📦 Archive Manager")
        dlg.resize(1000, 600)

        layout = QVBoxLayout(dlg)
        layout.addWidget(QLabel("<h2>📦 Archive Manager - Deleted Employees</h2>"))
        layout.addWidget(QLabel("Employees that have been deleted are stored here and can be restored."))

        # Toolbar - Phase 3: iOS frosted glass styling
        toolbar = QHBoxLayout()

        restore_btn = ModernAnimatedButton("↩️ Restore")
        apply_ios_style(restore_btn, 'green')
        restore_btn.clicked.connect(lambda: (
            self._restore_archived(table),
            self._load_archived(table, info_label, restore_btn, delete_btn)
        ))
        restore_btn.setEnabled(False)
        toolbar.addWidget(restore_btn)

        delete_btn = ModernAnimatedButton("🗑️ Permanently Delete")
        apply_ios_style(delete_btn, 'red')
        delete_btn.clicked.connect(lambda: (
            self._permanently_delete_archived(table),
            self._load_archived(table, info_label, restore_btn, delete_btn)
        ))
        delete_btn.setEnabled(False)
        toolbar.addWidget(delete_btn)

        toolbar.addStretch()



        layout.addLayout(toolbar)

        # Table
        from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QHeaderView
        table = QTableWidget()
        disable_cursor_changes(table)  # Remove hand cursor
        table.setColumnCount(9)
        table.setHorizontalHeaderLabels([
            "Employee ID", "Name", "Department", "Position",
            "Hire Date", "Archived Date", "Archived By", "Reason", "Status"
        ])
        table.horizontalHeader().setStretchLastSection(True)
        table.verticalHeader().setVisible(False)
        table.setSelectionBehavior(QTableView.SelectRows)
        table.setSelectionMode(QTableView.SingleSelection)
        table.setAlternatingRowColors(True)
        table.setEditTriggers(QTableView.NoEditTriggers)

        def on_selection():
            has_selection = len(table.selectedItems()) > 0
            restore_btn.setEnabled(has_selection)
            delete_btn.setEnabled(has_selection)

        table.itemSelectionChanged.connect(on_selection)

        layout.addWidget(table)

        # Info label
        info_label = QLabel()
        layout.addWidget(info_label)

        # Close button - Phase 3: iOS frosted glass styling
        close_btn = ModernAnimatedButton("Close")
        apply_ios_style(close_btn, 'gray')
        close_btn.clicked.connect(dlg.close)
        bottom_layout = QHBoxLayout()
        bottom_layout.addStretch()
        bottom_layout.addWidget(close_btn)
        layout.addLayout(bottom_layout)

        # Load archived employees
        self._load_archived(table, info_label, restore_btn, delete_btn)

        # Create an auto-refresh timer that is a child of the dialog
        # This ensures the timer is destroyed when the dialog is closed.
        refresh_timer = QTimer(dlg)
        refresh_timer.setInterval(5000) # 15 seconds (or your preference)

        # Connect the timer's timeout signal to the _load_archived function
        # We use a lambda to pass all the required arguments.
        refresh_timer.timeout.connect(lambda: self._load_archived(
            table, info_label, restore_btn, delete_btn
        ))

        # Start the timer
        refresh_timer.start()

        # --- END OF ADDED CODE ---

        # This line should already be here:
        dlg.exec()

    def _load_archived(self, table, info_label, restore_btn, delete_btn):
        """Load archived employees into table"""
        archived = self.db.get_archived_employees()
        table.setRowCount(len(archived))

        for i, emp in enumerate(archived):
            table.setItem(i, 0, QTableWidgetItem(emp.get('emp_id', '')))
            table.setItem(i, 1, QTableWidgetItem(emp.get('name', '')))
            table.setItem(i, 2, QTableWidgetItem(emp.get('department', '') or '—'))
            table.setItem(i, 3, QTableWidgetItem(emp.get('position', '') or '—'))
            table.setItem(i, 4, QTableWidgetItem(emp.get('hire_date', '') or '—'))
            table.setItem(i, 5, QTableWidgetItem(emp.get('archived_date', '')))
            table.setItem(i, 6, QTableWidgetItem(emp.get('archived_by', '')))
            table.setItem(i, 7, QTableWidgetItem(emp.get('archive_reason', '') or 'No reason'))

            # Status
            if emp.get('resign_date'):
                table.setItem(i, 8, QTableWidgetItem(f"Resigned - {emp.get('resign_date')}"))
            else:
                table.setItem(i, 8, QTableWidgetItem("Was Active"))

        table.resizeColumnsToContents()
        info_label.setText(f"📊 Total archived employees: <b>{len(archived)}</b>")

        restore_btn.setEnabled(False)
        delete_btn.setEnabled(False)

    def _restore_archived(self, table):
        """Restore selected archived employee"""
        selected = table.selectedItems()
        if not selected:
            return

        row = table.currentRow()
        emp_id = table.item(row, 0).text()
        name = table.item(row, 1).text()

        reply = QMessageBox.question(
            self,
            "Confirm Restore",
            f"Restore employee:\n\n"
            f"👤 {name} ({emp_id})\n\n"
            f"This will move the employee back to the active employees list.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            if self.db.restore_employee(emp_id, self.current_user):
                # Restore files
                archive_folder = os.path.join(FILES_DIR, "_archived", emp_id)
                if os.path.isdir(archive_folder):
                    dest_folder = os.path.join(FILES_DIR, emp_id)
                    try:
                        shutil.move(archive_folder, dest_folder)
                    except (OSError, shutil.Error):
                        # Folder may already exist or be locked
                        pass

                # Restore photo
                archive_photo = os.path.join(PHOTOS_DIR, "_archived", f"{emp_id}.png")
                if os.path.exists(archive_photo):
                    dest_photo = os.path.join(PHOTOS_DIR, f"{emp_id}.png")
                    try:
                        shutil.move(archive_photo, dest_photo)
                    except (OSError, shutil.Error):
                        # Photo may already exist or be locked
                        pass

                show_success_toast(self, f"Employee {name} has been restored!")
                self._refresh_all()
            else:
                show_warning_toast(
                    self, f"Unable to restore employee '{name}' (ID: {emp_id}).\n\n"
                    "Possible causes:\n"
                    "• Database is locked by another user\n"
                    "• Employee ID already exists in active records\n"
                    "• Insufficient permissions\n"
                    "• Database connection lost\n\n"
                    "Please try again or contact your administrator."
                )

    def _permanently_delete_archived(self, table):
        """Permanently delete selected archived employee"""
        selected = table.selectedItems()
        if not selected:
            return

        row = table.currentRow()
        emp_id = table.item(row, 0).text()
        name = table.item(row, 1).text()

        reply = show_warning_toast(
            self, f"PERMANENTLY DELETE employee:\n\n"
            f"👤 {name} ({emp_id})\n\n"
            f"⚠️ WARNING: This action CANNOT be undone!\n"
            f"The employee and all associated files will be permanently deleted.\n\n"
            f"Are you absolutely sure?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            # Second confirmation
            confirm = QInputDialog.getText(
                self,
                "Final Confirmation",
                f"Type the employee name to confirm permanent deletion:\n\n{name}"
            )

            if confirm[1] and confirm[0] == name:
                if self.db.permanently_delete_archived(emp_id, self.current_user):
                    # Delete archived files permanently
                    archive_folder = os.path.join(FILES_DIR, "_archived", emp_id)
                    if os.path.isdir(archive_folder):
                        shutil.rmtree(archive_folder, ignore_errors=True)

                    # Delete archived photo
                    archive_photo = os.path.join(PHOTOS_DIR, "_archived", f"{emp_id}.png")
                    if os.path.exists(archive_photo):
                        os.remove(archive_photo)

                    show_success_toast(self, f"Employee {name} has been permanently deleted.")
                else:
                    show_warning_toast(
                        self, f"Unable to permanently delete employee '{name}' (ID: {emp_id}).\n\n"
                        "Possible causes:\n"
                        "• Database is locked by another user\n"
                        "• Insufficient permissions\n"
                        "• Database connection lost\n"
                        "• Employee record is corrupted\n\n"
                        "Please try again or contact your administrator."
                    )
            else:
                show_success_toast(self, "Deletion cancelled - name did not match.")

    def _show_user_management(self):
        """Show user management dialog"""
        dlg = UserManagementDialog(self.db, self.current_user, self)
        dlg.exec()

    def _show_emp_id_swap(self):
        """Show employee ID swap dialog (admin only)"""
        from employee_vault.ui.dialogs.emp_id_swap import EmpIDSwapDialog
        dlg = EmpIDSwapDialog(self.db, self.current_user, self)
        if dlg.exec() == QDialog.Accepted:
            # Refresh employee list to show updated IDs
            self._refresh_all()

    def _show_letter_generation(self):
        """Show letter generation dialog"""
        dialog = LetterGenerationDialog(self.db, self.current_user, parent=self)
        dialog.exec()

    def _show_session_monitor(self):
        """Show active sessions monitor"""
        dialog = SessionMonitorDialog(self.db, parent=self)
        dialog.exec()

    def _show_store_management(self):
        """Show store management dialog"""
        dialog = StoreManagementDialog(self.db, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Refresh might be needed if stores are used elsewhere
            pass

    def _show_audit_log(self):
        """Show audit log viewer dialog"""
        # v4.4.1: Animated dialog for audit log
        from employee_vault.ui.widgets import SmoothAnimatedDialog
        dlg = SmoothAnimatedDialog(self, animation_style="slide")
        dlg.setWindowTitle("📜 Audit Log")
        dlg.resize(1000, 600)

        layout = QVBoxLayout(dlg)
        layout.addWidget(QLabel("<h2>📜 Audit Log - Activity History</h2>"))

        # Filters
        filter_box = QGroupBox("Filters")
        filter_layout = QHBoxLayout(filter_box)

        # Username filter
        filter_layout.addWidget(QLabel("Username:"))
        username_filter = NeumorphicGradientComboBox("All Users")
        username_filter.setMinimumHeight(70)
        username_filter.addItem("All Users")
        # Get unique usernames from audit log
        users = set()
        for log in self.db.get_audit_log(limit=1000):
            users.add(log['username'])
        for user in sorted(users):
            username_filter.addItem(user)
        filter_layout.addWidget(username_filter)

        # Action filter
        filter_layout.addWidget(QLabel("Action:"))
        action_filter = NeumorphicGradientComboBox("All Actions")
        action_filter.setMinimumHeight(70)
        action_filter.addItems(["All Actions", "LOGIN", "LOGOUT", "INSERT", "UPDATE", "DELETE"])
        filter_layout.addWidget(action_filter)

        # Record ID filter
        filter_layout.addWidget(QLabel("Employee ID:"))
        record_filter = QLineEdit()
        record_filter.setPlaceholderText("Filter by Employee ID")
        filter_layout.addWidget(record_filter)

        # Refresh button
        refresh_btn = ModernAnimatedButton("🔄 Refresh")
        apply_ios_style(refresh_btn, 'blue')
        filter_layout.addWidget(refresh_btn)

        layout.addWidget(filter_box)

        # Table for audit log
        from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QHeaderView
        table = QTableWidget()
        disable_cursor_changes(table)  # Remove hand cursor
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels(["ID", "Timestamp", "Username", "Action", "Table", "Record ID", "Old Value", "New Value", "Details"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        table.horizontalHeader().setStretchLastSection(True)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QAbstractItemView.SelectRows)

        def load_logs():
            """Load logs with current filters"""
            username = None if username_filter.currentText() == "All Users" else username_filter.currentText()
            action = None if action_filter.currentText() == "All Actions" else action_filter.currentText()
            record_id = record_filter.text().strip() or None

            logs = self.db.get_audit_log(limit=500, username=username, action=action, record_id=record_id)

            table.setRowCount(len(logs))
            for i, log in enumerate(logs):
                table.setItem(i, 0, QTableWidgetItem(str(log.get('id', ''))))
                table.setItem(i, 1, QTableWidgetItem(log.get('timestamp', '')))
                table.setItem(i, 2, QTableWidgetItem(log.get('username', '')))

                # Color code actions
                action_item = QTableWidgetItem(log.get('action', ''))
                action_text = log.get('action', '')
                if action_text == 'LOGIN':
                    action_item.setForeground(QColor('#4CAF50'))  # Green
                elif action_text == 'LOGOUT':
                    action_item.setForeground(QColor('#9E9E9E'))  # Gray
                elif action_text == 'INSERT':
                    action_item.setForeground(QColor('#2196F3'))  # Blue
                elif action_text == 'UPDATE':
                    action_item.setForeground(QColor('#FF9800'))  # Orange
                elif action_text == 'DELETE':
                    action_item.setForeground(QColor('#F44336'))  # Red
                table.setItem(i, 3, action_item)

                table.setItem(i, 4, QTableWidgetItem(log.get('table_name', '') or ''))
                table.setItem(i, 5, QTableWidgetItem(log.get('record_id', '') or ''))
                table.setItem(i, 6, QTableWidgetItem(log.get('old_value', '') or ''))
                table.setItem(i, 7, QTableWidgetItem(log.get('new_value', '') or ''))

                # Show details in a separate column
                details_text = log.get('details', '') or ''
                details_item = QTableWidgetItem(details_text[:100] + ('...' if len(details_text) > 100 else ''))
                details_item.setToolTip(details_text)  # Full text on hover
                table.setItem(i, 8, details_item)

            # Adjust column widths
            table.resizeColumnsToContents()
            table.setColumnWidth(8, 200)  # Details column wider

        # Initial load
        load_logs()

        # Connect refresh button
        refresh_btn.clicked.connect(load_logs)
        username_filter.currentTextChanged.connect(load_logs)
        action_filter.currentTextChanged.connect(load_logs)

        layout.addWidget(table)

        # Stats
        stats_label = QLabel()
        def update_stats():
            stats_label.setText(f"Showing {table.rowCount()} entries")
        update_stats()
        layout.addWidget(stats_label)

        # Export button
        export_btn = ModernAnimatedButton("📥 Export to JSON")
        apply_ios_style(export_btn, 'green')
        def export_logs():
            filename, _ = QFileDialog.getSaveFileName(dlg, "Save Audit Log", "audit_log.json", "JSON Files (*.json)")
            if filename:
                try:
                    logs = self.db.get_audit_log(limit=5000)
                    with open(filename, 'w') as f:
                        json.dump(logs, f, indent=2)
                    show_success_toast(dlg, f"Audit log exported to {filename}")
                except Exception as e:
                    show_error_toast(dlg, f"Failed to export: {str(e)}")
        export_btn.clicked.connect(export_logs)
        layout.addWidget(export_btn)

        # Close button
        close_btn = ModernAnimatedButton("Close")
        apply_ios_style(close_btn, 'gray')
        close_btn.clicked.connect(dlg.close)
        layout.addWidget(close_btn)

        dlg.exec()

    def _show_force_close_dialog(self):
        """Admin dialog to force close all users' applications for updates"""
        from employee_vault.ui.widgets import SmoothAnimatedDialog
        
        dlg = SmoothAnimatedDialog(self, animation_style="slide")
        dlg.setWindowTitle("🛑 Force Close All Users")
        dlg.resize(500, 350)
        
        layout = QVBoxLayout(dlg)
        layout.setSpacing(15)
        
        # Warning header
        header = QLabel("<h2>🛑 Force Close All Applications</h2>")
        header.setStyleSheet("color: #ef4444;")
        layout.addWidget(header)
        
        # Explanation
        info = QLabel(
            "This will send a shutdown signal to all running instances of Employee Vault.\n\n"
            "Use this feature when you need to:\n"
            "• Update the application for all users\n"
            "• Perform database maintenance\n"
            "• Deploy critical fixes\n\n"
            "All users will see a message and the application will close automatically."
        )
        info.setWordWrap(True)
        info.setStyleSheet("padding: 10px; background: rgba(239, 68, 68, 0.1); border-radius: 8px;")
        layout.addWidget(info)
        
        # Custom message input
        msg_group = QGroupBox("Custom Message (Optional)")
        msg_layout = QVBoxLayout(msg_group)
        message_input = QLineEdit()
        message_input.setPlaceholderText("e.g., 'Updating to version 5.4 - please restart in 5 minutes'")
        msg_layout.addWidget(message_input)
        layout.addWidget(msg_group)
        
        # Check current force close status
        current_status = self.db.check_force_close()
        if current_status:
            status_label = QLabel(f"⚠️ Force close is currently ACTIVE (requested by {current_status.get('requested_by', 'unknown')})")
            status_label.setStyleSheet("color: #f59e0b; font-weight: bold;")
            layout.addWidget(status_label)
            
            # Clear button
            clear_btn = ModernAnimatedButton("✅ Clear Force Close Flag")
            apply_ios_style(clear_btn, 'green')
            def clear_flag():
                self.db.clear_force_close()
                show_success_toast(dlg, "Force close flag cleared. Users can now use the application normally.")
                dlg.close()
            clear_btn.clicked.connect(clear_flag)
            layout.addWidget(clear_btn)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        force_btn = ModernAnimatedButton("🛑 Force Close All")
        apply_ios_style(force_btn, 'red')
        def trigger_force_close():
            from PySide6.QtWidgets import QMessageBox
            reply = QMessageBox.warning(
                dlg, "Confirm Force Close",
                "Are you sure you want to force close all user applications?\n\n"
                "This will immediately close the application for all connected users.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                message = message_input.text().strip() or "Application update in progress. Please restart the application."
                self.db.set_force_close(self.current_user, message)
                show_success_toast(dlg, "Force close signal sent! All users will be notified and their applications will close.")
                dlg.close()
        force_btn.clicked.connect(trigger_force_close)
        btn_layout.addWidget(force_btn)
        
        cancel_btn = ModernAnimatedButton("Cancel")
        apply_ios_style(cancel_btn, 'gray')
        cancel_btn.clicked.connect(dlg.close)
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
        dlg.exec()

    def _show_reports(self):
        """Show reports dialog"""
        # v4.4.1: Animated dialog for reports
        from employee_vault.ui.widgets import AnimatedDialogBase
        dlg = AnimatedDialogBase(self, animation_style="fade")
        dlg.setWindowTitle("📊 Reports")
        dlg.resize(500, 400)

        layout = QVBoxLayout(dlg)
        layout.addWidget(QLabel("<h2>📊 Generate Reports</h2>"))

        # Report options - Phase 3: iOS frosted glass styling
        reports = QGroupBox("Available Reports")
        reports_layout = QVBoxLayout(reports)

        btn_contract_expiry = ModernAnimatedButton("📄 Contract Expiry Report")
        apply_ios_style(btn_contract_expiry, 'blue')
        btn_contract_expiry.clicked.connect(lambda: self._generate_contract_report(dlg))
        reports_layout.addWidget(btn_contract_expiry)

        btn_department_summary = ModernAnimatedButton("🏢 Department Summary")
        apply_ios_style(btn_department_summary, 'blue')
        btn_department_summary.clicked.connect(lambda: self._generate_department_report(dlg))
        reports_layout.addWidget(btn_department_summary)

        btn_employee_list = ModernAnimatedButton("👥 Complete Employee List")
        apply_ios_style(btn_employee_list, 'blue')
        btn_employee_list.clicked.connect(lambda: self._generate_employee_list_report(dlg))
        reports_layout.addWidget(btn_employee_list)

        btn_agency_report = ModernAnimatedButton("🏢 Agency Report")
        apply_ios_style(btn_agency_report, 'blue')
        btn_agency_report.clicked.connect(lambda: self._generate_agency_report(dlg))
        reports_layout.addWidget(btn_agency_report)

        # NEW: Incomplete Records Report
        btn_incomplete_records = ModernAnimatedButton("⚠️ Incomplete Records Report")
        apply_ios_style(btn_incomplete_records, 'orange')
        btn_incomplete_records.clicked.connect(lambda: self._generate_incomplete_records_report(dlg))
        reports_layout.addWidget(btn_incomplete_records)

        # v5.3: PDF Export
        btn_pdf_export = ModernAnimatedButton("📥 Export to PDF")
        apply_ios_style(btn_pdf_export, 'green')
        btn_pdf_export.clicked.connect(lambda: self._show_pdf_export(dlg))
        reports_layout.addWidget(btn_pdf_export)

        # Salary analysis removed per user request

        layout.addWidget(reports)

        close_btn = ModernAnimatedButton("Close")
        apply_ios_style(close_btn, 'gray')
        close_btn.clicked.connect(dlg.close)
        layout.addWidget(close_btn)

        dlg.exec()

    def _generate_contract_report(self, parent):
        """Generate contract expiry report"""
        expired = [e for e in self.employees if (contract_days_left(e) or 999999) < 0]
        expiring_soon = [e for e in self.employees if 0 <= (contract_days_left(e) or 999999) <= ALERT_DAYS]
        valid = [e for e in self.employees if (contract_days_left(e) or 999999) > ALERT_DAYS]

        report = f"""
<h2>📄 Contract Expiry Report</h2>
<p><b>Generated:</b> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
<hr>

<h3 style='color:#ff6b6b;'>🔴 Expired Contracts ({len(expired)})</h3>
"""
        for e in expired:
            days = contract_days_left(e)
            report += f"<p>• <b>{e['name']}</b> ({e['emp_id']}) - Expired {-days} days ago<br>"
            report += f"&nbsp;&nbsp;Department: {e.get('department', 'N/A')} | Agency: {e.get('agency', 'N/A')}</p>"

        report += f"<h3 style='color:#ffcc66;'>🟠 Expiring Soon ({len(expiring_soon)})</h3>"
        for e in expiring_soon:
            days = contract_days_left(e)
            report += f"<p>• <b>{e['name']}</b> ({e['emp_id']}) - Expires in {days} days<br>"
            report += f"&nbsp;&nbsp;Department: {e.get('department', 'N/A')} | Agency: {e.get('agency', 'N/A')}</p>"

        report += f"<h3 style='color:#9ad17a;'>🟢 Valid Contracts ({len(valid)})</h3>"
        report += f"<p>{len(valid)} employees with valid contracts (>{ALERT_DAYS} days remaining)</p>"

        self._show_report_dialog(parent, "Contract Expiry Report", report)

    def _generate_department_report(self, parent):
        """Generate department summary"""
        dept_stats = {}
        for e in self.employees:
            dept = e.get('department', 'Unassigned')
            if dept not in dept_stats:
                dept_stats[dept] = {'total': 0, 'active': 0, 'resigned': 0, 'total_salary': 0}
            dept_stats[dept]['total'] += 1
            if not e.get('resign_date'):
                dept_stats[dept]['active'] += 1
            else:
                dept_stats[dept]['resigned'] += 1
            dept_stats[dept]['total_salary'] += e.get('salary', 0)

        report = f"""
<h2>🏢 Department Summary Report</h2>
<p><b>Generated:</b> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
<hr>
"""
        for dept, stats in sorted(dept_stats.items()):
            avg_salary = stats['total_salary'] / stats['total'] if stats['total'] > 0 else 0
            report += f"""
<h3>{dept}</h3>
<p>
• <b>Total Employees:</b> {stats['total']}<br>
• <b>Active:</b> {stats['active']}<br>
• <b>Resigned:</b> {stats['resigned']}<br>
• <b>Avg Salary/Day:</b> ₱{avg_salary:,.2f}<br>
• <b>Total Daily Payroll:</b> ₱{stats['total_salary']:,.2f}
</p>
"""

        self._show_report_dialog(parent, "Department Summary", report)

    def _generate_employee_list_report(self, parent):
        """Generate complete employee list"""
        active = [e for e in self.employees if not e.get('resign_date')]
        resigned = [e for e in self.employees if e.get('resign_date')]

        report = f"""
<h2>👥 Complete Employee List</h2>
<p><b>Generated:</b> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
<p><b>Total:</b> {len(self.employees)} | <b>Active:</b> {len(active)} | <b>Resigned:</b> {len(resigned)}</p>
<hr>

<h3>Active Employees ({len(active)})</h3>
"""
        for e in sorted(active, key=lambda x: x.get('name', '')):
            report += f"""
<p><b>{e['name']}</b> ({e['emp_id']})<br>
&nbsp;&nbsp;{e.get('position', 'N/A')} - {e.get('department', 'N/A')}<br>
&nbsp;&nbsp;Hired: {e.get('hire_date', 'N/A')}</p>
"""

        self._show_report_dialog(parent, "Employee List", report)

    def _generate_agency_report(self, parent):
        """Generate agency report"""
        agency_stats = {}
        for e in self.employees:
            agency = e.get('agency', 'Direct Hire')
            if agency not in agency_stats:
                agency_stats[agency] = []
            agency_stats[agency].append(e)

        report = f"""
<h2>🏢 Agency Report</h2>
<p><b>Generated:</b> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
<hr>
"""
        for agency, emps in sorted(agency_stats.items(), key=lambda x: x[0] or 'ZZZ_No_Agency'):
            active = sum(1 for e in emps if not e.get('resign_date'))
            report += f"""
<h3>{agency}</h3>
<p><b>Total:</b> {len(emps)} | <b>Active:</b> {active}</p>
<ul>
"""
            for e in emps:
                status = "✅" if not e.get('resign_date') else "❌"
                report += f"<li>{status} {e['name']} - {e.get('position', 'N/A')}</li>"
            report += "</ul>"

        self._show_report_dialog(parent, "Agency Report", report)

    def _generate_incomplete_records_report(self, parent):
        """Generate incomplete records report - employees missing required data"""
        incomplete = []
        
        for emp in self.employees:
            emp_id = emp.get('emp_id', '')
            issues = []
            
            # Check for missing photo
            photo_path = os.path.join(PHOTOS_DIR, f"{emp_id}.png")
            if not os.path.exists(photo_path):
                issues.append("📷 Missing photo")
            
            # Check for missing government IDs
            if not emp.get('sss_number'):
                issues.append("🪪 Missing SSS #")
            if not emp.get('tin_number'):
                issues.append("🪪 Missing TIN #")
            if not emp.get('philhealth_number'):
                issues.append("🪪 Missing PhilHealth #")
            if not emp.get('pagibig_number'):
                issues.append("🪪 Missing Pag-IBIG #")
            
            # Check for missing emergency contact
            if not emp.get('emergency_contact_name'):
                issues.append("📞 Missing emergency contact")
            
            # Check for missing contract info
            if not emp.get('contract_expiry') and not emp.get('resign_date'):
                issues.append("📋 Missing contract expiry")
            
            # Check for missing email/phone
            if not emp.get('email') and not emp.get('phone'):
                issues.append("📱 Missing contact info")
            
            if issues:
                incomplete.append({
                    'emp': emp,
                    'issues': issues,
                    'severity': len(issues)
                })
        
        # Sort by severity (most issues first)
        incomplete.sort(key=lambda x: -x['severity'])
        
        # Generate report
        report = f"""
<h2>⚠️ Incomplete Records Report</h2>
<p><b>Generated:</b> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
<p><b>Total Employees:</b> {len(self.employees)} | <b>Incomplete:</b> {len(incomplete)}</p>
<hr>
"""
        
        if not incomplete:
            report += """
<h3 style='color: #4CAF50;'>✅ All Records Complete!</h3>
<p>All employee records have the required information.</p>
"""
        else:
            # Summary by issue type
            issue_counts = {}
            for item in incomplete:
                for issue in item['issues']:
                    issue_counts[issue] = issue_counts.get(issue, 0) + 1
            
            report += "<h3>📊 Summary by Issue Type</h3><ul>"
            for issue, count in sorted(issue_counts.items(), key=lambda x: -x[1]):
                report += f"<li>{issue}: <b>{count}</b> employees</li>"
            report += "</ul><hr>"
            
            # Critical (4+ issues)
            critical = [x for x in incomplete if x['severity'] >= 4]
            if critical:
                report += f"<h3 style='color: #F44336;'>🔴 Critical ({len(critical)} employees - 4+ missing fields)</h3>"
                for item in critical:
                    emp = item['emp']
                    report += f"<p><b>{emp.get('name', 'Unknown')}</b> ({emp.get('emp_id', 'N/A')}) - {emp.get('department', 'N/A')}<br>"
                    report += f"<span style='color: #ffcc80;'>{', '.join(item['issues'])}</span></p>"
            
            # Warning (2-3 issues)
            warning = [x for x in incomplete if 2 <= x['severity'] < 4]
            if warning:
                report += f"<h3 style='color: #FF9800;'>🟠 Warning ({len(warning)} employees - 2-3 missing fields)</h3>"
                for item in warning:
                    emp = item['emp']
                    report += f"<p><b>{emp.get('name', 'Unknown')}</b> ({emp.get('emp_id', 'N/A')}) - {emp.get('department', 'N/A')}<br>"
                    report += f"<span style='color: #ffcc80;'>{', '.join(item['issues'])}</span></p>"
            
            # Minor (1 issue)
            minor = [x for x in incomplete if x['severity'] == 1]
            if minor:
                report += f"<h3 style='color: #FFEB3B;'>🟡 Minor ({len(minor)} employees - 1 missing field)</h3>"
                for item in minor[:20]:  # Limit to first 20
                    emp = item['emp']
                    report += f"<p><b>{emp.get('name', 'Unknown')}</b> ({emp.get('emp_id', 'N/A')}): {item['issues'][0]}</p>"
                if len(minor) > 20:
                    report += f"<p><i>...and {len(minor) - 20} more</i></p>"
        
        self._show_report_dialog(parent, "Incomplete Records Report", report)

    def _show_pdf_export(self, parent):
        """Show PDF export dialog - v5.3"""
        parent.close()  # Close the reports dialog
        from employee_vault.ui.dialogs.pdf_report import PDFExportDialog
        dlg = PDFExportDialog(self.db, self.employees, self)
        dlg.exec()

    def _generate_salary_report(self, parent):
        """Generate salary analysis"""
        salaries = [e.get('salary', 0) for e in self.employees if not e.get('resign_date')]
        if not salaries:
            show_success_toast(parent, "No salary data available.")
            return

        total = sum(salaries)
        avg = total / len(salaries)
        min_sal = min(salaries)
        max_sal = max(salaries)

        report = f"""
<h2>💰 Salary Analysis</h2>
<p><b>Generated:</b> {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
<p><b>Based on:</b> {len(salaries)} active employees</p>
<hr>

<h3>Summary Statistics</h3>
<p>
• <b>Total Daily Payroll:</b> ₱{total:,.2f}<br>
• <b>Average Salary/Day:</b> ₱{avg:,.2f}<br>
• <b>Lowest Salary:</b> ₱{min_sal:,.2f}<br>
• <b>Highest Salary:</b> ₱{max_sal:,.2f}<br>
• <b>Monthly Estimate:</b> ₱{total * 30:,.2f}
</p>

<h3>Salary Ranges</h3>
"""
        ranges = [
            (0, 500, "Under ₱500"),
            (500, 1000, "₱500 - ₱1,000"),
            (1000, 1500, "₱1,000 - ₱1,500"),
            (1500, 2000, "₱1,500 - ₱2,000"),
            (2000, 999999, "Over ₱2,000")
        ]

        for min_r, max_r, label in ranges:
            count = sum(1 for s in salaries if min_r <= s < max_r)
            pct = (count / len(salaries) * 100) if salaries else 0
            report += f"<p>• {label}: {count} employees ({pct:.1f}%)</p>"

        self._show_report_dialog(parent, "Salary Analysis", report)

    def _show_report_dialog(self, parent, title, html_content, report_data=None):
        """Show report in a dialog with preview and export options (HTML, PDF, Excel)"""
        # v4.4.1: Animated dialog for report display
        from employee_vault.ui.widgets import AnimatedDialogBase
        dlg = AnimatedDialogBase(parent, animation_style="fade")
        dlg.setWindowTitle(title)
        dlg.resize(850, 700)

        layout = QVBoxLayout(dlg)

        # Title bar with report name
        title_label = QLabel(f"<h2>📊 {title}</h2>")
        title_label.setStyleSheet("padding: 10px; background: rgba(74, 158, 255, 0.2); border-radius: 8px;")
        layout.addWidget(title_label)

        # Report content in scrollable area
        text_browser = QTextEdit()
        text_browser.setHtml(html_content)
        text_browser.setReadOnly(True)
        layout.addWidget(text_browser, 1)

        # Export buttons row
        btn_layout = QHBoxLayout()

        export_html_btn = ModernAnimatedButton("📄 Export HTML")
        apply_ios_style(export_html_btn, 'blue')
        export_html_btn.clicked.connect(lambda: self._export_report_as_html(html_content, title))
        btn_layout.addWidget(export_html_btn)

        export_pdf_btn = ModernAnimatedButton("📕 Export PDF")
        apply_ios_style(export_pdf_btn, 'orange')
        export_pdf_btn.clicked.connect(lambda: self._export_report_as_pdf(html_content, title))
        btn_layout.addWidget(export_pdf_btn)

        # Excel export only if we have structured data
        if report_data:
            export_excel_btn = ModernAnimatedButton("📊 Export Excel")
            apply_ios_style(export_excel_btn, 'green')
            export_excel_btn.clicked.connect(lambda: self._export_report_as_excel(report_data, title))
            btn_layout.addWidget(export_excel_btn)

        btn_layout.addStretch()

        # Print Preview button (mandatory preview before print)
        print_preview_btn = ModernAnimatedButton("🖨️ Print Preview")
        apply_ios_style(print_preview_btn, 'purple')
        print_preview_btn.clicked.connect(lambda: self._print_report_with_preview(text_browser))
        btn_layout.addWidget(print_preview_btn)

        close_btn = ModernAnimatedButton("Close")
        apply_ios_style(close_btn, 'gray')
        close_btn.clicked.connect(dlg.close)
        btn_layout.addWidget(close_btn)

        layout.addLayout(btn_layout)
        dlg.exec()

    def _print_report_with_preview(self, text_browser):
        """Print report with mandatory preview dialog using QPrintPreviewDialog"""
        from PySide6.QtPrintSupport import QPrinter, QPrintPreviewDialog
        
        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageOrientation(QPageLayout.Portrait)
        
        def handle_paint_request(printer):
            text_browser.document().print_(printer)
        
        preview_dialog = QPrintPreviewDialog(printer, self)
        preview_dialog.setWindowTitle("Print Preview - Report")
        preview_dialog.resize(900, 700)
        preview_dialog.paintRequested.connect(handle_paint_request)
        preview_dialog.exec()

    def _print_report(self, text_browser):
        """Print report with mandatory preview (redirects to preview)"""
        self._print_report_with_preview(text_browser)

    def _export_report_as_html(self, content, title):
        """Export report as HTML file"""
        filename = title.replace(" ", "_").lower() + ".html"
        path, _ = QFileDialog.getSaveFileName(self, "Export Report as HTML", filename, "HTML (*.html)")
        if not path:
            return

        html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; padding: 20px; background: #f5f5f5; }}
        h2, h3 {{ color: #333; }}
        p {{ margin: 8px 0; }}
        ul {{ margin: 10px 0; }}
        li {{ margin: 5px 0; }}
        hr {{ border: none; border-top: 1px solid #ccc; margin: 15px 0; }}
    </style>
</head>
<body>
{content}
</body>
</html>"""
        try:
            with open(path, 'w', encoding='utf-8') as f:
                f.write(html)
            show_success_toast(self, f"Report exported to:\n{path}")
        except Exception as e:
            show_error_toast(self, f"Failed to export report:\n{str(e)}")

    def _export_report_as_pdf(self, html_content, title):
        """Export report as PDF file"""
        filename = title.replace(" ", "_").lower() + ".pdf"
        path, _ = QFileDialog.getSaveFileName(self, "Export Report as PDF", filename, "PDF (*.pdf)")
        if not path:
            return

        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            
            doc = SimpleDocTemplate(path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=12)
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 12))
            
            import re
            text = html_content.replace('<br>', '\n').replace('<br/>', '\n').replace('</p>', '\n').replace('</li>', '\n')
            text = re.sub(r'<h[23][^>]*>', '\n\n### ', text)
            text = re.sub(r'</h[23]>', ' ###\n', text)
            text = re.sub(r'<[^>]+>', '', text)
            text = re.sub(r'&nbsp;', ' ', text)
            text = re.sub(r'\n\s*\n', '\n\n', text)
            
            normal_style = styles['Normal']
            for line in text.split('\n'):
                line = line.strip()
                if line:
                    if line.startswith('###') and line.endswith('###'):
                        header = line.replace('###', '').strip()
                        header_style = ParagraphStyle('Header', parent=styles['Heading2'], fontSize=14, spaceBefore=12)
                        story.append(Paragraph(header, header_style))
                    else:
                        story.append(Paragraph(line, normal_style))
                    story.append(Spacer(1, 4))
            
            doc.build(story)
            show_success_toast(self, f"PDF exported to:\n{path}")
            
        except ImportError:
            show_error_toast(self, "PDF export requires 'reportlab' library.\n\nInstall with: pip install reportlab")
        except Exception as e:
            show_error_toast(self, f"Failed to export PDF:\n{str(e)}")

    def _export_report_as_excel(self, report_data, title):
        """Export report data as Excel file"""
        filename = title.replace(" ", "_").lower() + ".xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Export Report as Excel", filename, "Excel (*.xlsx)")
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title[:31]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            if isinstance(report_data, list) and report_data:
                headers = list(report_data[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                for row, item in enumerate(report_data, 2):
                    for col, key in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=item.get(key, ''))
                
                for col in ws.columns:
                    max_length = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
            
            elif isinstance(report_data, dict):
                row = 1
                for category, items in report_data.items():
                    ws.cell(row=row, column=1, value=category)
                    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                    row += 1
                    
                    if isinstance(items, list):
                        for item in items:
                            ws.cell(row=row, column=2, value=str(item))
                            row += 1
                    else:
                        ws.cell(row=row, column=2, value=str(items))
                        row += 1
                    row += 1
            
            wb.save(path)
            show_success_toast(self, f"Excel exported to:\n{path}")
            
        except ImportError:
            show_error_toast(self, "Excel export requires 'openpyxl' library.\n\nInstall with: pip install openpyxl")
        except Exception as e:
            show_error_toast(self, f"Failed to export Excel:\n{str(e)}")

    def _toggle_theme(self): self.setStyleSheet("" if self.styleSheet() else APP_QSS)
    def _show_theme_selector(self):
        """Show theme selector dialog"""
        # v4.4.1: Animated dialog for theme selector
        from employee_vault.ui.widgets import AnimatedDialogBase
        dialog = AnimatedDialogBase(self, animation_style="fade")
        dialog.setWindowTitle("🎨 Choose Theme")
        dialog.resize(500, 400)
        
        layout = QVBoxLayout(dialog)
        
        # Title
        title = QLabel("<h2>Choose Your Theme</h2>")
        layout.addWidget(title)
        
        info = QLabel("Select a theme to change the appearance of the application.")
        info.setWordWrap(True)
        info.setStyleSheet("color: #888; margin-bottom: 10px;")
        layout.addWidget(info)
        
        # Theme list
        theme_list = QListWidget()
        disable_cursor_changes(theme_list)  # Remove hand cursor
        
        # Get current theme
        current_theme = load_theme_preference()
        
        for theme_id, theme_data in MODERN_THEMES.items():
            item = QListWidgetItem(f"🎨 {theme_data['name']}")
            item.setData(Qt.UserRole, theme_id)
            theme_list.addItem(item)
            
            # Select current theme
            if theme_id == current_theme:
                theme_list.setCurrentItem(item)
        
        layout.addWidget(theme_list)
        
        # Preview label
        preview_label = QLabel("<i>Click a theme to preview its colors</i>")
        preview_label.setStyleSheet("color: #888; font-size: 12px; padding: 10px;")
        layout.addWidget(preview_label)
        
        # Buttons
        btn_layout = QHBoxLayout()

        apply_btn = ModernAnimatedButton("✓ Apply Theme")
        apply_ios_style(apply_btn, 'green')
        apply_btn.clicked.connect(dialog.accept)

        cancel_btn = ModernAnimatedButton("Cancel")
        apply_ios_style(cancel_btn, 'gray')
        cancel_btn.clicked.connect(dialog.reject)

        btn_layout.addStretch()
        btn_layout.addWidget(apply_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
        
        # Execute dialog
        if dialog.exec() == QDialog.Accepted:
            selected_item = theme_list.currentItem()
            if selected_item:
                new_theme = selected_item.data(Qt.UserRole)
                
                # Save theme preference
                try:
                    with open(THEME_PREFERENCE_FILE, 'w') as f:
                        f.write(new_theme)
                except (IOError, OSError):
                    # Failed to save theme preference
                    pass
                
                # Ask for confirmation and restart
                reply = QMessageBox.question(
                    self,
                    "Apply Theme",
                    f"Apply {MODERN_THEMES[new_theme]['name']} theme?\n\n"
                    f"The application will restart automatically.",
                    QMessageBox.Yes | QMessageBox.No
                )
                
                if reply == QMessageBox.Yes:
                    # Restart the application
                    QTimer.singleShot(500, lambda: self._restart_application())
    
    def _restart_application(self):
        """Restart the application to apply theme"""
        import subprocess
        
        # Close database connection
        try:
            if hasattr(self, 'db'):
                self.db.conn.close()
        except (AttributeError, sqlite3.Error):
            # Database already closed or connection error
            pass
        
        # Restart using python
        python = sys.executable
        subprocess.Popen([python] + sys.argv)
        
        # Close current application
        QApplication.quit()

    def _show_about(self):
        """Show about dialog with version and system info"""
        from employee_vault.ui.widgets import AnimatedDialogBase, ModernAnimatedButton
        from employee_vault.ui.ios_button_styles import apply_ios_style

        stats = self.db.get_database_stats()

        # Get current theme name
        current_theme = load_theme_preference()
        theme_name = MODERN_THEMES[current_theme]['name']

        # Create custom animated dialog
        dialog = AnimatedDialogBase(self, animation_style="fade")
        dialog.setWindowTitle("About")
        dialog.resize(500, 600)

        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(30, 30, 30, 30)

        about_text = f"""
        <h2 style="color: #4a9eff;">{APP_TITLE}</h2>
        <p><b>Version:</b> {VERSION}<br>
        <b>Build Date:</b> {BUILD_DATE}<br>
        <b>Database Version:</b> {DATABASE_VERSION}</p>

        <hr style="border: 1px solid rgba(255, 255, 255, 0.1);">

        <h3 style="color: #4a9eff;">UI Theme</h3>
        <p><b>Current Theme:</b> {theme_name}<br>
        <b>Available Themes:</b> {len(MODERN_THEMES)}<br>
        <i>Click "Change Theme" in Settings to switch themes</i></p>

        <hr style="border: 1px solid rgba(255, 255, 255, 0.1);">

        <h3 style="color: #4a9eff;">System Information</h3>
        <p><b>Total Employees:</b> {stats.get('total_employees', 0)}<br>
        <b>Active Employees:</b> {stats.get('active_employees', 0)}<br>
        <b>Total Users:</b> {stats.get('total_users', 0)}<br>
        <b>Total Agencies:</b> {stats.get('total_agencies', 0)}<br>
        <b>Database Size:</b> {stats.get('db_size_mb', 0):.2f} MB</p>

        <hr style="border: 1px solid rgba(255, 255, 255, 0.1);">

        <p><b>Logged in as:</b> {self.current_user} ({self.user_row.get('role', 'user')})<br>
        <b>Security:</b> ✅ Bcrypt password hashing<br>
        <b>Database:</b> SQLite with WAL mode<br>
        <b>UI:</b> Modern with glassmorphism & animations</p>

        <p style="color: #888; font-size: 11px;">
        © 2025 Cuddly International Corporation<br>
        Employee Information Management System - Modern UI Edition
        </p>
        """

        # Text label (no background box)
        text_label = QLabel(about_text)
        text_label.setTextFormat(Qt.RichText)
        text_label.setWordWrap(True)
        text_label.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        text_label.setStyleSheet("background: transparent; padding: 10px;")
        layout.addWidget(text_label)

        layout.addStretch()

        # Close button
        close_btn = ModernAnimatedButton("✓ Close")
        apply_ios_style(close_btn, 'blue')
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)

        dialog.exec()
    
    def _load_user_photo(self):
        """v4.0.1: Load user photo with perfect circular clipping (like employee form)"""
        user_photo_path = os.path.join(PHOTOS_DIR, f"user_{self.current_user}.png")
        
        # If the user hasn't uploaded a photo, fall back to a default icon.
        if not os.path.exists(user_photo_path):
            self.user_photo_label.setPixmap(QPixmap())
            self.user_photo_label.setText("👤")
            self.user_photo_label.setAlignment(Qt.AlignCenter)
            return

        pix = QPixmap(user_photo_path)
        # Determine the size based on the current label dimensions.  This
        # allows the avatar to scale automatically if the UI designer
        # changes its size.  We pick the smaller of width and height to
        # ensure a square canvas for the circular mask.
        size = min(self.user_photo_label.width(), self.user_photo_label.height())

        # If the pixmap couldn't be loaded, show the default icon
        if pix.isNull():
            self.user_photo_label.setPixmap(QPixmap())
            self.user_photo_label.setText("👤")
            self.user_photo_label.setAlignment(Qt.AlignCenter)
            return

        # Scale the image to fit entirely within the circle.  Using Qt.KeepAspectRatio
        # avoids cropping any part of the photo, instead letterboxing inside the
        # circular mask.  This solves cases where a tall portrait photo is
        # clipped at the top/bottom when using KeepAspectRatioByExpanding.
        scaled = pix.scaled(size, size, Qt.KeepAspectRatio, Qt.SmoothTransformation)

        # Create a transparent canvas to draw the final circular image on
        result = QPixmap(size, size)
        result.fill(Qt.transparent)

        painter = QPainter(result)
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setRenderHint(QPainter.SmoothPixmapTransform, True)

        # Define a circular clip path
        path = QPainterPath()
        path.addEllipse(0, 0, size, size)
        painter.setClipPath(path)

        # Center the scaled image inside the circle by computing offsets.
        x_offset = (size - scaled.width()) // 2
        y_offset = (size - scaled.height()) // 2
        painter.drawPixmap(x_offset, y_offset, scaled)
        painter.end()

        # Set the processed pixmap and clear any text on the label
        self.user_photo_label.setPixmap(result)
        self.user_photo_label.setText("")
    
    def eventFilter(self, obj, event):
        """Handle hover events for sidebar expand/collapse"""
        if obj == self.sidebar:
            # Only handle hover if sidebar is collapsed
            if self.is_sidebar_collapsed:
                if event.type() == QEvent.Type.Enter:
                    # Mouse entered sidebar - temporarily expand
                    self._hover_expand_sidebar()
                    return False
                elif event.type() == QEvent.Type.Leave:
                    # Mouse left sidebar - collapse back
                    self._hover_collapse_sidebar()
                    return False
        if obj == getattr(self, "logo_label", None) and hasattr(self, "_logo_glow_animation"):
            if event.type() == QEvent.Type.Enter:
                self._animate_logo_glow(18)
            elif event.type() == QEvent.Type.Leave:
                self._animate_logo_glow(12)
        return super().eventFilter(obj, event)

    def _animate_logo_glow(self, target_blur: int):
        """Animate the header logo glow on hover."""
        try:
            self._logo_glow_animation.stop()
            current = self._logo_glow_effect.blurRadius()
            self._logo_glow_animation.setStartValue(current)
            self._logo_glow_animation.setEndValue(target_blur)
            if self.anim_manager:
                self._logo_glow_animation.setDuration(self.anim_manager.get_theme_duration("hover"))
                self._logo_glow_animation.setEasingCurve(self.anim_manager.get_theme_easing())
            self._logo_glow_animation.start()
        except Exception as e:
            logging.debug(f"Logo glow animation skipped: {e}")

    def _resolve_header_logo_path(self) -> str:
        """Find the best available header logo, prioritizing the provided Cuddly asset."""
        candidates = [
            resource_path(os.path.join("assets", "cuddly_logo.png")),
            resource_path(os.path.join("assets", "cuddly_logo (1).png")),
            resource_path("cuddly_logo.png"),
            resource_path("company_logo.png"),
            os.path.join("assets", "cuddly_logo.png"),
            "cuddly_logo.png",
            "company_logo.png",
        ]
        for path in candidates:
            try:
                if path and os.path.exists(path):
                    return path
            except Exception:
                continue
        return ""

    def _hover_expand_sidebar(self):
        """Option D: Temporarily expand sidebar on hover to show full text"""
        if not self.sidebar_hover_expanded:
            self.sidebar_hover_expanded = True
            # Animate to expanded width
            self.sidebar_animation = QPropertyAnimation(self.sidebar, b"maximumWidth", self)
            self.sidebar_animation.setDuration(200)
            self.sidebar_animation.setStartValue(self.sidebar.width())
            self.sidebar_animation.setEndValue(self.sidebar_expanded_width)
            self.sidebar_animation.setEasingCurve(QEasingCurve.Type.OutCubic)
            self.sidebar_animation.start()
            self.sidebar.setFixedWidth(self.sidebar_expanded_width)

            # Restore full text mode for sections
            if hasattr(self, "sidebar_sections"):
                for section in self.sidebar_sections:
                    icon_text = f"{section.icon} " if section.icon else ""
                    section.toggle_button.setText(f"{icon_text}{section.title}")
                    section.toggle_button.setToolTip("")
                    section.toggle_button.setStyleSheet(f"""
                        QPushButton {{
                            text-align: left;
                            padding: 12px 20px;
                            font-weight: 600;
                            font-size: 12px;
                            text-transform: uppercase;
                            letter-spacing: 0.5px;
                            color: white;
                            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 rgba(255, 255, 255, 0.15),
                                stop:0.5 rgba({self._hex_to_rgb(section.color)}, 0.35),
                                stop:1 rgba({self._hex_to_rgb(section.color)}, 0.55));
                            border: 1px solid rgba(255, 255, 255, 0.18);
                            border-top: 1px solid rgba(255, 255, 255, 0.3);
                            border-radius: 22px;
                            margin: 4px 6px;
                        }}
                        QPushButton:hover {{
                            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                stop:0 rgba(255, 255, 255, 0.22),
                                stop:0.5 rgba({self._hex_to_rgb(section.color)}, 0.5),
                                stop:1 rgba({self._hex_to_rgb(section.color)}, 0.7));
                        }}
                    """)

    def _hover_collapse_sidebar(self):
        """Option D: Collapse sidebar back to icons after hover"""
        if self.sidebar_hover_expanded:
            self.sidebar_hover_expanded = False
            # Animate back to collapsed width
            self.sidebar_animation = QPropertyAnimation(self.sidebar, b"maximumWidth", self)
            self.sidebar_animation.setDuration(200)
            self.sidebar_animation.setStartValue(self.sidebar.width())
            self.sidebar_animation.setEndValue(self.sidebar_collapsed_width)
            self.sidebar_animation.setEasingCurve(QEasingCurve.Type.InCubic)
            self.sidebar_animation.start()
            self.sidebar.setFixedWidth(self.sidebar_collapsed_width)

            # Restore icon-only mode for sections
            if hasattr(self, "sidebar_sections"):
                for section in self.sidebar_sections:
                    icon_char = self.section_icons.get(section, "")
                    section.toggle_button.setText(icon_char)
                    section.toggle_button.setToolTip(f"{section.full_title} - Click for menu")
                    section.toggle_button.setStyleSheet(f"""
                        QPushButton {{
                            text-align: center;
                            padding: 14px 8px;
                            font-size: 20px;
                            color: rgba(255, 255, 255, 0.7);
                            background: transparent;
                            border: none;
                            border-radius: 10px;
                            margin: 2px 6px;
                        }}
                        QPushButton:hover {{
                            color: white;
                            background: rgba(255, 255, 255, 0.1);
                        }}
                    """)

    def _init_sidebar_styles(self):
        """Phase 1.5: Pre-generate all sidebar button stylesheets to avoid expensive regeneration during animations"""
        self._sidebar_styles_cache = {}

        # Icon-only style (same for all sections)
        icon_only_style = """
            QPushButton {
                text-align: center;
                padding: 14px 8px;
                font-size: 20px;
                color: rgba(255, 255, 255, 0.7);
                background: transparent;
                border: none;
                border-radius: 10px;
                margin: 2px 6px;
            }
            QPushButton:hover {
                color: white;
                background: rgba(255, 255, 255, 0.1);
            }
        """

        for section in self.sidebar_sections:
            rgb = self._hex_to_rgb(section.color)

            self._sidebar_styles_cache[section] = {
                'icon_only': icon_only_style,
                'expanded': f"""
                    QPushButton {{
                        text-align: left;
                        padding: 12px 20px;
                        font-weight: 600;
                        font-size: 12px;
                        text-transform: uppercase;
                        letter-spacing: 0.5px;
                        color: white;
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 rgba(255, 255, 255, 0.15),
                            stop:0.5 rgba({rgb}, 0.35),
                            stop:1 rgba({rgb}, 0.55));
                        border: 1px solid rgba(255, 255, 255, 0.18);
                        border-top: 1px solid rgba(255, 255, 255, 0.3);
                        border-radius: 22px;
                        margin: 4px 6px;
                    }}
                    QPushButton:hover {{
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                            stop:0 rgba(255, 255, 255, 0.22),
                            stop:0.5 rgba({rgb}, 0.5),
                            stop:1 rgba({rgb}, 0.7));
                        border: 1px solid rgba(255, 255, 255, 0.35);
                        border-top: 1px solid rgba(255, 255, 255, 0.45);
                    }}
                    QPushButton:pressed {{
                        background: rgba({rgb}, 0.6);
                        border-top: 1px solid rgba(255, 255, 255, 0.15);
                    }}
                """
            }

    def _toggle_sidebar(self):
        """Option D: Animate sidebar between expanded (220px) and collapsed/icons-only (70px).
        When collapsed, clicking an icon shows a popup menu with the section's items."""
        # Performance instrumentation (Phase 0)
        start = time.perf_counter()

        # Phase 1.1: Pause table updates to prevent lag during animation
        if hasattr(self, 'employees_page'):
            self.employees_page.pause_updates_for_animation()

        # Determine target width based on current state
        if self.is_sidebar_collapsed:
            target_width = self.sidebar_expanded_width  # 220px
        else:
            target_width = self.sidebar_collapsed_width  # 70px

        # Create smooth animation on both minimum and maximum width
        from PySide6.QtCore import QParallelAnimationGroup

        self.sidebar_animation_group = QParallelAnimationGroup(self)

        # Animate minimum width
        min_width_anim = QPropertyAnimation(self.sidebar, b"minimumWidth", self)
        min_width_anim.setDuration(150)
        min_width_anim.setStartValue(self.sidebar.width())
        min_width_anim.setEndValue(target_width)
        min_width_anim.setEasingCurve(QEasingCurve.Type.InOutQuad)

        # Animate maximum width
        max_width_anim = QPropertyAnimation(self.sidebar, b"maximumWidth", self)
        max_width_anim.setDuration(150)
        max_width_anim.setStartValue(self.sidebar.width())
        max_width_anim.setEndValue(target_width)
        max_width_anim.setEasingCurve(QEasingCurve.Type.InOutQuad)

        # Add both animations to the group
        self.sidebar_animation_group.addAnimation(min_width_anim)
        self.sidebar_animation_group.addAnimation(max_width_anim)

        # Update collapsed state flag BEFORE animation starts
        self.is_sidebar_collapsed = not self.is_sidebar_collapsed
        # Reset hover state when manually toggling
        self.sidebar_hover_expanded = False

        # Connect finished signal - DEFER section updates until animation finishes
        def on_animation_finished():
            # First set fixed width
            self.sidebar.setFixedWidth(target_width)

            # OPTIMIZATION: Batch all visibility updates in single repaint cycle
            self.sidebar.setUpdatesEnabled(False)
            try:
                # Phase 1.5: Update each collapsible section using cached stylesheets
                if hasattr(self, "sidebar_sections"):
                    for section in self.sidebar_sections:
                        icon_char = self.section_icons.get(section, "")
                        if self.is_sidebar_collapsed:
                            # Icon-only mode - use cached stylesheet
                            section.toggle_button.setText(icon_char)
                            section.toggle_button.setToolTip(f"{section.full_title} - Click for menu")
                            section.toggle_button.setStyleSheet(self._sidebar_styles_cache[section]['icon_only'])
                            # Collapse content area
                            section.is_collapsed = True
                            section.content_area.setMaximumHeight(0)
                            section.content_area.setVisible(False)
                        else:
                            # Full text mode - use cached stylesheet
                            icon_text = f"{section.icon} " if section.icon else ""
                            section.toggle_button.setText(f"{icon_text}{section.title}")
                            section.toggle_button.setToolTip("")
                            section.toggle_button.setStyleSheet(self._sidebar_styles_cache[section]['expanded'])
                            # Show content if section was expanded before collapse
                            if not section.is_collapsed:
                                section.content_area.setVisible(True)
                                section.content_area.setMaximumHeight(16777215)

                # Force a layout update to avoid artifacts
                self.sidebar.updateGeometry()
            finally:
                # Single repaint for all changes
                self.sidebar.setUpdatesEnabled(True)

            # Phase 1.1: Resume table updates after animation
            if hasattr(self, 'employees_page'):
                self.employees_page.resume_updates_after_animation()

        self.sidebar_animation_group.finished.connect(on_animation_finished)

        # Start animation
        self.sidebar_animation_group.start()

        # Performance instrumentation (Phase 0)
        print(f"[PERF] Sidebar toggle: {(time.perf_counter() - start)*1000:.2f}ms")

    def _upload_user_photo(self):
        """v3.9: Upload user photo - clickable with photo editor"""
        try:
            user_photo_path = os.path.join(PHOTOS_DIR, f"user_{self.current_user}.png")
            has_photo = os.path.exists(user_photo_path)
            
            if has_photo:
                # Show menu with options: Edit current or Upload new
                menu = QMenu(self)
                menu.setStyleSheet("""
                    QMenu {
                        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                   stop:0 #2d3748, stop:1 #1a202c);
                        border: 1px solid rgba(74, 158, 255, 0.5);
                        border-radius: 8px;
                        padding: 8px;
                    }
                    QMenu::item {
                        color: white;
                        padding: 10px 20px;
                        border-radius: 6px;
                        margin: 2px 4px;
                    }
                    QMenu::item:selected {
                        background: rgba(74, 158, 255, 0.3);
                    }
                """)
                
                edit_action = menu.addAction("✏️ Edit Current Photo")
                upload_action = menu.addAction("📤 Upload New Photo")
                remove_action = menu.addAction("🗑️ Remove Photo")
                
                # Show menu at cursor position
                action = menu.exec(QCursor.pos())
                
                if action == edit_action:
                    self._edit_user_photo(user_photo_path)
                elif action == upload_action:
                    self._select_and_upload_user_photo()
                elif action == remove_action:
                    self._remove_user_photo()
            else:
                # No photo - go directly to upload
                self._select_and_upload_user_photo()
            
        except Exception as e:
            show_error_toast(self, f"Unable to manage photo:\n{str(e)}")
            logging.error(f"User photo error: {e}")

    def _select_and_upload_user_photo(self):
        """Select a new photo file and open editor"""
        try:
            fn, _ = QFileDialog.getOpenFileName(
                self,
                "📷 Select Your Photo",
                "",
                "Images (*.png *.jpg *.jpeg *.bmp)"
            )
            
            if not fn:
                return
            
            if not os.path.exists(fn):
                show_warning_toast(
                    self, f"The selected file could not be found:\n{fn}\n\n"
                    "The file may have been moved, deleted, or is on a disconnected network drive.\n\n"
                    "Please select a valid image file and try again."
                )
                return

            pix = QPixmap(fn)
            if pix.isNull():
                show_warning_toast(
                    self, "Failed to load the selected image.\n\n"
                    "Possible causes:\n"
                    "• File is corrupt or damaged\n"
                    "• File format is not supported\n"
                    "• File is not an image\n\n"
                    "Supported formats: PNG, JPG, JPEG, BMP, GIF\n\n"
                    "Please select a valid image file and try again."
                )
                return

            file_size = os.path.getsize(fn) / (1024 * 1024)
            if file_size > 5:
                show_warning_toast(
                    self, f"The selected image is {file_size:.1f}MB, which exceeds the maximum allowed size of 5MB.\n\n"
                    "To reduce file size:\n"
                    "1. Use an image editor to resize the image\n"
                    "2. Compress the image quality\n"
                    "3. Convert to JPG format\n\n"
                    "Recommended image size: 800x800 pixels or smaller"
                )
                return
            
            # Open photo editor for new photo
            self._edit_user_photo(fn)
            
        except Exception as e:
            show_error_toast(
                self, f"Unable to upload photo:\n{str(e)}\n\n"
                "Please check your system and try again."
            )
            logging.error(f"User photo upload error: {e}")

    def _remove_user_photo(self):
        """Remove the current user's profile photo"""
        try:
            user_photo_path = os.path.join(PHOTOS_DIR, f"user_{self.current_user}.png")
            if os.path.exists(user_photo_path):
                os.remove(user_photo_path)
                self._load_user_photo()  # Reload to show placeholder
                show_success_toast(self, "✅ Photo removed successfully!")
                logging.info(f"User photo removed for {self.current_user}")
        except Exception as e:
            show_error_toast(self, f"Unable to remove photo:\n{str(e)}")
            logging.error(f"User photo removal error: {e}")

    def _edit_user_photo(self, photo_path: str):
        """Open photo editor dialog for user profile photo with optional background removal"""
        try:
            from employee_vault.ui.dialogs.photo_editor import PhotoEditorDialog
            
            # Show remove background option for user profile photos
            dialog = PhotoEditorDialog(image_path=photo_path, parent=self, show_remove_bg=True)
            if dialog.exec() == QDialog.Accepted:
                result = dialog.get_result()
                if result and not result.isNull():
                    # Save as user photo
                    dest_path = os.path.join(PHOTOS_DIR, f"user_{self.current_user}.png")
                    result.save(dest_path, "PNG")
                    
                    # Remove background if requested
                    if dialog.should_remove_background():
                        try:
                            from employee_vault.utils import remove_background
                            remove_background(dest_path, dest_path)
                            logging.info(f"Background removed from user photo for {self.current_user}")
                            show_success_toast(self, "✅ Photo updated with background removed!")
                        except Exception as bg_err:
                            logging.warning(f"Background removal failed: {bg_err}")
                            show_success_toast(self, "✅ Photo updated! (Background removal unavailable)")
                    else:
                        show_success_toast(self, "✅ Photo updated successfully!")
                    
                    # Reload photo
                    self._load_user_photo()
                    logging.info(f"User photo updated for {self.current_user}")
        except Exception as e:
            show_error_toast(
                self, f"Unable to edit photo:\n{str(e)}\n\n"
                "Please try again."
            )
            logging.error(f"User photo edit error: {e}")

    def _logout(self):
        # Log logout action
        self.db.log_action(username=self.current_user, action="LOGOUT", details=f"User logged out")

        # Clean up session
        try:
            self.db.end_session(self.current_user)
        except Exception as e:
            logging.warning(f"Could not end session: {e}")

        # Stop all timers
        if hasattr(self, 'session_timer'):
            self.session_timer.stop()
        if hasattr(self, 'idle_timer'):
            self.idle_timer.stop()

        # Create fade-out animation for main window using window opacity
        from PySide6.QtCore import QPropertyAnimation
        from employee_vault.animation_manager import AnimationManager

        anim_manager = AnimationManager()

        # Store animation reference to prevent garbage collection
        self._logout_animation = QPropertyAnimation(self, b"windowOpacity")
        self._logout_animation.setDuration(anim_manager.get_theme_duration("transition"))
        self._logout_animation.setStartValue(1.0)
        self._logout_animation.setEndValue(0.0)
        self._logout_animation.setEasingCurve(anim_manager.get_theme_easing())

        def on_fade_out_finished():
            # Hide main window after fade-out
            self.hide()
            # Reset opacity for next time
            self.setWindowOpacity(1.0)

            # Use QTimer.singleShot to defer dialog creation/show
            # This ensures event loop processes properly before showing login
            def show_login_dialog():
                from employee_vault.ui.dialogs.login import LoginDialog
                login = LoginDialog(self.db)

                # Show dialog - hover animation will work normally
                if login.exec() == QDialog.Accepted:
                    # User logged in again, create new main window
                    u = login.username.text().strip()
                    row = self.db.get_user(u)
                    self._relogging_in = True
                    self.close()

                    # Create new main window
                    new_window = MainWindow(self.db, u, row)

                    # Show immediately without fade animation to avoid flickering
                    new_window.show()

                    # Store reference to prevent garbage collection
                    QApplication.instance().main_window = new_window
                else:
                    # User cancelled login, close app completely
                    self._should_quit = True
                    self.close()
                    QApplication.instance().quit()
            
            # Defer to next event loop iteration
            QTimer.singleShot(50, show_login_dialog)

        self._logout_animation.finished.connect(on_fade_out_finished)
        self._logout_animation.start()

    def closeEvent(self, event):
        """Handle window close button - show confirmation dialog"""
        from PySide6.QtWidgets import QMessageBox

        # If we're re-logging in, just accept
        if hasattr(self, '_relogging_in') and self._relogging_in:
            event.accept()
            return

        # If we already decided to quit, just accept
        if hasattr(self, '_should_quit') and self._should_quit:
            event.accept()
            return

        # Prevent duplicate close events
        if hasattr(self, '_closing') and self._closing:
            event.ignore()
            return

        # Show confirmation dialog
        msg = QMessageBox(self)
        msg.setWindowTitle("Close Application")
        msg.setText("What would you like to do?")
        msg.setIcon(QMessageBox.Question)

        logout_btn = msg.addButton("Back to Login", QMessageBox.ActionRole)
        exit_btn = msg.addButton("Exit Program", QMessageBox.DestructiveRole)
        cancel_btn = msg.addButton("Cancel", QMessageBox.RejectRole)

        msg.exec()
        clicked = msg.clickedButton()

        if clicked == logout_btn:
            # Mark that we're closing
            self._closing = True
            event.ignore()
            # Go back to login
            self._logout()
        elif clicked == exit_btn:
            # Exit program completely - checkpoint database before closing
            try:
                if hasattr(self, 'db') and self.db:
                    logging.info("Checkpointing database before exit...")
                    self.db.checkpoint_database()
                    self.db.close()
                    logging.info("Database closed with checkpoint")
            except Exception as e:
                logging.error(f"Error during database cleanup: {e}")
            self._should_quit = True
            event.accept()
            QApplication.instance().quit()
        else:
            # Cancel - don't close
            event.ignore()

    def changeEvent(self, event):
        """Handle window state changes - auto-refresh when window gains focus"""
        from PySide6.QtCore import QEvent
        if event.type() == QEvent.Type.ActivationChange:
            if self.isActiveWindow():
                # Window gained focus - check for database changes
                try:
                    current_mtime = db_latest_mtime(DB_FILE)
                    if hasattr(self, '_last_db_mtime') and current_mtime > self._last_db_mtime:
                        self._last_db_mtime = current_mtime
                        self._refresh_all()
                        logging.debug("Focus-refresh: Database updated while window was inactive")
                except Exception as e:
                    logging.debug(f"Focus-refresh check failed: {e}")
        super().changeEvent(event)

    def _auto_refresh(self):
        """
        Enhanced auto-refresh for real-time multi-user synchronization.

        v2.1.2 MULTI-USER IMPROVEMENTS:
        - Optimized for 5-7 concurrent users
        - Detects database changes before refreshing (performance boost)
        - Better handling of concurrent writes
        - Notifies user when data changes from other users
        - Auto-refresh on window focus (v5.1)
        """
        try:
            # Check if database has been modified by another user
            current_mtime = db_latest_mtime(DB_FILE)
            if not hasattr(self, '_last_db_mtime'):
                self._last_db_mtime = current_mtime

            # Only refresh if database actually changed (performance optimization)
            if current_mtime > self._last_db_mtime:
                self._last_db_mtime = current_mtime

                # Store current selection to restore after refresh
                current_page_index = self.stack.currentIndex()

                # Refresh data
                self._refresh_all()

                # Restore page selection
                self.stack.setCurrentIndex(current_page_index)

                logging.debug("Auto-refresh: Database updated by another user")

                # Optional: Show subtle notification
                if hasattr(self, 'statusBar'):
                    self.statusBar().showMessage("Data updated", 2000)
            else:
                logging.debug("Auto-refresh: No changes detected (skipped refresh)")

        except sqlite3.OperationalError as e:
            if "locked" in str(e).lower():
                # Database locked is normal during write operations
                # Use exponential backoff - wait longer on next attempt
                logging.debug("Database locked during refresh - will retry next cycle")
            else:
                logging.error(f"Auto-refresh database error: {e}")
        except Exception as e:
            # Unexpected errors should be logged with full traceback
            logging.error(f"Unexpected error in auto-refresh: {e}", exc_info=True)

    def _apply_user_permissions(self):
        """Apply UI restrictions based on user permissions - v2.0 FIXED"""
        perms = self.user_permissions

        # Debug logging
        logging.info(f"Applying permissions for user: {self.current_user}")
        logging.info(f"User role: {self.user_row.get('role', 'UNKNOWN')}")
        logging.info(f"User permissions: {perms}")

        # If admin, skip restrictions (admin has all permissions)
        if self.user_row.get('role') == 'admin':
            logging.info("Admin user - all permissions granted")
            return

        logging.info(f"Non-admin user - applying permission restrictions")

        # Note: Most UI elements are already conditionally shown based on role
        # This method provides fine-grained control for non-admin users

        # The sidebar sections already check user_row.get('role') and self.user_permissions
        # for conditional display, so permissions are already being enforced

        # Additional runtime checks can be added here for dynamic UI updates
        # For now, permissions are enforced at creation time (in __init__)

        logging.info(f"Permissions applied successfully for {self.current_user}")

    # ============================================================================
    # WEEK 2 FEATURE #1: SESSION IDLE TIMEOUT (AUTO-LOCK)
    # ============================================================================

    def eventFilter(self, obj, event):
        """Track user activity for idle timeout and handle sidebar hover + background resize"""
        # Track mouse and keyboard events as activity
        if event.type() in [QEvent.MouseMove, QEvent.MouseButtonPress,
                           QEvent.KeyPress, QEvent.Wheel]:
            self.last_activity_time = datetime.now()
        
        # Resize animated background when content area resizes
        if hasattr(self, 'content_area') and obj == self.content_area and event.type() == QEvent.Resize:
            if hasattr(self, 'animated_background'):
                self.animated_background.setGeometry(self.content_area.rect())
                
        return super().eventFilter(obj, event)

    def _periodic_checkpoint(self):
        """Periodically checkpoint WAL for multi-PC synchronization (runs in background thread)"""
        try:
            if hasattr(self, 'db') and self.db:
                # Use background worker to prevent UI freezes
                if not hasattr(self, '_checkpoint_worker') or not self._checkpoint_worker.isRunning():
                    self._checkpoint_worker = CheckpointWorker(self.db, self)
                    self._checkpoint_worker.start()
        except Exception as e:
            logging.warning(f"Periodic checkpoint failed: {e}")

    def _check_idle_timeout(self):
        """Check if user has been idle and auto-close program after 30 minutes"""
        idle_minutes = (datetime.now() - self.last_activity_time).total_seconds() / 60

        # Warning at 5 minutes before timeout - show countdown dialog
        if idle_minutes >= (self.idle_timeout_minutes - 5) and idle_minutes < self.idle_timeout_minutes:
            if not hasattr(self, '_warning_shown') or not self._warning_shown:
                self._warning_shown = True
                self._show_idle_warning_dialog()

        # Auto-close program after timeout
        if idle_minutes >= self.idle_timeout_minutes:
            # Only auto-close if warning dialog is not active (user didn't extend)
            if not hasattr(self, '_idle_warning_active') or not self._idle_warning_active:
                logging.info(f"Auto-close: User {self.current_user} idle for {idle_minutes:.1f} minutes - closing program")
                self._auto_close_program()

    def _show_idle_warning_dialog(self):
        """Show a countdown warning dialog before session timeout with extend option"""
        self._idle_warning_active = True
        
        # Create warning dialog
        dialog = QDialog(self)
        dialog.setWindowTitle("Session Timeout Warning")
        dialog.setModal(True)
        dialog.setFixedSize(400, 200)
        dialog.setWindowFlags(dialog.windowFlags() | Qt.WindowStaysOnTopHint)
        
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)
        
        # Warning icon and message
        warning_label = QLabel("⚠️ Session Timeout Warning")
        warning_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #FFA726;")
        warning_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(warning_label)
        
        # Countdown label
        remaining_seconds = int((self.idle_timeout_minutes * 60) - (datetime.now() - self.last_activity_time).total_seconds())
        countdown_label = QLabel(f"Program will close in {remaining_seconds} seconds due to inactivity")
        countdown_label.setStyleSheet("font-size: 14px; color: white;")
        countdown_label.setAlignment(Qt.AlignCenter)
        countdown_label.setWordWrap(True)
        layout.addWidget(countdown_label)
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(12)
        
        extend_btn = ModernAnimatedButton("🔄 Extend Session (15 min)")
        extend_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(76, 175, 80, 0.9),
                                           stop:1 rgba(56, 142, 60, 0.7));
                border: 2px solid rgba(76, 175, 80, 0.8);
                border-radius: 16px;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(102, 187, 106, 1.0),
                                           stop:1 rgba(76, 175, 80, 0.85));
            }
        """)
        
        logout_btn = ModernAnimatedButton("🚪 Logout Now")
        logout_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(244, 67, 54, 0.9),
                                           stop:1 rgba(211, 47, 47, 0.7));
                border: 2px solid rgba(244, 67, 54, 0.8);
                border-radius: 16px;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(239, 83, 80, 1.0),
                                           stop:1 rgba(244, 67, 54, 0.85));
            }
        """)
        
        btn_layout.addWidget(extend_btn)
        btn_layout.addWidget(logout_btn)
        layout.addLayout(btn_layout)
        
        # Countdown timer
        countdown_timer = QTimer(dialog)
        
        def update_countdown():
            nonlocal remaining_seconds
            remaining_seconds -= 1
            if remaining_seconds > 0:
                countdown_label.setText(f"Program will close in {remaining_seconds} seconds due to inactivity")
            else:
                countdown_timer.stop()
                dialog.reject()
                self._idle_warning_active = False
                self._auto_close_program()
        
        countdown_timer.timeout.connect(update_countdown)
        countdown_timer.start(1000)  # Update every second
        
        def extend_session():
            countdown_timer.stop()
            self.last_activity_time = datetime.now()
            self._warning_shown = False
            self._idle_warning_active = False
            dialog.accept()
            show_success_toast(self, "Session extended by 15 minutes")
            logging.info(f"Session extended by user {self.current_user}")
        
        def logout_now():
            countdown_timer.stop()
            dialog.reject()
            self._idle_warning_active = False
            self._auto_close_program()
        
        extend_btn.clicked.connect(extend_session)
        logout_btn.clicked.connect(logout_now)
        
        # Apply dialog styling
        dialog.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(40, 40, 60, 0.98),
                                           stop:1 rgba(30, 30, 50, 0.95));
                border: 2px solid rgba(255, 167, 38, 0.6);
                border-radius: 16px;
            }
        """)
        
        dialog.exec()

    def _auto_close_program(self):
        """Auto-close the program due to inactivity"""
        from PySide6.QtWidgets import QApplication
        
        # Stop all timers
        self.idle_timer.stop()
        if hasattr(self, 'timer'):
            self.timer.stop()
        if hasattr(self, 'session_timer'):
            self.session_timer.stop()
        if hasattr(self, 'checkpoint_timer'):
            self.checkpoint_timer.stop()

        # Log the auto-close
        try:
            self.db.log_action(
                username=self.current_user, 
                action="AUTO_LOGOUT", 
                details=f"Program auto-closed after {self.idle_timeout_minutes} minutes of inactivity"
            )
        except:
            pass

        # Close database properly with checkpoint
        try:
            if hasattr(self, 'db') and self.db:
                logging.info("Checkpointing database before auto-close...")
                self.db.checkpoint_database()
                self.db.close()
                logging.info("Database closed with checkpoint")
        except Exception as e:
            logging.error(f"Error during database cleanup: {e}")

        # Close the application
        logging.info("Program auto-closed due to inactivity")
        self._should_quit = True
        QApplication.instance().quit()

    def _lock_session(self):
        """Lock the session and require re-authentication"""
        # Stop timers
        self.idle_timer.stop()
        if hasattr(self, 'timer'):
            self.timer.stop()

        # Hide main window
        self.hide()

        # Show login dialog for re-authentication
        login = LoginDialog(self.db)
        login.setWindowTitle("Session Locked - Re-authenticate")
        login.info.setText(f"Session locked due to inactivity.\nPlease re-authenticate as: {self.current_user}")
        login.username.setText(self.current_user)
        login.username.setReadOnly(True)
        login.password.setFocus()

        if login.exec() == QDialog.Accepted:
            # Re-authenticated successfully
            self.last_activity_time = datetime.now()
            self._warning_shown = False
            self.idle_timer.start()
            if hasattr(self, 'timer'):
                self.timer.start()
            self.show()
            logging.info(f"Session unlocked: {self.current_user} re-authenticated")
        else:
            # User cancelled - logout
            self._logout()

    def _notify_contracts(self):
        expired=[e for e in self.employees if (contract_days_left(e) or 999999) < 0]
        soon=[e for e in self.employees if 0 <= (contract_days_left(e) or 999999) <= ALERT_DAYS]
        if expired: self.tray.showMessage("Contract expired", f"{len(expired)} contract(s) expired.", QSystemTrayIcon.Information, 8000)
        if soon: self.tray.showMessage("Contract expiring", f"{len(soon)} contract(s) will expire ≤ {ALERT_DAYS} days.", QSystemTrayIcon.Information, 8000)
        
        # v5.3: Update notification center
        if hasattr(self, 'notification_center'):
            self.notification_center.check_contract_expiry()
            self._update_notification_badge()
