"""
Employee Form Widget
"""

import os
import shutil
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any

from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *

from employee_vault.config import *
from employee_vault.database import DB
from employee_vault.utils import remove_background
from employee_vault.validators import *
from employee_vault.models import *
from employee_vault.ui.widgets import *
from employee_vault.ui.widgets.animated_input import NeumorphicGradientLineEdit
from employee_vault.ui.modern_ui_helper import show_success_toast, show_error_toast, show_warning_toast, show_info_toast
from employee_vault.ui.widgets import disable_cursor_changes
from employee_vault.ui.widgets import ModernAnimatedButton, PulseButton, titlecase, normalize_ph_phone
from employee_vault.ui.ios_button_styles import apply_ios_style

# Import fuzzy string matching for duplicate name detection
try:
    from rapidfuzz import fuzz
except ImportError:
    try:
        from fuzzywuzzy import fuzz
    except ImportError:
        # Fallback: create a simple fuzz module with basic ratio function
        class SimpleFuzz:
            @staticmethod
            def ratio(s1, s2):
                """Simple string similarity (exact match or length-based)"""
                if s1 == s2:
                    return 100
                # Basic similarity based on common characters
                s1_lower = s1.lower()
                s2_lower = s2.lower()
                if s1_lower == s2_lower:
                    return 100
                # Simple length-based similarity
                max_len = max(len(s1), len(s2))
                if max_len == 0:
                    return 100
                common = sum(1 for a, b in zip(s1_lower, s2_lower) if a == b)
                return int((common / max_len) * 100)

            @staticmethod
            def token_sort_ratio(s1, s2):
                """Token sort ratio - sorts words before comparing"""
                words1 = sorted(s1.lower().split())
                words2 = sorted(s2.lower().split())
                return SimpleFuzz.ratio(' '.join(words1), ' '.join(words2))

        fuzz = SimpleFuzz()


# ============================================================================
# BACKGROUND WORKER FOR PHOTO PROCESSING
# ============================================================================
class PhotoProcessingWorker(QThread):
    """Background thread worker for processing photos (background removal)"""
    finished = Signal(str, bool)  # (result_path, success)
    progress = Signal(str)  # status message
    error = Signal(str)  # error message
    
    def __init__(self, source_path, parent=None):
        super().__init__(parent)
        self.source_path = source_path
        self.result_path = None
        self._cancelled = False
    
    def cancel(self):
        self._cancelled = True
    
    def run(self):
        try:
            if self._cancelled:
                return
            
            import tempfile
            import shutil
            
            self.progress.emit("Preparing image...")
            
            # Create temp file for background-removed version
            temp_dir = tempfile.mkdtemp()
            temp_bg_removed = os.path.join(temp_dir, "bg_removed_" + os.path.splitext(os.path.basename(self.source_path))[0] + ".png")
            shutil.copy2(self.source_path, temp_bg_removed)
            
            if self._cancelled:
                shutil.rmtree(temp_dir, ignore_errors=True)
                return
            
            self.progress.emit("Removing background...")
            
            # Remove background
            logging.info(f"Removing background from employee photo: {self.source_path}")
            remove_background(temp_bg_removed, temp_bg_removed)
            logging.info(f"Background removed successfully")
            
            if self._cancelled:
                shutil.rmtree(temp_dir, ignore_errors=True)
                return
            
            self.result_path = temp_bg_removed
            self.finished.emit(temp_bg_removed, True)
            
        except Exception as e:
            logging.warning(f"Background removal failed: {e}")
            self.error.emit(str(e))
            self.finished.emit(self.source_path, False)  # Return original on failure


# ============================================================================
# DRAFT AUTO-SAVE CONSTANTS
# ============================================================================
DRAFT_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), "employee_draft.json")


class EmployeeForm(QWidget):
    # Agencies helpers
    def _reload_agencies_into_combo(self):
        self.agency_combo.blockSignals(True)  # Prevent signals during update
        self.agency_combo.clear()
        self.agency_combo.addItem("— Select agency —", None)
        for a in self.db.get_agencies():
            self.agency_combo.addItem(a, a)
        self.agency_combo.blockSignals(False)  # Re-enable signals
    def _add_agency_dialog(self):
        text, ok = QInputDialog.getText(self, "Add Agency", "Agency name:")
        if not ok:
            return
        name = titlecase((text or "").strip())
        if not name:
            return
        try:
            self.db.add_agency(name)
        except Exception as e:
            show_error_toast(
                self, f"Unable to add agency '{name}':\n{e}\n\n"
                "Possible causes:\n"
                "• Agency name already exists\n"
                "• Database is locked by another user\n"
                "• Insufficient permissions\n\n"
                "Please try again or contact your administrator."
            )
            return
        self._reload_agencies_into_combo()
        idx = self.agency_combo.findText(name, Qt.MatchFixedString)
        if idx >= 0:
            self.agency_combo.setCurrentIndex(idx)

    def _on_department_changed(self, dept):
        """Handle department selection, show store picker if Store selected"""
        if dept == "Store":
            # Show store picker dialog
            store = self._show_store_picker()
            if store:
                # Set full department as "Store - [Store Name]" (with dash)
                full_dept = f"Store - {store}"
                self.entries["department"].setText(full_dept)
                self.department_display.setText(f"✓ {full_dept}")
                self.department_display.show()
            else:
                # User cancelled, reset to default
                self.department_combo.combo_box.setCurrentIndex(0)
                self.entries["department"].setText("")
                self.department_display.hide()
        elif dept == "— Select Department —":
            self.entries["department"].setText("")
            self.department_display.hide()
        else:
            # Office or Warehouse
            self.entries["department"].setText(dept)
            self.department_display.setText(f"✓ {dept}")
            self.department_display.show()

    def _show_store_picker(self):
        """Show dialog to pick store location with Excel import"""
        # Load stores from settings or use defaults
        stores = self._load_stores_list()

        # v4.4.1: Use animated dialog for store picker
        from employee_vault.ui.widgets import AnimatedDialogBase
        dlg = AnimatedDialogBase(self, animation_style="fade")
        dlg.setWindowTitle("Select Store")
        dlg.resize(500, 600)

        layout = QVBoxLayout(dlg)
        layout.addWidget(QLabel("<h3>📍 Select Store Location</h3>"))
        layout.addWidget(QLabel("Choose the store where this employee works:"))

        # Search box - NeumorphicGradientLineEdit for login-style look
        search_box = NeumorphicGradientLineEdit("Search stores...")
        layout.addWidget(search_box)

        # Store list with context menu for deletion
        store_list = QListWidget()
        disable_cursor_changes(store_list)  # Remove hand cursor
        store_list.addItems(stores)
        store_list.setCurrentRow(0)
        store_list.setContextMenuPolicy(Qt.CustomContextMenu)

        # Add context menu for deleting stores
        def show_store_context_menu(pos):
            item = store_list.itemAt(pos)
            if not item:
                return

            menu = QMenu(store_list)
            delete_action = menu.addAction("🗑️ Delete Store")

            action = menu.exec(store_list.mapToGlobal(pos))
            if action == delete_action:
                # Confirm deletion
                reply = QMessageBox.question(
                    dlg,
                    "Confirm Delete",
                    f"Delete '{item.text()}' from the store list?\n\nThis will remove it from future selections.",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )

                if reply == QMessageBox.StandardButton.Yes:
                    # Remove from list widget
                    row = store_list.row(item)
                    store_list.takeItem(row)

                    # Update stores list and save to database
                    updated_stores = [store_list.item(i).text() for i in range(store_list.count())]
                    self._save_stores_list(updated_stores)
                    show_success_toast(self, f"Store '{item.text()}' deleted successfully!")

        store_list.customContextMenuRequested.connect(show_store_context_menu)
        layout.addWidget(store_list)

        # Search functionality
        def filter_stores():
            search_text = search_box.text().lower()
            for i in range(store_list.count()):
                item = store_list.item(i)
                item.setHidden(search_text not in item.text().lower())

        search_box.textChanged.connect(filter_stores)

        # Custom store input - NeumorphicGradientLineEdit for login-style look
        custom_layout = QVBoxLayout()
        custom_layout.addWidget(QLabel("Or enter custom store name:"))
        custom_store = NeumorphicGradientLineEdit("Enter store name...")
        custom_layout.addWidget(custom_store)
        layout.addLayout(custom_layout)

        # Import/Export/Delete buttons - Phase 3: iOS frosted glass
        import_export_layout = QHBoxLayout()
        btn_import = ModernAnimatedButton("📥 Import")
        apply_ios_style(btn_import, 'blue')
        btn_export = ModernAnimatedButton("📤 Export")
        apply_ios_style(btn_export, 'green')
        btn_delete = ModernAnimatedButton("🗑️ Delete")
        apply_ios_style(btn_delete, 'red')
        btn_delete.setToolTip("Delete selected store")

        def delete_selected_store():
            current_item = store_list.currentItem()
            if not current_item:
                show_info_toast(self, "Please select a store to delete.")
                return

            # Confirm deletion
            reply = QMessageBox.question(
                dlg,
                "Confirm Delete",
                f"Delete '{current_item.text()}' from the store list?\n\nThis will remove it from future selections.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )

            if reply == QMessageBox.StandardButton.Yes:
                # Remove from list widget
                row = store_list.row(current_item)
                store_list.takeItem(row)

                # Update stores list and save to database
                updated_stores = [store_list.item(i).text() for i in range(store_list.count())]
                self._save_stores_list(updated_stores)
                show_success_toast(self, f"Store '{current_item.text()}' deleted!")

        btn_import.clicked.connect(lambda: self._import_stores(store_list))
        btn_export.clicked.connect(lambda: self._export_stores(stores))
        btn_delete.clicked.connect(delete_selected_store)
        import_export_layout.addWidget(btn_import)
        import_export_layout.addWidget(btn_export)
        import_export_layout.addWidget(btn_delete)
        layout.addLayout(import_export_layout)

        # Buttons - Phase 3: iOS frosted glass
        btn_layout = QHBoxLayout()
        btn_ok = ModernAnimatedButton("✓ Select")
        apply_ios_style(btn_ok, 'green')
        btn_cancel = ModernAnimatedButton("✗ Cancel")
        apply_ios_style(btn_cancel, 'gray')
        btn_ok.clicked.connect(dlg.accept)
        btn_cancel.clicked.connect(dlg.reject)
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        layout.addLayout(btn_layout)

        if dlg.exec() == QDialog.Accepted:
            # Check custom input first
            custom = custom_store.text().strip()
            if custom:
                return custom

            # Then check list selection
            if store_list.currentItem():
                selected = store_list.currentItem().text()
                if selected != "Other":
                    return selected

        return None

    def _load_stores_list(self):
        """Load stores from settings or return defaults"""
        try:
            result = self.db.conn.execute("SELECT value FROM settings WHERE key='stores_list'").fetchone()
            if result:
                import json
                return json.loads(result[0])
        except:
            pass

        # Default stores
        return [

        ]

    def _save_stores_list(self, stores):
        """Save stores list to settings"""
        try:
            import json
            stores_json = json.dumps(stores)
            self.db.conn.execute("INSERT OR REPLACE INTO settings(key, value) VALUES('stores_list', ?)", (stores_json,))
            self.db.conn.commit()
        except Exception as e:
            show_warning_toast(
                self, f"Unable to save stores list to database:\n{e}\n\n"
                "Your stores list has been updated in the form,\n"
                "but the changes were not saved to the database.\n\n"
                "The changes will be lost when you close the application."
            )

    def _import_stores(self, store_list):
        """Import stores from Excel file"""
        from openpyxl import load_workbook

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Stores from Excel",
            "",
            "Excel Files (*.xlsx *.xls)"
        )

        if not file_path:
            return

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            new_stores = []
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                if row[0]:  # If first column has value
                    store_name = str(row[0]).strip()
                    if store_name and store_name not in new_stores:
                        new_stores.append(store_name)

            if new_stores:
                # Add to list
                for store in new_stores:
                    if not any(store_list.item(i).text() == store for i in range(store_list.count())):
                        store_list.addItem(store)

                # Save to settings
                all_stores = [store_list.item(i).text() for i in range(store_list.count())]
                self._save_stores_list(all_stores)

                show_success_toast(self, f"Imported {len(new_stores)} stores!")
            else:
                show_warning_toast(self, "No stores found in Excel file.")

        except Exception as e:
            show_error_toast(
                self, f"Unable to import stores from Excel file:\n{e}\n\n"
                "Possible causes:\n"
                "• Excel file format is invalid or corrupt\n"
                "• File is currently open in Excel\n"
                "• Missing required Python package (openpyxl)\n"
                "• File contains no data\n\n"
                "Please check the file format and try again.\n"
                "Expected format: First column should contain store names, starting from row 2."
            )

    def _export_stores(self, stores):
        """Export stores to Excel file"""
        from openpyxl import Workbook

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Stores to Excel",
            "stores.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Stores"

            # Header
            ws.append(["Store Name"])

            # Data
            for store in stores:
                ws.append([store])

            wb.save(file_path)
            show_success_toast(self, f"Exported {len(stores)} stores to {file_path}")

        except Exception as e:
            show_error_toast(
                self, f"Unable to export stores to Excel file:\n{e}\n\n"
                "Possible causes:\n"
                "• Destination folder does not exist or is read-only\n"
                "• Insufficient disk space\n"
                "• Missing required Python package (openpyxl)\n"
                "• File path contains invalid characters\n\n"
                "Please check the destination and try again."
            )

    def validate_inputs(self):
        errors = []

        if not self.name_edit.text().strip():
            errors.append("Name is required")

        if not self.emp_id_edit.text().strip():
            errors.append("Employee ID is required")

        email = self.email_edit.text().strip()
        if email and not EMAIL_RE.match(email):
            errors.append("Invalid email format")

        try:
            salary = float(self.salary_edit.text() or 0)
            if salary < 0:
                errors.append("Salary cannot be negative")
        except ValueError:
            errors.append("Invalid salary")

        # Check for duplicate government IDs
        duplicate_errors = self._check_duplicate_government_ids()
        errors.extend(duplicate_errors)

        return errors

    def _check_duplicate_government_ids(self):
        """
        Check for duplicate government IDs (SSS, TIN, PhilHealth, Pag-IBIG)
        Returns list of error messages if duplicates found
        """
        errors = []

        # Get all government ID fields and their values
        gov_ids = {
            'SSS Number': self.sss_edit.text().strip(),
            'TIN Number': self.tin_edit.text().strip(),
            'PhilHealth Number': self.philhealth_edit.text().strip(),
            'Pag-IBIG Number': self.pagibig_edit.text().strip()
        }

        # Check each government ID for duplicates
        for id_type, id_value in gov_ids.items():
            if not id_value:  # Skip empty values
                continue

            # Get the column name for database query
            column_map = {
                'SSS Number': 'sss_number',
                'TIN Number': 'tin_number',
                'PhilHealth Number': 'philhealth_number',
                'Pag-IBIG Number': 'pagibig_number'
            }
            column_name = column_map[id_type]

            # Query database for duplicate
            try:
                existing = self.db.conn.execute(
                    f"SELECT emp_id, name FROM employees WHERE {column_name} = ?",
                    (id_value,)
                ).fetchone()

                if existing:
                    # If editing, check if it's the same employee
                    if self.current_employee and existing['emp_id'] == self.current_employee.get('emp_id'):
                        continue  # Same employee, not a duplicate

                    # Found duplicate
                    errors.append(
                        f"❌ Duplicate {id_type}: '{id_value}' is already assigned to "
                        f"{existing['name']} (ID: {existing['emp_id']})"
                    )
            except Exception as e:
                logging.error(f"Error checking duplicate {id_type}: {e}")

        return errors

    def _check_duplicate_name(self, full_name: str) -> List[Dict[str, Any]]:
        """
        Check for duplicate or similar employee names using fuzzy matching

        Args:
            full_name: Full name to check (e.g., "John Doe")

        Returns:
            List of similar employees with similarity scores
        """
        # Get all employees
        all_employees = self.db.all_employees()

        # Find similar names
        similar = []
        for emp in all_employees:
            # Skip if editing the same employee
            if self.current_employee and emp.get('emp_id') == self.current_employee.get('emp_id'):
                continue

            emp_name = emp.get('name', '')
            if not emp_name:
                continue

            # Calculate similarity ratio (0-100)
            ratio = fuzz.ratio(full_name.lower(), emp_name.lower())

            # Also check token sort ratio (ignores word order)
            token_ratio = fuzz.token_sort_ratio(full_name.lower(), emp_name.lower())

            # Use the higher of the two
            best_ratio = max(ratio, token_ratio)

            # If similarity is 85% or higher, it's a potential duplicate
            if best_ratio >= 85:
                similar.append({
                    'employee': emp,
                    'similarity': best_ratio,
                    'name': emp_name
                })

        # Sort by similarity (highest first)
        similar.sort(key=lambda x: x['similarity'], reverse=True)

        return similar

    def _show_duplicate_warning(self, full_name: str, similar_employees: List[Dict[str, Any]]) -> bool:
        """
        Show warning dialog for duplicate employee names

        Args:
            full_name: Name that was entered
            similar_employees: List of similar employees

        Returns:
            True if user wants to proceed, False if cancelled
        """
        from employee_vault.ui.widgets import AnimatedDialogBase

        dlg = AnimatedDialogBase(self, animation_style="fade")
        dlg.setWindowTitle("⚠️ Possible Duplicate Detected")
        dlg.resize(650, 500)

        layout = QVBoxLayout(dlg)

        # Warning header
        header = QLabel("<h2>⚠️ Similar Employee Names Found</h2>")
        header.setStyleSheet("color: #ffaa00; padding: 10px;")
        layout.addWidget(header)

        # Message
        msg = QLabel(
            f"<p>You are trying to add/update an employee with the name:</p>"
            f"<p style='font-weight: bold; font-size: 14px; color: #4a9eff;'>{full_name}</p>"
            f"<p>The following similar names were found in the database:</p>"
        )
        msg.setWordWrap(True)
        layout.addWidget(msg)

        # List of similar employees
        list_widget = QListWidget()
        list_widget.setStyleSheet("""
            QListWidget {
                background-color: rgba(45, 45, 48, 0.95);
                border: 1px solid rgba(255, 255, 255, 0.1);
                border-radius: 8px;
                padding: 8px;
                font-size: 13px;
            }
            QListWidget::item {
                padding: 12px;
                border-bottom: 1px solid rgba(255, 255, 255, 0.05);
                border-radius: 4px;
            }
            QListWidget::item:hover {
                background-color: rgba(74, 158, 255, 0.2);
            }
        """)

        for item in similar_employees[:5]:  # Show top 5 matches
            emp = item['employee']
            similarity = item['similarity']

            # Create detailed item text
            emp_id = emp.get('emp_id', 'N/A')
            dept = emp.get('department', 'N/A')
            status = 'Active' if not emp.get('resign_date') else 'Resigned'

            item_text = (
                f"👤 {emp.get('name', '')} ({emp_id})\n"
                f"   Department: {dept}  •  Status: {status}  •  Match: {similarity}%"
            )

            list_item = QListWidgetItem(item_text)

            # Color code by similarity
            if similarity >= 95:
                list_item.setForeground(QColor("#ff6b6b"))  # Red for very high match
            elif similarity >= 90:
                list_item.setForeground(QColor("#ffaa00"))  # Orange for high match
            else:
                list_item.setForeground(QColor("#ffdd66"))  # Yellow for medium match

            list_widget.addItem(list_item)

        layout.addWidget(list_widget)

        # Question
        question = QLabel(
            "<p style='margin-top: 10px;'>Do you want to proceed anyway?</p>"
        )
        question.setStyleSheet("color: #ffaa00; font-weight: bold;")
        layout.addWidget(question)

        # Buttons
        btn_layout = QHBoxLayout()

        btn_cancel = ModernAnimatedButton("✗ Cancel")
        apply_ios_style(btn_cancel, 'red')
        btn_cancel.clicked.connect(dlg.reject)

        btn_proceed = ModernAnimatedButton("✓ Proceed Anyway")
        apply_ios_style(btn_proceed, 'green')
        btn_proceed.clicked.connect(dlg.accept)

        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_proceed)
        layout.addLayout(btn_layout)

        # Return True if user clicked Proceed
        return dlg.exec() == QDialog.Accepted

    def _save(self):
        # Call validation before saving
        errors = self.validate_inputs()
        if errors:
            show_warning_toast(
                self, "\n".join(f"• {e}" for e in errors)
            )
            return

    def __init__(self, db, current_user, on_saved_callback):
        super().__init__()
        self.db=db
        self.on_saved=on_saved_callback
        # Ensure current_user is never None or empty
        self.current_user = current_user if current_user else "system"
        if not self.current_user or self.current_user == "?":
            self.current_user = "system"
        
        # v4.5.0: Get user role for permission checks (e.g., department edit restriction)
        self.user_role = "user"  # Default to user
        try:
            user_info = self.db.get_user(self.current_user)
            if user_info:
                self.user_role = user_info.get('role', 'user')
        except Exception:
            pass  # Keep default role
        
        self.current_employee=None
        self.entries={}
        self.date_entries={}
        self.files_list=None
        self.photo_label=None
        self.photo_path=None  # Initialize photo path to prevent cross-contamination bug
        self.still_working=None
        self.has_unsaved_changes=False  # Track if form has unsaved changes

        # Record locking for multi-user concurrent editing
        self.current_lock_id = None  # Track current record lock
        self.lock_refresh_timer = None  # Timer to refresh lock every 5 minutes

        self.agency_combo=QComboBox()
        # Add this exact line below the line above
        self.agency_combo.setFocusPolicy(Qt.ClickFocus)
        self.agency_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.agency_combo.setMinimumWidth(260)

        self.contract_user_changed=False
        self.resign_user_changed=False

        root=QVBoxLayout(self); root.addWidget(QLabel("<h2>➕ Add Employee</h2>"))
        top = QHBoxLayout()
        # Avatar column with modern solid gradient card
        photo_col=QVBoxLayout()
        
        # Photo container with accent border and glow - NO blue background when photo is set
        self.photo_label=QLabel("📷", alignment=Qt.AlignCenter)
        self.photo_label.setFixedSize(160,160)
        self.photo_label.setStyleSheet("""
            QLabel {
                border-radius: 80px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(60, 60, 80, 0.5),
                                           stop:1 rgba(40, 40, 60, 0.3));
                border: 4px solid rgba(100, 100, 120, 0.6);
                font-size: 56px;
                color: rgba(255, 255, 255, 0.5);
            }
            QLabel:hover {
                border: 4px solid rgba(100, 181, 246, 0.8);
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(70, 70, 90, 0.6),
                                           stop:1 rgba(50, 50, 70, 0.4));
            }
        """)
        # Make photo clickable to edit/view
        self.photo_label.mousePressEvent = lambda e: self._on_photo_clicked()
        
        # Modern solid gradient buttons with 22px pill shape
        uphoto=ModernAnimatedButton("📤 Upload Photo")
        uphoto.setFixedWidth(160)
        uphoto.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(33, 150, 243, 0.9),
                                           stop:1 rgba(25, 118, 210, 0.7));
                border: 2px solid rgba(33, 150, 243, 0.8);
                border-top: 2px solid rgba(255, 255, 255, 0.3);
                border-radius: 22px;
                color: white;
                font-weight: bold;
                padding: 10px 16px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(66, 165, 245, 1.0),
                                           stop:1 rgba(33, 150, 243, 0.85));
                border: 2px solid rgba(100, 181, 246, 1.0);
            }
        """)
        uphoto.clicked.connect(self._upload_photo)

        # Remove Photo button - red accent
        remove_photo=ModernAnimatedButton("🗑️ Remove Photo")
        remove_photo.setFixedWidth(160)
        remove_photo.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(244, 67, 54, 0.9),
                                           stop:1 rgba(211, 47, 47, 0.7));
                border: 2px solid rgba(244, 67, 54, 0.8);
                border-top: 2px solid rgba(255, 255, 255, 0.3);
                border-radius: 22px;
                color: white;
                font-weight: bold;
                padding: 10px 16px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(239, 83, 80, 1.0),
                                           stop:1 rgba(244, 67, 54, 0.85));
                border: 2px solid rgba(239, 83, 80, 1.0);
            }
        """)
        remove_photo.clicked.connect(self._remove_photo)

        # Background removal toggle checkbox (DISABLED - user request)
        self.remove_bg_checkbox = QCheckBox("Remove Background")
        self.remove_bg_checkbox.setChecked(False)  # Disabled by default
        self.remove_bg_checkbox.setVisible(False)  # Hidden from UI
        self.remove_bg_checkbox.setFixedWidth(160)
        self.remove_bg_checkbox.setStyleSheet("""
            QCheckBox {
                color: white;
                font-size: 12px;
                font-weight: 500;
                padding: 4px 8px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
                border-radius: 4px;
                border: 2px solid rgba(33, 150, 243, 0.8);
                background: rgba(0, 0, 0, 0.2);
            }
            QCheckBox::indicator:checked {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(33, 150, 243, 0.9),
                                           stop:1 rgba(25, 118, 210, 0.7));
                border: 2px solid rgba(33, 150, 243, 1.0);
            }
            QCheckBox::indicator:hover {
                border: 2px solid rgba(100, 181, 246, 1.0);
            }
        """)

        photo_col.addWidget(self.photo_label, alignment=Qt.AlignTop); photo_col.addSpacing(8)
        photo_col.addWidget(uphoto, alignment=Qt.AlignTop); photo_col.addSpacing(4)
        photo_col.addWidget(remove_photo, alignment=Qt.AlignTop); photo_col.addSpacing(4)
        photo_col.addWidget(self.remove_bg_checkbox, alignment=Qt.AlignTop); photo_col.addStretch(1)
        top.addLayout(photo_col); top.addSpacing(28)

        # Helper functions for creating field widgets
        def le(ph="", ro=False):
            # Use NeumorphicGradientLineEdit from animated_input to guarantee setCompleter is present
            e = NeumorphicGradientLineEdit(ph, float_y=7.0, rest_y=25.0, input_y=8)
            e.setReadOnly(ro)
            e.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            e.setMinimumWidth(260)  # Wider fields
            e.setMinimumHeight(70)  # Standard neumorphic height
            return e
        def de():
            # Use new theme-aware DatePicker with clickable field + popup calendar
            d = DatePicker()
            d.setFocusPolicy(Qt.ClickFocus)
            d.setDisplayFormat("MM-dd-yyyy")
            d.setDate(QDate.currentDate())
            d.setMinimumDate(QDate(1900,1,1))
            d.setMaximumDate(QDate(2100,12,31))
            d.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            d.setMinimumWidth(260)
            return d

        # Helper to add field row to a grid with required field markers
        def add_row(grid, r, labL, widL, labR, widR):
            grid.addWidget(QLabel(labL), r, 0, alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            grid.addWidget(widL, r, 1)
            grid.addWidget(QLabel(labR), r, 2, alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            grid.addWidget(widR, r, 3)

        # Create vertical layout for all form sections
        sections_layout = QVBoxLayout()
        sections_layout.setSpacing(16)

        # ============================================================
        # SECTION 1: PERSONAL INFORMATION
        # ============================================================
        personal_section = FormSection("PERSONAL INFORMATION", icon="📝", start_collapsed=False)
        personal_grid = QGridLayout()
        personal_grid.setHorizontalSpacing(12); personal_grid.setVerticalSpacing(12)
        personal_grid.setColumnStretch(0,0); personal_grid.setColumnStretch(1,1); personal_grid.setColumnStretch(2,0); personal_grid.setColumnStretch(3,1)

        # Name fields
        self.entries["first_name"] = le("First Name")
        self.entries["middle_name"] = le("Middle Name")
        self.entries["last_name"] = le("Last Name")
        self.no_middle_name_check = QCheckBox("No Middle Name")

        add_row(personal_grid, 0, "First Name *", self.entries["first_name"], "Middle Name", self.entries["middle_name"])
        self.entries["email"] = le("name@example.com")
        add_row(personal_grid, 1, "Last Name *", self.entries["last_name"], "Email", self.entries["email"])

        # Phone and emergency contact fields
        self.entries["phone"] = le("+63 9XX XXX XXXX")
        self.entries["emergency_contact_name"] = le("Emergency Contact Name")
        add_row(personal_grid, 2, "Phone", self.entries["phone"], "Emergency Contact", self.entries["emergency_contact_name"])

        self.entries["emergency_contact_phone"] = le("+63 9XX XXX XXXX")
        personal_grid.addWidget(QLabel("Emergency Phone"), 3, 0, alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        personal_grid.addWidget(self.entries["emergency_contact_phone"], 3, 1)

        personal_section.add_layout(personal_grid)
        sections_layout.addWidget(personal_section)

        # ============================================================
        # SECTION 2: JOB DETAILS
        # ============================================================
        job_section = FormSection("JOB DETAILS", icon="💼", start_collapsed=False)
        job_grid = QGridLayout()
        job_grid.setHorizontalSpacing(12); job_grid.setVerticalSpacing(12)
        job_grid.setColumnStretch(0,0); job_grid.setColumnStretch(1,1); job_grid.setColumnStretch(2,0); job_grid.setColumnStretch(3,1)

        # Employee ID and Position
        self.entries["emp_id"] = le("ID Number", ro=True)
        self.entries["position"] = le("Position")
        self._setup_autocomplete(self.entries["position"], self._get_common_positions())
        add_row(job_grid, 0, "Employee ID *", self.entries["emp_id"], "Position *", self.entries["position"])

        # Department dropdown with store picker
        dept_container = QWidget()
        dept_layout = QVBoxLayout(dept_container)
        dept_layout.setContentsMargins(0, 0, 0, 0)
        dept_layout.setSpacing(4)

        self.department_combo = NeumorphicGradientComboBox(
            "— Select Department —", 
            float_y=7.0,   # Adjusts the top "Floating" label position
            rest_y=25.0    # Adjusts the middle "Resting" label position
        )
        self.department_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.department_combo.setMinimumWidth(260)
        self.department_combo.setMinimumHeight(70)
        self.department_combo.combo_box.setFocusPolicy(Qt.ClickFocus)
        self.department_combo.addItems(["— Select Department —", "Office", "Warehouse", "Store"])
        self.department_combo.combo_box.currentTextChanged.connect(self._on_department_changed)

        # Removed custom styling - using neumorphic widget styling
        if False:  # Keep old stylesheet code for reference
            self.department_combo.setStyleSheet("""
            QComboBox {
                padding: 10px 35px 10px 12px;
                border: 2px solid rgba(33, 150, 243, 0.6);
                border-radius: 12px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 rgba(55, 65, 85, 0.95),
                                           stop:1 rgba(45, 55, 75, 0.95));
                color: white;
                font-size: 13px;
                font-weight: 500;
                min-height: 36px;
            }
            QComboBox:hover {
                border: 2px solid rgba(66, 165, 245, 0.9);
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 rgba(65, 75, 95, 0.95),
                                           stop:1 rgba(55, 65, 85, 0.95));
            }
            QComboBox:focus {
                border: 2px solid rgba(100, 181, 246, 1.0);
            }
            QComboBox::drop-down {
                border: none;
                width: 35px;
            }
            QComboBox::down-arrow {
                image: none;
                border: 2px solid rgba(255, 255, 255, 0.8);
                width: 8px;
                height: 8px;
                border-top: none;
                border-left: none;
                margin-right: 10px;
            }
            QComboBox QAbstractItemView {
                background: rgba(45, 55, 75, 0.98);
                border: 2px solid rgba(66, 165, 245, 0.7);
                border-radius: 12px;
                selection-background-color: rgba(33, 150, 243, 0.4);
                color: white;
                padding: 6px;
            }
        """)

        # Install wheel event filter to prevent accidental scrolling
        self.wheel_filter = WheelEventFilter()
        self.department_combo.installEventFilter(self.wheel_filter)

        dept_layout.addWidget(self.department_combo)

        # Visible department display (shows "Store - SM Novaliches")
        self.department_display = QLabel("")
        self.department_display.setStyleSheet("""
            QLabel {
                color: #4a9eff;
                font-weight: bold;
                font-size: 12px;
                padding: 4px 8px;
                background: rgba(74, 158, 255, 0.1);
                border-radius: 4px;
            }
        """)
        self.department_display.hide()
        dept_layout.addWidget(self.department_display)

        # Hidden field to store full department (e.g., "Store, SM Novaliches")
        self.entries["department"] = le()
        self.entries["department"].setReadOnly(True)
        self.entries["department"].hide()  # We'll use combo + display label

        # Department and Hire Date
        self.date_entries["hire_date"] = de()
        self.date_entries["hire_date"].setFocusPolicy(Qt.ClickFocus)
        add_row(job_grid, 1, "Department *", dept_container, "Hire Date *", self.date_entries["hire_date"])

        # Salary and Resign Date
        self.entries["salary"] = le("Salary/Day")
        self.date_entries["resign_date"] = de()
        self.date_entries["resign_date"].setFocusPolicy(Qt.ClickFocus)
        self.date_entries["resign_date"].setToolTip("Only enable if employee has left the company")
        add_row(job_grid, 2, "Salary/Day", self.entries["salary"], "Resign/Fired Date", self.date_entries["resign_date"])

        job_section.add_layout(job_grid)
        sections_layout.addWidget(job_section)

        # ============================================================
        # SECTION 3: GOVERNMENT IDs
        # ============================================================
        govt_section = FormSection("GOVERNMENT IDs", icon="🆔", start_collapsed=False)
        govt_grid = QGridLayout()
        govt_grid.setHorizontalSpacing(12); govt_grid.setVerticalSpacing(12)
        govt_grid.setColumnStretch(0,0); govt_grid.setColumnStretch(1,1); govt_grid.setColumnStretch(2,0); govt_grid.setColumnStretch(3,1)

        # SSS and TIN
        self.entries["sss_number"] = le("XX-XXXXXXX-X")
        self.entries["sss_number"].textChanged.connect(lambda: self._format_sss(self.entries["sss_number"]))
        self.entries["sss_number"].setToolTip("Format: XX-XXXXXXX-X (auto-formatted)")
        self.entries["tin_number"] = le("XXX-XXX-XXX-XXX")
        self.entries["tin_number"].textChanged.connect(lambda: self._format_tin(self.entries["tin_number"]))
        self.entries["tin_number"].setToolTip("Format: XXX-XXX-XXX-XXX (auto-formatted)")
        add_row(govt_grid, 0, "SSS #", self.entries["sss_number"], "TIN", self.entries["tin_number"])

        # PhilHealth and Pag-IBIG
        self.entries["philhealth_number"] = le("XX-XXXXXXXXX-X")
        self.entries["philhealth_number"].textChanged.connect(lambda: self._format_philhealth(self.entries["philhealth_number"]))
        self.entries["philhealth_number"].setToolTip("Format: XX-XXXXXXXXX-X (auto-formatted)")
        self.entries["pagibig_number"] = le("XXXX-XXXX-XXXX")
        self.entries["pagibig_number"].textChanged.connect(lambda: self._format_pagibig(self.entries["pagibig_number"]))
        self.entries["pagibig_number"].setToolTip("Format: XXXX-XXXX-XXXX (auto-formatted)")
        add_row(govt_grid, 1, "PhilHealth #", self.entries["philhealth_number"], "Pag-IBIG #", self.entries["pagibig_number"])

        govt_section.add_layout(govt_grid)
        sections_layout.addWidget(govt_section)

        # ============================================================
        # SECTION 4: CONTRACT INFORMATION
        # ============================================================
        contract_section = FormSection("CONTRACT INFORMATION", icon="📋", start_collapsed=False)
        contract_grid = QGridLayout()
        contract_grid.setHorizontalSpacing(12); contract_grid.setVerticalSpacing(12)
        contract_grid.setColumnStretch(0,0); contract_grid.setColumnStretch(1,1); contract_grid.setColumnStretch(2,0); contract_grid.setColumnStretch(3,1)

        # Contract Start and Duration
        self.date_entries["contract_start_date"] = de()
        self.date_entries["contract_start_date"].setFocusPolicy(Qt.ClickFocus)
        self.date_entries["contract_start_date"].dateChanged.connect(self._calculate_contract_expiry)

        self.contract_months_combo = NeumorphicGradientComboBox(
            "— Select Duration —",
            float_y=7.0, 
            rest_y=25.0    
        )
        self.contract_months_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.contract_months_combo.setMinimumWidth(260)
        self.contract_months_combo.setMinimumHeight(70)
        self.contract_months_combo.combo_box.setFocusPolicy(Qt.ClickFocus)
        self.contract_months_combo.addItem("— Select Duration —", None)
        for months in [1, 2, 3, 6, 12, 18, 24, 36]:
            label = f"{months} month" if months == 1 else f"{months} months"
            if months >= 12:
                years = months // 12
                remaining_months = months % 12
                if remaining_months == 0:
                    label = f"{years} year" if years == 1 else f"{years} years"
                else:
                    label = f"{years}y {remaining_months}m"
            self.contract_months_combo.addItem(label, months)
        self.contract_months_combo.combo_box.currentIndexChanged.connect(self._calculate_contract_expiry)

        add_row(contract_grid, 0, "Contract Start", self.date_entries["contract_start_date"], "Contract Duration", self.contract_months_combo)

        # Contract Expiry (auto-calculated)
        self.date_entries["contract_expiry"] = de()
        self.date_entries["contract_expiry"].setFocusPolicy(Qt.ClickFocus)
        self.date_entries["contract_expiry"].setToolTip("Auto-calculated from start date + duration, or manually set")
        contract_grid.addWidget(QLabel("Contract Expiry"), 1, 0, alignment=Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        contract_grid.addWidget(self.date_entries["contract_expiry"], 1, 1)

        # Help text
        contract_help = QLabel("(Auto-calculated from start + duration)")
        contract_help.setStyleSheet("color: #888; font-size: 11px;")
        contract_grid.addWidget(contract_help, 1, 2, 1, 2)

        contract_section.add_layout(contract_grid)
        sections_layout.addWidget(contract_section)

        # ============================================================
        # SECTION 5: AGENCY
        # ============================================================
        agency_section = FormSection("AGENCY", icon="🏢", start_collapsed=False)
        agency_layout = QHBoxLayout()

        # Agency combo
        self._reload_agencies_into_combo()
        agency_layout.addWidget(QLabel("Agency"))
        agency_layout.addWidget(self.agency_combo, 1)

        # Add Agency button
        add_ag_btn = ModernAnimatedButton("➕ Add Agency")
        apply_ios_style(add_ag_btn, 'green')
        add_ag_btn.setToolTip("Add new agency")
        add_ag_btn.setMinimumSize(180, 50)
        add_ag_btn.clicked.connect(self._add_agency_dialog)
        agency_layout.addWidget(add_ag_btn)

        agency_section.add_layout(agency_layout)
        sections_layout.addWidget(agency_section)

        # ============================================================
        # SECTION 6: ADDITIONAL INFORMATION (starts collapsed)
        # ============================================================
        additional_section = FormSection("ADDITIONAL INFORMATION", icon="📎", start_collapsed=True)
        additional_layout = QVBoxLayout()

        # "Still working" checkbox
        sw_row = QHBoxLayout()
        self.still_working = QCheckBox("Still working")
        self.still_working.setChecked(True)
        self.still_working.toggled.connect(self._toggle_resign)
        sw_row.addWidget(self.still_working)
        sw_row.addStretch(1)
        additional_layout.addLayout(sw_row)

        # Contract hint
        self.contract_hint = QLabel("")
        self.contract_hint.setStyleSheet("color:#ffcc66;")
        additional_layout.addWidget(self.contract_hint)

        # Notes
        additional_layout.addWidget(QLabel("<b>Notes</b>"))
        self.notes = NeumorphicGradientTextEdit("Additional notes...", min_height=100)
        self.notes.setFixedHeight(120)
        self.notes.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        additional_layout.addWidget(self.notes)

        # Files section will be added to additional_layout below (after files_box creation)

        # Add sections layout to top layout, then top to root
        top.addLayout(sections_layout, 1)
        root.addLayout(top)

        # Validators & behaviors
        # v5.1: Remove salary limit - allow any positive number (no max limit)
        # Using a regex validator instead of QDoubleValidator to avoid locale issues
        self.entries["salary"].setValidator(QRegularExpressionValidator(QRegularExpression(r"^\d*\.?\d{0,2}$")))
        self.entries["email"].setValidator(QRegularExpressionValidator(QRegularExpression(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")))
        for k in ["first_name", "middle_name", "last_name", "department", "position", "emergency_contact_name"]:
            self.entries[k].editingFinished.connect(lambda k=k: self.entries[k].setText(titlecase(self.entries[k].text())))
        self.entries["phone"].editingFinished.connect(self._normalize_phone_field)
        self.entries["emergency_contact_phone"].editingFinished.connect(self._normalize_emergency_phone_field)
        self.date_entries["hire_date"].dateChanged.connect(lambda *_: self._update_emp_id_preview())
        self.entries["email"].editingFinished.connect(self._validate_email_field)
        self.entries["salary"].editingFinished.connect(self._validate_salary_field)
        self.date_entries["contract_expiry"].dateChanged.connect(self._contract_date_changed)

        self._wheel_guard = WheelGuard(self)

        # Guard the dropdown and all date fields
        for w in [
            self.agency_combo,
            self.contract_months_combo,
            self.date_entries["hire_date"],
            self.date_entries["resign_date"],
            self.date_entries["contract_expiry"],
            self.date_entries["contract_start_date"],
        ]:
            w.setFocusPolicy(Qt.ClickFocus)
            w.installEventFilter(self._wheel_guard)

        # v5.1: Tabbed files section - moved to Additional Information section
        files_box=QGroupBox("📎 Attached Files")
        files_box.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                color: white;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(55, 65, 85, 0.85),
                                           stop:1 rgba(45, 55, 75, 0.7));
                border: 2px solid rgba(33, 150, 243, 0.5);
                border-top: 2px solid rgba(255, 255, 255, 0.2);
                border-left: 2px solid rgba(255, 255, 255, 0.15);
                border-radius: 22px;
                margin-top: 12px;
                padding: 16px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 16px;
                padding: 0 8px;
                color: white;
            }
        """)
        fb=QVBoxLayout(files_box)
        
        # Create tab widget for photos/documents separation - modern solid gradient
        from PySide6.QtWidgets import QTabWidget
        self.files_tabs = QTabWidget()
        self.files_tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 2px solid rgba(33, 150, 243, 0.5);
                border-radius: 16px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 rgba(45, 55, 75, 0.9),
                                           stop:1 rgba(35, 45, 65, 0.9));
                margin-top: -1px;
            }
            QTabBar::tab {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 rgba(55, 65, 85, 0.8),
                                           stop:1 rgba(45, 55, 75, 0.8));
                color: rgba(255, 255, 255, 0.7);
                padding: 10px 20px;
                margin-right: 4px;
                border-top-left-radius: 12px;
                border-top-right-radius: 12px;
                border: 2px solid rgba(255, 255, 255, 0.1);
                border-bottom: none;
                font-weight: 500;
            }
            QTabBar::tab:selected {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 rgba(33, 150, 243, 0.8),
                                           stop:1 rgba(25, 118, 210, 0.6));
                color: white;
                border: 2px solid rgba(66, 165, 245, 0.8);
                border-bottom: none;
            }
            QTabBar::tab:hover:!selected {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 rgba(65, 75, 95, 0.9),
                                           stop:1 rgba(55, 65, 85, 0.9));
                color: white;
            }
        """)
        
        # Photos tab
        photos_widget = QWidget()
        photos_layout = QVBoxLayout(photos_widget)
        photos_layout.setContentsMargins(4, 4, 4, 4)
        self.photos_list = QListWidget()
        disable_cursor_changes(self.photos_list)
        self.photos_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.photos_list.customContextMenuRequested.connect(lambda pos: self._files_context_menu(pos, 'photo'))
        self.photos_list.itemDoubleClicked.connect(self._open_file)
        photos_layout.addWidget(self.photos_list, 1)
        
        # Gallery button for photos - purple accent with solid gradient
        gallery_btn = ModernAnimatedButton("🖼️ View Gallery")
        gallery_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(156, 39, 176, 0.9),
                                           stop:1 rgba(123, 31, 162, 0.7));
                border: 2px solid rgba(156, 39, 176, 0.8);
                border-top: 2px solid rgba(255, 255, 255, 0.3);
                border-radius: 22px;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(186, 104, 200, 1.0),
                                           stop:1 rgba(156, 39, 176, 0.85));
                border: 2px solid rgba(186, 104, 200, 1.0);
            }
        """)
        gallery_btn.clicked.connect(self._show_photo_gallery)
        photos_layout.addWidget(gallery_btn)
        
        self.files_tabs.addTab(photos_widget, "📷 Photos")
        
        # Documents tab
        docs_widget = QWidget()
        docs_layout = QVBoxLayout(docs_widget)
        docs_layout.setContentsMargins(4, 4, 4, 4)
        self.files_list = QListWidget()  # Keep this name for backward compatibility
        disable_cursor_changes(self.files_list)
        self.files_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.files_list.customContextMenuRequested.connect(lambda pos: self._files_context_menu(pos, 'document'))
        self.files_list.itemDoubleClicked.connect(self._open_file)
        docs_layout.addWidget(self.files_list, 1)
        self.files_tabs.addTab(docs_widget, "📄 Documents")
        
        fb.addWidget(self.files_tabs, 1)
        
        # Modern solid gradient Upload Files button
        up_files=ModernAnimatedButton("📤 Upload Files")
        up_files.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(33, 150, 243, 0.9),
                                           stop:1 rgba(25, 118, 210, 0.7));
                border: 2px solid rgba(33, 150, 243, 0.8);
                border-top: 2px solid rgba(255, 255, 255, 0.3);
                border-radius: 22px;
                color: white;
                font-weight: bold;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(66, 165, 245, 1.0),
                                           stop:1 rgba(33, 150, 243, 0.85));
                border: 2px solid rgba(100, 181, 246, 1.0);
            }
        """)
        up_files.clicked.connect(self._upload_files)
        fb.addWidget(up_files)

        # Add files_box to Additional Information section
        additional_layout.addWidget(files_box)
        additional_section.add_layout(additional_layout)
        sections_layout.addWidget(additional_section)

        # Modern solid gradient Save and Cancel buttons with 22px pill shape
        actions=QHBoxLayout()
        self.save_btn=ModernAnimatedButton("💾 Save Employee")
        self.save_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(76, 175, 80, 0.95),
                                           stop:0.5 rgba(67, 160, 71, 0.8),
                                           stop:1 rgba(56, 142, 60, 0.65));
                border: 2px solid rgba(76, 175, 80, 0.9);
                border-top: 2px solid rgba(255, 255, 255, 0.3);
                border-left: 2px solid rgba(255, 255, 255, 0.2);
                border-radius: 22px;
                color: white;
                font-weight: bold;
                font-size: 14px;
                padding: 12px 24px;
                min-height: 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(102, 187, 106, 1.0),
                                           stop:0.5 rgba(76, 175, 80, 0.9),
                                           stop:1 rgba(67, 160, 71, 0.75));
                border: 2px solid rgba(102, 187, 106, 1.0);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(56, 142, 60, 0.95),
                                           stop:1 rgba(46, 125, 50, 0.8));
            }
        """)
        
        cancel=ModernAnimatedButton("✖ Cancel")
        cancel.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(120, 120, 130, 0.8),
                                           stop:1 rgba(90, 90, 100, 0.6));
                border: 2px solid rgba(150, 150, 160, 0.7);
                border-top: 2px solid rgba(255, 255, 255, 0.2);
                border-radius: 22px;
                color: white;
                font-weight: bold;
                font-size: 14px;
                padding: 12px 24px;
                min-height: 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(140, 140, 150, 0.9),
                                           stop:1 rgba(110, 110, 120, 0.7));
                border: 2px solid rgba(180, 180, 190, 0.9);
            }
        """)
        actions.addWidget(self.save_btn,2); actions.addWidget(cancel,1); root.addLayout(actions)
        self.save_btn.clicked.connect(self._save_employee); cancel.clicked.connect(self._cancel)

        # Prevent multiple rapid clicks on save button
        self._save_in_progress = False

        # QUICK WIN #3: Keyboard shortcuts for employee form
        from PySide6.QtGui import QShortcut, QKeySequence
        self.shortcut_save = QShortcut(QKeySequence("Ctrl+S"), self)
        self.shortcut_save.activated.connect(self._save_employee)
        self.shortcut_cancel = QShortcut(QKeySequence("Esc"), self)
        self.shortcut_cancel.activated.connect(self._cancel)

        # Connect all form fields to mark as changed
        for entry in self.entries.values():
            entry.textChanged.connect(self._mark_as_changed)
        for date_entry in self.date_entries.values():
            date_entry.dateChanged.connect(self._mark_as_changed)
        self.notes.textChanged.connect(self._mark_as_changed)
        self.agency_combo.currentTextChanged.connect(self._mark_as_changed)
        self.contract_months_combo.currentTextChanged.connect(self._mark_as_changed)
        self.department_combo.currentTextChanged.connect(self._mark_as_changed)

        # Auto-save draft timer (debounced - saves 2 seconds after last change)
        self._draft_save_timer = QTimer(self)
        self._draft_save_timer.setSingleShot(True)
        self._draft_save_timer.timeout.connect(self._save_draft)
        
        # Connect all fields to trigger draft save
        for entry in self.entries.values():
            entry.textChanged.connect(self._schedule_draft_save)
        for date_entry in self.date_entries.values():
            date_entry.dateChanged.connect(self._schedule_draft_save)
        self.notes.textChanged.connect(self._schedule_draft_save)
        self.agency_combo.currentTextChanged.connect(self._schedule_draft_save)
        self.contract_months_combo.currentTextChanged.connect(self._schedule_draft_save)
        self.department_combo.currentTextChanged.connect(self._schedule_draft_save)

        self._toggle_resign(True)
        
        # Check for and restore draft on startup (delayed to allow UI to initialize)
        QTimer.singleShot(500, self._check_for_draft)

    # Utility UI helpers
    def _set_photo_pixmap(self, pix: QPixmap, size=160):
        """Set photo with circular mask and update styling"""
        if pix.isNull(): return
        scaled=pix.scaled(size,size,Qt.KeepAspectRatioByExpanding,Qt.SmoothTransformation); result=QPixmap(size,size); result.fill(Qt.transparent)
        painter=QPainter(result); painter.setRenderHint(QPainter.Antialiasing,True); path=QPainterPath(); path.addEllipse(0,0,size,size); painter.setClipPath(path); painter.drawPixmap(0,0,scaled); painter.end()
        self.photo_label.setPixmap(result); self.photo_label.setText("")
        # No border, no background - just the circular photo
        self.photo_label.setStyleSheet("""
            QLabel {
                border-radius: 80px;
                background: transparent;
                border: none;
            }
        """)
    
    def _clear_photo_display(self):
        """Clear photo and reset to placeholder style"""
        self.photo_label.clear()
        self.photo_label.setText("📷")
        self.photo_label.setStyleSheet("""
            QLabel {
                border-radius: 80px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(60, 60, 80, 0.5),
                                           stop:1 rgba(40, 40, 60, 0.3));
                border: 4px solid rgba(100, 100, 120, 0.6);
                font-size: 56px;
                color: rgba(255, 255, 255, 0.5);
            }
            QLabel:hover {
                border: 4px solid rgba(100, 181, 246, 0.8);
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(70, 70, 90, 0.6),
                                           stop:1 rgba(50, 50, 70, 0.4));
            }
        """)
    def _contract_date_changed(self): self.contract_user_changed=True; self._update_contract_hint()
    def _update_contract_hint(self):
        if not self.contract_user_changed: self.contract_hint.setText(""); return
        d=self.date_entries["contract_expiry"].date(); days=(d.toPython()-datetime.now().date()).days
        if days<0: self.contract_hint.setStyleSheet("color:#ff6b6b;"); self.contract_hint.setText(f"Contract expired {-days} day(s) ago")
        elif days==0: self.contract_hint.setStyleSheet("color:#ff6b6b;"); self.contract_hint.setText("Contract expires today")
        elif days<=ALERT_DAYS: self.contract_hint.setStyleSheet("color:#ffcc66;"); self.contract_hint.setText(f"Contract expires in {days} day(s)")
        else: self.contract_hint.setStyleSheet("color:#9ad17a;"); self.contract_hint.setText(f"Contract valid ({days} day(s) left)")
    def _set_invalid(self, w, msg): w.setStyleSheet(INVALID_QSS); show_tooltip(w, msg)
    def _set_valid(self, w): w.setStyleSheet(VALID_QSS); w.setToolTip("")
    def _validate_email_field(self):
        w=self.entries["email"]; t=w.text().strip()
        if not t: self._set_valid(w); return
        self._set_valid(w) if EMAIL_RE.match(t) else self._set_invalid(w, "Invalid email format")
    def _validate_salary_field(self):
        w=self.entries["salary"]; t=w.text().strip()
        if t=="": self._set_valid(w); return
        try: val=float(t)
        except Exception: self._set_invalid(w,"Salary must be a number"); return
        if val<0: self._set_invalid(w,"Salary must be positive"); return
        w.setText(f"{val:.2f}"); self._set_valid(w)
    def _normalize_phone_field(self): self.entries["phone"].setText(normalize_ph_phone(self.entries["phone"].text()))

    def _format_sss(self, field):
        """Auto-format SSS number: XX-XXXXXXX-X"""
        text = field.text().replace("-", "").replace(" ", "")
        if not text:
            return
        # Only allow digits
        digits = ''.join(c for c in text if c.isdigit())
        if len(digits) > 10:
            digits = digits[:10]

        # Format: XX-XXXXXXX-X
        formatted = ""
        if len(digits) > 0:
            formatted = digits[:2]
        if len(digits) > 2:
            formatted += "-" + digits[2:9]
        if len(digits) > 9:
            formatted += "-" + digits[9]

        if formatted != field.text():
            cursor_pos = field.cursorPosition()
            field.blockSignals(True)
            field.setText(formatted)
            field.setCursorPosition(len(formatted))
            field.blockSignals(False)

    def _format_tin(self, field):
        """Auto-format TIN: XXX-XXX-XXX-XXX"""
        text = field.text().replace("-", "").replace(" ", "")
        if not text:
            return
        digits = ''.join(c for c in text if c.isdigit())
        if len(digits) > 12:
            digits = digits[:12]

        # Format: XXX-XXX-XXX-XXX
        formatted = ""
        if len(digits) > 0:
            formatted = digits[:3]
        if len(digits) > 3:
            formatted += "-" + digits[3:6]
        if len(digits) > 6:
            formatted += "-" + digits[6:9]
        if len(digits) > 9:
            formatted += "-" + digits[9:12]

        if formatted != field.text():
            cursor_pos = field.cursorPosition()
            field.blockSignals(True)
            field.setText(formatted)
            field.setCursorPosition(len(formatted))
            field.blockSignals(False)

    def _format_pagibig(self, field):
        """Auto-format Pag-IBIG: XXXX-XXXX-XXXX"""
        text = field.text().replace("-", "").replace(" ", "")
        if not text:
            return
        digits = ''.join(c for c in text if c.isdigit())
        if len(digits) > 12:
            digits = digits[:12]

        # Format: XXXX-XXXX-XXXX
        formatted = ""
        if len(digits) > 0:
            formatted = digits[:4]
        if len(digits) > 4:
            formatted += "-" + digits[4:8]
        if len(digits) > 8:
            formatted += "-" + digits[8:12]

        if formatted != field.text():
            cursor_pos = field.cursorPosition()
            field.blockSignals(True)
            field.setText(formatted)
            field.setCursorPosition(len(formatted))
            field.blockSignals(False)

    def _format_philhealth(self, field):
        """Auto-format PhilHealth: XX-XXXXXXXXX-X"""
        text = field.text().replace("-", "").replace(" ", "")
        if not text:
            return
        digits = ''.join(c for c in text if c.isdigit())
        if len(digits) > 12:
            digits = digits[:12]

        # Format: XX-XXXXXXXXX-X
        formatted = ""
        if len(digits) > 0:
            formatted = digits[:2]
        if len(digits) > 2:
            formatted += "-" + digits[2:11]
        if len(digits) > 11:
            formatted += "-" + digits[11]

        if formatted != field.text():
            cursor_pos = field.cursorPosition()
            field.blockSignals(True)
            field.setText(formatted)
            field.setCursorPosition(len(formatted))
            field.blockSignals(False)

    def _normalize_emergency_phone_field(self):
        self.entries["emergency_contact_phone"].setText(normalize_ph_phone(self.entries["emergency_contact_phone"].text()))

    def _calculate_contract_expiry(self):
        """Auto-calculate contract expiry based on start date and duration"""
        months = self.contract_months_combo.combo_box.currentData()
        if months is None:
            return

        start_date = self.date_entries["contract_start_date"].date()
        # Calculate expiry date by adding months
        expiry_date = start_date.addMonths(months)

        # Set the contract expiry date
        self.date_entries["contract_expiry"].setDate(expiry_date)
        self.contract_user_changed = True
        self._update_contract_hint()
    def _toggle_resign(self, checked):
        en=not checked; self.date_entries["resign_date"].setEnabled(en)
        if en and not self.resign_user_changed: self.date_entries["resign_date"].setDate(QDate.currentDate())
    def edit_employee(self, emp):
        # Release any existing lock before acquiring new one
        self._release_current_lock()

        self.current_employee=emp
        for e in self.entries.values(): e.clear(); self._set_valid(e)
        # v5.1: Clear both files lists (photos + documents)
        self.notes.clear(); self.files_list.clear()
        if hasattr(self, 'photos_list'): self.photos_list.clear()
        self.contract_hint.clear(); self.still_working.setChecked(True); self.photo_label.setText("📷"); self.photo_label.setPixmap(QPixmap()); self.photo_path = None  # Clear photo path to prevent cross-contamination
        self.contract_user_changed=False; self.resign_user_changed=False; self._reload_agencies_into_combo()
        self.date_entries["hire_date"].setDate(QDate.currentDate()); self.date_entries["resign_date"].setDate(QDate.currentDate()); self.date_entries["contract_expiry"].setDate(QDate.currentDate())
        
        # Reset dropdowns to default state for ADD mode
        self.department_combo.setCurrentIndex(0)  # "— Select Department —"
        self.department_display.hide()
        self.agency_combo.setCurrentIndex(0)
        
        # Inside edit_employee
        if emp:
            # Try to acquire lock for this employee
            emp_id = emp.get("emp_id", "")
            if emp_id and not self._acquire_lock(emp_id):
                # Lock acquisition failed - someone else is editing
                return  # Don't load the employee data
            for k in ["emp_id","email","phone","position","salary","sss_number","emergency_contact_name","emergency_contact_phone","tin_number","pagibig_number","philhealth_number"]:
                if k in self.entries: self.entries[k].setText(str(emp.get(k,"")) if emp.get(k) is not None else "")

            # Handle department dropdown
            dept = emp.get("department", "")
            self.entries["department"].setText(dept)  # Store full value
            if dept.startswith("Store"):  # Handle both "Store, X" and "Store - X"
                self.department_combo.combo_box.setCurrentText("Store")
                self.department_display.setText(f"✓ {dept}")
                self.department_display.show()
            elif dept in ["Office", "Warehouse"]:
                self.department_combo.combo_box.setCurrentText(dept)
                self.department_display.setText(f"✓ {dept}")
                self.department_display.show()
            else:
                self.department_combo.combo_box.setCurrentIndex(0)
                self.department_display.hide()

            # Parse the full name and set the individual fields
            full_name = emp.get("name", "")
            parts = full_name.split()
            first = parts[0] if parts else ""
            last = parts[-1] if len(parts) > 1 else ""
            middle = " ".join(parts[1:-1]) if len(parts) > 2 else ""

            self.entries["first_name"].setText(first)
            self.entries["last_name"].setText(last)
            self.entries["middle_name"].setText(middle)

            # Set the checkbox state based on whether a middle name exists
            self.no_middle_name_check.setChecked(not middle)
            ag=emp.get("agency","") or ""
            if ag and self.agency_combo.findText(ag)==-1: self.agency_combo.addItem(ag, ag)
            idx=self.agency_combo.findText(ag) if ag else 0; self.agency_combo.setCurrentIndex(idx if idx>=0 else 0)

            # Set contract months dropdown
            contract_months = emp.get("contract_months")
            if contract_months:
                idx = self.contract_months_combo.findData(contract_months)
                if idx >= 0:
                    self.contract_months_combo.combo_box.setCurrentIndex(idx)

            def sdate(v,key):
                if v:
                    try: d=datetime.strptime(v,"%m-%d-%Y"); self.date_entries[key].setDate(QDate(d.year,d.month,d.day))
                    except Exception: pass
            sdate(emp.get("hire_date",""),"hire_date")
            if emp.get("resign_date"): self.still_working.setChecked(False); self.date_entries["resign_date"].setEnabled(True); sdate(emp.get("resign_date",""),"resign_date"); self.resign_user_changed=True
            else: self.still_working.setChecked(True); self.date_entries["resign_date"].setEnabled(False)
            if emp.get("contract_expiry"): sdate(emp.get("contract_expiry",""),"contract_expiry"); self.contract_user_changed=True
            if emp.get("contract_start_date"): sdate(emp.get("contract_start_date",""),"contract_start_date")
            # v5.2: Check new folder structure first, then legacy location
            photos = get_employee_photos(emp.get('emp_id', ''))
            if photos:
                # Use first photo as profile photo
                self._set_photo_pixmap(QPixmap(photos[0]), 160)
            else:
                # Legacy: check old PHOTOS_DIR location
                p=os.path.join(PHOTOS_DIR,f"{emp.get('emp_id','')}.png")
                if os.path.exists(p): self._set_photo_pixmap(QPixmap(p),160)
            self._refresh_files_list(emp.get("emp_id",""))
        else: self._update_emp_id_preview()
        self._update_contract_hint()

        # Reset unsaved changes flag when loading employee
        self.has_unsaved_changes = False

    def _mark_as_changed(self):
        """Mark form as having unsaved changes"""
        self.has_unsaved_changes = True

    # ============================================================================
    # AUTO-SAVE DRAFT FUNCTIONALITY
    # ============================================================================
    
    def _schedule_draft_save(self):
        """Schedule a draft save (debounced - waits 2 seconds after last change)"""
        # Only save drafts for new employees, not when editing existing ones
        if self.current_employee:
            return
        self._draft_save_timer.start(2000)  # 2 second delay
    
    def _save_draft(self, force=False):
        """Save current form data as a draft
        
        Args:
            force: If True, save even when editing existing employee (used for force close)
        """
        try:
            # Skip if editing existing employee (unless force saving for force close)
            if self.current_employee and not force:
                return
            
            # Only save if there's meaningful content
            has_content = False
            for key, entry in self.entries.items():
                if entry.text().strip():
                    has_content = True
                    break
            
            if not has_content and not self.notes.text_edit.toPlainText().strip():
                return  # Don't save empty drafts
            
            # Collect all form data
            draft_data = {
                'saved_at': datetime.now().isoformat(),
                'user': self.current_user,
                'is_edit': bool(self.current_employee),  # Track if editing existing employee
                'emp_id': self.current_employee.get('emp_id') if self.current_employee else None,
                'fields': {},
                'dates': {},
                'dropdowns': {},
                'notes': self.notes.text_edit.toPlainText(),
                'photo_path': self.photo_path if hasattr(self, 'photo_path') else None,
                'still_working': self.still_working.isChecked() if self.still_working else True,
                'no_middle_name': self.no_middle_name_check.isChecked() if hasattr(self, 'no_middle_name_check') else False
            }
            
            # Save text fields
            for key, entry in self.entries.items():
                draft_data['fields'][key] = entry.text()
            
            # Save date fields
            for key, date_entry in self.date_entries.items():
                draft_data['dates'][key] = date_entry.date().toString("yyyy-MM-dd")
            
            # Save dropdown selections
            draft_data['dropdowns']['agency'] = self.agency_combo.currentIndex()
            draft_data['dropdowns']['department'] = self.department_combo.combo_box.currentIndex()
            draft_data['dropdowns']['contract_months'] = self.contract_months_combo.combo_box.currentIndex()
            
            # Write to file
            import json
            with open(DRAFT_FILE, 'w', encoding='utf-8') as f:
                json.dump(draft_data, f, indent=2, ensure_ascii=False)
            
            logging.info(f"Draft saved at {draft_data['saved_at']} (edit={draft_data['is_edit']})")
            
        except Exception as e:
            logging.warning(f"Failed to save draft: {e}")
    
    def _check_for_draft(self):
        """Check if there's a saved draft and offer to restore it"""
        try:
            if not os.path.exists(DRAFT_FILE):
                return
            
            import json
            with open(DRAFT_FILE, 'r', encoding='utf-8') as f:
                draft_data = json.load(f)
            
            # Check if draft has meaningful content
            has_content = False
            for key, value in draft_data.get('fields', {}).items():
                if value and value.strip():
                    has_content = True
                    break
            
            if not has_content:
                self._delete_draft()
                return
            
            # Get employee name for display
            first_name = draft_data.get('fields', {}).get('first_name', '')
            last_name = draft_data.get('fields', {}).get('last_name', '')
            employee_name = f"{first_name} {last_name}".strip() or "Unknown"
            
            # Check if this was editing an existing employee
            is_edit = draft_data.get('is_edit', False)
            emp_id = draft_data.get('emp_id', '')
            
            # Get saved time
            saved_at = draft_data.get('saved_at', '')
            try:
                saved_dt = datetime.fromisoformat(saved_at)
                saved_str = saved_dt.strftime("%b %d, %Y at %I:%M %p")
            except:
                saved_str = "Unknown time"
            
            # Build description based on type
            if is_edit and emp_id:
                type_info = f"<p><b>Type:</b> Editing existing employee (ID: {emp_id})</p>"
            else:
                type_info = "<p><b>Type:</b> New employee</p>"
            
            # Show restore dialog
            msg = QMessageBox(self)
            msg.setWindowTitle("⚠️ Unsaved Draft Found")
            msg.setIcon(QMessageBox.Question)
            msg.setText(f"<b>You have an unsaved employee form from your last session.</b>")
            msg.setInformativeText(
                f"<p><b>Employee:</b> {employee_name}</p>"
                f"{type_info}"
                f"<p><b>Last edited:</b> {saved_str}</p>"
                f"<p>Would you like to restore this draft?</p>"
            )
            msg.setStyleSheet("""
                QMessageBox {
                    background: rgba(45, 55, 75, 0.98);
                    color: white;
                }
                QLabel {
                    color: white;
                    font-size: 13px;
                }
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                               stop:0 rgba(33, 150, 243, 0.9),
                                               stop:1 rgba(25, 118, 210, 0.7));
                    border: 2px solid rgba(33, 150, 243, 0.8);
                    border-radius: 16px;
                    color: white;
                    font-weight: bold;
                    padding: 8px 16px;
                    min-width: 100px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                               stop:0 rgba(66, 165, 245, 1.0),
                                               stop:1 rgba(33, 150, 243, 0.85));
                }
            """)
            
            restore_btn = msg.addButton("Restore Draft", QMessageBox.AcceptRole)
            discard_btn = msg.addButton("Discard & Start New", QMessageBox.RejectRole)
            
            msg.exec()
            
            if msg.clickedButton() == restore_btn:
                self._restore_draft(draft_data)
                show_success_toast(self, "✓ Draft restored successfully!")
            else:
                self._delete_draft()
                show_info_toast(self, "Draft discarded. Starting fresh.")
                
        except Exception as e:
            logging.warning(f"Failed to check for draft: {e}")
            self._delete_draft()
    
    def _restore_draft(self, draft_data):
        """Restore form data from a draft"""
        try:
            # Block signals to prevent triggering draft save during restore
            for entry in self.entries.values():
                entry.blockSignals(True)
            for date_entry in self.date_entries.values():
                date_entry.blockSignals(True)
            self.notes.blockSignals(True)
            self.agency_combo.blockSignals(True)
            self.department_combo.blockSignals(True)
            self.contract_months_combo.blockSignals(True)
            
            # Restore text fields
            for key, value in draft_data.get('fields', {}).items():
                if key in self.entries:
                    self.entries[key].setText(value or '')
            
            # Restore date fields
            for key, value in draft_data.get('dates', {}).items():
                if key in self.date_entries and value:
                    try:
                        date = QDate.fromString(value, "yyyy-MM-dd")
                        if date.isValid():
                            self.date_entries[key].setDate(date)
                    except:
                        pass
            
            # Restore dropdowns
            dropdowns = draft_data.get('dropdowns', {})
            if 'agency' in dropdowns:
                self.agency_combo.setCurrentIndex(dropdowns['agency'])
            if 'department' in dropdowns:
                self.department_combo.combo_box.setCurrentIndex(dropdowns['department'])
                # Trigger department display update
                dept_text = self.department_combo.combo_box.currentText()
                if dept_text and dept_text != "— Select Department —":
                    self.department_display.setText(f"✓ {self.entries.get('department', {}).text() or dept_text}")
                    self.department_display.show()
            if 'contract_months' in dropdowns:
                self.contract_months_combo.combo_box.setCurrentIndex(dropdowns['contract_months'])
            
            # Restore notes
            self.notes.text_edit.setPlainText(draft_data.get('notes', ''))
            
            # Restore checkboxes
            if self.still_working:
                self.still_working.setChecked(draft_data.get('still_working', True))
            if hasattr(self, 'no_middle_name_check'):
                self.no_middle_name_check.setChecked(draft_data.get('no_middle_name', False))
            
            # Restore photo if exists
            photo_path = draft_data.get('photo_path')
            if photo_path and os.path.exists(photo_path):
                pix = QPixmap(photo_path)
                if not pix.isNull():
                    self._set_photo_pixmap(pix, 160)
                    self.photo_path = photo_path
            
            # Re-enable signals
            for entry in self.entries.values():
                entry.blockSignals(False)
            for date_entry in self.date_entries.values():
                date_entry.blockSignals(False)
            self.notes.blockSignals(False)
            self.agency_combo.blockSignals(False)
            self.department_combo.blockSignals(False)
            self.contract_months_combo.blockSignals(False)
            
            # Mark as having changes
            self.has_unsaved_changes = True
            
            # Delete the draft file after successful restore
            self._delete_draft()
            
            logging.info("Draft restored successfully")
            
        except Exception as e:
            logging.error(f"Failed to restore draft: {e}")
            show_error_toast(self, f"Failed to restore draft: {e}")
    
    def _delete_draft(self):
        """Delete the draft file"""
        try:
            if os.path.exists(DRAFT_FILE):
                os.remove(DRAFT_FILE)
                logging.info("Draft file deleted")
        except Exception as e:
            logging.warning(f"Failed to delete draft: {e}")

    def _check_unsaved_changes(self):
        """Check if form has any unsaved input and return True if user wants to continue"""
        if not self.has_unsaved_changes:
            return True

        # Check if any field has content
        has_content = False
        for entry in self.entries.values():
            if entry.text().strip():
                has_content = True
                break

        if not has_content and not self.photo_path:
            return True

        # Show confirmation dialog
        from employee_vault.ui.modern_ui_helper import show_warning_toast
        reply = QMessageBox.question(
            self,
            "Unsaved Changes",
            "You have unsaved changes. Do you want to discard them?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        return reply == QMessageBox.StandardButton.Yes

    def _employee_folder(self, emp_id): 
        """Get employee's files subfolder using new structure"""
        return get_employee_folder(emp_id, 'files')
    def _upload_photo(self):
        """Upload employee photo with error handling"""
        try:
            fn, _ = QFileDialog.getOpenFileName(
                self,
                "Select Photo",
                "",
                "Images (*.png *.jpg *.jpeg *.bmp)"
            )

            if not fn:
                return

            # Validate file exists
            if not os.path.exists(fn):
                show_warning_toast(
                    self, f"The selected file could not be found:\n{fn}\n\n"
                    "The file may have been moved, deleted, or is on a disconnected network drive.\n\n"
                    "Please select a valid image file and try again."
                )
                return

            # Try to load image
            pix = QPixmap(fn)
            if pix.isNull():
                show_warning_toast(
                    self, "Failed to load the selected image.\n\n"
                    "Possible causes:\n"
                    "• File is corrupt or damaged\n"
                    "• File format is not supported\n"
                    "• File is not an image\n\n"
                    "Supported formats: PNG, JPG, JPEG, BMP, GIF\n\n"
                    "Please select a valid image file and try again."
                )
                return

            # Check file size (e.g., max 5MB)
            file_size = os.path.getsize(fn) / (1024 * 1024)  # MB
            if file_size > 5:
                # PHASE 5: Auto-compress large images instead of rejecting them
                from employee_vault.utils import compress_image
                import tempfile
                import shutil

                # Create a temporary copy to compress
                temp_dir = tempfile.mkdtemp()
                temp_file = os.path.join(temp_dir, "compressed_" + os.path.basename(fn))
                shutil.copy2(fn, temp_file)

                # Try to compress the image
                try:
                    compress_image(temp_file, max_size_kb=500, quality=85)
                    new_size = os.path.getsize(temp_file) / (1024 * 1024)
                    show_success_toast(
                        self, f"Image automatically compressed from {file_size:.1f}MB to {new_size:.1f}MB!\n\n"
                        "The compressed image will be used for the employee photo."
                    )
                    fn = temp_file
                    pix = QPixmap(fn)  # Reload the compressed image
                except Exception as compress_error:
                    # Compression failed, show original warning
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    show_warning_toast(
                        self, f"The selected image is {file_size:.1f}MB, which exceeds the maximum allowed size of 5MB.\n\n"
                        f"Automatic compression failed: {str(compress_error)}\n\n"
                        "Please manually reduce the file size:\n"
                        "1. Use an image editor to resize the image\n"
                        "2. Compress the image quality\n"
                        "3. Convert to JPG format\n\n"
                        "Recommended image size: 800x800 pixels or smaller"
                    )
                    return
            elif file_size > 0.5:  # If over 500KB, compress it anyway for optimization
                from employee_vault.utils import compress_image
                import tempfile
                import shutil

                # Create a temporary copy to compress
                temp_dir = tempfile.mkdtemp()
                temp_file = os.path.join(temp_dir, "compressed_" + os.path.basename(fn))
                shutil.copy2(fn, temp_file)

                # Silently compress for optimization
                try:
                    compress_image(temp_file, max_size_kb=500, quality=85)
                    fn = temp_file
                    pix = QPixmap(fn)  # Reload the compressed image
                except Exception as e:
                    # If compression fails, just use original image
                    logging.warning(f"Image compression failed (non-critical): {e}")
                    shutil.rmtree(temp_dir, ignore_errors=True)

            # Show original photo immediately for preview
            self._set_photo_pixmap(pix, 160)
            self.photo_path = fn
            
            # Only remove background if checkbox is checked
            if hasattr(self, 'remove_bg_checkbox') and self.remove_bg_checkbox.isChecked():
                # Auto-remove background in background thread (non-blocking)
                self._start_background_removal(fn)

        except Exception as e:
            show_error_toast(
                self, f"Unable to upload photo:\n{str(e)}\n\n"
                "Possible causes:\n"
                "• Insufficient disk space\n"
                "• Photo directory is read-only\n"
                "• Network drive is disconnected\n"
                "• File path contains invalid characters\n\n"
                "Please check your system and try again."
            )

    def _start_background_removal(self, photo_path):
        """Start background removal in a worker thread with progress dialog"""
        # Create progress dialog
        self.bg_progress_dialog = QProgressDialog(
            "Processing photo...\nRemoving background...",
            "Cancel",
            0, 0,  # Indeterminate progress
            self
        )
        self.bg_progress_dialog.setWindowTitle("🔄 Processing Photo")
        self.bg_progress_dialog.setWindowModality(Qt.WindowModal)
        self.bg_progress_dialog.setMinimumDuration(0)
        self.bg_progress_dialog.setAutoClose(True)
        self.bg_progress_dialog.setAutoReset(True)
        self.bg_progress_dialog.setStyleSheet("""
            QProgressDialog {
                background: rgba(45, 55, 75, 0.98);
                border: 2px solid rgba(33, 150, 243, 0.7);
                border-radius: 16px;
                color: white;
                min-width: 350px;
                padding: 20px;
            }
            QLabel {
                color: white;
                font-size: 14px;
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(244, 67, 54, 0.9),
                                           stop:1 rgba(211, 47, 47, 0.7));
                border: 2px solid rgba(244, 67, 54, 0.8);
                border-radius: 22px;
                color: white;
                font-weight: bold;
                padding: 8px 20px;
                min-width: 80px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                                           stop:0 rgba(239, 83, 80, 1.0),
                                           stop:1 rgba(244, 67, 54, 0.85));
            }
        """)
        
        # Create and start worker
        self.photo_worker = PhotoProcessingWorker(photo_path)
        self.photo_worker.progress.connect(self._on_bg_progress)
        self.photo_worker.finished.connect(self._on_bg_finished)
        self.photo_worker.error.connect(self._on_bg_error)
        
        # Connect cancel button
        self.bg_progress_dialog.canceled.connect(self._on_bg_cancelled)
        
        self.photo_worker.start()
        self.bg_progress_dialog.show()
    
    def _on_bg_progress(self, message):
        """Update progress dialog with status message"""
        if hasattr(self, 'bg_progress_dialog') and self.bg_progress_dialog:
            self.bg_progress_dialog.setLabelText(f"Processing photo...\n{message}")
    
    def _on_bg_finished(self, result_path, success):
        """Handle background removal completion"""
        if hasattr(self, 'bg_progress_dialog') and self.bg_progress_dialog:
            self.bg_progress_dialog.close()
        
        if success and result_path:
            # Update photo with background-removed version
            pix = QPixmap(result_path)
            if not pix.isNull():
                self._set_photo_pixmap(pix, 160)
                self.photo_path = result_path
                show_success_toast(self, "✓ Photo processed successfully!\nBackground removed.")
            else:
                logging.warning("Failed to load processed photo")
        else:
            # Keep original photo
            show_info_toast(self, "Photo uploaded (background removal skipped)")
    
    def _on_bg_error(self, error_msg):
        """Handle background removal error"""
        logging.warning(f"Background removal error: {error_msg}")
        # Don't show error toast - just keep original photo silently
    
    def _on_bg_cancelled(self):
        """Handle user cancellation of background removal"""
        if hasattr(self, 'photo_worker') and self.photo_worker:
            self.photo_worker.cancel()
            self.photo_worker.wait(1000)  # Wait up to 1 second for thread to stop
        show_info_toast(self, "Background removal cancelled.\nOriginal photo will be used.")

    def _remove_photo(self):
        """Remove employee photo with confirmation"""
        # Check if there's a photo to remove
        has_temp_photo = hasattr(self, 'photo_path') and self.photo_path
        
        # v5.2: Check new folder structure first, then legacy
        emp_id = self.current_employee.get('emp_id', '') if self.current_employee else ''
        photos_in_new_location = get_employee_photos(emp_id) if emp_id else []
        has_saved_photo_new = len(photos_in_new_location) > 0
        has_saved_photo_legacy = self.current_employee and os.path.exists(
            os.path.join(PHOTOS_DIR, f"{emp_id}.png")
        )
        has_saved_photo = has_saved_photo_new or has_saved_photo_legacy

        if not has_temp_photo and not has_saved_photo:
            show_info_toast(self, "No photo to remove.\n\nThis employee doesn't have a photo uploaded.")
            return

        # Confirm deletion
        reply = QMessageBox.question(
            self,
            "Confirm Remove Photo",
            "Are you sure you want to remove this employee's photo?\n\nThis action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            # Remove temporary photo (not yet saved)
            if has_temp_photo:
                self.photo_path = None

            # Remove saved photo from disk
            if has_saved_photo:
                # v5.2: Remove from new location
                for photo_path in photos_in_new_location:
                    if os.path.exists(photo_path):
                        os.remove(photo_path)
                        logging.info(f"Photo removed from new location: {photo_path}")
                
                # Also remove from legacy location if exists
                photo_file = os.path.join(PHOTOS_DIR, f"{self.current_employee['emp_id']}.png")
                if os.path.exists(photo_file):
                    os.remove(photo_file)
                    logging.info(f"Photo removed from legacy location for {self.current_employee['emp_id']} by {self.current_user}")

                    # Log the action
                    if hasattr(self.db, 'log_action'):
                        self.db.log_action(
                            username=self.current_user,
                            action="DELETE_PHOTO",
                            table_name="employees",
                            record_id=self.current_employee['emp_id'],
                            details=f"Removed employee photo"
                        )

            # Reset photo display using the helper method
            self._clear_photo_display()
            self.photo_path = None
            show_success_toast(self, "Photo removed successfully!")

        except PermissionError:
            show_error_toast(self, "Cannot remove photo (permission denied).\n\nPlease check file permissions.")
            logging.error(f"Permission denied removing photo")
        except Exception as e:
            show_error_toast(self, f"Unable to remove photo:\n{str(e)}")
            logging.error(f"Error removing photo: {e}")

    def _view_photo_fullscreen(self):
        """View employee photo in a screen-fitting preview dialog"""
        try:
            # Check if there's a photo to view
            if hasattr(self, 'photo_path') and self.photo_path and os.path.exists(self.photo_path):
                from employee_vault.ui.dialogs.photo_editor import PhotoPreviewDialog
                dialog = PhotoPreviewDialog(image_path=self.photo_path, parent=self)
                dialog.exec()
            elif self.photo_label.pixmap() and not self.photo_label.pixmap().isNull():
                from employee_vault.ui.dialogs.photo_editor import PhotoPreviewDialog
                dialog = PhotoPreviewDialog(pixmap=self.photo_label.pixmap(), parent=self)
                dialog.exec()
            else:
                show_info_toast(
                    self, "No photo uploaded yet.\nClick 'Upload Photo' to add a photo."
                )
        except Exception as e:
            show_warning_toast(self, f"Unable to view photo:\n{str(e)}")

    def _on_photo_clicked(self):
        """Handle photo avatar click - edit if photo exists, otherwise prompt to upload"""
        try:
            # Check if there's a photo to edit
            if hasattr(self, 'photo_path') and self.photo_path and os.path.exists(self.photo_path):
                self._edit_photo(self.photo_path)
            elif self.photo_label.pixmap() and not self.photo_label.pixmap().isNull():
                # Photo in memory but not saved yet
                self._edit_photo_pixmap(self.photo_label.pixmap())
            else:
                # No photo - trigger upload
                self._upload_photo()
        except Exception as e:
            show_warning_toast(self, f"Unable to edit photo:\n{str(e)}")

    def _edit_photo(self, photo_path: str):
        """Open photo editor dialog for cropping and positioning"""
        from employee_vault.ui.dialogs.photo_editor import PhotoEditorDialog
        
        dialog = PhotoEditorDialog(image_path=photo_path, parent=self)
        if dialog.exec() == QDialog.Accepted:
            result = dialog.get_result()
            if result and not result.isNull():
                self._set_photo_pixmap(result, 160)
                # Save to temp location for later save
                import tempfile
                temp_dir = tempfile.mkdtemp()
                temp_path = os.path.join(temp_dir, "edited_photo.png")
                result.save(temp_path, "PNG")
                self.photo_path = temp_path
                show_success_toast(self, "Photo edited successfully!")

    def _edit_photo_pixmap(self, pixmap: QPixmap):
        """Open photo editor dialog for a pixmap in memory"""
        from employee_vault.ui.dialogs.photo_editor import PhotoEditorDialog
        
        dialog = PhotoEditorDialog(pixmap=pixmap, parent=self)
        if dialog.exec() == QDialog.Accepted:
            result = dialog.get_result()
            if result and not result.isNull():
                self._set_photo_pixmap(result, 160)
                # Save to temp location for later save
                import tempfile
                temp_dir = tempfile.mkdtemp()
                temp_path = os.path.join(temp_dir, "edited_photo.png")
                result.save(temp_path, "PNG")
                self.photo_path = temp_path
                show_success_toast(self, "Photo edited successfully!")

    def _upload_files(self):
        emp_id=self.entries["emp_id"].text().strip()
        if not emp_id:
            show_warning_toast(
                self, "You must set an Employee ID before uploading files.\n\n"
                "Please enter an Employee ID in the form and try again."
            )
            return
        files,_=QFileDialog.getOpenFileNames(self,"Select Files")
        if not files: return

        # Image extensions for photos
        image_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.tiff', '.ico'}

        # File size limit: 10MB
        MAX_FILE_SIZE_MB = 10
        MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

        try:
            photos_uploaded = 0
            docs_uploaded = 0
            skipped_files = []

            for fp in files:
                filename = os.path.basename(fp)
                ext = os.path.splitext(filename)[1].lower()

                # Validate file size
                file_size = os.path.getsize(fp)
                if file_size > MAX_FILE_SIZE_BYTES:
                    size_mb = file_size / (1024 * 1024)
                    skipped_files.append(f"{filename} ({size_mb:.1f}MB)")
                    continue

                # Route photos to photos folder, documents to files folder
                if ext in image_extensions:
                    folder = get_employee_folder(emp_id, 'photos')
                    file_type = "photo"
                    photos_uploaded += 1
                else:
                    folder = get_employee_folder(emp_id, 'files')
                    file_type = "document"
                    docs_uploaded += 1

                # Validate filename and create safe path
                safe_dest = safe_file_path(folder, filename, allow_subdirs=False)
                shutil.copy2(fp, safe_dest)
                logging.info(f"{file_type.capitalize()} uploaded: {filename} for employee {emp_id}")

                # QUICK WIN #5: Enhanced audit logging - log file uploads
                try:
                    self.db.log_action(
                        username=self.current_user,
                        action="FILE_UPLOADED",
                        table_name="employee_files",
                        record_id=emp_id,
                        details=f"Uploaded {file_type}: {filename}"
                    )
                except:
                    pass  # Don't fail upload if logging fails

            self._refresh_files_list(emp_id)

            # Show detailed success message
            msg_parts = []
            if photos_uploaded > 0:
                msg_parts.append(f"{photos_uploaded} photo(s)")
            if docs_uploaded > 0:
                msg_parts.append(f"{docs_uploaded} document(s)")

            if msg_parts:
                show_success_toast(self, f"{' and '.join(msg_parts)} uploaded securely!")

            # Warn about skipped files
            if skipped_files:
                show_warning_toast(
                    self,
                    f"Skipped {len(skipped_files)} file(s) exceeding {MAX_FILE_SIZE_MB}MB limit:\n" +
                    "\n".join(skipped_files[:5]) +
                    (f"\n...and {len(skipped_files)-5} more" if len(skipped_files) > 5 else "")
                )

        except ValueError as e:
            show_error_toast(self, f"Invalid file path: {e}")
            logging.error(f"File upload security error: {e}")
        except Exception as e:
            show_error_toast(self, f"Failed to upload files: {e}")
            logging.error(f"File upload failed: {e}")
            
    def _refresh_files_list(self, emp_id):
        """Refresh both photos and documents lists (v5.2 new folder structure)"""
        # Clear both lists
        self.files_list.clear()
        if hasattr(self, 'photos_list'):
            self.photos_list.clear()
        
        if not emp_id: 
            return
        
        # v5.2: Use new folder structure with photos/ and files/ subdirectories
        try:
            photos_folder = get_employee_folder(emp_id, 'photos')
            files_folder = get_employee_folder(emp_id, 'files')
        except ValueError:
            return  # Invalid emp_id
        
        # Image extensions for photos tab
        image_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.tiff', '.ico'}
        
        # Load photos from photos/ subfolder
        if os.path.exists(photos_folder) and hasattr(self, 'photos_list'):
            for file in os.listdir(photos_folder):
                file_path = os.path.join(photos_folder, file)
                ext = os.path.splitext(file)[1].lower()
                if ext in image_extensions:
                    icon = "🖼️"
                    item = QListWidgetItem(f"{icon} {file}")
                    item.setData(Qt.UserRole, file_path)
                    item.setData(Qt.UserRole + 1, 'photo')
                    self.photos_list.addItem(item)
        
        # Also check legacy PHOTOS_DIR location
        legacy_photo = os.path.join(PHOTOS_DIR, f"{emp_id}.png")
        if os.path.exists(legacy_photo) and hasattr(self, 'photos_list'):
            icon = "🖼️"
            item = QListWidgetItem(f"{icon} {emp_id}.png (legacy)")
            item.setData(Qt.UserRole, legacy_photo)
            item.setData(Qt.UserRole + 1, 'photo')
            self.photos_list.addItem(item)
        
        # Load documents from files/ subfolder (exclude images - they belong in photos tab)
        if os.path.exists(files_folder):
            for file in os.listdir(files_folder):
                file_path = os.path.join(files_folder, file)
                ext = os.path.splitext(file)[1].lower()

                # Skip images - they should only appear in photos tab
                if ext in image_extensions:
                    continue

                # Choose icon based on extension
                if ext == '.pdf':
                    icon = "📕"
                elif ext in {'.doc', '.docx'}:
                    icon = "📝"
                elif ext in {'.xls', '.xlsx'}:
                    icon = "📊"
                elif ext in {'.txt', '.log'}:
                    icon = "📃"
                elif ext in {'.zip', '.rar', '.7z'}:
                    icon = "📦"
                else:
                    icon = "📄"
                item = QListWidgetItem(f"{icon} {file}")
                item.setData(Qt.UserRole, file_path)
                item.setData(Qt.UserRole + 1, 'document')
                self.files_list.addItem(item)
        
        # Update tab labels with counts
        if hasattr(self, 'files_tabs'):
            photos_count = self.photos_list.count() if hasattr(self, 'photos_list') else 0
            docs_count = self.files_list.count()
            self.files_tabs.setTabText(0, f"📷 Photos ({photos_count})")
            self.files_tabs.setTabText(1, f"📄 Documents ({docs_count})")
    
    def _show_photo_gallery(self):
        """Show animated gallery preview for employee photos"""
        emp_id = self.entries["emp_id"].text().strip()
        if not emp_id:
            show_warning_toast(self, "No employee selected.")
            return
        
        # v5.2: Use new folder structure helper
        image_paths = get_employee_photos(emp_id)
        
        # Also check legacy location
        legacy_photo = os.path.join(PHOTOS_DIR, f"{emp_id}.png")
        if os.path.exists(legacy_photo) and legacy_photo not in image_paths:
            image_paths.insert(0, legacy_photo)
        
        if not image_paths:
            show_warning_toast(self, "No photos found for this employee.\n\nUpload some photos first!")
            return
        
        # Show gallery dialog
        try:
            from employee_vault.ui.widgets.stacked_card_gallery import GalleryPreviewDialog
            emp_name = self.entries.get("name", {})
            name_text = emp_name.text() if hasattr(emp_name, 'text') else "Employee"
            dialog = GalleryPreviewDialog(
                image_paths, 
                parent=self, 
                title=f"📷 {name_text}'s Photos"
            )
            dialog.exec()
        except ImportError as e:
            logging.error(f"Could not import gallery: {e}")
            show_error_toast(self, "Gallery feature not available.")
    
    def _files_context_menu(self, pos, file_type='document'):
        """Context menu for file actions (v5.1: supports both photos and documents)"""
        # Determine which list widget to use
        if file_type == 'photo' and hasattr(self, 'photos_list'):
            list_widget = self.photos_list
        else:
            list_widget = self.files_list
            
        item = list_widget.itemAt(pos)
        if not item: 
            return
        path = item.data(Qt.UserRole)

        # Create context menu
        m = QMenu(self)
        act_open = m.addAction("📂 Open")
        
        # For photos, add view in gallery option
        if file_type == 'photo':
            act_gallery = m.addAction("🖼️ View in Gallery")
        else:
            act_gallery = None
            
        act_reveal = m.addAction("📁 Reveal in Folder")
        m.addSeparator()
        act_delete = m.addAction("🗑️ Delete File")
        act_delete.setEnabled(True)

        # Execute menu
        a = m.exec(list_widget.mapToGlobal(pos))

        # Handle actions
        if a == act_open:
            self._open_file(item)
        elif act_gallery and a == act_gallery:
            self._show_photo_gallery()
        elif a == act_reveal:
            folder = os.path.dirname(path)
            try:
                if sys.platform.startswith("win"): 
                    os.startfile(folder)
                elif sys.platform == "darwin": 
                    os.system(f'open "{folder}"')
                else: 
                    os.system(f'xdg-open "{folder}"')
            except Exception: 
                pass
        elif a == act_delete:
            self._delete_file(item, path, list_widget)
    def _open_file(self, item):
        path=item.data(Qt.UserRole); ext=os.path.splitext(path)[1].lower()
        if (ext=='.pdf' and PDF_AVAILABLE) or ext in {'.txt','.log','.json','.xml','.csv','.py','.html','.css','.js','.png','.jpg','.jpeg','.gif','.bmp'}:
            FileViewerDialog(path,self).exec()
        else:
            from PySide6.QtCore import QUrl; QDesktopServices.openUrl(QUrl.fromLocalFile(path))

    def _delete_file(self, item, path, list_widget=None):
        """Delete attached file with confirmation (v5.1: supports both lists)"""
        if list_widget is None:
            list_widget = self.files_list
            
        filename = os.path.basename(path)

        # Confirm deletion
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Are you sure you want to delete this file?\n\n{filename}\n\nThis action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )

        if reply != QMessageBox.StandardButton.Yes:
            return

        try:
            # Delete the physical file
            if os.path.exists(path):
                os.remove(path)
                logging.info(f"File deleted: {path} by {self.current_user}")

            # Remove from database if using employee_files table
            emp_id = self.emp_id_edit.text().strip()
            if emp_id and hasattr(self.db, 'delete_employee_file'):
                try:
                    self.db.delete_employee_file(emp_id, filename)
                except:
                    pass  # Table might not exist in older versions

            # Remove from the correct list widget
            row = list_widget.row(item)
            list_widget.takeItem(row)
            
            # Update tab counts
            self._update_file_tab_counts()

            # Log the action
            if hasattr(self.db, 'log_action'):
                self.db.log_action(
                    username=self.current_user,
                    action="DELETE_FILE",
                    table_name="employee_files",
                    record_id=emp_id,
                    details=f"Deleted file: {filename}"
                )

            show_success_toast(self, f"File deleted:\n{filename}")

        except PermissionError:
            show_error_toast(self, f"Cannot delete file (permission denied):\n{filename}")
            logging.error(f"Permission denied deleting file: {path}")
        except Exception as e:
            show_error_toast(
                self, f"Unable to delete file '{filename}':\n{str(e)}\n\n"
                "Possible causes:\n"
                "• File is currently open in another program\n"
                "• Insufficient permissions\n"
                "• File is on a read-only network drive\n"
                "• File has been moved or deleted already\n\n"
                "Please close any programs using this file and try again."
            )
            logging.error(f"Error deleting file {path}: {e}")

    def _update_file_tab_counts(self):
        """Update the tab labels with file counts"""
        if hasattr(self, 'files_tabs') and hasattr(self, 'photos_list'):
            photos_count = self.photos_list.count()
            docs_count = self.files_list.count()
            self.files_tabs.setTabText(0, f"📷 Photos ({photos_count})")
            self.files_tabs.setTabText(1, f"📄 Documents ({docs_count})")

    def _prefix_from_department(self, dept: str) -> str:
        d = (dept or "").strip()
        if not d: return "X"
        # "Store - SM Novaliches" -> take first letter of the store name ("S")
        if d.lower().startswith("store"):
            store = d.split("-", 1)[1].strip() if "-" in d else d[5:].strip()
            for ch in store:
                if ch.isalpha():
                    return ch.upper()
            return "X"
        # Otherwise first letter of department
        for ch in d:
            if ch.isalpha():
                return ch.upper()
        return "X"

    def _update_emp_id_preview(self):
        if self.current_employee is not None:
            return
        dept = self.entries["department"].text().strip()
        year_2 = self.date_entries["hire_date"].date().year() % 100
        prefix = self._prefix_from_department(dept)
        seq = self.db.next_sequence()
        self.entries["emp_id"].setText(f"{prefix}-{seq:03d}-{year_2:02d}")

    def _cancel(self): self.on_saved(cancel_only=True)
    def _save_employee(self):
        logging.info(f"Save employee button clicked by {self.current_user}")

        # Prevent multiple rapid clicks
        if self._save_in_progress:
            logging.warning("Save already in progress, ignoring duplicate click")
            return

        self._save_in_progress = True
        self.save_btn.setEnabled(False)
        self.save_btn.setText("⏳ Saving...")

        try:
            self._perform_save()
        finally:
            self._save_in_progress = False
            self.save_btn.setEnabled(True)
            self.save_btn.setText("💾 Save Employee")

    def _perform_save(self):
        logging.info("Starting _perform_save")
        first_name = titlecase(self.entries["first_name"].text().strip())
        middle_name = titlecase(self.entries["middle_name"].text().strip())
        last_name = titlecase(self.entries["last_name"].text().strip())
        logging.info(f"Names parsed: {first_name=}, {middle_name=}, {last_name=}")

        if not first_name:
            logging.warning("Validation failed: First name required")
            self._set_invalid(self.entries["first_name"], "First Name is required!")
            return
        if not last_name:
            logging.warning("Validation failed: Last name required")
            self._set_invalid(self.entries["last_name"], "Last Name is required!")
            return
        logging.info("Name validation passed")

        # PHASE 1 FIX: Validate department is selected
        department = self.entries["department"].text().strip()
        logging.info(f"Department check: {department=}")
        if not department or department == "— Select Department —":
            logging.warning("Validation failed: Department not selected")
            show_warning_toast(self, "Please select a Department!")
            return
        logging.info("Department validation passed")

        # Combine the names into a single string, handling the optional middle name
        name_parts = [first_name, middle_name, last_name]
        name = " ".join(part for part in name_parts if part)
        logging.info(f"Full name: {name}")

        # DATA QUALITY FEATURE: Check for duplicate employee names
        logging.info("Checking for duplicate names...")
        similar_employees = self._check_duplicate_name(name)
        logging.info(f"Duplicate check complete: {len(similar_employees) if similar_employees else 0} similar found")
        if similar_employees:
            # Show warning dialog
            logging.info("Showing duplicate warning dialog...")
            if not self._show_duplicate_warning(name, similar_employees):
                # User cancelled, don't save
                logging.info("User cancelled due to duplicate warning")
                return
            logging.info("User confirmed save despite duplicates")
        logging.info("Processing dates...")
        hire=self.date_entries["hire_date"].date().toString("MM-dd-yyyy")
        resign=None
        if not self.still_working.isChecked(): self.resign_user_changed=True; resign=self.date_entries["resign_date"].date().toString("MM-dd-yyyy")
        logging.info(f"Dates: hire={hire}, resign={resign}")

        logging.info("Validating email...")
        email=self.entries["email"].text().strip()
        if email and not EMAIL_RE.match(email):
            logging.warning(f"Invalid email format: {email}")
            self._set_invalid(self.entries["email"],"Invalid email format");
            return
        logging.info(f"Email validation passed: {email}")

        logging.info("Processing phone...")
        phone=normalize_ph_phone(self.entries["phone"].text());
        self.entries["phone"].setText(phone)
        logging.info(f"Phone: {phone}")

        logging.info("Processing salary...")
        salary_txt=self.entries["salary"].text().strip()
        if salary_txt:
            try: salary=float(salary_txt)
            except Exception:
                logging.warning(f"Invalid salary: {salary_txt}")
                self._set_invalid(self.entries["salary"],"Salary must be a number");
                return
        else: salary=0.0
        logging.info(f"Salary: {salary}")
        logging.info("Processing employee ID...")
        emp_id=self.entries["emp_id"].text().strip()
        if not self.current_employee:
            logging.info("New employee - generating employee ID")
            dept_full = self.entries["department"].text()
            year_2 = self.date_entries["hire_date"].date().year() % 100
            prefix = self._prefix_from_department(dept_full)
            n = self.db.next_sequence()
            emp_id = f"{prefix}-{n:03d}-{year_2:02d}"
            while self.db.employee_exists(emp_id):
                n += 1
                emp_id = f"{prefix}-{n:03d}-{year_2:02d}"
            self.entries["emp_id"].setText(emp_id)
            logging.info(f"Generated employee ID: {emp_id}")
        else:
            logging.info(f"Updating existing employee: {emp_id}")
        contract=self.date_entries["contract_expiry"].date().toString("MM-dd-yyyy") if self.contract_user_changed else ""
        contract_start = self.date_entries["contract_start_date"].date().toString("MM-dd-yyyy") if self.contract_user_changed else ""
        contract_months = self.contract_months_combo.combo_box.currentData()
        agency = None if self.agency_combo.currentIndex()==0 else self.agency_combo.currentText().strip()

        # Get new fields
        sss_number = self.entries["sss_number"].text().strip() or None
        emergency_contact_name = titlecase(self.entries["emergency_contact_name"].text().strip()) or None
        emergency_contact_phone = normalize_ph_phone(self.entries["emergency_contact_phone"].text()) or None

        # Get government ID fields with validation
        tin_number = self.entries["tin_number"].text().strip() or None
        pagibig_number = self.entries["pagibig_number"].text().strip() or None
        philhealth_number = self.entries["philhealth_number"].text().strip() or None

        # Validate government IDs if provided
        logging.info("Validating government IDs...")
        if sss_number:
            # SSS format: XX-XXXXXXX-X (10 digits, 2-7-1 format)
            sss_clean = sss_number.replace("-", "")
            if not sss_clean.isdigit() or len(sss_clean) != 10:
                logging.warning(f"Invalid SSS number: {sss_number}")
                self._set_invalid(self.entries["sss_number"], "SSS# must be 10 digits (XX-XXXXXXX-X)")
                return

        if tin_number:
            # TIN format: XXX-XXX-XXX-XXX (12 digits)
            tin_clean = tin_number.replace("-", "")
            if not tin_clean.isdigit() or len(tin_clean) != 12:
                logging.warning(f"Invalid TIN number: {tin_number}")
                self._set_invalid(self.entries["tin_number"], "TIN must be 12 digits (XXX-XXX-XXX-XXX)")
                return

        if pagibig_number:
            # Pag-IBIG format: XXXX-XXXX-XXXX (12 digits)
            pagibig_clean = pagibig_number.replace("-", "")
            if not pagibig_clean.isdigit() or len(pagibig_clean) != 12:
                logging.warning(f"Invalid Pag-IBIG number: {pagibig_number}")
                self._set_invalid(self.entries["pagibig_number"], "Pag-IBIG# must be 12 digits (XXXX-XXXX-XXXX)")
                return

        if philhealth_number:
            # PhilHealth format: XX-XXXXXXXXX-X (12 digits, 2-9-1 format)
            philhealth_clean = philhealth_number.replace("-", "")
            if not philhealth_clean.isdigit() or len(philhealth_clean) != 12:
                logging.warning(f"Invalid PhilHealth number: {philhealth_number}")
                self._set_invalid(self.entries["philhealth_number"], "PhilHealth# must be 12 digits (XX-XXXXXXXXX-X)")
                return
        logging.info("Government ID validation passed")

        # Ensure modified_by is always set to a valid username
        modified_by_user = self.current_user
        if not modified_by_user or modified_by_user == "?":
            modified_by_user = "system"

        logging.info("Building employee data dictionary...")
        data={"emp_id":emp_id,"name":name,"email":email,"phone":phone,"department":titlecase(self.entries["department"].text()),"position":titlecase(self.entries["position"].text()),
              "hire_date":hire,"resign_date":resign,"salary":salary,"notes":self.notes.text_edit.toPlainText(),
              "modified":datetime.now().strftime("%Y-%m-%d %H:%M:%S"),"modified_by":modified_by_user,
              "contract_expiry":contract,"agency":agency,
              "sss_number":sss_number,"emergency_contact_name":emergency_contact_name,
              "emergency_contact_phone":emergency_contact_phone,
              "contract_start_date":contract_start,"contract_months":contract_months,
              "tin_number":tin_number,"pagibig_number":pagibig_number,"philhealth_number":philhealth_number}
        logging.info(f"Data dictionary complete: emp_id={emp_id}, name={name}")
        try:
            # PHASE 1 FIX: Wrap multi-step operation in transaction for data integrity
            # Note: SQLite in Python already uses implicit transactions, no need for explicit BEGIN

            try:
                logging.info("Entering database transaction...")
                if self.current_employee:
                    logging.info(f"Updating employee {self.current_employee['emp_id']}...")
                    self.db.update_employee(self.current_employee["emp_id"], data)
                    logging.info("Database update complete")

                    # v3.6: Save photo if uploaded
                    if hasattr(self, 'photo_path') and self.photo_path:
                        logging.info(f"Saving photo from {self.photo_path}...")
                        try:
                            # v5.2: Save to new folder structure with standardized filename
                            photos_folder = get_employee_folder(self.current_employee['emp_id'], 'photos')
                            # Use standardized filename 'profile.png' for easy retrieval
                            dest_path = os.path.join(photos_folder, "profile.png")
                            # Convert and save as PNG (background already removed during upload)
                            pixmap = QPixmap(self.photo_path)
                            pixmap.save(dest_path, "PNG")
                            logging.info(f"Photo saved to {dest_path}")
                        except Exception as e:
                            logging.error(f"Failed to save photo: {e}")
                            # Photo save failure is non-critical, continue
                    else:
                        logging.info("No photo to save")

                    # Commit transaction if all succeeded + checkpoint for multi-PC sync
                    logging.info("Committing database transaction with checkpoint...")
                    self.db.commit_and_checkpoint()
                    logging.info("Transaction committed with checkpoint")

                    # Show success message with employee name and automatically return to list
                    employee_name = name
                    logging.info(f"Update successful for {employee_name}")
                    show_success_toast(self, f"✓ Employee Updated Successfully!\n\n{employee_name}\n\nReturning to employee list...")
                    self.current_employee=None
                    self.has_unsaved_changes = False  # Reset unsaved changes flag
                    self._delete_draft()  # Delete any draft after successful save
                    self._release_current_lock()  # Release lock after successful save
                    self.on_saved()  # Switch back to list after updating
                else:
                    logging.info(f"Inserting new employee {emp_id}...")
                    self.db.insert_employee(data)
                    logging.info("Database insert complete")

                    # v3.6: Save photo if uploaded (CRITICAL FIX)
                    if hasattr(self, 'photo_path') and self.photo_path and emp_id:
                        logging.info(f"Saving photo from {self.photo_path}...")
                        try:
                            # v5.2: Save to new folder structure with standardized filename
                            photos_folder = get_employee_folder(emp_id, 'photos')
                            # Use standardized filename 'profile.png' for easy retrieval
                            dest_path = os.path.join(photos_folder, "profile.png")
                            # Convert and save as PNG (background already removed during upload)
                            pixmap = QPixmap(self.photo_path)
                            if not pixmap.isNull():
                                pixmap.save(dest_path, "PNG")
                                logging.info(f"Photo saved to {dest_path}")
                            else:
                                logging.warning("Photo pixmap is null, not saving")
                        except Exception as e:
                            logging.error(f"Failed to save photo: {e}")
                            # Photo save failure is non-critical, continue
                    else:
                        logging.info("No photo to save")

                    # Commit transaction if all succeeded + checkpoint for multi-PC sync
                    logging.info("Committing database transaction with checkpoint...")
                    self.db.commit_and_checkpoint()
                    logging.info("Transaction committed with checkpoint")

                    logging.info("Insert successful, showing success message")
                    show_success_toast(self, "Employee saved successfully!")

                    # Delete draft after successful save
                    self._delete_draft()

                    # Clear form for new entry but stay on form page
                    self.current_employee = None
                    self.has_unsaved_changes = False  # Reset unsaved changes flag
                    for e in self.entries.values():
                        e.clear()
                        self._set_valid(e)
                    self.notes.clear()
                    # v5.1: Clear both files lists
                    self.files_list.clear()
                    if hasattr(self, 'photos_list'): self.photos_list.clear()
                    self._update_file_tab_counts()
                    self.contract_hint.clear()
                    self.still_working.setChecked(True)
                    self.photo_label.setText("📷")
                    self.photo_label.setPixmap(QPixmap())
                    self.photo_path = None  # v3.6: Clear photo path
                    self.contract_user_changed = False
                    self.resign_user_changed = False
                    self.date_entries["hire_date"].setDate(QDate.currentDate())
                    self.date_entries["resign_date"].setDate(QDate.currentDate())
                    self.date_entries["contract_expiry"].setDate(QDate.currentDate())
                    self.date_entries["contract_start_date"].setDate(QDate.currentDate())
                    self.contract_months_combo.combo_box.setCurrentIndex(0)
                    self.no_middle_name_check.setChecked(False)
                    # Reset all dropdowns to default "Select..." state
                    self.agency_combo.setCurrentIndex(0)
                    self.department_combo.combo_box.setCurrentIndex(0)
                    self.department_display.hide()
                    self._reload_agencies_into_combo()
                    self._update_emp_id_preview()
                    self._update_contract_hint()
                    # Refresh data in background but don't switch pages
                    self.on_saved(cancel_only=False, switch_page=False)

            except Exception as e:
                # Rollback transaction on any error
                self.db.conn.rollback()
                logging.error(f"Transaction rolled back due to error: {e}")
                raise  # Re-raise to be caught by outer exception handler
        except PermissionError as ex:
            show_error_toast(self, f"Access denied: {ex}\n\nPlease contact your administrator.")
            logging.error(f"Permission error in _save_employee: {ex}")
            return
        except ValueError as ex:
            show_error_toast(self, f"Invalid data: {ex}")
            logging.error(f"Validation error in _save_employee: {ex}")
            return
        except Exception as ex:
            import traceback
            error_details = traceback.format_exc()
            logging.error(f"Error saving employee: {error_details}")
            show_error_toast(self, f"Could not save employee.\n\nError: {str(ex)}\n\nPlease check:\n• All required fields are filled\n• Employee ID is valid\n• You have permission to save")
            return

    # ============================================================================
    # AUTO-COMPLETE FUNCTIONALITY
    # ============================================================================

    def _setup_autocomplete(self, line_edit, suggestions):
        """Setup autocomplete for a QLineEdit field"""
        completer = QCompleter(suggestions, line_edit)
        completer.setCaseSensitivity(Qt.CaseInsensitive)
        completer.setFilterMode(Qt.MatchContains)
        line_edit.setCompleter(completer)

    def _get_common_positions(self):
        """Get list of common positions from database"""
        try:
            # Get unique positions from existing employees
            positions = set()
            employees = self.db.all_employees()
            for emp in employees:
                pos = emp.get('position', '').strip()
                if pos:
                    positions.add(pos)

            # Add common default positions if database is empty
            if not positions:
                positions = {
                    "Manager", "Assistant Manager", "Supervisor",
                    "Sales Associate", "Cashier", "Stock Clerk",
                    "Security Guard", "Janitor", "Admin Staff",
                    "HR Officer", "Accounting Staff", "IT Support",
                    "Warehouse Staff", "Driver", "Customer Service",
                    "Team Leader", "Store Keeper", "Maintenance"
                }

            return sorted(list(positions))
        except:
            # Fallback to default list
            return [
                "Manager", "Assistant Manager", "Supervisor",
                "Sales Associate", "Cashier", "Stock Clerk",
                "Security Guard", "Janitor", "Admin Staff"
            ]

    def _get_common_agencies(self):
        """Get list of common agencies from database"""
        try:
            # Get unique agencies from existing employees
            agencies = set()
            employees = self.db.all_employees()
            for emp in employees:
                agency = emp.get('agency', '').strip()
                if agency and agency != 'Regular':
                    agencies.add(agency)

            # Add common default agencies if database is empty
            if not agencies:
                agencies = {
                    "ABC Manpower Services",
                    "XYZ Staffing Solutions",
                    "Professional Outsourcing Inc",
                    "Elite Employment Agency",
                    "Premier Workforce Services"
                }

            return sorted(list(agencies))
        except:
            # Fallback to default list
            return [
                "ABC Manpower Services",
                "XYZ Staffing Solutions",
                "Professional Outsourcing Inc"
            ]

    # ==================== RECORD LOCKING METHODS ====================

    def _acquire_lock(self, emp_id: str) -> bool:
        """Acquire lock for editing employee. Returns True if successful, False if locked by another user."""
        try:
            from PySide6.QtCore import QTimer
            from PySide6.QtWidgets import QMessageBox

            record_id = f"employee:{emp_id}"

            # Try to acquire lock
            if self.db.acquire_lock(record_id, self.current_user):
                # Lock acquired successfully
                self.current_lock_id = record_id

                # Set up lock refresh timer (every 5 minutes)
                if self.lock_refresh_timer:
                    self.lock_refresh_timer.stop()

                self.lock_refresh_timer = QTimer(self)
                self.lock_refresh_timer.timeout.connect(lambda: self._refresh_lock())
                self.lock_refresh_timer.start(300000)  # 5 minutes

                return True
            else:
                # Lock is held by someone else
                lock_info = self.db.get_lock_info(record_id)
                if lock_info and lock_info.get('locked'):
                    locked_by = lock_info.get('locked_by', 'another user')
                    computer = lock_info.get('computer_name', '')
                    computer_info = f" on {computer}" if computer else ""

                    # Show warning dialog
                    reply = QMessageBox.warning(
                        self,
                        "Employee Locked",
                        f"This employee is currently being edited by {locked_by}{computer_info}.\n\n"
                        f"Please wait for them to finish or contact them to release the lock.\n\n"
                        f"Would you like to open in read-only mode instead?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.No
                    )

                    if reply == QMessageBox.StandardButton.Yes:
                        # User wants read-only mode
                        # Set all fields to read-only
                        for entry in self.entries.values():
                            if hasattr(entry, 'setReadOnly'):
                                entry.setReadOnly(True)
                        # Disable save buttons
                        # (This would need access to the save button widgets)
                        return True  # Allow loading in read-only mode

                return False

        except Exception as e:
            import logging
            logging.error(f"Error acquiring lock for {emp_id}: {e}")
            return True  # Allow editing on error (fail open)

    def _refresh_lock(self):
        """Refresh the current lock to prevent expiry"""
        if self.current_lock_id:
            self.db.refresh_lock(self.current_lock_id, self.current_user)

    def _release_current_lock(self):
        """Release the current record lock"""
        if self.current_lock_id:
            self.db.release_lock(self.current_lock_id, self.current_user)
            self.current_lock_id = None

        if self.lock_refresh_timer:
            self.lock_refresh_timer.stop()
            self.lock_refresh_timer = None

    def closeEvent(self, event):
        """Release lock when form is closed"""
        self._release_current_lock()
        super().closeEvent(event)

    # ==================== END RECORD LOCKING METHODS ====================


