"""
UI Widgets for Employee Vault
Contains custom widget classes used throughout the application
"""

import os
import re
import math
import logging
import subprocess
from datetime import datetime, date
from typing import Optional, Callable, Tuple, List, Dict, Any

from PySide6.QtWidgets import (
    QWidget, QLabel, QLineEdit, QPushButton, QVBoxLayout, QHBoxLayout,
    QDialog, QTextEdit, QMessageBox, QFrame, QScrollArea, QCalendarWidget,
    QDateEdit, QTableWidget, QTableWidgetItem, QHeaderView, QMenu, QGridLayout,
    QSizePolicy, QApplication, QToolButton, QSpinBox, QStackedWidget, QGraphicsOpacityEffect,
    QComboBox, QAbstractSpinBox, QListWidget, QListWidgetItem, QGraphicsDropShadowEffect,
    QGraphicsBlurEffect
)
from PySide6.QtCore import (
    Qt, Signal, QSize, QRect, QRectF, QDate, QObject, QEvent, QPoint, QTimer,
    QPropertyAnimation, QEasingCurve, Property
)
from PySide6.QtGui import (
    QPainter, QPainterPath, QColor, QPalette, QTextCursor, QPixmap,
    QFont, QLinearGradient, QPen, QBrush, QRadialGradient
)

from employee_vault.config import PASSWORD_MIN_LENGTH, validate_password_strength, load_theme_preference, MODERN_THEMES
from employee_vault.database import DB

# ======================================================================
# GLOBAL CURSOR UTILITIES - Remove Distracting Hand Cursor
# ======================================================================

class NoCursorEventFilter(QObject):
    """
    Global event filter to prevent cursor changes - keeps arrow cursor always
    Apply this to any widget viewport to prevent hand cursor on hover
    """
    def eventFilter(self, obj, event):
        if event.type() in (QEvent.Enter, QEvent.HoverMove, QEvent.MouseMove):
            # Force arrow cursor on hover/mouse events
            obj.setCursor(Qt.ArrowCursor)
        return super().eventFilter(obj, event)


def disable_cursor_changes(widget):
    """
    Helper function to disable cursor changes on any widget
    Applies NoCursorEventFilter to widget and its viewport (if it has one)

    Usage:
        table = QTableView()
        disable_cursor_changes(table)
    """
    # Create and apply filter
    cursor_filter = NoCursorEventFilter()
    widget.installEventFilter(cursor_filter)

    # Also apply to viewport if widget has one (for tables/list widgets)
    if hasattr(widget, 'viewport'):
        widget.viewport().installEventFilter(cursor_filter)

    # Set initial cursor
    widget.setCursor(Qt.ArrowCursor)
    if hasattr(widget, 'viewport'):
        widget.viewport().setCursor(Qt.ArrowCursor)

    # Store filter as attribute to prevent garbage collection
    if not hasattr(widget, '_cursor_filters'):
        widget._cursor_filters = []
    widget._cursor_filters.append(cursor_filter)


def remove_focus_rectangle(widget):
    """
    Remove the distracting focus rectangle (rounded border) from table/list widgets
    v4.5.0: Global fix for item focus rectangle

    This removes the visual border that appears around the "current" cell
    while still allowing keyboard navigation (arrow keys still work)

    Usage:
        table = QTableView()
        remove_focus_rectangle(table)
    """
    # Get existing stylesheet
    existing_style = widget.styleSheet()

    # Focus rectangle removal CSS
    focus_fix_css = """
        QTableView {
            outline: 0;
        }
        QTableView::item:focus {
            border: none;
            outline: none;
        }
        QTableWidget {
            outline: 0;
        }
        QTableWidget::item:focus {
            border: none;
            outline: none;
        }
        QListWidget {
            outline: 0;
        }
        QListWidget::item:focus {
            border: none;
            outline: none;
        }
    """

    # Combine existing styles with focus fix
    if existing_style:
        widget.setStyleSheet(existing_style + "\n" + focus_fix_css)
    else:
        widget.setStyleSheet(focus_fix_css)


def apply_table_fixes(widget):
    """
    Apply all table/list widget fixes globally
    v4.5.0: Combined fix for cursor and focus rectangle

    Usage:
        table = QTableView()
        apply_table_fixes(table)
    """
    disable_cursor_changes(widget)
    remove_focus_rectangle(widget)

# ======================================================================
# ValidatedLineEdit
# ======================================================================

class ValidatedLineEdit(QLineEdit):
    """QLineEdit with real-time validation and visual feedback"""

    validationChanged = Signal(bool, str)  # Signal: (is_valid, error_message)

    def __init__(self, validator_func=None, auto_format_func=None, placeholder="", parent=None):
        super().__init__(parent)
        self.validator_func = validator_func
        self.auto_format_func = auto_format_func
        self.error_label = None
        self.setPlaceholderText(placeholder)

        if validator_func:
            self.textChanged.connect(self._validate)

        if auto_format_func:
            self.textChanged.connect(self._auto_format)

    def set_error_label(self, label: QLabel):
        """Set the error label to show validation messages"""
        self.error_label = label

    def _auto_format(self):
        """Auto-format text as user types"""
        if not self.auto_format_func:
            return

        current_text = self.text()
        cursor_pos = self.cursorPosition()

        formatted = self.auto_format_func(current_text)

        if formatted != current_text:
            self.blockSignals(True)
            self.setText(formatted)
            self.setCursorPosition(min(cursor_pos + (len(formatted) - len(current_text)), len(formatted)))
            self.blockSignals(False)

    def _validate(self):
        """Perform validation and update UI"""
        if not self.validator_func:
            return

        text = self.text().strip()
        is_valid, error_msg = self.validator_func(text)

        # Update visual feedback
        if not text:
            # Empty field - neutral state
            self.setProperty("validation", "")
            if self.error_label:
                self.error_label.setText("")
                self.error_label.hide()
        elif is_valid:
            # Valid input
            self.setProperty("validation", "valid")
            if self.error_label:
                self.error_label.setText("")
                self.error_label.hide()
        else:
            # Invalid input
            self.setProperty("validation", "invalid")
            if self.error_label:
                self.error_label.setText(f"⚠️ {error_msg}")
                self.error_label.setProperty("error", "true")
                self.error_label.show()

        # Refresh style
        self.style().unpolish(self)
        self.style().polish(self)

        # Emit signal
        self.validationChanged.emit(is_valid or not text, error_msg)

    def is_valid(self) -> Tuple[bool, str]:
        """Check if current value is valid"""
        if not self.validator_func:
            return True, ""

        text = self.text().strip()
        if not text:
            return True, ""  # Empty is valid unless required

        return self.validator_func(text)


# ==================== USER-FRIENDLY ERROR DIALOGS (Quick Win #2) ====================

class UserFriendlyMessageBox:
    """Show user-friendly error messages with helpful suggestions"""

    @staticmethod
    def show_error(parent, title: str, message: str, details: str = "", suggestion: str = ""):
        """Show error dialog with optional technical details and suggestions"""
        msg = QMessageBox(parent)
        msg.setIcon(QMessageBox.Critical)
        msg.setWindowTitle(f"❌ {title}")

        # Main message (user-friendly)
        msg.setText(message)

        # Add suggestion if provided
        if suggestion:
            msg.setInformativeText(f"\n💡 Suggestion: {suggestion}")

        # Add technical details if provided (hidden by default)
        if details:
            msg.setDetailedText(f"Technical details:\n{details}")

        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec()

        # Log the error
        logging.error(f"{title}: {message} | Details: {details}")

    @staticmethod
    def show_warning(parent, title: str, message: str, suggestion: str = ""):
        """Show warning dialog with suggestion"""
        msg = QMessageBox(parent)
        msg.setIcon(QMessageBox.Warning)
        msg.setWindowTitle(f"⚠️ {title}")
        msg.setText(message)

        if suggestion:
            msg.setInformativeText(f"\n💡 Suggestion: {suggestion}")

        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec()

        logging.warning(f"{title}: {message}")

    @staticmethod
    def show_success(parent, title: str, message: str):
        """Show success notification"""
        msg = QMessageBox(parent)
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle(f"✅ {title}")
        msg.setText(message)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec()

    @staticmethod
    def confirm_action(parent, title: str, message: str, danger: bool = False) -> bool:
        """Ask for confirmation before potentially destructive action"""
        msg = QMessageBox(parent)
        msg.setIcon(QMessageBox.Question if not danger else QMessageBox.Warning)
        msg.setWindowTitle(f"❓ {title}")
        msg.setText(message)

        if danger:
            msg.setInformativeText("\n⚠️ This action cannot be undone!")
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msg.setDefaultButton(QMessageBox.No)
        else:
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            msg.setDefaultButton(QMessageBox.Yes)

        return msg.exec() == QMessageBox.Yes


# ==================== DATA INTEGRITY CHECKER (Quick Win #3) ====================

class DataIntegrityChecker:
    """Check and repair data integrity issues"""

    def __init__(self, db):
        self.db = db
        self.issues = []

    def check_all(self) -> List[Dict[str, Any]]:
        """Run all integrity checks"""
        self.issues = []

        self.issues.extend(self._check_duplicates())
        self.issues.extend(self._check_orphaned_files())
        self.issues.extend(self._check_invalid_dates())
        self.issues.extend(self._check_database_corruption())

        return self.issues

    def _check_duplicates(self) -> List[Dict[str, Any]]:
        """Check for duplicate employees by name or ID"""
        issues = []

        try:
            # Check duplicate names
            cursor = self.db.conn.execute("""
                SELECT fullname, COUNT(*) as count
                FROM employees
                WHERE fullname IS NOT NULL AND fullname != ''
                GROUP BY LOWER(fullname)
                HAVING count > 1
            """)

            for row in cursor.fetchall():
                issues.append({
                    'type': 'duplicate_name',
                    'severity': 'warning',
                    'message': f"Duplicate name found: '{row[0]}' appears {row[1]} times",
                    'suggestion': "Check if these are the same person or need unique identifiers"
                })

            # Check duplicate SSS numbers
            cursor = self.db.conn.execute("""
                SELECT sss_number, COUNT(*) as count
                FROM employees
                WHERE sss_number IS NOT NULL AND sss_number != ''
                GROUP BY sss_number
                HAVING count > 1
            """)

            for row in cursor.fetchall():
                issues.append({
                    'type': 'duplicate_sss',
                    'severity': 'error',
                    'message': f"Duplicate SSS number: '{row[0]}' used {row[1]} times",
                    'suggestion': "SSS numbers should be unique. Please verify and correct."
                })

            # Check duplicate TIN numbers
            cursor = self.db.conn.execute("""
                SELECT tin_number, COUNT(*) as count
                FROM employees
                WHERE tin_number IS NOT NULL AND tin_number != ''
                GROUP BY tin_number
                HAVING count > 1
            """)

            for row in cursor.fetchall():
                issues.append({
                    'type': 'duplicate_tin',
                    'severity': 'error',
                    'message': f"Duplicate TIN number: '{row[0]}' used {row[1]} times",
                    'suggestion': "TIN numbers should be unique. Please verify and correct."
                })

        except Exception as e:
            logging.error(f"Error checking duplicates: {e}")

        return issues

    def _check_orphaned_files(self) -> List[Dict[str, Any]]:
        """Check for files in the filesystem not referenced in DB"""
        issues = []

        try:
            # Get all file references from database
            cursor = self.db.conn.execute("SELECT photo_path FROM employees WHERE photo_path IS NOT NULL")
            db_photos = set(row[0] for row in cursor.fetchall())

            cursor = self.db.conn.execute("""
                SELECT attached_files FROM employees
                WHERE attached_files IS NOT NULL AND attached_files != ''
            """)
            db_files = set()
            for row in cursor.fetchall():
                if row[0]:
                    try:
                        files = json.loads(row[0])
                        db_files.update(files)
                    except:
                        pass

            # Check actual files
            orphaned_photos = []
            if os.path.exists(PHOTOS_DIR):
                for filename in os.listdir(PHOTOS_DIR):
                    filepath = os.path.join(PHOTOS_DIR, filename)
                    if filename not in db_photos and os.path.isfile(filepath):
                        orphaned_photos.append(filename)

            if orphaned_photos:
                issues.append({
                    'type': 'orphaned_photos',
                    'severity': 'info',
                    'message': f"Found {len(orphaned_photos)} photo(s) not linked to any employee",
                    'suggestion': "These files can be safely deleted to save space",
                    'files': orphaned_photos
                })

            orphaned_files = []
            if os.path.exists(FILES_DIR):
                for filename in os.listdir(FILES_DIR):
                    filepath = os.path.join(FILES_DIR, filename)
                    if filename not in db_files and os.path.isfile(filepath):
                        orphaned_files.append(filename)

            if orphaned_files:
                issues.append({
                    'type': 'orphaned_files',
                    'severity': 'info',
                    'message': f"Found {len(orphaned_files)} file(s) not linked to any employee",
                    'suggestion': "These files can be safely deleted to save space",
                    'files': orphaned_files
                })

        except Exception as e:
            logging.error(f"Error checking orphaned files: {e}")

        return issues

    def _check_invalid_dates(self) -> List[Dict[str, Any]]:
        """Check for invalid or suspicious dates"""
        issues = []

        try:
            # Check for contracts ending before they started
            cursor = self.db.conn.execute("""
                SELECT id, fullname, contract_start, contract_end
                FROM employees
                WHERE contract_start IS NOT NULL
                AND contract_end IS NOT NULL
                AND contract_end < contract_start
            """)

            for row in cursor.fetchall():
                issues.append({
                    'type': 'invalid_date_range',
                    'severity': 'error',
                    'message': f"{row[1]}: Contract ends before it starts ({row[2]} to {row[3]})",
                    'suggestion': "Please check and correct the contract dates",
                    'employee_id': row[0]
                })

            # Check for birthdates in the future
            today = datetime.now().strftime("%Y-%m-%d")
            cursor = self.db.conn.execute("""
                SELECT id, fullname, birthdate
                FROM employees
                WHERE birthdate IS NOT NULL
                AND birthdate > ?
            """, (today,))

            for row in cursor.fetchall():
                issues.append({
                    'type': 'future_birthdate',
                    'severity': 'error',
                    'message': f"{row[1]}: Birthdate is in the future ({row[2]})",
                    'suggestion': "Please correct the birthdate",
                    'employee_id': row[0]
                })

            # Check for suspiciously old ages (> 100 years)
            old_date = (datetime.now() - timedelta(days=365*100)).strftime("%Y-%m-%d")
            cursor = self.db.conn.execute("""
                SELECT id, fullname, birthdate
                FROM employees
                WHERE birthdate IS NOT NULL
                AND birthdate < ?
            """, (old_date,))

            for row in cursor.fetchall():
                issues.append({
                    'type': 'suspicious_birthdate',
                    'severity': 'warning',
                    'message': f"{row[1]}: Age appears to be over 100 years ({row[2]})",
                    'suggestion': "Please verify this birthdate is correct",
                    'employee_id': row[0]
                })

        except Exception as e:
            logging.error(f"Error checking dates: {e}")

        return issues

    def _check_database_corruption(self) -> List[Dict[str, Any]]:
        """Check for database corruption"""
        issues = []

        try:
            # Run SQLite integrity check
            cursor = self.db.conn.execute("PRAGMA integrity_check")
            result = cursor.fetchone()[0]

            if result != "ok":
                issues.append({
                    'type': 'database_corruption',
                    'severity': 'critical',
                    'message': "Database integrity check failed",
                    'suggestion': "Please contact support immediately. A backup restore may be needed.",
                    'details': result
                })

        except Exception as e:
            logging.error(f"Error checking database integrity: {e}")
            issues.append({
                'type': 'integrity_check_failed',
                'severity': 'error',
                'message': "Could not perform database integrity check",
                'suggestion': "Please contact support",
                'details': str(e)
                })

        return issues

    def repair_orphaned_files(self, orphaned_list: List[str], file_type: str) -> int:
        """Delete orphaned files"""
        deleted_count = 0
        directory = PHOTOS_DIR if file_type == 'photos' else FILES_DIR

        for filename in orphaned_list:
            try:
                filepath = os.path.join(directory, filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                    deleted_count += 1
                    logging.info(f"Deleted orphaned file: {filepath}")
            except Exception as e:
                logging.error(f"Failed to delete {filepath}: {e}")

        return deleted_count


class IntegrityCheckDialog(QDialog):
    """Dialog to show integrity check results and offer repairs"""

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.checker = DataIntegrityChecker(db)
        self.setWindowTitle("🔍 Data Integrity Check")
        self.setMinimumSize(700, 500)

        layout = QVBoxLayout(self)

        header = QLabel("<h2>🔍 Database Integrity Check</h2>")
        layout.addWidget(header)

        info = QLabel("Checking for data consistency issues...")
        layout.addWidget(info)

        self.progress = QProgressBar()
        self.progress.setRange(0, 0)
        layout.addWidget(self.progress)

        self.results_widget = QWidget()
        self.results_layout = QVBoxLayout(self.results_widget)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.results_widget)
        layout.addWidget(scroll, 1)

        btn_layout = QHBoxLayout()
        self.refresh_btn = QPushButton("🔄 Recheck")
        self.refresh_btn.clicked.connect(self._run_check)
        self.repair_btn = QPushButton("🔧 Repair Issues")
        self.repair_btn.setEnabled(False)
        self.repair_btn.clicked.connect(self._repair_issues)
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)

        btn_layout.addWidget(self.refresh_btn)
        btn_layout.addWidget(self.repair_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        QTimer.singleShot(500, self._run_check)

    def _run_check(self):
        self.progress.setRange(0, 0)
        self.refresh_btn.setEnabled(False)
        self.repair_btn.setEnabled(False)

        while self.results_layout.count():
            item = self.results_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        import threading
        def check_thread():
            issues = self.checker.check_all()
            QTimer.singleShot(0, lambda: self._show_results(issues))

        threading.Thread(target=check_thread, daemon=True).start()

    def _show_results(self, issues):
        self.progress.setRange(0, 1)
        self.progress.setValue(1)
        self.refresh_btn.setEnabled(True)

        if not issues:
            success_label = QLabel("✅ <b>No integrity issues found!</b>")
            success_label.setStyleSheet("color: #52c41a; font-size: 16px; padding: 20px;")
            self.results_layout.addWidget(success_label)

            detail_label = QLabel("Your database is in good condition. All checks passed successfully.")
            self.results_layout.addWidget(detail_label)
        else:
            critical = [i for i in issues if i['severity'] == 'critical']
            errors = [i for i in issues if i['severity'] == 'error']
            warnings = [i for i in issues if i['severity'] == 'warning']
            info = [i for i in issues if i['severity'] == 'info']

            summary = QLabel(f"<b>Found {len(issues)} issue(s):</b> "
                           f"{len(critical)} critical, {len(errors)} errors, "
                           f"{len(warnings)} warnings, {len(info)} info")
            summary.setStyleSheet("font-size: 14px; padding: 10px;")
            self.results_layout.addWidget(summary)

            for issue in critical + errors + warnings + info:
                self._add_issue_widget(issue)

            repairable = [i for i in issues if i['type'] in ['orphaned_photos', 'orphaned_files']]
            if repairable:
                self.repair_btn.setEnabled(True)

        self.results_layout.addStretch()

    def _add_issue_widget(self, issue):
        frame = QFrame()
        frame.setFrameShape(QFrame.StyledPanel)

        colors = {'critical': '#ff4d4f', 'error': '#ff7875', 'warning': '#ffa940', 'info': '#4a8cff'}
        color = colors.get(issue['severity'], '#666')
        frame.setStyleSheet(f"QFrame {{ border-left: 4px solid {color}; padding: 10px; margin: 5px; }}")

        layout = QVBoxLayout(frame)

        severity_icons = {'critical': '🔴', 'error': '❌', 'warning': '⚠️', 'info': 'ℹ️'}
        icon = severity_icons.get(issue['severity'], '•')

        title = QLabel(f"{icon} <b>{issue['severity'].upper()}</b>: {issue['message']}")
        title.setWordWrap(True)
        layout.addWidget(title)

        if issue.get('suggestion'):
            suggestion = QLabel(f"💡 {issue['suggestion']}")
            suggestion.setStyleSheet("color: #888; margin-left: 20px;")
            suggestion.setWordWrap(True)
            layout.addWidget(suggestion)

        self.results_layout.addWidget(frame)

    def _repair_issues(self):
        if not UserFriendlyMessageBox.confirm_action(
            self, "Repair Data Issues",
            "This will delete orphaned files that are not linked to any employee.\n\nAre you sure?",
            danger=False
        ):
            return

        issues = self.checker.check_all()
        repaired_count = 0

        for issue in issues:
            if issue['type'] == 'orphaned_photos':
                deleted = self.checker.repair_orphaned_files(issue.get('files', []), 'photos')
                repaired_count += deleted
            elif issue['type'] == 'orphaned_files':
                deleted = self.checker.repair_orphaned_files(issue.get('files', []), 'files')
                repaired_count += deleted

        if repaired_count > 0:
            UserFriendlyMessageBox.show_success(self, "Repair Complete",
                f"Successfully deleted {repaired_count} orphaned file(s).")
            self._run_check()
        else:
            UserFriendlyMessageBox.show_warning(self, "Nothing to Repair", "No repairable issues found.")



# Continue in next part...

# ==================== END QUICK WINS ENHANCEMENTS ====================

def contract_days_left(emp: dict):
    ce = (emp.get("contract_expiry") or "").strip()
    if not ce: return None
    try:
        dt = datetime.strptime(ce, "%m-%d-%Y").date()
        return (dt - datetime.now().date()).days
    except Exception:
        return None

def titlecase(text: str) -> str:
    def tc(w): return w[:1].upper() + w[1:].lower() if w else w
    parts = re.split(r"(\s+|-|'|/)", text.strip())
    return ''.join(tc(p) if i % 2 == 0 else p for i, p in enumerate(parts))


# ============================================================================
# v3.6: GLOBAL PRINTER SELECTION HELPER
# ============================================================================

def select_printer(parent=None):
    """
    Show printer selection dialog and return configured printer.
    Returns None if user cancels.
    """
    from PySide6.QtPrintSupport import QPrinter, QPrintDialog
    
    try:
        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageOrientation(QPrinter.PageOrientation.Portrait)
        
        dialog = QPrintDialog(printer, parent)
        dialog.setWindowTitle("🖨️ Select Printer")
        
        if dialog.exec() == QPrintDialog.Accepted:
            return printer
        else:
            return None
            
    except Exception as e:
        logging.error(f"Printer selection error: {e}")
        return None


# ============================================================================
# END PRINTER HELPER
# ============================================================================

# ============================================================================
# v2.1: SECURE PASSWORD HASHING - REMOVED (use from config.py instead)
# ============================================================================
# Password hashing functions have been consolidated in employee_vault.config
# Import with: from employee_vault.config import _hash_pwd, _verify_pwd, _needs_password_rehash

# ============================================================================
# v2.1: DATABASE RETRY DECORATOR
# ============================================================================

import time
from functools import wraps

def retry_on_lock(max_attempts=3, delay=0.5):
    """
    Retry database operations if database is locked.

    Args:
        max_attempts: Maximum number of retry attempts
        delay: Base delay between retries (exponential backoff)

    Returns:
        Decorated function with retry logic
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_attempts):
                try:
                    return func(*args, **kwargs)
                except sqlite3.OperationalError as e:
                    if "locked" in str(e).lower() and attempt < max_attempts - 1:
                        wait_time = delay * (2 ** attempt)  # Exponential backoff
                        logging.warning(f"Database locked, retrying in {wait_time}s... (attempt {attempt + 1}/{max_attempts})")
                        time.sleep(wait_time)
                        continue
                    raise
            return None
        return wrapper
    return decorator

# ============================================================================
# WEEK 1 PRIORITY #3: DB-LAYER PERMISSION CHECKS - REMOVED (use from utils/decorators.py instead)
# ============================================================================
# Permission checking decorator has been consolidated in employee_vault.utils.decorators
# Import with: from employee_vault.utils import check_permission
#
# Database layer applies this decorator to enforce permissions (defense in depth)
# Never trust UI-only checks - always validate at DB layer

def normalize_ph_phone(raw: str) -> str:
    s = re.sub(r'\D', '', raw or '')
    if not s: return ""
    if len(s)==11 and s.startswith('09'): s='63'+s[1:]
    elif len(s)==10 and s.startswith('9'): s='63'+s
    elif len(s)==12 and s.startswith('63') and s[2]=='9': pass
    if len(s)==12 and s.startswith('63') and s[2]=='9': return f"+63 {s[2]}{s[3]}{s[4]} {s[5]}{s[6]}{s[7]} {s[8]}{s[9]}{s[10]}{s[11]}"
    if s.startswith('63') and len(s) in (10,11): return "+{} {}".format(s[:2], s[2:])
    return raw.strip()

def show_tooltip(widget, message: str):
    widget.setToolTip(message)
    QToolTip.showText(widget.mapToGlobal(widget.rect().bottomLeft()) + QPoint(0, 6), message, widget, widget.rect(), 3000)


def db_latest_mtime(db_path: str) -> float:
    """Get the latest modification time of the database file"""
    try:
        if os.path.exists(db_path):
            return os.path.getmtime(db_path)
        return 0.0
    except Exception:
        return 0.0

# ============================================================================
# PHASE 1: ENHANCED VALIDATION FUNCTIONS
# ============================================================================

def validate_email(email: str) -> Tuple[bool, str]:
    """
    Enhanced email validation with detailed error messages

    Args:
        email: Email address to validate

    Returns:
        Tuple of (is_valid, error_message)
    """
    if not email or not email.strip():
        return True, ""  # Empty is valid (optional field)

    email = email.strip()

    # Check basic format
    if '@' not in email:
        return False, "Email must contain @"

    if email.count('@') > 1:
        return False, "Email can only have one @"

    # Check pattern
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(pattern, email):
        return False, "Invalid email format (e.g., user@example.com)"

    return True, ""

def validate_phone(phone: str) -> Tuple[bool, str]:
    """
    Enhanced phone validation for Philippine numbers

    Args:
        phone: Phone number to validate

    Returns:
        Tuple of (is_valid, error_message)
    """
    if not phone or not phone.strip():
        return True, ""  # Empty is valid (optional field)

    # Remove common separators
    cleaned = re.sub(r'[\s\-\(\)\+]', '', phone)

    if not cleaned.isdigit():
        return False, "Phone must contain only numbers"

    if len(cleaned) < 7:
        return False, "Phone number too short (min 7 digits)"

    if len(cleaned) > 15:
        return False, "Phone number too long (max 15 digits)"

    return True, ""

# ============================================================================
# WEEK 1 PRIORITY #1: PASSWORD STRENGTH VALIDATION (CRITICAL SECURITY)
# ============================================================================

def validate_password_strength(password: str) -> Tuple[bool, str, str]:
    """
    Validate password strength with detailed feedback.

    CRITICAL SECURITY: Enforces strong passwords (changed from 4 to 10+ chars)

    Requirements:
    - Minimum 10 characters (was 4 - DANGEROUS!)
    - At least one uppercase letter
    - At least one lowercase letter
    - At least one number
    - At least one special character

    Args:
        password: Password to validate

    Returns:
        Tuple of (is_valid, error_message, strength_label)
        strength_label: "None", "Weak", "Medium", "Strong", "Very Strong"
    """
    if not password:
        return False, "Password is required", "None"

    MIN_PASSWORD_LENGTH = 10  # CRITICAL CHANGE: was 4, now 10

    if len(password) < MIN_PASSWORD_LENGTH:
        return False, f"Password must be at least {MIN_PASSWORD_LENGTH} characters", "Weak"

    has_upper = bool(re.search(r'[A-Z]', password))
    has_lower = bool(re.search(r'[a-z]', password))
    has_digit = bool(re.search(r'[0-9]', password))
    has_special = bool(re.search(r'[!@#$%^&*()_+\-=\[\]{};:,.<>?/\\|`~]', password))

    missing = []
    if not has_upper: missing.append("uppercase letter (A-Z)")
    if not has_lower: missing.append("lowercase letter (a-z)")
    if not has_digit: missing.append("number (0-9)")
    if not has_special: missing.append("special character (!@#$%^&* etc.)")

    if missing:
        return False, f"Password must contain at least one: {', '.join(missing)}", "Weak"

    # Calculate strength score
    score = len(password) // 2 + sum([has_upper, has_lower, has_digit, has_special]) * 5

    if score < 20: strength = "Medium"
    elif score < 30: strength = "Strong"
    else: strength = "Very Strong"

    return True, "", strength

def check_password_strength_ui(password: str) -> str:
    """
    Check password strength for real-time UI feedback (no error messages).

    Args:
        password: Password to check

    Returns:
        Strength label: "None", "Weak", "Medium", "Strong", "Very Strong"
    """
    if not password: return "None"
    if len(password) < 6: return "Weak"

    has_upper = bool(re.search(r'[A-Z]', password))
    has_lower = bool(re.search(r'[a-z]', password))
    has_digit = bool(re.search(r'[0-9]', password))
    has_special = bool(re.search(r'[!@#$%^&*()_+\-=\[\]{};:,.<>?/\\|`~]', password))

    score = len(password) + sum([has_upper, has_lower, has_digit, has_special]) * 5

    if len(password) < 10: return "Weak"
    elif score < 20: return "Medium"
    elif score < 30: return "Strong"
    else: return "Very Strong"

# ============================================================================
# WEEK 1 PRIORITY #2: FILE PATH TRAVERSAL PROTECTION (CRITICAL SECURITY)
# ============================================================================

def safe_file_path(base_dir: str, filename: str, allow_subdirs: bool = True) -> str:
    """
    Sanitize file path to prevent directory traversal attacks.

    CRITICAL SECURITY: Prevents ../../../etc/passwd style attacks

    Args:
        base_dir: Base directory that file must be within
        filename: Requested filename (may contain path components)
        allow_subdirs: If True, allows subdirectories within base_dir

    Returns:
        Safe absolute path within base_dir

    Raises:
        ValueError: If path attempts to escape base_dir or contains illegal characters
    """
    # Remove any null bytes (common attack vector)
    filename = filename.replace('\x00', '')

    # Block absolute paths
    if os.path.isabs(filename):
        raise ValueError("Absolute paths not allowed")

    # Block dangerous patterns
    dangerous = ['..', '~', '$', '`', '<', '>', '|', ';', '&', '\n', '\r']
    for danger in dangerous:
        if danger in filename:
            raise ValueError(f"Illegal character in filename: {danger}")

    # Normalize path and ensure it's within base_dir
    safe_base = os.path.abspath(base_dir)
    requested_path = os.path.normpath(os.path.join(safe_base, filename))

    # CRITICAL CHECK: Ensure resolved path is still within base_dir
    if not requested_path.startswith(safe_base + os.sep) and requested_path != safe_base:
        raise ValueError("Path traversal detected - attempted escape from base directory")

    # If subdirectories not allowed, ensure file is directly in base_dir
    if not allow_subdirs:
        if os.path.dirname(requested_path) != safe_base:
            raise ValueError("Subdirectories not allowed")

    return requested_path

def validate_id_number(id_num: str, field_name: str = "ID") -> Tuple[bool, str]:
    """
    Enhanced ID number validation

    Args:
        id_num: ID number to validate
        field_name: Name of the field for error messages

    Returns:
        Tuple of (is_valid, error_message)
    """
    if not id_num or not id_num.strip():
        return True, ""  # Empty is valid (optional field)

    id_num = id_num.strip()

    if len(id_num) < 5:
        return False, f"{field_name} too short (min 5 characters)"

    if len(id_num) > 20:
        return False, f"{field_name} too long (max 20 characters)"

    # Allow alphanumeric and dashes only
    if not re.match(r'^[A-Za-z0-9\-]+$', id_num):
        return False, f"{field_name} can only contain letters, numbers, and dashes"

    return True, ""

def validate_salary(salary: str) -> Tuple[bool, Optional[float], str]:
    """
    Enhanced salary validation with range checking

    Args:
        salary: Salary value to validate

    Returns:
        Tuple of (is_valid, parsed_value, error_message)
    """
    if not salary or not salary.strip():
        return True, 0.0, ""  # Empty is valid

    try:
        # Remove commas and parse
        value = float(salary.replace(',', '').strip())

        if value < 0:
            return False, None, "Salary cannot be negative"

        if value > 10000000:  # 10 million max
            return False, None, "Salary exceeds maximum (10,000,000)"

        return True, value, ""
    except ValueError:
        return False, None, "Salary must be a valid number"

def validate_required_field(value: str, field_name: str) -> Tuple[bool, str]:
    """
    Validate required fields

    Args:
        value: Field value to validate
        field_name: Name of the field for error messages

    Returns:
        Tuple of (is_valid, error_message)
    """
    if not value or not value.strip():
        return False, f"{field_name} is required"

    if len(value.strip()) < 2:
        return False, f"{field_name} must be at least 2 characters"

    return True, ""

# ============================================================================
# PHASE 1: SAFE FILE OPERATION WRAPPER
# ============================================================================

def safe_file_operation(operation: str, func, *args, **kwargs) -> Any:
    """
    Wrapper for safe file operations with comprehensive error handling

    Args:
        operation: Description of the operation for logging
        func: Function to execute
        *args: Arguments for the function
        **kwargs: Keyword arguments for the function

    Returns:
        Result of the function or None on error
    """
    try:
        result = func(*args, **kwargs)
        logging.info(f"{operation} - Success")
        return result
    except FileNotFoundError as e:
        logging.error(f"{operation} - File not found: {e}")
        QMessageBox.warning(None, "File Not Found", f"File not found:\n{e}")
        return None
    except PermissionError as e:
        logging.error(f"{operation} - Permission denied: {e}")
        QMessageBox.warning(None, "Permission Denied",
                          f"Cannot access file (permission denied):\n{e}")
        return None
    except IOError as e:
        logging.error(f"{operation} - IO error: {e}")
        QMessageBox.warning(None, "File Error", f"File operation failed:\n{e}")
        return None
    except Exception as e:
        logging.error(f"{operation} - Unexpected error: {e}", exc_info=True)
        QMessageBox.warning(None, "Error", f"Operation failed:\n{e}")
        return None



# ============================================================================
# PHASE 2: EXCEL/CSV EXPORT FUNCTIONS
# ============================================================================

def export_to_csv(employees: list, filename: str) -> bool:
    """Export employees to CSV file"""
    try:
        import csv
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            if not employees:
                return False

            # Get all field names from first employee
            fieldnames = list(employees[0].keys())
            writer = csv.DictWriter(f, fieldnames=fieldnames)

            writer.writeheader()
            writer.writerows(employees)

        logging.info(f"Exported {len(employees)} employees to CSV: {filename}")
        return True
    except Exception as e:
        logging.error(f"Error exporting to CSV: {e}")
        return False

def export_to_excel(employees: list, filename: str) -> bool:
    """Export employees to Excel file"""
    try:
        # Try to use openpyxl if available, otherwise use xlsxwriter
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            use_openpyxl = True
        except ImportError:
            try:
                import xlsxwriter
                use_openpyxl = False
            except ImportError:
                logging.warning("Neither openpyxl nor xlsxwriter found, falling back to CSV")
                # Fallback to CSV
                csv_filename = filename.replace('.xlsx', '.csv')
                return export_to_csv(employees, csv_filename)

        if not employees:
            return False

        if use_openpyxl:
            # Using openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Employees"

            # Get headers
            headers = list(employees[0].keys())

            # Write headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header.replace('_', ' ').title())
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Write data
            for row_idx, emp in enumerate(employees, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=str(emp.get(header, '')))

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            wb.save(filename)
        else:
            # Using xlsxwriter
            workbook = xlsxwriter.Workbook(filename)
            worksheet = workbook.add_worksheet("Employees")

            # Formats
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#366092',
                'font_color': 'white',
                'align': 'center'
            })

            # Get headers
            headers = list(employees[0].keys())

            # Write headers
            for col, header in enumerate(headers):
                worksheet.write(0, col, header.replace('_', ' ').title(), header_format)

            # Write data
            for row_idx, emp in enumerate(employees, 1):
                for col_idx, header in enumerate(headers):
                    worksheet.write(row_idx, col_idx, str(emp.get(header, '')))

            workbook.close()

        logging.info(f"Exported {len(employees)} employees to Excel: {filename}")
        return True
    except Exception as e:
        logging.error(f"Error exporting to Excel: {e}")
        return False

# ============================================================================
# PHASE 2: PAGINATION CLASS
# ============================================================================

class PaginationWidget(QWidget):
    """Pagination control widget"""
    def __init__(self, items_per_page: int = 50):
        super().__init__()
        self.items_per_page = items_per_page
        self.current_page = 1
        self.total_items = 0
        self.total_pages = 1

        self._init_ui()

    def _init_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # First page button
        self.first_btn = QPushButton("⏮ First")
        self.first_btn.clicked.connect(lambda: self.goto_page(1))
        layout.addWidget(self.first_btn)

        # Previous page button
        self.prev_btn = QPushButton("◀ Prev")
        self.prev_btn.clicked.connect(self.prev_page)
        layout.addWidget(self.prev_btn)

        # Page info
        self.page_label = QLabel("Page 1 of 1")
        self.page_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.page_label)

        # Next page button
        self.next_btn = QPushButton("Next ▶")
        self.next_btn.clicked.connect(self.next_page)
        layout.addWidget(self.next_btn)

        # Last page button
        self.last_btn = QPushButton("Last ⏭")
        self.last_btn.clicked.connect(lambda: self.goto_page(self.total_pages))
        layout.addWidget(self.last_btn)

        # Items per page
        layout.addWidget(QLabel("Items per page:"))
        self.items_spinner = QSpinBox()
        self.items_spinner.setRange(10, 500)
        self.items_spinner.setSingleStep(10)
        self.items_spinner.setValue(self.items_per_page)
        self.items_spinner.valueChanged.connect(self.change_items_per_page)
        layout.addWidget(self.items_spinner)

        layout.addStretch()

    def set_total_items(self, count: int):
        """Set total number of items"""
        self.total_items = count
        self.total_pages = max(1, (count + self.items_per_page - 1) // self.items_per_page)
        self.current_page = min(self.current_page, self.total_pages)
        self._update_ui()

    def get_page_range(self) -> tuple:
        """Get the start and end indices for current page"""
        start = (self.current_page - 1) * self.items_per_page
        end = min(start + self.items_per_page, self.total_items)
        return start, end

    def next_page(self):
        """Go to next page"""
        if self.current_page < self.total_pages:
            self.current_page += 1
            self._update_ui()

    def prev_page(self):
        """Go to previous page"""
        if self.current_page > 1:
            self.current_page -= 1
            self._update_ui()

    def goto_page(self, page: int):
        """Go to specific page"""
        if 1 <= page <= self.total_pages:
            self.current_page = page
            self._update_ui()

    def change_items_per_page(self, value: int):
        """Change items per page"""
        self.items_per_page = value
        self.set_total_items(self.total_items)

    def _update_ui(self):
        """Update UI elements"""
        self.page_label.setText(f"Page {self.current_page} of {self.total_pages} ({self.total_items} items)")
        self.first_btn.setEnabled(self.current_page > 1)
        self.prev_btn.setEnabled(self.current_page > 1)
        self.next_btn.setEnabled(self.current_page < self.total_pages)
        self.last_btn.setEnabled(self.current_page < self.total_pages)


# ======================================================================
# WheelEventFilter
# ======================================================================

class WheelEventFilter(QObject):
    """Event filter to prevent mouse wheel scrolling on QComboBox when not focused"""
    def eventFilter(self, obj, event):
        if event.type() == QEvent.Wheel:
            if isinstance(obj, QComboBox) and not obj.hasFocus():
                event.ignore()
                return True
        return super().eventFilter(obj, event)

# ============================================================================
# PHASE 1: ENHANCED LOGGING SETUP
# ============================================================================
from logging.handlers import RotatingFileHandler

LOG_FILE = "employee_vault.log"
LOG_MAX_BYTES = 10 * 1024 * 1024  # 10 MB
LOG_BACKUP_COUNT = 5

def setup_logging() -> None:
    """Configure enhanced logging with rotation"""
    try:
        os.makedirs('logs', exist_ok=True)

        # Create formatter
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )

        # File handler with rotation
        file_handler = RotatingFileHandler(
            f'logs/{LOG_FILE}',
            maxBytes=LOG_MAX_BYTES,
            backupCount=LOG_BACKUP_COUNT,
            encoding='utf-8'
        )
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(formatter)

        # Console handler for errors only
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.ERROR)
        console_handler.setFormatter(formatter)

        # Root logger configuration
        root_logger = logging.getLogger()
        root_logger.setLevel(logging.INFO)

        # Clear existing handlers
        root_logger.handlers.clear()

        # Add handlers
        root_logger.addHandler(file_handler)
        root_logger.addHandler(console_handler)

        logging.info("=" * 70)
        logging.info("Employee Vault Application Started")
        logging.info("=" * 70)
    except Exception as e:
        # Fallback to stderr if logging setup fails
        import sys
        sys.stderr.write(f"Warning: Could not set up logging: {e}\n")

# Initialize enhanced logging
setup_logging()

# ======================================================================
# PasswordRequirementsWidget
# ======================================================================

class PasswordRequirementsWidget(QLabel):
    """
    Real-time password requirements display widget (v2.3.1)
    Shows password requirements with visual feedback as user types
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setTextFormat(Qt.RichText)
        self.setWordWrap(True)
        self.setAlignment(Qt.AlignLeft | Qt.AlignTop)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        self.update_requirements("")
    
    def update_requirements(self, password: str):
        """Update requirements display based on current password"""
        has_length = len(password) >= PASSWORD_MIN_LENGTH
        has_upper = bool(re.search(r'[A-Z]', password))
        has_lower = bool(re.search(r'[a-z]', password))
        has_digit = bool(re.search(r'\d', password))
        has_special = bool(re.search(r'[!@#$%^&*(),.?":{}|<>]', password))
        
        def format_requirement(met: bool, text: str) -> str:
            icon = "✓" if met else "●"
            color = "#4ade80" if met else "#f87171"
            return f'<div style="color: {color}; margin: 3px 0; font-size: 12px;">{icon} {text}</div>'
        
        html = f'''
        <div style="background: #2a2a2d; border: 1px solid #3b3b3b; border-radius: 8px; padding: 10px; margin: 4px 0;">
            <div style="font-weight: 600; color: #4a8cff; font-size: 13px; margin-bottom: 6px;">
                🔒 Password Requirements:
            </div>
            {format_requirement(has_length, f"At least {PASSWORD_MIN_LENGTH} characters")}
            {format_requirement(has_upper, "At least 1 uppercase letter (A-Z)")}
            {format_requirement(has_lower, "At least 1 lowercase letter (a-z)")}
            {format_requirement(has_digit, "At least 1 number (0-9)")}
            {format_requirement(has_special, "At least 1 special character (!@#$%^&*...)")}
        </div>
        '''
        self.setText(html)

# ======================================================================
# CollapsibleSection
# ======================================================================

class CollapsibleSection(QWidget):
    """A collapsible section widget for sidebar menus - MODERNIZED with popup support"""
    def __init__(self, title, parent=None, start_collapsed=True, color="#4a9eff", icon="", main_window=None, theme_colors=None):
        super().__init__(parent)

        self.title = title
        self.icon = icon
        self.color = color
        self.main_window = main_window  # Reference to check if sidebar is collapsed
        self.menu_items = []  # Store menu items for popup
        self.theme_colors = theme_colors if theme_colors else {
            'surface': '#2d2d2d', 'text_primary': 'white'
        }  # Fallback to defaults

        # Modern button header with PILL SHAPE - 22px TRENDY
        # Include icon in button text (no arrows per user request)
        icon_text = f"{icon} " if icon else ""
        self.toggle_button = QPushButton(f"{icon_text}{title}")
        self.toggle_button.setFlat(False)  # Not flat for modern look
        self.toggle_button.setStyleSheet(f"""
            QPushButton {{
                text-align: left;
                padding: 12px 20px;
                font-weight: 600;
                font-size: 12px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
                color: white;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 rgba(255, 255, 255, 0.15),
                                           stop:0.5 rgba({self._get_rgb_from_hex(color)}, 0.35),
                                           stop:1 rgba({self._get_rgb_from_hex(color)}, 0.55));
                border: 1px solid rgba(255, 255, 255, 0.18);
                border-top: 1px solid rgba(255, 255, 255, 0.3);
                border-radius: 22px;
                margin: 4px 6px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                           stop:0 rgba(255, 255, 255, 0.22),
                                           stop:0.5 rgba({self._get_rgb_from_hex(color)}, 0.5),
                                           stop:1 rgba({self._get_rgb_from_hex(color)}, 0.7));
                border: 1px solid rgba(255, 255, 255, 0.35);
                border-top: 1px solid rgba(255, 255, 255, 0.45);
            }}
            QPushButton:pressed {{
                background: rgba({self._get_rgb_from_hex(color)}, 0.6);
                border-top: 1px solid rgba(255, 255, 255, 0.15);
            }}
            QPushButton:focus {{
                outline: none;
            }}
        """)
        # Cursor removed per user request
        # self.toggle_button.setCursor(Qt.PointingHandCursor)
        # v4.5.0: Remove focus box/outline when clicking sidebar menu
        self.toggle_button.setFocusPolicy(Qt.NoFocus)
        self.toggle_button.clicked.connect(self.on_button_clicked)

        self.content_area = QWidget()
        self.content_layout = QVBoxLayout(self.content_area)
        self.content_layout.setContentsMargins(8, 4, 8, 4)
        self.content_layout.setSpacing(2)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 4)
        layout.setSpacing(0)
        layout.addWidget(self.toggle_button)
        layout.addWidget(self.content_area)

        # Start collapsed or expanded
        self.is_collapsed = start_collapsed

        # Get AnimationManager for theme-aware animations
        from employee_vault.animation_manager import AnimationManager
        self.anim_manager = AnimationManager()

        # Use a QPropertyAnimation to animate the expansion/collapse of the
        # content area.  Animating the maximumHeight property gives a smooth
        # slide effect rather than an abrupt show/hide.  We use theme-specific
        # easing curve for consistency with overall theme feel.
        self.animation = QPropertyAnimation(self.content_area, b"maximumHeight", self)
        # v4.5.0: Use shorter duration for snappier feel (especially on network drives)
        # Original was theme-based ~300ms, reduced to 150ms for responsiveness
        self.animation.setDuration(150)  # Fast, responsive animation
        self.animation.setEasingCurve(QEasingCurve.OutQuad)  # Simple, fast easing

        # Set the initial maximum height based on collapsed state
        initial_height = 0 if self.is_collapsed else self.content_layout.sizeHint().height()
        self.content_area.setMaximumHeight(initial_height)
        # If collapsed, hide the content area initially
        self.content_area.setVisible(not self.is_collapsed)

    def _get_rgb_from_hex(self, hex_color):
        """Convert hex color to RGB string for rgba()"""
        hex_color = hex_color.lstrip('#')
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        return f"{r}, {g}, {b}"

    def _lighten_color(self, color):
        """Lighten a color for hover effect"""
        if color == "#4a9eff": return "#6bb0ff"
        if color == "#ff9800": return "#ffad33"
        if color == "#9c27b0": return "#b44bc4"
        if color == "#ff5722": return "#ff7043"  # TOOLS section
        return color

    def _darken_color(self, color):
        """Darken a color for gradient"""
        if color == "#4a9eff": return "#3d85d9"
        if color == "#ff9800": return "#e68900"
        if color == "#9c27b0": return "#7b1f8f"
        if color == "#ff5722": return "#e64a19"  # TOOLS section
        return color

    def on_button_clicked(self):
        """Handle button click - show popup if sidebar collapsed, else toggle"""
        # Check if sidebar is collapsed
        if self.main_window and hasattr(self.main_window, 'is_sidebar_collapsed') and self.main_window.is_sidebar_collapsed:
            # Show popup menu
            self.show_popup_menu()
        else:
            # Normal toggle behavior
            self.toggle()

    def show_popup_menu(self):
        """Show popup menu when sidebar is collapsed (now theme-aware)"""
        if not self.menu_items:
            return

        # Create popup menu with modern styling
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{
                background-color: {self.theme_colors['surface']};
                border: 2px solid {self.color};
                border-radius: 12px;
                padding: 8px;
            }}
            QMenu::item {{
                padding: 12px 32px 12px 14px;
                border-radius: 8px;
                color: {self.theme_colors['text_primary']};
                margin: 2px;
            }}
            QMenu::item:selected {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba({self._get_rgb_from_hex(self.color)}, 0.7),
                    stop:1 rgba({self._get_rgb_from_hex(self.color)}, 0.9));
                color: white;
            }}
        """)

        # Add menu items
        for item_text, item_callback in self.menu_items:
            action = menu.addAction(item_text)
            action.triggered.connect(item_callback)

        # Position menu to the right of the button
        button_pos = self.toggle_button.mapToGlobal(self.toggle_button.rect().topRight())
        menu.exec(button_pos)

    def toggle(self):
        """Toggle collapsed/expanded state"""
        # Toggle collapsed state
        self.is_collapsed = not self.is_collapsed
        # Update button text (icon + title, no arrows per user request)
        icon_text = f"{self.icon} " if self.icon else ""
        self.toggle_button.setText(f"{icon_text}{self.title}")

        # Determine start and end heights for the animation
        # Use current height as start; for collapse the end height is 0,
        # for expand it's the size hint of the content layout
        start_height = self.content_area.height()
        end_height = 0 if self.is_collapsed else self.content_layout.sizeHint().height()

        # Ensure the content area is visible during the animation
        self.content_area.setVisible(True)

        # Set up and start animation
        self.animation.stop()
        self.animation.setStartValue(start_height)
        self.animation.setEndValue(end_height)
        self.animation.start()

        # After collapsing, hide the content area completely to remove it from
        # the layout.  We connect to the finished signal and disconnect
        # afterwards to avoid accumulating connections.
        if self.is_collapsed:
            def hide_on_finish():
                self.content_area.setVisible(False)
                try:
                    self.animation.finished.disconnect(hide_on_finish)
                except Exception:
                    pass
            self.animation.finished.connect(hide_on_finish)

    def add_button(self, text, callback):
        """Add button to section (stores for popup menu too)"""
        # Store for popup menu
        self.menu_items.append((text, callback))

        # Create normal button with modern pill hover effect
        btn = QPushButton(text)
        btn.setFlat(True)
        btn.setStyleSheet(f"""
            QPushButton {{
                text-align: left;
                padding: 10px 14px 10px 22px;
                border-radius: 16px;
                background: transparent;
                color: rgba(255, 255, 255, 0.9);
                border-left: 4px solid transparent;
                font-size: 12px;
                font-weight: 500;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba({self._get_rgb_from_hex(self.color)}, 0.25),
                    stop:1 rgba({self._get_rgb_from_hex(self.color)}, 0.1));
                border-left: 4px solid {self.color};
                padding-left: 26px;
                color: white;
            }}
            QPushButton:pressed {{
                background: rgba({self._get_rgb_from_hex(self.color)}, 0.4);
            }}
        """)
        # Cursor removed per user request
        # btn.setCursor(Qt.PointingHandCursor)
        btn.clicked.connect(callback)
        self.content_layout.addWidget(btn)
        return btn


# ======================================================================
# Form Section Widget (for Employee Form grouping)
# ======================================================================
class FormSection(QWidget):
    """Collapsible section widget optimized for form field grouping"""
    def __init__(self, title, icon="", start_collapsed=False, parent=None):
        super().__init__(parent)

        self.title = title
        self.icon = icon
        self.is_collapsed = start_collapsed

        # Main layout
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 8)
        layout.setSpacing(0)

        # Header button with icon
        icon_text = f"{icon} " if icon else ""
        arrow = "▶" if start_collapsed else "▼"
        self.toggle_button = QPushButton(f"{icon_text}{arrow} {title}")
        self.toggle_button.setFlat(True)
        self.toggle_button.setStyleSheet("""
            QPushButton {
                text-align: left;
                padding: 12px 16px;
                font-weight: 600;
                font-size: 13px;
                color: white;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                           stop:0 rgba(33, 150, 243, 0.3),
                                           stop:1 rgba(33, 150, 243, 0.1));
                border-left: 4px solid rgba(33, 150, 243, 0.8);
                border-radius: 8px;
                margin-bottom: 8px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                                           stop:0 rgba(66, 165, 245, 0.4),
                                           stop:1 rgba(66, 165, 245, 0.2));
                border-left: 4px solid rgba(66, 165, 245, 1.0);
            }
            QPushButton:pressed {
                background: rgba(33, 150, 243, 0.5);
            }
        """)
        self.toggle_button.setFocusPolicy(Qt.NoFocus)
        self.toggle_button.clicked.connect(self.toggle)
        layout.addWidget(self.toggle_button)

        # Content area
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(12, 0, 0, 0)
        self.content_layout.setSpacing(12)
        layout.addWidget(self.content_widget)

        # Animation for smooth expand/collapse
        self.animation = QPropertyAnimation(self.content_widget, b"maximumHeight")
        self.animation.setDuration(200)
        self.animation.setEasingCurve(QEasingCurve.InOutQuad)

        # Set initial state
        if start_collapsed:
            self.content_widget.setMaximumHeight(0)
            self.content_widget.setVisible(False)
        else:
            self.content_widget.setMaximumHeight(16777215)  # QWIDGETSIZE_MAX

    def toggle(self):
        """Toggle collapsed/expanded state"""
        self.is_collapsed = not self.is_collapsed

        # Update arrow
        icon_text = f"{self.icon} " if self.icon else ""
        arrow = "▶" if self.is_collapsed else "▼"
        self.toggle_button.setText(f"{icon_text}{arrow} {self.title}")

        # Animate height change
        if self.is_collapsed:
            # Collapse
            self.animation.setStartValue(self.content_widget.height())
            self.animation.setEndValue(0)
            self.content_widget.setVisible(True)
            self.animation.start()
            # Hide after animation completes
            def hide_after():
                self.content_widget.setVisible(False)
                try:
                    self.animation.finished.disconnect(hide_after)
                except:
                    pass
            self.animation.finished.connect(hide_after)
        else:
            # Expand
            self.content_widget.setVisible(True)
            self.animation.setStartValue(0)
            self.animation.setEndValue(self.content_layout.sizeHint().height())
            self.animation.start()

    def add_widget(self, widget):
        """Add a widget to this section"""
        self.content_layout.addWidget(widget)

    def add_layout(self, layout):
        """Add a layout to this section"""
        self.content_layout.addLayout(layout)


# ======================================================================
# Loading Spinner Widget
# ======================================================================

class LoadingSpinner(QWidget):
    """Animated loading spinner with theme support"""
    def __init__(self, parent=None, size=40, color=None, theme_name=None):
        super().__init__(parent)
        self.size = size
        
        # Use theme color if available, fallback to provided color or theme default
        if color is None:
            if theme_name is None:
                theme_name = load_theme_preference()
            theme = MODERN_THEMES.get(theme_name, MODERN_THEMES.get("default", {}))
            self.color = theme.get("primary", "#4a8fd9")
        else:
            self.color = color
            
        self.angle = 0
        self.setFixedSize(size, size)

        # Timer for animation
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.rotate)
    
    def set_theme(self, theme_name):
        """Update spinner color based on theme"""
        theme = MODERN_THEMES.get(theme_name, MODERN_THEMES.get("default", {}))
        self.color = theme.get("primary", "#4a8fd9")
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # Draw spinning arc
        rect = self.rect().adjusted(5, 5, -5, -5)
        painter.setPen(QPen(QColor(self.color), 4, Qt.SolidLine, Qt.RoundCap))
        painter.drawArc(rect, self.angle * 16, 270 * 16)

    def rotate(self):
        self.angle = (self.angle + 10) % 360
        self.update()

    def start(self):
        self.timer.start(30)  # ~33 FPS
        self.show()

    def stop(self):
        self.timer.stop()
        self.hide()


# ======================================================================
# Toast Notification Widget
# ======================================================================

class ToastNotification(QWidget):
    """Modern toast notification with theme support"""
    def __init__(self, parent, message, duration=3000, toast_type="info", theme_name=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Tool | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setAttribute(Qt.WA_ShowWithoutActivating)
        
        # Get theme colors
        if theme_name is None:
            theme_name = load_theme_preference()
        theme = MODERN_THEMES.get(theme_name, MODERN_THEMES.get("default", {}))
        
        # Theme-aware colors
        theme_colors = {
            "success": theme.get("success", "#3d9a5f"),
            "error": theme.get("error", "#c95a5a"),
            "warning": theme.get("warning", "#c9943a"),
            "info": theme.get("info", "#5b8ec9")
        }
        surface_color = theme.get("surface", "#22262e")
        text_color = theme.get("text_primary", "#e8eaed")

        # Setup UI
        layout = QHBoxLayout(self)
        layout.setContentsMargins(15, 10, 15, 10)

        # Icon and colors based on type (using theme colors)
        icons = {
            "success": ("✓", theme_colors["success"]),
            "error": ("✗", theme_colors["error"]),
            "warning": ("⚠", theme_colors["warning"]),
            "info": ("ℹ", theme_colors["info"])
        }
        icon, color = icons.get(toast_type, icons["info"])

        # Icon label
        icon_label = QLabel(icon)
        icon_label.setStyleSheet(f"color: {color}; font-size: 20px; font-weight: bold;")
        layout.addWidget(icon_label)

        # Message label (using theme text color)
        msg_label = QLabel(message)
        msg_label.setStyleSheet(f"color: {text_color}; font-size: 13px;")
        msg_label.setWordWrap(True)
        layout.addWidget(msg_label, 1)

        # Style (using theme surface color)
        self.setStyleSheet(f"""
            ToastNotification {{
                background-color: {surface_color};
                border-left: 4px solid {color};
                border-radius: 8px;
            }}
        """)

        # Position
        self.adjustSize()
        parent_rect = parent.rect()
        x = parent_rect.width() - self.width() - 20
        y = parent_rect.height() - self.height() - 20
        self.move(x, y)

        # Auto-hide timer
        QTimer.singleShot(duration, self.fade_out)

        # Fade in animation
        self.opacity_effect = QGraphicsOpacityEffect(self)
        self.setGraphicsEffect(self.opacity_effect)
        self.fade_in()

    def fade_in(self):
        self.anim = QPropertyAnimation(self.opacity_effect, b"opacity")
        self.anim.setDuration(300)
        self.anim.setStartValue(0)
        self.anim.setEndValue(1)
        self.anim.setEasingCurve(QEasingCurve.InOutQuad)
        self.anim.start()

    def fade_out(self):
        self.anim = QPropertyAnimation(self.opacity_effect, b"opacity")
        self.anim.setDuration(300)
        self.anim.setStartValue(1)
        self.anim.setEndValue(0)
        self.anim.setEasingCurve(QEasingCurve.InOutQuad)
        self.anim.finished.connect(self.close)
        self.anim.start()


# Helper function to show toast from anywhere
def show_toast(parent, message, toast_type="info", duration=3000):
    """Show a toast notification"""
    toast = ToastNotification(parent, message, duration, toast_type)
    toast.show()
    return toast


# ======================================================================
# WheelGuard
# ======================================================================

class WheelGuard(QObject):
    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.Wheel:
            # Combos: allow only when dropdown is open (already in your code)
            if isinstance(obj, QComboBox):
                view = obj.view() if hasattr(obj, "view") else None
                return not (view and view.isVisible())

            # Spin-like widgets (QDateEdit/QSpinBox): allow only when focused
            if isinstance(obj, QAbstractSpinBox):
                # If calendar popup is visible, always block
                cal = getattr(obj, "calendarWidget", lambda: None)()
                if cal and cal.isVisible():
                    return True
                return not obj.hasFocus()

            # Calendar popup: always block month scrolling via wheel
            if isinstance(obj, QCalendarWidget):
                return True

        return False
        
# --- ADD THIS NEW FUNCTION HERE ---
def create_circular_pixmap(pix: QPixmap, size: int) -> QPixmap:
    """Helper function to create a circular pixmap."""
    if pix.isNull():
        # Return an empty transparent pixmap if the source is null
        result = QPixmap(size, size)
        result.fill(Qt.transparent)
        return result

    # Scale the pixmap to be a square of the given size,
    # covering the area to avoid black borders (Qt.KeepAspectRatioByExpanding)
    scaled = pix.scaled(size, size, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)

    # Create the final circular pixmap
    result = QPixmap(size, size)
    result.fill(Qt.transparent) # Fill with transparency

    # Use QPainter to clip
    painter = QPainter(result)
    painter.setRenderHint(QPainter.Antialiasing, True)
    path = QPainterPath()
    path.addEllipse(0, 0, size, size) # The circular clip path
    painter.setClipPath(path)

    # Draw the scaled image onto the clipped pixmap
    painter.drawPixmap(0, 0, scaled)
    painter.end()


# ============================================================================
# PHASE 2 FIX: PROGRESS DIALOG FOR LONG OPERATIONS
# ============================================================================

class SimpleProgressDialog:
    """
    Simple progress dialog for long-running operations.

    Usage:
        with SimpleProgressDialog(parent, "Processing...", "Please wait...") as progress:
            # Do long operation
            progress.set_value(50)  # Update progress
    """

    def __init__(self, parent, title: str, message: str, max_value: int = 100):
        from PySide6.QtWidgets import QProgressDialog
        from PySide6.QtCore import Qt

        self.dialog = QProgressDialog(message, None, 0, max_value, parent)
        self.dialog.setWindowTitle(title)
        self.dialog.setWindowModality(Qt.WindowModal)
        self.dialog.setMinimumDuration(0)  # Show immediately
        self.dialog.setCancelButton(None)  # No cancel button
        self.dialog.setAutoClose(True)
        self.dialog.setAutoReset(True)

    def __enter__(self):
        self.dialog.show()
        QApplication.processEvents()  # Force UI update
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.dialog.close()
        return False

    def set_value(self, value: int):
        """Update progress value"""
        self.dialog.setValue(value)
        QApplication.processEvents()  # Force UI update

    def set_label(self, text: str):
        """Update progress message"""
        self.dialog.setLabelText(text)
        QApplication.processEvents()  # Force UI update


# ============================================================================
# CUSTOM MODERN CALENDAR WIDGET WITH YEAR DROPDOWN
# ============================================================================
# MODERN 3D WHEEL PICKER (iOS-style drum picker with perspective)
# ============================================================================


class WheelColumn(QWidget):
    """
    iOS-style 3D drum wheel with infinite loop scrolling.
    Uses perspective transforms for realistic 3D effect.
    """
    valueChanged = Signal(object)

    def __init__(self, items, label="", parent=None):
        super().__init__(parent)
        self.items = list(items)
        self.label = label
        self._current_index = 0
        self._offset = 0.0  # Fractional offset for smooth scrolling
        self._velocity = 0.0
        self._last_y = 0
        self._last_time = 0
        self._dragging = False

        # Animation settings
        self._inertia_timer = QTimer(self)
        self._inertia_timer.timeout.connect(self._apply_inertia)

        self._snap_anim = QPropertyAnimation(self, b"offset")
        self._snap_anim.setDuration(180)
        self._snap_anim.setEasingCurve(QEasingCurve.OutCubic)

        # Visual settings
        self._item_height = 40
        self._visible_items = 5
        self.setFixedSize(80, self._item_height * self._visible_items)
        self.setMouseTracking(True)

    def get_offset(self):
        return self._offset

    def set_offset(self, value):
        self._offset = value
        self.update()

    offset = Property(float, get_offset, set_offset)

    def current_value(self):
        """Return the currently selected value"""
        if not self.items:
            return None
        return self.items[self._current_index % len(self.items)]

    def set_value(self, value):
        """Set the current value"""
        try:
            idx = self.items.index(value)
            self._current_index = idx
            self._offset = 0.0
            self.update()
        except ValueError:
            pass

    def paintEvent(self, event):
        """Paint the 3D wheel with perspective transform"""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # Get theme colors
        theme_name = load_theme_preference()
        theme = MODERN_THEMES.get(theme_name, MODERN_THEMES.get("default", {}))
        primary = theme.get('primary', '#4a9eff')

        center_y = self.height() / 2
        half_visible = self._visible_items // 2

        # Draw subtle selection indicator lines (top and bottom of center)
        line_color = QColor(255, 255, 255, 25)
        painter.setPen(QPen(line_color, 1))
        top_line_y = center_y - self._item_height / 2
        bottom_line_y = center_y + self._item_height / 2
        painter.drawLine(8, int(top_line_y), self.width() - 8, int(top_line_y))
        painter.drawLine(8, int(bottom_line_y), self.width() - 8, int(bottom_line_y))

        # Draw items with 3D perspective (drum/wheel effect)
        for i in range(-half_visible - 1, half_visible + 2):
            # Calculate actual item index with infinite loop wrapping
            item_offset = i - self._offset
            actual_index = (self._current_index + i) % len(self.items)

            # Normalized position for curve calculations (-1 to 1 range)
            normalized_pos = item_offset / (half_visible + 1)
            distance = abs(normalized_pos)
            distance = min(distance, 1.0)
            
            # Y position - linear spacing with slight curve at edges
            # Base position using item height
            base_y = center_y + item_offset * self._item_height
            # Slight curve - edge items move slightly closer (subtle wheel effect)
            curve_offset = normalized_pos * distance * 8  # Subtle bunching at edges
            item_center_y = base_y - curve_offset

            # Skip if completely outside view
            if item_center_y < -self._item_height or item_center_y > self.height() + self._item_height:
                continue

            # Perspective scale - edge items slightly smaller
            scale = 1.0 - (distance * 0.3)

            # Rotation angle (simulates drum rotation)
            angle = distance * 60

            # Alpha based on distance (fade at edges)
            alpha = int(255 * (1.0 - distance * 0.5))

            # Font size based on scale - smaller at edges
            font_size = int(18 * scale)

            # Draw the item
            painter.save()

            # Center item text
            item_text = str(self.items[actual_index])

            # Set font
            font = painter.font()
            font.setPixelSize(max(font_size, 9))
            font.setWeight(QFont.DemiBold if distance < 0.2 else QFont.Normal)
            painter.setFont(font)

            # Calculate text dimensions for centering
            metrics = painter.fontMetrics()
            text_width = metrics.horizontalAdvance(item_text)
            text_ascent = metrics.ascent()

            # Vertical compression based on rotation angle
            v_scale = math.cos(math.radians(angle))

            # Apply vertical compression transform around item center
            painter.translate(self.width() / 2, item_center_y)
            painter.scale(1.0, v_scale)

            # Text position - centered horizontally and vertically
            x = -text_width / 2
            text_y = text_ascent / 2

            # Color based on distance from center
            if distance < 0.15:
                # Center item - highlighted
                color = QColor(primary)
                color.setAlpha(255)
            else:
                # Other items - white with fade
                color = QColor(255, 255, 255, alpha)

            painter.setPen(color)
            painter.drawText(int(x), int(text_y), item_text)

            painter.restore()

    def mousePressEvent(self, event):
        """Start dragging"""
        self._dragging = True
        self._last_y = event.position().y()
        self._last_time = event.timestamp()
        self._velocity = 0.0
        self._inertia_timer.stop()
        self._snap_anim.stop()

    def mouseMoveEvent(self, event):
        """Handle drag scrolling"""
        if not self._dragging:
            return

        y = event.position().y()
        delta = y - self._last_y
        current_time = event.timestamp()

        # Calculate velocity for inertia
        time_delta = max(current_time - self._last_time, 1)
        self._velocity = delta / time_delta * 16  # Scale for 60fps

        # Update offset
        self._offset -= delta / self._item_height

        # Wrap offset and update index
        while self._offset >= 0.5:
            self._offset -= 1.0
            self._current_index = (self._current_index + 1) % len(self.items)
        while self._offset <= -0.5:
            self._offset += 1.0
            self._current_index = (self._current_index - 1) % len(self.items)

        self._last_y = y
        self._last_time = current_time
        self.update()

    def mouseReleaseEvent(self, event):
        """End dragging - start inertia or snap"""
        self._dragging = False

        # If velocity is significant, start inertia
        if abs(self._velocity) > 0.5:
            self._inertia_timer.start(16)  # ~60fps
        else:
            self._snap_to_nearest()

    def wheelEvent(self, event):
        """Handle mouse wheel scrolling"""
        delta = event.angleDelta().y()
        steps = -1 if delta > 0 else 1

        self._current_index = (self._current_index + steps) % len(self.items)
        self._offset = 0.0
        self.update()
        self.valueChanged.emit(self.current_value())

    def _apply_inertia(self):
        """Apply inertia deceleration"""
        # Update offset based on velocity
        self._offset -= self._velocity / self._item_height

        # Wrap offset and update index
        while self._offset >= 0.5:
            self._offset -= 1.0
            self._current_index = (self._current_index + 1) % len(self.items)
        while self._offset <= -0.5:
            self._offset += 1.0
            self._current_index = (self._current_index - 1) % len(self.items)

        # Apply friction
        self._velocity *= 0.92

        self.update()

        # Stop when velocity is low enough
        if abs(self._velocity) < 0.3:
            self._inertia_timer.stop()
            self._snap_to_nearest()

    def _snap_to_nearest(self):
        """Animate snap to nearest item"""
        self._snap_anim.stop()
        self._snap_anim.setStartValue(self._offset)
        self._snap_anim.setEndValue(0.0)
        self._snap_anim.finished.connect(self._on_snap_finished)
        self._snap_anim.start()

    def _on_snap_finished(self):
        """Emit value changed when snap completes"""
        try:
            self._snap_anim.finished.disconnect(self._on_snap_finished)
        except RuntimeError:
            pass
        self.valueChanged.emit(self.current_value())


class WheelDatePickerPopup(QWidget):
    """
    iOS-style 3D wheel date picker with Month, Day, Year drums.
    All wheels support infinite loop scrolling.
    """
    dateSelected = Signal(QDate)

    def __init__(self, parent=None):
        super().__init__(parent, Qt.Popup | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setFixedSize(320, 340)

        # Date constraints
        self._min_date = QDate(1950, 1, 1)
        self._max_date = QDate(2100, 12, 31)
        self._selected_date = QDate.currentDate()

        self._setup_ui()
        self._update_wheels()

    def _setup_ui(self):
        """Setup the wheel picker UI"""
        # Get theme colors
        theme_name = load_theme_preference()
        self.theme = MODERN_THEMES.get(theme_name, MODERN_THEMES.get("default", {}))
        primary = self.theme.get('primary', '#4a9eff')

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # Main container with heavy iOS frosted glass effect
        container = QWidget()
        container.setObjectName("wheelPickerContainer")
        container.setStyleSheet(f"""
            QWidget#wheelPickerContainer {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(60, 60, 65, 0.92),
                    stop:0.5 rgba(45, 45, 50, 0.94),
                    stop:1 rgba(35, 35, 40, 0.96));
                border: 1.5px solid rgba(255, 255, 255, 0.25);
                border-top: 1.5px solid rgba(255, 255, 255, 0.4);
                border-radius: 20px;
            }}
        """)
        
        # Add strong drop shadow for frosted depth effect
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(50)
        shadow.setColor(QColor(0, 0, 0, 200))
        shadow.setOffset(0, 12)
        container.setGraphicsEffect(shadow)

        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(16, 12, 16, 16)
        container_layout.setSpacing(8)

        # Title
        title = QLabel("Select Date")
        title.setStyleSheet("""
            background: transparent;
            color: white;
            font-size: 18px;
            font-weight: 700;
            padding: 6px 4px 10px 4px;
        """)
        title.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(title)

        # Wheels container - properly centered
        wheels_frame = QFrame()
        wheels_frame.setStyleSheet("background: transparent;")
        wheels_layout = QHBoxLayout(wheels_frame)
        wheels_layout.setContentsMargins(0, 0, 0, 0)
        wheels_layout.setSpacing(12)

        # Month wheel
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        self.month_wheel = WheelColumn(months, "Month")
        self.month_wheel.valueChanged.connect(self._on_month_changed)

        # Day wheel (will be updated based on month/year)
        self.day_wheel = WheelColumn(list(range(1, 32)), "Day")
        self.day_wheel.valueChanged.connect(self._on_value_changed)

        # Year wheel
        years = list(range(self._min_date.year(), self._max_date.year() + 1))
        self.year_wheel = WheelColumn(years, "Year")
        self.year_wheel.valueChanged.connect(self._on_year_changed)

        # Labels
        month_label = self._create_label("Month")
        day_label = self._create_label("Day")
        year_label = self._create_label("Year")

        # Month column
        month_col = QVBoxLayout()
        month_col.setSpacing(4)
        month_col.addWidget(month_label, 0, Qt.AlignHCenter)
        month_col.addWidget(self.month_wheel, 0, Qt.AlignHCenter)

        # Day column
        day_col = QVBoxLayout()
        day_col.setSpacing(4)
        day_col.addWidget(day_label, 0, Qt.AlignHCenter)
        day_col.addWidget(self.day_wheel, 0, Qt.AlignHCenter)

        # Year column
        year_col = QVBoxLayout()
        year_col.setSpacing(4)
        year_col.addWidget(year_label, 0, Qt.AlignHCenter)
        year_col.addWidget(self.year_wheel, 0, Qt.AlignHCenter)

        wheels_layout.addStretch(1)
        wheels_layout.addLayout(month_col)
        wheels_layout.addLayout(day_col)
        wheels_layout.addLayout(year_col)
        wheels_layout.addStretch(1)

        container_layout.addWidget(wheels_frame)

        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(12)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(100, 100, 110, 0.6),
                    stop:1 rgba(70, 70, 80, 0.8));
                border: 1.5px solid rgba(255, 255, 255, 0.15);
                border-radius: 20px;
                color: white;
                font-weight: 600;
                padding: 10px 24px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(120, 120, 130, 0.8),
                    stop:1 rgba(90, 90, 100, 0.9));
            }}
        """)
        cancel_btn.clicked.connect(self.hide)

        select_btn = QPushButton("Select")
        select_btn.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(255, 255, 255, 0.1),
                    stop:0.5 {primary}cc,
                    stop:1 {primary});
                border: 1.5px solid {primary}80;
                border-radius: 20px;
                color: white;
                font-weight: 600;
                padding: 10px 24px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(255, 255, 255, 0.2),
                    stop:0.5 {primary},
                    stop:1 {primary});
            }}
        """)
        select_btn.clicked.connect(self._on_select)

        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(select_btn)
        container_layout.addLayout(btn_layout)

        layout.addWidget(container)

    def _create_label(self, text):
        """Create a styled label for wheel columns"""
        primary = self.theme.get('primary', '#4a9eff')
        label = QLabel(text)
        label.setStyleSheet(f"""
            color: {primary};
            font-size: 13px;
            font-weight: 700;
            letter-spacing: 1px;
            padding: 4px 8px;
            text-transform: uppercase;
        """)
        label.setAlignment(Qt.AlignCenter)
        return label

    def _update_wheels(self):
        """Update wheel values to match selected date"""
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        self.month_wheel.set_value(months[self._selected_date.month() - 1])
        self.day_wheel.set_value(self._selected_date.day())
        self.year_wheel.set_value(self._selected_date.year())

    def _update_day_wheel(self):
        """Update day wheel based on current month and year"""
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        month_val = self.month_wheel.current_value()
        year_val = self.year_wheel.current_value()

        if month_val and year_val:
            month_idx = months.index(month_val) + 1
            days_in_month = QDate(year_val, month_idx, 1).daysInMonth()

            current_day = self.day_wheel.current_value()
            self.day_wheel.items = list(range(1, days_in_month + 1))

            # Clamp day if needed
            if current_day and current_day > days_in_month:
                self.day_wheel.set_value(days_in_month)

            self.day_wheel.update()

    def _on_month_changed(self, value):
        """Handle month change"""
        self._update_day_wheel()
        self._on_value_changed(value)

    def _on_year_changed(self, value):
        """Handle year change"""
        self._update_day_wheel()
        self._on_value_changed(value)

    def _on_value_changed(self, value):
        """Handle any value change"""
        pass  # Could add live preview here

    def _on_select(self):
        """Handle selection confirmation"""
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        month_val = self.month_wheel.current_value()
        day_val = self.day_wheel.current_value()
        year_val = self.year_wheel.current_value()

        if month_val and day_val and year_val:
            month_idx = months.index(month_val) + 1
            date = QDate(year_val, month_idx, day_val)
            if date.isValid():
                self._selected_date = date
                self.dateSelected.emit(date)
                self.hide()

    def setSelectedDate(self, date):
        """Set the selected date"""
        self._selected_date = date
        self._update_wheels()

    def setMinimumDate(self, date):
        """Set minimum selectable date"""
        self._min_date = date
        years = list(range(self._min_date.year(), self._max_date.year() + 1))
        self.year_wheel.items = years
        self.year_wheel.update()

    def setMaximumDate(self, date):
        """Set maximum selectable date"""
        self._max_date = date
        years = list(range(self._min_date.year(), self._max_date.year() + 1))
        self.year_wheel.items = years
        self.year_wheel.update()


# ============================================================================
# MODERN WHEEL PICKER (iOS-style infinite scroll picker - legacy)
# ============================================================================

class ModernWheelPicker(QDialog):
    """iOS-style wheel picker with infinite loop scrolling"""

    def __init__(self, items, current_value, title="Select", parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.selected_value = current_value
        self.items = items
        self.is_infinite = len(items) <= 50  # Enable infinite scroll for small lists

        # Window styling
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setFixedSize(280, 320)

        self._setup_ui()

    def _setup_ui(self):
        """Setup the wheel picker UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # Container with iOS frosted glass
        container = QWidget()
        container.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(50, 50, 55, 0.98),
                    stop:1 rgba(35, 35, 40, 0.98));
                border: 2px solid rgba(255, 255, 255, 0.2);
                border-radius: 20px;
            }
        """)
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(16, 16, 16, 16)

        # Title
        title_label = QLabel(self.windowTitle())
        title_label.setStyleSheet("""
            color: white;
            font-size: 18px;
            font-weight: 700;
            padding: 8px;
        """)
        title_label.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(title_label)

        # List widget for wheel picker
        self.list_widget = QListWidget()
        self.list_widget.setVerticalScrollMode(QListWidget.ScrollPerPixel)
        self.list_widget.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.list_widget.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        # Populate list (triple for infinite scroll effect if enabled)
        if self.is_infinite:
            # Add items 3 times for loop effect
            for _ in range(3):
                for item in self.items:
                    list_item = QListWidgetItem(str(item))
                    list_item.setData(Qt.UserRole, item)
                    list_item.setTextAlignment(Qt.AlignCenter)
                    self.list_widget.addItem(list_item)

            # Scroll to middle section
            middle_index = len(self.items)
            try:
                current_index = self.items.index(self.selected_value)
                self.list_widget.setCurrentRow(middle_index + current_index)
            except ValueError:
                self.list_widget.setCurrentRow(middle_index)
        else:
            # Normal list for large datasets
            for item in self.items:
                list_item = QListWidgetItem(str(item))
                list_item.setData(Qt.UserRole, item)
                list_item.setTextAlignment(Qt.AlignCenter)
                self.list_widget.addItem(list_item)

            try:
                current_index = self.items.index(self.selected_value)
                self.list_widget.setCurrentRow(current_index)
            except ValueError:
                self.list_widget.setCurrentRow(0)

        # Styling
        self.list_widget.setStyleSheet("""
            QListWidget {
                background: transparent;
                border: none;
                outline: none;
                font-size: 16px;
                font-weight: 600;
            }
            QListWidget::item {
                color: rgba(255, 255, 255, 0.5);
                padding: 12px 16px;
                border-radius: 8px;
                margin: 2px 8px;
            }
            QListWidget::item:selected {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(74, 158, 255, 0.7),
                    stop:1 rgba(25, 118, 210, 0.7));
                color: white;
                font-weight: 700;
                font-size: 18px;
            }
            QListWidget::item:hover {
                background: rgba(74, 158, 255, 0.3);
                color: rgba(255, 255, 255, 0.8);
            }
        """)

        container_layout.addWidget(self.list_widget)

        # Connect infinite scroll
        if self.is_infinite:
            self.list_widget.verticalScrollBar().valueChanged.connect(self._handle_infinite_scroll)

        # Buttons
        btn_layout = QHBoxLayout()

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(100, 100, 105, 0.6),
                    stop:1 rgba(70, 70, 75, 0.8));
                border: 1.5px solid rgba(255, 255, 255, 0.15);
                border-top: 1.5px solid rgba(255, 255, 255, 0.25);
                border-radius: 22px;
                color: white;
                font-weight: 600;
                padding: 12px 24px;
                font-size: 14px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(120, 120, 125, 0.8),
                    stop:1 rgba(90, 90, 95, 0.9));
                border: 1.5px solid rgba(255, 255, 255, 0.35);
            }
        """)
        cancel_btn.clicked.connect(self.reject)

        select_btn = QPushButton("Select")
        select_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(255, 255, 255, 0.15),
                    stop:0.5 rgba(74, 158, 255, 0.8),
                    stop:1 rgba(25, 118, 210, 0.9));
                border: 1.5px solid rgba(74, 158, 255, 0.5);
                border-top: 1.5px solid rgba(255, 255, 255, 0.35);
                border-radius: 22px;
                color: white;
                font-weight: 600;
                padding: 12px 24px;
                font-size: 14px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(255, 255, 255, 0.2),
                    stop:0.5 rgba(74, 158, 255, 0.9),
                    stop:1 rgba(25, 118, 210, 1));
                border: 1.5px solid rgba(74, 158, 255, 0.8);
            }
        """)
        select_btn.clicked.connect(self._on_select)

        btn_layout.addWidget(cancel_btn)
        btn_layout.addWidget(select_btn)
        container_layout.addLayout(btn_layout)

        layout.addWidget(container)

    def _handle_infinite_scroll(self, value):
        """Handle infinite scrolling by repositioning when reaching edges"""
        if not self.is_infinite:
            return

        scrollbar = self.list_widget.verticalScrollBar()
        max_val = scrollbar.maximum()

        # If scrolled to bottom, jump to second section
        if value >= max_val - 10:
            scrollbar.setValue(max_val // 3)
        # If scrolled to top, jump to second section
        elif value <= 10:
            scrollbar.setValue((max_val * 2) // 3)

    def _on_select(self):
        """Handle selection"""
        current_item = self.list_widget.currentItem()
        if current_item:
            self.selected_value = current_item.data(Qt.UserRole)
            self.accept()

    def get_selected_value(self):
        """Return the selected value"""
        return self.selected_value


# ============================================================================

class ModernCalendarWidget(QCalendarWidget):
    """Custom calendar widget with iOS frosted glass styling and modern wheel pickers"""

    def __init__(self, parent=None):
        super().__init__(parent)

        # Get current theme colors
        theme_name = load_theme_preference()
        self.theme = MODERN_THEMES.get(theme_name, MODERN_THEMES.get("default", MODERN_THEMES[list(MODERN_THEMES.keys())[0]]))

        # Set minimum size for calendar popup
        self.setMinimumSize(380, 350)

        # Enable navigation bar
        self.setNavigationBarVisible(True)
        self.setGridVisible(True)
        self.setVerticalHeaderFormat(QCalendarWidget.VerticalHeaderFormat.NoVerticalHeader)

        # Apply modern styling
        self._apply_modern_style()

        # Replace month/year buttons with wheel pickers
        QTimer.singleShot(0, self._setup_wheel_pickers)

    def _setup_wheel_pickers(self):
        """Setup modern wheel pickers for month and year selection"""
        try:
            # Find the navigation buttons
            month_button = self.findChild(QToolButton, "qt_calendar_monthbutton")
            year_button = self.findChild(QToolButton, "qt_calendar_yearbutton")
            spinbox = self.findChild(QSpinBox)

            # Hide spinbox if exists
            if spinbox:
                spinbox.hide()

            # Setup month picker
            if month_button:
                # Disconnect default menu
                month_button.setMenu(None)
                month_button.setPopupMode(QToolButton.DelayedPopup)
                month_button.clicked.connect(self._show_month_picker)

            # Setup year picker
            if year_button:
                year_button.clicked.connect(self._show_year_picker)

        except Exception as e:
            import logging
            logging.error(f"Error setting up wheel pickers: {e}")

    def _show_month_picker(self):
        """Show modern wheel picker for month selection"""
        try:
            months = [
                "January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"
            ]
            current_month = months[self.selectedDate().month() - 1]

            picker = ModernWheelPicker(months, current_month, "Select Month", self)
            if picker.exec() == QDialog.Accepted:
                selected_month = picker.get_selected_value()
                month_index = months.index(selected_month) + 1

                current_date = self.selectedDate()
                new_date = QDate(current_date.year(), month_index, min(current_date.day(), 28))
                self.setSelectedDate(new_date)

                # Update button text
                month_button = self.findChild(QToolButton, "qt_calendar_monthbutton")
                if month_button:
                    month_button.setText(selected_month)

        except Exception as e:
            import logging
            logging.error(f"Error showing month picker: {e}")

    def _show_year_picker(self):
        """Show modern wheel picker for year selection"""
        try:
            # Generate year list (1900-2100)
            years = list(range(1900, 2101))
            current_year = self.selectedDate().year()

            picker = ModernWheelPicker(years, current_year, "Select Year", self)
            result = picker.exec()

            if result == QDialog.Accepted:
                selected_year = picker.get_selected_value()

                if selected_year is not None and isinstance(selected_year, int):
                    current_date = self.selectedDate()
                    new_date = QDate(selected_year, current_date.month(), current_date.day())
                    self.setSelectedDate(new_date)

                    # Update button text
                    year_button = self.findChild(QToolButton, "qt_calendar_yearbutton")
                    if year_button:
                        year_button.setText(str(selected_year))

        except Exception as e:
            import logging
            logging.error(f"Error showing year picker: {e}")
            import traceback
            logging.error(traceback.format_exc())

    def _apply_modern_style(self):
        """Apply modern iOS-style frosted glass styling to the calendar"""
        primary = self.theme.get('primary', '#4a9eff')
        primary_light = self.theme.get('primary_light', '#6bb3ff')
        primary_dark = self.theme.get('primary_dark', '#1976D2')

        # Apply iOS frosted glass stylesheet
        self.setStyleSheet(f"""
            /* Main calendar widget - iOS frosted glass */
            QCalendarWidget {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(50, 50, 55, 0.95),
                    stop:1 rgba(35, 35, 40, 0.95));
                border: 2px solid rgba(255, 255, 255, 0.15);
                border-radius: 16px;
                padding: 12px;
            }}

            /* Navigation bar - iOS gradient header */
            QCalendarWidget QWidget#qt_calendar_navigationbar {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(74, 158, 255, 0.8),
                    stop:0.5 rgba(102, 126, 234, 0.8),
                    stop:1 rgba(118, 75, 162, 0.8));
                border: 1.5px solid rgba(255, 255, 255, 0.3);
                border-radius: 12px;
                padding: 8px;
                min-height: 48px;
            }}

            /* Month/Year buttons - iOS style */
            QCalendarWidget QToolButton {{
                color: white;
                background: transparent;
                border: none;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 14px;
                font-weight: 600;
                min-width: 40px;
            }}

            QCalendarWidget QToolButton:hover {{
                background: rgba(255, 255, 255, 0.25);
            }}

            QCalendarWidget QToolButton:pressed {{
                background: rgba(255, 255, 255, 0.15);
            }}

            /* Navigation arrows - iOS style */
            QCalendarWidget QToolButton#qt_calendar_prevmonth {{
                qproperty-text: "◀";
                font-size: 18px;
                min-width: 36px;
                padding: 8px;
            }}

            QCalendarWidget QToolButton#qt_calendar_nextmonth {{
                qproperty-text: "▶";
                font-size: 18px;
                min-width: 36px;
                padding: 8px;
            }}

            /* Month button */
            QCalendarWidget QToolButton#qt_calendar_monthbutton {{
                min-width: 100px;
                text-align: center;
            }}

            /* Year button */
            QCalendarWidget QToolButton#qt_calendar_yearbutton {{
                min-width: 80px;
                text-align: center;
            }}

            /* Menu dropdowns - iOS frosted */
            QCalendarWidget QMenu {{
                background: rgba(45, 45, 48, 0.95);
                border: 1.5px solid rgba(255, 255, 255, 0.2);
                border-radius: 12px;
                padding: 8px;
            }}

            QCalendarWidget QMenu::item {{
                color: white;
                padding: 10px 24px;
                border-radius: 8px;
                margin: 2px;
            }}

            QCalendarWidget QMenu::item:selected {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(74, 158, 255, 0.6),
                    stop:1 rgba(25, 118, 210, 0.6));
            }}

            /* Week day headers - iOS style */
            QCalendarWidget QWidget {{
                alternate-background-color: transparent;
            }}

            /* Table view - date cells */
            QCalendarWidget QAbstractItemView {{
                color: rgba(255, 255, 255, 0.95);
                background: transparent;
                selection-background-color: transparent;
                border: none;
                outline: none;
                font-size: 13px;
                font-weight: 500;
            }}

            QCalendarWidget QAbstractItemView::item {{
                padding: 8px;
                border-radius: 8px;
                margin: 1px;
            }}

            QCalendarWidget QAbstractItemView::item:hover {{
                background: rgba(74, 158, 255, 0.3);
            }}

            QCalendarWidget QAbstractItemView::item:selected {{
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(74, 158, 255, 0.8),
                    stop:1 rgba(25, 118, 210, 0.8));
                color: white;
                font-weight: 700;
            }}

            QCalendarWidget QAbstractItemView::item:disabled {{
                color: rgba(255, 255, 255, 0.25);
            }}

            /* Header cells (day names) */
            QCalendarWidget QAbstractItemView {{
                gridline-color: rgba(255, 255, 255, 0.05);
            }}
        """)


# ======================================================================
# CalendarPopup - Modern glassmorphism calendar with animations
# ======================================================================

class CalendarPopup(QWidget):
    """
    Modern popup calendar with glassmorphism effects and smooth animations.
    Features: frosted glass background, micro-animations, selection ripple.
    Theme-aware: adapts blur/opacity based on light/dark mode.
    """
    dateSelectedCustom = Signal(QDate)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.Popup | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)  # Enable transparency for glassmorphism
        self.setFixedSize(280, 300)  # Slightly larger for effects
        
        # Get theme colors and detect light/dark mode
        theme_name = load_theme_preference()
        self.theme = MODERN_THEMES.get(theme_name, MODERN_THEMES.get("default", MODERN_THEMES[list(MODERN_THEMES.keys())[0]]))
        self._is_dark_theme = theme_name.lower() in ['dark', 'midnight', 'ocean', 'default']
        
        # Theme-aware glassmorphism settings
        self._blur_radius = 8 if self._is_dark_theme else 12  # Lighter blur for dark, heavier for light
        self._glass_opacity = 0.85 if self._is_dark_theme else 0.75
        
        # Animation settings (150-180ms for snappy feel per Material Design)
        self._anim_duration_fast = 120  # Fast micro-animations
        self._anim_duration_normal = 150  # Standard transitions
        self._anim_duration_slow = 200  # Complex animations
        
        self._selected_date = QDate.currentDate()
        self._min_date = QDate(1900, 1, 1)
        self._max_date = QDate(2100, 12, 31)
        self._header_collapsed = True  # iOS-style: start collapsed
        self._animating = False  # Track animation state
        self._slide_direction = 0  # -1=left, 1=right
        self._last_clicked_btn = None  # For ripple effect
        
        self._setup_ui()
        self._apply_style()
        self._sync_header()
        self._update_calendar()

    def _setup_ui(self):
        # Main container for glassmorphism effect
        self.glass_container = QWidget(self)
        self.glass_container.setObjectName("glassContainer")
        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(self.glass_container)
        
        layout = QVBoxLayout(self.glass_container)
        layout.setContentsMargins(12, 12, 12, 12)  # Comfortable margins
        layout.setSpacing(6)
        
        # Header with iOS-style collapsible month/year
        header = QHBoxLayout()
        header.setSpacing(8)
        
        # Previous month button - circular with chevron and hover animation
        self.prev_btn = QToolButton()
        self.prev_btn.setText("‹")  # Thin chevron
        self.prev_btn.setFixedSize(32, 32)  # Circular size
        self.prev_btn.setCursor(Qt.PointingHandCursor)
        self.prev_btn.clicked.connect(self._prev_month)
        self.prev_btn.installEventFilter(self)  # For hover animations
        
        # COLLAPSED STATE: Single "Month Year" button (iOS style pill)
        self.header_btn = QPushButton("December 2025")
        self.header_btn.setFixedHeight(26)
        self.header_btn.setCursor(Qt.PointingHandCursor)
        self.header_btn.clicked.connect(self._toggle_header)
        
        # EXPANDED STATE: Separate Month and Year dropdowns (hidden initially)
        self.month_combo = QComboBox()
        self.month_combo.setFixedHeight(26)
        self.month_combo.setCursor(Qt.PointingHandCursor)
        self.month_combo.hide()
        
        self.year_combo = QComboBox()
        self.year_combo.setFixedHeight(26)
        self.year_combo.setCursor(Qt.PointingHandCursor)
        self.year_combo.hide()
        
        # Year grid button (shows year picker overlay) - hidden when collapsed
        self.year_grid_btn = QToolButton()
        self.year_grid_btn.setText("📅")
        self.year_grid_btn.setFixedSize(26, 26)
        self.year_grid_btn.setCursor(Qt.PointingHandCursor)
        self.year_grid_btn.setToolTip("Show year grid")
        self.year_grid_btn.hide()
        self.year_grid_btn.clicked.connect(self._show_year_grid)
        
        # Populate month names
        months = ["January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"]
        for i, m in enumerate(months, 1):
            self.month_combo.addItem(m, i)
        
        # Populate years (current year ±50)
        current_year = QDate.currentDate().year()
        for y in range(current_year - 50, current_year + 51):
            self.year_combo.addItem(str(y), y)
        
        self.month_combo.currentIndexChanged.connect(self._on_combo_changed)
        self.year_combo.currentIndexChanged.connect(self._on_combo_changed)
        
        # Next month button - circular with chevron and hover animation
        self.next_btn = QToolButton()
        self.next_btn.setText("›")  # Thin chevron
        self.next_btn.setFixedSize(32, 32)  # Circular size
        self.next_btn.setCursor(Qt.PointingHandCursor)
        self.next_btn.clicked.connect(self._next_month)
        self.next_btn.installEventFilter(self)  # For hover animations
        
        header.addWidget(self.prev_btn)
        header.addStretch(1)
        header.addWidget(self.header_btn)
        header.addWidget(self.month_combo)
        header.addWidget(self.year_combo)
        header.addWidget(self.year_grid_btn)
        header.addStretch(1)
        header.addWidget(self.next_btn)
        layout.addLayout(header)
        
        # Day names header - theme-aware styling
        days_layout = QHBoxLayout()
        days_layout.setSpacing(2)
        self.day_labels = []
        day_names = ["S", "M", "T", "W", "T", "F", "S"]  # Single letters
        day_color = "rgba(255, 255, 255, 0.5)" if self._is_dark_theme else "rgba(0, 0, 0, 0.45)"
        for day in day_names:
            lbl = QLabel(day)
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setFixedSize(34, 18)
            lbl.setStyleSheet(f"color: {day_color}; font-weight: 600; font-size: 11px;")
            days_layout.addWidget(lbl)
            self.day_labels.append(lbl)
        layout.addLayout(days_layout)
        
        # Calendar grid wrapper for clipping during slide
        self.grid_wrapper = QWidget()
        self.grid_wrapper.setStyleSheet("background: transparent;")
        grid_wrapper_layout = QVBoxLayout(self.grid_wrapper)
        grid_wrapper_layout.setContentsMargins(0, 0, 0, 0)
        
        # Calendar grid container (for slide + fade animation)
        self.grid_container = QWidget()
        self.grid_container.setStyleSheet("background: transparent;")
        self.grid_layout = QGridLayout(self.grid_container)
        self.grid_layout.setSpacing(3)  # Slightly increased for hover effects
        self.grid_layout.setContentsMargins(0, 0, 0, 0)
        self.day_buttons = []
        
        grid_wrapper_layout.addWidget(self.grid_container)
        
        for row in range(6):
            row_buttons = []
            for col in range(7):
                btn = QPushButton("")
                btn.setFixedSize(34, 34)  # Square buttons
                btn.setCursor(Qt.PointingHandCursor)
                btn.clicked.connect(lambda checked=False, r=row, c=col: self._on_day_clicked(r, c))
                btn.installEventFilter(self)  # For hover micro-animations
                self.grid_layout.addWidget(btn, row, col)
                row_buttons.append(btn)
            self.day_buttons.append(row_buttons)
        
        # Opacity effect for fade transition
        self.grid_opacity = QGraphicsOpacityEffect(self.grid_container)
        self.grid_opacity.setOpacity(1.0)
        self.grid_container.setGraphicsEffect(self.grid_opacity)
        
        layout.addWidget(self.grid_wrapper)

    def eventFilter(self, obj, event):
        """Handle hover micro-animations for buttons"""
        from PySide6.QtCore import QEvent

        if event.type() == QEvent.Enter:
            # Scale up slightly on hover
            if hasattr(obj, 'setStyleSheet') and obj in [self.prev_btn, self.next_btn]:
                self._animate_button_scale(obj, 1.08)
            elif isinstance(obj, QPushButton) and obj.isEnabled():
                # Day button hover - subtle scale
                self._animate_button_scale(obj, 1.05)
        elif event.type() == QEvent.Leave:
            # Scale back to normal
            if hasattr(obj, 'setStyleSheet'):
                self._animate_button_scale(obj, 1.0)

        return super().eventFilter(obj, event)

    def keyPressEvent(self, event):
        """Handle keyboard navigation for calendar"""
        from PySide6.QtCore import Qt

        key = event.key()

        # Arrow key navigation
        if key == Qt.Key_Left:
            # Move to previous day
            new_date = self._selected_date.addDays(-1)
            if new_date >= self._min_date:
                self._selected_date = new_date
                self._sync_header()
                self._update_calendar()
        elif key == Qt.Key_Right:
            # Move to next day
            new_date = self._selected_date.addDays(1)
            if new_date <= self._max_date:
                self._selected_date = new_date
                self._sync_header()
                self._update_calendar()
        elif key == Qt.Key_Up:
            # Move to previous week
            new_date = self._selected_date.addDays(-7)
            if new_date >= self._min_date:
                self._selected_date = new_date
                self._sync_header()
                self._update_calendar()
        elif key == Qt.Key_Down:
            # Move to next week
            new_date = self._selected_date.addDays(7)
            if new_date <= self._max_date:
                self._selected_date = new_date
                self._sync_header()
                self._update_calendar()
        elif key in (Qt.Key_Return, Qt.Key_Enter, Qt.Key_Space):
            # Select current date and close
            self.dateSelectedCustom.emit(self._selected_date)
            self.close()
        elif key == Qt.Key_Escape:
            # Close without selection
            self.close()
        elif key == Qt.Key_PageUp:
            # Previous month
            self._prev_month()
        elif key == Qt.Key_PageDown:
            # Next month
            self._next_month()
        elif key == Qt.Key_Home:
            # Go to today
            today = QDate.currentDate()
            if self._min_date <= today <= self._max_date:
                self._selected_date = today
                self._sync_header()
                self._update_calendar()
        else:
            super().keyPressEvent(event)
    
    def _animate_button_scale(self, button, scale):
        """Animate button scale with smooth transition"""
        if not hasattr(button, '_scale_anim'):
            button._scale_anim = None
            button._original_geometry = button.geometry()

        # Stop any existing animation
        if button._scale_anim and button._scale_anim.state() == QPropertyAnimation.Running:
            button._scale_anim.stop()

        # Store original geometry if not set
        if not hasattr(button, '_original_geometry') or button._original_geometry is None:
            button._original_geometry = button.geometry()

        # Calculate scaled geometry
        orig = button._original_geometry
        if scale == 1.0:
            target_geom = orig
        else:
            # Calculate the size increase
            size_increase = int((scale - 1.0) * orig.width())
            target_geom = orig.adjusted(-size_increase//2, -size_increase//2,
                                       size_increase//2, size_increase//2)

        # Create smooth scale animation
        button._scale_anim = QPropertyAnimation(button, b"geometry")
        button._scale_anim.setDuration(150)
        button._scale_anim.setStartValue(button.geometry())
        button._scale_anim.setEndValue(target_geom)
        button._scale_anim.setEasingCurve(QEasingCurve.OutCubic)
        button._scale_anim.start()

    def _toggle_header(self):
        """Toggle between collapsed (single button) and expanded (separate dropdowns) - iOS style"""
        self._header_collapsed = not self._header_collapsed
        if self._header_collapsed:
            self.header_btn.show()
            self.month_combo.hide()
            self.year_combo.hide()
            self.year_grid_btn.hide()
        else:
            self.header_btn.hide()
            self.month_combo.show()
            self.year_combo.show()
            self.year_grid_btn.show()
        self._sync_header()

    def _sync_header(self):
        """Sync header button text and combo selections with current date"""
        months = ["January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"]
        
        month = self._selected_date.month()
        year = self._selected_date.year()
        
        # Update collapsed button text
        self.header_btn.setText(f"{months[month-1]} {year}")
        
        # Update month combo
        self.month_combo.blockSignals(True)
        idx_month = self.month_combo.findData(month)
        if idx_month >= 0:
            self.month_combo.setCurrentIndex(idx_month)
        self.month_combo.blockSignals(False)
        
        # Update year combo
        self.year_combo.blockSignals(True)
        idx_year = self.year_combo.findData(year)
        if idx_year >= 0:
            self.year_combo.setCurrentIndex(idx_year)
        self.year_combo.blockSignals(False)

    def _on_combo_changed(self):
        """Called when month or year combo changes"""
        month = self.month_combo.currentData()
        year = self.year_combo.currentData()
        if month and year:
            current_day = self._selected_date.day()
            new_date = QDate(year, month, 1)
            # Ensure day is valid for new month
            if current_day > new_date.daysInMonth():
                current_day = new_date.daysInMonth()
            self._selected_date = QDate(year, month, current_day)
            self._sync_header()
            self._update_calendar()

    def _prev_month(self):
        new_date = self._selected_date.addMonths(-1)
        if new_date >= self._min_date and not self._animating:
            self._slide_direction = 1  # Slide right (prev month comes from left)
            self._animate_month_change(new_date)

    def _next_month(self):
        new_date = self._selected_date.addMonths(1)
        if new_date <= self._max_date and not self._animating:
            self._slide_direction = -1  # Slide left (next month comes from right)
            self._animate_month_change(new_date)

    def _animate_month_change(self, new_date):
        """Animate month transition with slide + fade effect"""
        self._animating = True
        
        # Store original position
        original_pos = self.grid_container.pos()
        slide_distance = 30  # Pixels to slide
        
        # Calculate slide direction
        start_x = original_pos.x()
        end_x_out = start_x + (slide_distance * self._slide_direction)
        end_x_in = start_x - (slide_distance * self._slide_direction)
        
        # Phase 1: Fade out + slide out
        self.fade_out = QPropertyAnimation(self.grid_opacity, b"opacity")
        self.fade_out.setDuration(self._anim_duration_fast)
        self.fade_out.setStartValue(1.0)
        self.fade_out.setEndValue(0.0)
        self.fade_out.setEasingCurve(QEasingCurve.OutCubic)
        
        def on_fade_out_done():
            # Update the calendar while invisible
            self._selected_date = new_date
            self._sync_header()
            self._update_calendar()
            
            # Reset position for slide in from opposite direction
            from PySide6.QtCore import QPoint
            self.grid_container.move(QPoint(int(end_x_in), original_pos.y()))
            
            # Phase 2: Fade in + slide in
            self.fade_in = QPropertyAnimation(self.grid_opacity, b"opacity")
            self.fade_in.setDuration(self._anim_duration_normal)
            self.fade_in.setStartValue(0.0)
            self.fade_in.setEndValue(1.0)
            self.fade_in.setEasingCurve(QEasingCurve.OutCubic)
            
            # Slide animation for phase 2
            self.slide_in = QPropertyAnimation(self.grid_container, b"pos")
            self.slide_in.setDuration(self._anim_duration_normal)
            self.slide_in.setStartValue(QPoint(int(end_x_in), original_pos.y()))
            self.slide_in.setEndValue(original_pos)
            self.slide_in.setEasingCurve(QEasingCurve.OutCubic)
            
            self.fade_in.finished.connect(lambda: setattr(self, '_animating', False))
            self.fade_in.start()
            self.slide_in.start()
        
        self.fade_out.finished.connect(on_fade_out_done)
        self.fade_out.start()

    def _update_calendar(self):
        """Update the calendar grid for current month"""
        year = self._selected_date.year()
        month = self._selected_date.month()
        
        first_of_month = QDate(year, month, 1)
        # Qt: dayOfWeek() returns 1=Monday to 7=Sunday
        # We need Sunday=0 for our grid layout
        qt_dow = first_of_month.dayOfWeek()
        start_day = qt_dow % 7  # Sun(7)->0, Mon(1)->1, ..., Sat(6)->6
        days_in_month = first_of_month.daysInMonth()
        
        # Previous month
        prev_month = first_of_month.addMonths(-1)
        days_in_prev = prev_month.daysInMonth()
        
        today = QDate.currentDate()
        
        day = 1
        next_day = 1
        
        for row in range(6):
            for col in range(7):
                btn = self.day_buttons[row][col]
                cell_index = row * 7 + col
                
                if cell_index < start_day:
                    # Previous month days
                    d = days_in_prev - start_day + cell_index + 1
                    btn.setText(str(d))
                    btn.setEnabled(False)
                    btn.setStyleSheet(self._get_day_style(False, False, False))
                elif day <= days_in_month:
                    # Current month days
                    btn.setText(str(day))
                    current_date = QDate(year, month, day)
                    is_today = (current_date == today)
                    is_selected = (current_date == self._selected_date)
                    btn.setEnabled(True)
                    btn.setStyleSheet(self._get_day_style(True, is_today, is_selected))
                    day += 1
                else:
                    # Next month days
                    btn.setText(str(next_day))
                    btn.setEnabled(False)
                    btn.setStyleSheet(self._get_day_style(False, False, False))
                    next_day += 1

    def _get_day_style(self, enabled, is_today, is_selected):
        primary = self.theme.get('primary', '#4a9eff')

        # Theme-aware colors with shadcn-style enhancements
        if self._is_dark_theme:
            text_color = "white"
            disabled_color = "rgba(255, 255, 255, 0.35)"
            hover_bg = f"{primary}35"
            press_bg = f"{primary}50"
            tile_bg = "rgba(255, 255, 255, 0.06)"
            # shadcn-style subtle shadows
            hover_shadow = "0 2px 8px rgba(0, 0, 0, 0.3)"
            selected_shadow = f"0 4px 12px {primary}60"
        else:
            text_color = "#1a1a2e"
            disabled_color = "rgba(0, 0, 0, 0.3)"
            hover_bg = f"{primary}25"
            press_bg = f"{primary}40"
            tile_bg = "rgba(0, 0, 0, 0.04)"
            # shadcn-style subtle shadows for light theme
            hover_shadow = "0 2px 6px rgba(0, 0, 0, 0.1)"
            selected_shadow = f"0 4px 12px {primary}40"

        # Common transition for smooth animations
        transition = "all 150ms cubic-bezier(0.4, 0, 0.2, 1)"

        if is_selected:
            # Selected: solid accent with glow effect and smooth transitions
            return f"""
                QPushButton {{
                    background: {primary};
                    border: none;
                    border-radius: 17px;
                    color: white;
                    font-weight: 600;
                    font-size: 12px;
                }}
                QPushButton:hover {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 {primary}, stop:1 {self.theme.get('primary_dark', primary)});
                }}
                QPushButton:pressed {{
                    background: {self.theme.get('primary_dark', primary)};
                }}
            """
        elif is_today:
            # Today: ring outline with shadcn-style hover effect
            return f"""
                QPushButton {{
                    background: transparent;
                    border: 2px solid {primary};
                    border-radius: 17px;
                    color: {primary};
                    font-weight: 600;
                    font-size: 12px;
                }}
                QPushButton:hover {{
                    background: {hover_bg};
                    border: 2px solid {primary};
                }}
                QPushButton:pressed {{
                    background: {press_bg};
                }}
            """
        elif enabled:
            # Normal day: shadcn-style clean tile with smooth transitions
            return f"""
                QPushButton {{
                    background: {tile_bg};
                    border: 1px solid transparent;
                    border-radius: 17px;
                    color: {text_color};
                    font-size: 12px;
                    font-weight: 500;
                }}
                QPushButton:hover {{
                    background: {hover_bg};
                    border: 1px solid {primary}30;
                }}
                QPushButton:pressed {{
                    background: {press_bg};
                }}
            """
        else:
            # Disabled: other-month days, very subtle
            return f"""
                QPushButton {{
                    background: transparent;
                    border: none;
                    border-radius: 17px;
                    color: {disabled_color};
                    font-size: 11px;
                    font-weight: 400;
                }}
            """

    def _on_day_clicked(self, row, col):
        btn = self.day_buttons[row][col]
        if btn.isEnabled():
            day = int(btn.text())
            self._selected_date = QDate(self._selected_date.year(), self._selected_date.month(), day)
            
            # Animate selection with brief pulse effect before closing
            self._animate_selection(btn)
    
    def _animate_selection(self, btn):
        """Animate selection with ripple-like pulse effect"""
        primary = self.theme.get('primary', '#4a9eff')
        
        # Create opacity effect for pulse
        pulse_effect = QGraphicsOpacityEffect(btn)
        btn.setGraphicsEffect(pulse_effect)
        
        # Pulse animation: quick bright flash then fade
        self.pulse_anim = QPropertyAnimation(pulse_effect, b"opacity")
        self.pulse_anim.setDuration(self._anim_duration_fast)
        self.pulse_anim.setStartValue(0.7)
        self.pulse_anim.setEndValue(1.0)
        self.pulse_anim.setEasingCurve(QEasingCurve.OutCubic)
        
        def on_pulse_done():
            btn.setGraphicsEffect(None)  # Remove effect
            self.dateSelectedCustom.emit(self._selected_date)
            self.close()
        
        self.pulse_anim.finished.connect(on_pulse_done)
        self.pulse_anim.start()

    def _apply_style(self):
        primary = self.theme.get('primary', '#4a9eff')
        primary_dark = self.theme.get('primary_dark', '#1976D2')
        
        # Theme-aware glassmorphism colors
        if self._is_dark_theme:
            glass_bg = f"rgba(30, 35, 45, {self._glass_opacity})"
            border_color = "rgba(255, 255, 255, 0.15)"
            text_color = "white"
            nav_bg = "rgba(255, 255, 255, 0.08)"
            nav_hover = f"{primary}50"
            header_bg = "rgba(255, 255, 255, 0.1)"
            header_hover = "rgba(255, 255, 255, 0.18)"
            shadow_color = "rgba(0, 0, 0, 0.4)"
        else:
            glass_bg = f"rgba(255, 255, 255, {self._glass_opacity})"
            border_color = "rgba(0, 0, 0, 0.1)"
            text_color = "#1a1a2e"
            nav_bg = "rgba(0, 0, 0, 0.05)"
            nav_hover = f"{primary}30"
            header_bg = "rgba(0, 0, 0, 0.06)"
            header_hover = "rgba(0, 0, 0, 0.1)"
            shadow_color = "rgba(0, 0, 0, 0.15)"
        
        # Glassmorphism container with frosted glass effect
        # Note: Qt doesn't support backdrop-filter, so we simulate with semi-transparent bg + shadow
        self.setStyleSheet(f"""
            CalendarPopup {{
                background: transparent;
            }}
        """)
        
        self.glass_container.setStyleSheet(f"""
            #glassContainer {{
                background: {glass_bg};
                border: 1px solid {border_color};
                border-radius: 16px;
            }}
        """)
        
        # Add drop shadow for depth (glassmorphism effect)
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(24)
        shadow.setColor(QColor(0, 0, 0, 80 if self._is_dark_theme else 40))
        shadow.setOffset(0, 8)
        self.glass_container.setGraphicsEffect(shadow)
        
        # Circular navigation buttons with smooth hover transition
        nav_style = f"""
            QToolButton {{
                background: {nav_bg};
                border: none;
                border-radius: 16px;
                color: {text_color};
                font-size: 20px;
                font-weight: 300;
            }}
            QToolButton:hover {{
                background: {nav_hover};
            }}
            QToolButton:pressed {{
                background: {primary}60;
            }}
        """
        self.prev_btn.setStyleSheet(nav_style)
        self.next_btn.setStyleSheet(nav_style)
        
        # Header button (collapsed state) - modern pill with subtle animation
        self.header_btn.setStyleSheet(f"""
            QPushButton {{
                background: {header_bg};
                border: none;
                border-radius: 13px;
                color: {text_color};
                font-size: 14px;
                font-weight: 600;
                padding: 4px 16px;
            }}
            QPushButton:hover {{
                background: {header_hover};
            }}
            QPushButton:pressed {{
                background: {primary}30;
            }}
        """)
        
        # Theme-aware dropdown colors
        if self._is_dark_theme:
            combo_bg = "rgba(255, 255, 255, 0.1)"
            combo_hover = "rgba(255, 255, 255, 0.18)"
            combo_text = "white"
            dropdown_bg = "rgb(35, 40, 50)"
            dropdown_border = "rgba(255, 255, 255, 0.12)"
            arrow_color = "rgba(255, 255, 255, 0.6)"
        else:
            combo_bg = "rgba(0, 0, 0, 0.06)"
            combo_hover = "rgba(0, 0, 0, 0.1)"
            combo_text = "#1a1a2e"
            dropdown_bg = "rgb(252, 252, 255)"
            dropdown_border = "rgba(0, 0, 0, 0.1)"
            arrow_color = "rgba(0, 0, 0, 0.5)"
        
        # Month/Year dropdowns (expanded state) - modern glassmorphic style
        combo_style = f"""
            QComboBox {{
                background: {combo_bg};
                border: none;
                border-radius: 10px;
                color: {combo_text};
                font-size: 12px;
                font-weight: 600;
                padding: 4px 22px 4px 10px;
            }}
            QComboBox:hover {{
                background: {combo_hover};
            }}
            QComboBox::drop-down {{
                border: none;
                width: 20px;
            }}
            QComboBox::down-arrow {{
                image: none;
                border: 2px solid {arrow_color};
                width: 6px;
                height: 6px;
                border-top: none;
                border-left: none;
            }}
            QComboBox QAbstractItemView {{
                background: {dropdown_bg};
                border: 1px solid {dropdown_border};
                border-radius: 10px;
                selection-background-color: {primary}50;
                color: {combo_text};
                padding: 6px;
                outline: none;
            }}
            QComboBox QAbstractItemView::item {{
                padding: 6px 10px;
                border-radius: 6px;
                margin: 2px;
            }}
            QComboBox QAbstractItemView::item:hover {{
                background: {primary}25;
            }}
            QComboBox QAbstractItemView::item:selected {{
                background: {primary}45;
            }}
        """
        self.month_combo.setStyleSheet(combo_style)
        self.year_combo.setStyleSheet(combo_style)
        
        # Style for year grid button
        year_grid_btn_style = f"""
            QToolButton {{
                background: {primary}30;
                border: 1px solid {primary}60;
                border-radius: 8px;
                color: {combo_text};
                font-size: 14px;
            }}
            QToolButton:hover {{
                background: {primary}50;
                border: 1px solid {primary};
            }}
        """
        self.year_grid_btn.setStyleSheet(year_grid_btn_style)
    
    def _show_year_grid(self):
        """Show a year grid overlay for quick year selection"""
        if hasattr(self, '_year_grid_overlay') and self._year_grid_overlay.isVisible():
            self._year_grid_overlay.hide()
            return
        
        # Create year grid overlay
        self._year_grid_overlay = QWidget(self.glass_container)
        self._year_grid_overlay.setObjectName("yearGridOverlay")
        
        # Position and size to cover the calendar grid
        grid_rect = self.day_grid.geometry()
        self._year_grid_overlay.setGeometry(0, grid_rect.y() - 20, self.glass_container.width(), grid_rect.height() + 40)
        
        # Style the overlay
        bg_color = self.theme.get("surface", "#22262e")
        text_color = self.theme.get("text_primary", "#e8eaed")
        primary = self.theme.get("primary", "#4a8fd9")
        
        self._year_grid_overlay.setStyleSheet(f"""
            QWidget#yearGridOverlay {{
                background: {bg_color};
                border-radius: 12px;
            }}
        """)
        
        overlay_layout = QVBoxLayout(self._year_grid_overlay)
        overlay_layout.setContentsMargins(8, 8, 8, 8)
        overlay_layout.setSpacing(4)
        
        # Year grid (4 columns x 5 rows = 20 years)
        year_grid_layout = QGridLayout()
        year_grid_layout.setSpacing(4)
        
        current_year = self._selected_date.year()
        start_year = (current_year // 10) * 10 - 5  # Center around current decade
        
        for i in range(20):
            year = start_year + i
            btn = QPushButton(str(year))
            btn.setFixedSize(50, 30)
            btn.setCursor(Qt.PointingHandCursor)
            
            is_selected = year == current_year
            btn_bg = f"{primary}80" if is_selected else "transparent"
            btn_border = f"2px solid {primary}" if is_selected else "1px solid rgba(255,255,255,0.2)"
            font_weight = "700" if is_selected else "500"
            
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {btn_bg};
                    color: {text_color};
                    border: {btn_border};
                    border-radius: 8px;
                    font-size: 12px;
                    font-weight: {font_weight};
                }}
                QPushButton:hover {{
                    background: {primary}40;
                    border: 2px solid {primary};
                }}
            """)
            btn.clicked.connect(lambda checked, y=year: self._select_year_from_grid(y))
            year_grid_layout.addWidget(btn, i // 4, i % 4)
        
        overlay_layout.addLayout(year_grid_layout)
        
        # Navigation buttons
        nav_layout = QHBoxLayout()
        nav_layout.setSpacing(8)
        
        prev_decade = QPushButton("◀ Earlier")
        prev_decade.setCursor(Qt.PointingHandCursor)
        prev_decade.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                color: {primary};
                border: none;
                padding: 5px 10px;
                font-size: 11px;
            }}
            QPushButton:hover {{
                text-decoration: underline;
            }}
        """)
        prev_decade.clicked.connect(lambda: self._shift_year_grid(-20))
        
        close_btn = QPushButton("✕")
        close_btn.setFixedSize(24, 24)
        close_btn.setCursor(Qt.PointingHandCursor)
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background: rgba(255,255,255,0.1);
                color: {text_color};
                border: none;
                border-radius: 12px;
                font-size: 12px;
            }}
            QPushButton:hover {{
                background: rgba(255,0,0,0.3);
            }}
        """)
        close_btn.clicked.connect(lambda: self._year_grid_overlay.hide())
        
        next_decade = QPushButton("Later ▶")
        next_decade.setCursor(Qt.PointingHandCursor)
        next_decade.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                color: {primary};
                border: none;
                padding: 5px 10px;
                font-size: 11px;
            }}
            QPushButton:hover {{
                text-decoration: underline;
            }}
        """)
        next_decade.clicked.connect(lambda: self._shift_year_grid(20))
        
        nav_layout.addWidget(prev_decade)
        nav_layout.addStretch()
        nav_layout.addWidget(close_btn)
        nav_layout.addStretch()
        nav_layout.addWidget(next_decade)
        overlay_layout.addLayout(nav_layout)
        
        # Store grid start year for shifting
        self._year_grid_start = start_year
        
        # Show with fade-in animation
        self._year_grid_overlay.show()
        self._year_grid_overlay.raise_()
        
        # Opacity animation
        opacity = QGraphicsOpacityEffect(self._year_grid_overlay)
        self._year_grid_overlay.setGraphicsEffect(opacity)
        anim = QPropertyAnimation(opacity, b"opacity", self)
        anim.setDuration(150)
        anim.setStartValue(0)
        anim.setEndValue(1)
        anim.setEasingCurve(QEasingCurve.OutCubic)
        anim.start()
        self._year_grid_anim = anim  # Keep reference
    
    def _select_year_from_grid(self, year):
        """Select a year from the grid and close overlay"""
        current_month = self._selected_date.month()
        current_day = min(self._selected_date.day(), QDate(year, current_month, 1).daysInMonth())
        self._selected_date = QDate(year, current_month, current_day)
        self._sync_header()
        self._update_calendar()
        if hasattr(self, '_year_grid_overlay'):
            self._year_grid_overlay.hide()
    
    def _shift_year_grid(self, delta):
        """Shift the year grid by delta years and rebuild"""
        if hasattr(self, '_year_grid_overlay'):
            self._year_grid_overlay.hide()
            self._year_grid_start = getattr(self, '_year_grid_start', self._selected_date.year() - 5) + delta
            # Temporarily adjust selected date to center the grid
            saved_date = self._selected_date
            self._selected_date = QDate(self._year_grid_start + 10, saved_date.month(), 1)
            self._show_year_grid()
            self._selected_date = saved_date

    def setSelectedDate(self, date: QDate):
        self._selected_date = date
        self._sync_header()
        self._update_calendar()

    def setMinimumDate(self, date: QDate):
        self._min_date = date

    def setMaximumDate(self, date: QDate):
        self._max_date = date


# ======================================================================
# DatePicker - Styled like department dropdown
# ======================================================================

class DatePicker(QWidget):
    """
    Custom date picker styled with modern theme-aware design.
    Clickable field that opens an iOS-style 3D wheel picker popup.
    """
    dateChanged = Signal(QDate)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._date = QDate.currentDate()
        self._format = "MM-dd-yyyy"
        self._min_date = QDate(1950, 1, 1)
        self._max_date = QDate(2100, 12, 31)
        
        # Get theme colors and detect light/dark mode
        theme_name = load_theme_preference()
        self.theme = MODERN_THEMES.get(theme_name, MODERN_THEMES.get("default", MODERN_THEMES[list(MODERN_THEMES.keys())[0]]))
        self._is_dark_theme = theme_name.lower() in ['dark', 'midnight', 'ocean', 'default']
        
        # Use QComboBox-like widget for consistent styling with department
        self.field = QLineEdit(self)
        self.field.setReadOnly(True)
        self.field.setCursor(Qt.PointingHandCursor)
        self.field.setFocusPolicy(Qt.ClickFocus)
        self.field.installEventFilter(self)
        
        # Layout - no margins, let stylesheet handle sizing like Department dropdown
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        layout.addWidget(self.field)
        
        # Create wheel picker popup (iOS-style 3D drums)
        self.wheel_picker = WheelDatePickerPopup(self)
        self.wheel_picker.dateSelected.connect(self._on_date_selected)
        
        # Apply styling to match department dropdown
        self._apply_style()
        
        # Set initial date display
        self._update_display()

    def eventFilter(self, obj, event):
        """Handle click on the date field"""
        if obj == self.field and event.type() == QEvent.MouseButtonPress:
            self.show_calendar()
            return True
        return super().eventFilter(obj, event)

    def _apply_style(self):
        """Apply theme-aware styling to match Department dropdown exactly"""
        primary = self.theme.get('primary', '#4a9eff')
        
        if self._is_dark_theme:
            # Dark theme styling - match Department dropdown exactly
            self.field.setStyleSheet("""
                QLineEdit {
                    padding: 10px 12px;
                    border: 2px solid rgba(33, 150, 243, 0.6);
                    border-radius: 12px;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                               stop:0 rgba(55, 65, 85, 0.95),
                                               stop:1 rgba(45, 55, 75, 0.95));
                    color: white;
                    font-size: 13px;
                    font-weight: 500;
                    min-height: 44px;
                }
                QLineEdit:hover {
                    border: 2px solid rgba(66, 165, 245, 0.9);
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                               stop:0 rgba(65, 75, 95, 0.95),
                                               stop:1 rgba(55, 65, 85, 0.95));
                }
                QLineEdit:focus {
                    border: 2px solid rgba(100, 181, 246, 1.0);
                }
                QLineEdit:disabled {
                    background: rgba(60, 60, 60, 0.5);
                    color: rgba(255, 255, 255, 0.4);
                    border: 2px solid rgba(100, 100, 100, 0.3);
                }
            """)
        else:
            # Light theme styling
            self.field.setStyleSheet(f"""
                QLineEdit {{
                    padding: 10px 12px;
                    border: 2px solid {primary}60;
                    border-radius: 12px;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                               stop:0 rgba(255, 255, 255, 0.98),
                                               stop:1 rgba(248, 250, 252, 0.98));
                    color: #1a1a2e;
                    font-size: 13px;
                    font-weight: 500;
                    min-height: 44px;
                }}
                QLineEdit:hover {{
                    border: 2px solid {primary}90;
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                               stop:0 rgba(255, 255, 255, 1.0),
                                               stop:1 rgba(252, 253, 255, 1.0));
                }}
                QLineEdit:focus {{
                    border: 2px solid {primary};
                }}
                QLineEdit:disabled {{
                    background: rgba(240, 240, 240, 0.5);
                    color: rgba(0, 0, 0, 0.35);
                    border: 2px solid rgba(0, 0, 0, 0.1);
                }}
            """)

    def show_calendar(self):
        """Show the wheel picker popup below the widget, ensuring it stays on screen"""
        if not self.isEnabled():
            return
        # Update wheel picker limits and date
        self.wheel_picker.setMinimumDate(self._min_date)
        self.wheel_picker.setMaximumDate(self._max_date)
        self.wheel_picker.setSelectedDate(self._date)
        
        # Calculate position - try below first, then above if needed
        global_pos = self.mapToGlobal(QPoint(0, self.height() + 4))
        
        # Get screen geometry
        screen = QApplication.primaryScreen()
        if screen:
            screen_rect = screen.availableGeometry()
            picker_height = self.wheel_picker.height()
            picker_width = self.wheel_picker.width()
            
            # If picker would go off bottom, show it above the field
            if global_pos.y() + picker_height > screen_rect.bottom():
                global_pos = self.mapToGlobal(QPoint(0, -picker_height - 4))
            
            # If picker would go off right, shift it left
            if global_pos.x() + picker_width > screen_rect.right():
                global_pos.setX(screen_rect.right() - picker_width - 10)
            
            # Ensure not off left side
            if global_pos.x() < screen_rect.left():
                global_pos.setX(screen_rect.left() + 10)
        
        self.wheel_picker.move(global_pos)
        self.wheel_picker.show()

    def _on_date_selected(self, date):
        """Handle date selection from calendar"""
        self._date = date
        self._update_display()
        self.dateChanged.emit(date)

    def _update_display(self):
        """Update the display field with current date"""
        self.field.setText(self._date.toString(self._format))

    def date(self) -> QDate:
        """Get the current date"""
        return self._date

    def setDate(self, date: QDate):
        """Set the current date"""
        if date.isValid():
            self._date = date
            self._update_display()
            self.wheel_picker.setSelectedDate(date)

    def setMinimumDate(self, date: QDate):
        """Set minimum selectable date"""
        self._min_date = date

    def setMaximumDate(self, date: QDate):
        """Set maximum selectable date"""
        self._max_date = date

    def setDisplayFormat(self, format_str: str):
        """Set the date display format"""
        self._format = format_str
        self._update_display()

    def setEnabled(self, enabled: bool):
        """Enable/disable the widget"""
        super().setEnabled(enabled)
        self.field.setEnabled(enabled)
        self.field.setCursor(Qt.PointingHandCursor if enabled else Qt.ArrowCursor)

    def setFocusPolicy(self, policy):
        super().setFocusPolicy(policy)
        self.field.setFocusPolicy(policy)

    def setSizePolicy(self, *args):
        super().setSizePolicy(*args)
        self.field.setSizePolicy(*args)

    def setFixedWidth(self, width: int):
        super().setFixedWidth(width)
        self.field.setFixedWidth(width)

    def setFixedHeight(self, height: int):
        """Set fixed height for both widget and field"""
        super().setFixedHeight(height)
        self.field.setFixedHeight(height)

    def setMinimumWidth(self, width: int):
        """Set minimum width"""
        super().setMinimumWidth(width)
        self.field.setMinimumWidth(width)


# ======================================================================
# WheelColumn - REMOVED (unused iOS-style wheel picker)
# WheelDatePickerPopup - REMOVED (unused)
# WheelDatePicker - REMOVED (unused)
# ======================================================================
# NOTE: These widgets were fully implemented but never used in the app.
# The app uses CalendarPopup and DatePicker instead.
# Removed ~420 lines of dead code for maintainability.
# If needed in future, check git history for full implementation.


# ======================================================================
# ModernDateEdit (Legacy - kept for compatibility)
# ======================================================================

class ModernDateEdit(QDateEdit):
    """Custom QDateEdit with premium iOS frosted glass styling - Click only, no typing"""
    def __init__(self, parent=None):
        super().__init__(parent)

        # Essential Settings
        self.setCalendarPopup(True)
        self.setMinimumWidth(160)
        self.setFixedHeight(40)  # Taller for modern touch functionality
        
        # Make it behave like a clickable dropdown
        self.setCursor(Qt.PointingHandCursor)
        
        # Prevent keyboard focus to stop typing
        self.setButtonSymbols(QDateEdit.NoButtons)
        self.lineEdit().setReadOnly(True)  # Prevent typing in line edit only

        # Attach the Modern Calendar (Pop-up)
        calendar = ModernCalendarWidget()
        self.setCalendarWidget(calendar)

        # Apply Premium iOS Frosted Glass Styling
        self._apply_premium_glass_style()
    
    def keyPressEvent(self, event):
        """Ignore all key presses to prevent typing - only allow calendar click"""
        event.ignore()
    
    def mousePressEvent(self, event):
        """Open calendar popup on any click"""
        # Show calendar popup when clicking anywhere on the widget
        super().mousePressEvent(event)
        if self.calendarPopup():
            self.calendarWidget().show()

    def _apply_premium_glass_style(self):
        """Apply premium iOS frosted glass styling to the date picker input box"""
        self.setStyleSheet("""
            QDateEdit {
                /* Premium Glass Background - Subtle gradient for depth */
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(255, 255, 255, 0.12),
                    stop:1 rgba(255, 255, 255, 0.08));

                /* Glass Borders - Multi-layer for 3D effect */
                border: 1px solid rgba(255, 255, 255, 0.2);
                border-top: 1px solid rgba(255, 255, 255, 0.3);    /* Highlight top edge */
                border-bottom: 1px solid rgba(0, 0, 0, 0.2);       /* Shadow bottom edge */

                border-radius: 10px;
                padding-left: 12px;
                color: white;
                font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif;
                font-size: 14px;
                font-weight: 500;
            }

            /* Hover State - Slightly brighter glass */
            QDateEdit:hover {
                background: rgba(255, 255, 255, 0.15);
                border: 1px solid rgba(255, 255, 255, 0.3);
                border-top: 1px solid rgba(255, 255, 255, 0.4);
            }

            /* Focus State - Blue Glow (iOS accent) */
            QDateEdit:focus {
                background: rgba(255, 255, 255, 0.18);
                border: 1.5px solid #4a9eff;
                border-top: 1.5px solid rgba(74, 158, 255, 0.8);
            }

            /* Dropdown Button Area */
            QDateEdit::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 32px;
                border-left-width: 0px;
                border-top-right-radius: 10px;
                border-bottom-right-radius: 10px;
                background: transparent;
            }

            /* Custom White Arrow Icon */
            QDateEdit::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 6px solid rgba(255, 255, 255, 0.7);
                margin-right: 10px;
                height: 0;
                width: 0;
            }

            QDateEdit::down-arrow:hover {
                border-top: 6px solid #4a9eff; /* Blue arrow on hover */
            }

            /* Disabled State */
            QDateEdit:disabled {
                background: rgba(255, 255, 255, 0.05);
                border: 1px solid rgba(255, 255, 255, 0.1);
                color: rgba(255, 255, 255, 0.3);
            }
        """)


# ============================================================================
# END CUSTOM MODERN CALENDAR
# ============================================================================


# ============================================================================
# MODERN ANIMATED WIDGETS
# ============================================================================

class AnimatedButton(QPushButton):
    """
    Button with smooth hover animations
    - Grows slightly on hover
    - Smooth transitions

    Usage:
        btn = AnimatedButton("Click Me")
        btn.clicked.connect(some_function)
    """

    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.default_height = 36
        self.setMinimumHeight(self.default_height)

        # Hover animation
        self.hover_animation = QPropertyAnimation(self, b"minimumHeight")
        self.hover_animation.setDuration(200)
        self.hover_animation.setEasingCurve(QEasingCurve.OutCubic)

    def enterEvent(self, event):
        """Mouse enter - grow button"""
        self.hover_animation.stop()
        self.hover_animation.setStartValue(self.height())
        self.hover_animation.setEndValue(self.default_height + 4)
        self.hover_animation.start()
        super().enterEvent(event)

    def leaveEvent(self, event):
        """Mouse leave - shrink button"""
        self.hover_animation.stop()
        self.hover_animation.setStartValue(self.height())
        self.hover_animation.setEndValue(self.default_height)
        self.hover_animation.start()
        super().leaveEvent(event)


class LoadingOverlay(QWidget):
    """
    Full-screen loading overlay with spinner and message

    Usage:
        overlay = LoadingOverlay(parent, "Processing...")
        overlay.show()
        # ... do work ...
        overlay.hide()
    """

    def __init__(self, parent=None, message="Loading..."):
        super().__init__(parent)
        self.setAttribute(Qt.WA_TransparentForMouseEvents, False)
        self.setAttribute(Qt.WA_TranslucentBackground)

        # Layout
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)

        # Spinner
        self.spinner = LoadingSpinner(self, 60)
        layout.addWidget(self.spinner, alignment=Qt.AlignCenter)

        # Message
        self.label = QLabel(message)
        self.label.setStyleSheet("""
            QLabel {
                color: white;
                font-size: 14px;
                font-weight: bold;
                background-color: rgba(0, 0, 0, 150);
                padding: 10px 20px;
                border-radius: 8px;
            }
        """)
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label, alignment=Qt.AlignCenter)

        # Start spinner
        self.spinner.start()

        # Match parent size
        if parent:
            self.setGeometry(parent.rect())

    def paintEvent(self, event):
        """Draw semi-transparent background"""
        painter = QPainter(self)
        painter.fillRect(self.rect(), QColor(0, 0, 0, 100))

    def set_message(self, message):
        """Update loading message"""
        self.label.setText(message)


# ============================================================================
# HELPER FUNCTIONS FOR MODERN UI
# ============================================================================

def create_circular_pixmap(pixmap, size=160):
    """
    Convert square pixmap to circular (like profile photos)

    Usage:
        circular_photo = create_circular_pixmap(original_photo, 160)
        label.setPixmap(circular_photo)

    Args:
        pixmap: QPixmap to make circular
        size: Diameter of circle in pixels

    Returns:
        Circular QPixmap with transparent background
    """
    if pixmap.isNull():
        return QPixmap()

    # Scale to size
    scaled = pixmap.scaled(size, size, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)

    # Create circular mask
    result = QPixmap(size, size)
    result.fill(Qt.transparent)

    painter = QPainter(result)
    painter.setRenderHint(QPainter.Antialiasing, True)

    path = QPainterPath()
    path.addEllipse(0, 0, size, size)
    painter.setClipPath(path)
    painter.drawPixmap(0, 0, scaled)
    painter.end()

    return result


# ============================================================================
# ADVANCED ANIMATED WIDGETS - PHASE 10
# ============================================================================

class PageTransitionWidget(QStackedWidget):
    """
    QStackedWidget with smooth page transitions

    Usage:
        pages = PageTransitionWidget()
        pages.addWidget(page1)
        pages.addWidget(page2)
        pages.transition_to(1)  # Smooth fade to page 1
    """

    def __init__(self, parent=None, animation_type="fade"):
        super().__init__(parent)
        self.animation_type = animation_type  # "fade", "slide_left", "slide_right"
        self.current_animation = None

    def transition_to(self, index, duration=300):
        """Transition to page with animation"""
        if index == self.currentIndex():
            return

        if self.animation_type == "fade":
            self._fade_transition(index, duration)
        elif self.animation_type == "slide_left":
            self._slide_transition(index, duration, direction="left")
        elif self.animation_type == "slide_right":
            self._slide_transition(index, duration, direction="right")
        else:
            self.setCurrentIndex(index)

    def _fade_transition(self, index, duration):
        """Fade transition between pages"""
        old_widget = self.currentWidget()
        new_widget = self.widget(index)

        if not old_widget or not new_widget:
            self.setCurrentIndex(index)
            return

        # Setup opacity effects
        old_effect = QGraphicsOpacityEffect(old_widget)
        old_widget.setGraphicsEffect(old_effect)

        new_effect = QGraphicsOpacityEffect(new_widget)
        new_widget.setGraphicsEffect(new_effect)
        new_effect.setOpacity(0.0)

        # Show new widget
        self.setCurrentIndex(index)

        # Fade out old, fade in new
        fade_out = QPropertyAnimation(old_effect, b"opacity")
        fade_out.setDuration(duration)
        fade_out.setStartValue(1.0)
        fade_out.setEndValue(0.0)
        fade_out.setEasingCurve(QEasingCurve.InOutQuad)

        fade_in = QPropertyAnimation(new_effect, b"opacity")
        fade_in.setDuration(duration)
        fade_in.setStartValue(0.0)
        fade_in.setEndValue(1.0)
        fade_in.setEasingCurve(QEasingCurve.InOutQuad)

        # Cleanup after animation
        fade_in.finished.connect(lambda: new_widget.setGraphicsEffect(None))
        fade_out.finished.connect(lambda: old_widget.setGraphicsEffect(None))

        fade_out.start()
        fade_in.start()

        self.current_animation = (fade_out, fade_in)

    def _slide_transition(self, index, duration, direction="left"):
        """Slide transition between pages"""
        # For now, use fade as slide requires more complex setup
        self._fade_transition(index, duration)


class CollapsibleSidebar(QWidget):
    """
    Animated collapsible sidebar like modern apps

    Usage:
        sidebar = CollapsibleSidebar()
        sidebar.add_button("Dashboard", icon, callback)
        sidebar.add_button("Employees", icon, callback)
        sidebar.toggle()  # Collapse/expand
    """

    toggled = Signal(bool)  # Emits True when expanded, False when collapsed

    def __init__(self, parent=None):
        super().__init__(parent)
        self.expanded = True
        # Per reference dashboard: 240px expanded, 80px collapsed
        self.expanded_width = 240
        self.collapsed_width = 80

        self.setFixedWidth(self.expanded_width)

        # Main layout
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)

        # Toggle button at top
        self.toggle_button = QPushButton("☰")
        self.toggle_button.setFixedHeight(50)
        self.toggle_button.clicked.connect(self.toggle)
        self.toggle_button.setStyleSheet("""
            QPushButton {
                font-size: 20px;
                border: none;
                text-align: center;
            }
            QPushButton:hover {
                background-color: rgba(255, 255, 255, 0.1);
            }
        """)
        self.main_layout.addWidget(self.toggle_button)

        # Button container
        self.button_layout = QVBoxLayout()
        self.button_layout.setSpacing(2)
        self.main_layout.addLayout(self.button_layout)
        self.main_layout.addStretch()

        # Animation
        self.animation = QPropertyAnimation(self, b"minimumWidth")
        self.animation.setDuration(300)
        self.animation.setEasingCurve(QEasingCurve.InOutQuart)

        self.animation2 = QPropertyAnimation(self, b"maximumWidth")
        self.animation2.setDuration(300)
        self.animation2.setEasingCurve(QEasingCurve.InOutQuart)

        self.buttons = []

    def add_button(self, text, icon=None, callback=None):
        """Add a button to sidebar"""
        btn = QPushButton(text)
        btn.setFixedHeight(50)
        btn.setStyleSheet("""
            QPushButton {
                text-align: left;
                padding-left: 15px;
                border: none;
                font-size: 13px;
            }
            QPushButton:hover {
                background-color: rgba(255, 255, 255, 0.1);
            }
        """)

        if callback:
            btn.clicked.connect(callback)

        self.button_layout.addWidget(btn)
        self.buttons.append((btn, text))

        return btn

    def toggle(self):
        """Toggle sidebar expanded/collapsed"""
        if self.expanded:
            # Collapse
            self.animation.setStartValue(self.expanded_width)
            self.animation.setEndValue(self.collapsed_width)
            self.animation2.setStartValue(self.expanded_width)
            self.animation2.setEndValue(self.collapsed_width)

            # Hide text on buttons
            for btn, text in self.buttons:
                btn.setText("")
        else:
            # Expand
            self.animation.setStartValue(self.collapsed_width)
            self.animation.setEndValue(self.expanded_width)
            self.animation2.setStartValue(self.collapsed_width)
            self.animation2.setEndValue(self.expanded_width)

            # Show text on buttons
            for btn, text in self.buttons:
                btn.setText(text)

        self.expanded = not self.expanded
        self.animation.start()
        self.animation2.start()
        self.toggled.emit(self.expanded)


class SkeletonLoader(QWidget):
    """
    Skeleton loading screen (like Facebook, LinkedIn)
    Shows animated placeholder while content loads

    Usage:
        skeleton = SkeletonLoader(rows=5)
        skeleton.start()
        # ... load data ...
        skeleton.stop()
    """

    def __init__(self, parent=None, rows=3):
        super().__init__(parent)
        self.rows = rows
        self.animation_offset = 0

        # Setup UI
        layout = QVBoxLayout(self)

        for i in range(rows):
            row = self._create_skeleton_row()
            layout.addWidget(row)

        # Timer for shimmer animation
        self.timer = QTimer(self)
        self.timer.timeout.connect(self._update_shimmer)

    def _create_skeleton_row(self):
        """Create one skeleton row"""
        frame = QFrame()
        frame.setFixedHeight(60)
        frame.setStyleSheet("""
            QFrame {
                background-color: rgba(128, 128, 128, 0.1);
                border-radius: 8px;
                margin: 5px;
            }
        """)
        return frame

    def _update_shimmer(self):
        """Update shimmer animation"""
        self.animation_offset = (self.animation_offset + 5) % 100
        self.update()

    def start(self):
        """Start skeleton animation"""
        self.show()
        self.timer.start(50)  # 20 FPS

    def stop(self):
        """Stop skeleton animation"""
        self.timer.stop()
        self.hide()

    def paintEvent(self, event):
        """Draw shimmer effect"""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # Draw shimmer gradient across all rows
        gradient = QLinearGradient(0, 0, self.width(), 0)
        pos = (self.animation_offset % 100) / 100.0

        gradient.setColorAt(max(0, pos - 0.3), QColor(200, 200, 200, 30))
        gradient.setColorAt(pos, QColor(255, 255, 255, 80))
        gradient.setColorAt(min(1, pos + 0.3), QColor(200, 200, 200, 30))

        painter.fillRect(self.rect(), gradient)


class RippleButton(QPushButton):
    """
    Button with Material Design ripple effect

    Usage:
        btn = RippleButton("Click Me")
        btn.clicked.connect(callback)
    """

    def __init__(self, text, parent=None):
        super().__init__(text, parent)
        self.ripple_radius = 0
        self.ripple_pos = None
        self.ripple_animation = None

    def mousePressEvent(self, event):
        """Create ripple effect on click"""
        self.ripple_pos = event.pos()
        self.ripple_radius = 0

        # Calculate max radius
        max_radius = max(
            self.ripple_pos.x(),
            self.ripple_pos.y(),
            self.width() - self.ripple_pos.x(),
            self.height() - self.ripple_pos.y()
        ) * 2

        # Animate ripple
        self.ripple_animation = QPropertyAnimation(self, b"ripple_radius")
        self.ripple_animation.setDuration(500)
        self.ripple_animation.setStartValue(0)
        self.ripple_animation.setEndValue(max_radius)
        self.ripple_animation.setEasingCurve(QEasingCurve.OutCubic)
        self.ripple_animation.valueChanged.connect(lambda: self.update())
        self.ripple_animation.start()

        super().mousePressEvent(event)

    def paintEvent(self, event):
        """Draw button with ripple"""
        super().paintEvent(event)

        if self.ripple_pos and self.ripple_radius > 0:
            painter = QPainter(self)
            painter.setRenderHint(QPainter.Antialiasing)

            # Draw ripple
            color = QColor(255, 255, 255, 100)
            painter.setBrush(color)
            painter.setPen(Qt.NoPen)
            painter.drawEllipse(
                self.ripple_pos,
                int(self.ripple_radius),
                int(self.ripple_radius)
            )

    @property
    def ripple_radius_value(self):
        return self.ripple_radius

    @ripple_radius_value.setter
    def ripple_radius_value(self, value):
        self.ripple_radius = value
        self.update()


class SlideInWidget(QWidget):
    """
    Widget that slides in from the side

    Usage:
        widget = SlideInWidget(content_widget, direction="left")
        widget.slide_in()
    """

    def __init__(self, child_widget, direction="left", parent=None):
        super().__init__(parent)
        self.child_widget = child_widget
        self.direction = direction

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(child_widget)

    def slide_in(self, duration=400):
        """Slide widget in from edge"""
        # Get parent geometry
        parent_width = self.parent().width() if self.parent() else 800
        parent_height = self.parent().height() if self.parent() else 600

        # Set start position (off-screen)
        if self.direction == "left":
            start_x = -self.width()
            end_x = 0
        elif self.direction == "right":
            start_x = parent_width
            end_x = parent_width - self.width()
        elif self.direction == "top":
            start_y = -self.height()
            end_y = 0
        elif self.direction == "bottom":
            start_y = parent_height
            end_y = parent_height - self.height()

        # Animate position
        if self.direction in ["left", "right"]:
            animation = QPropertyAnimation(self, b"pos")
            animation.setDuration(duration)
            animation.setStartValue(QPoint(start_x, self.y()))
            animation.setEndValue(QPoint(end_x, self.y()))
        else:
            animation = QPropertyAnimation(self, b"pos")
            animation.setDuration(duration)
            animation.setStartValue(QPoint(self.x(), start_y))
            animation.setEndValue(QPoint(self.x(), end_y))

        animation.setEasingCurve(QEasingCurve.OutCubic)
        animation.start()

        self.animation = animation  # Keep reference


class SimpleChartWidget(QWidget):
    """
    Simple bar/line chart for dashboard

    Usage:
        chart = SimpleChartWidget("Monthly Stats")
        chart.set_data([10, 25, 15, 40, 30])
    """

    def __init__(self, title="Chart", chart_type="bar", parent=None):
        super().__init__(parent)
        self.title = title
        self.chart_type = chart_type  # "bar" or "line"
        self.data = []
        self.labels = []
        self.color = QColor("#2196F3")

        self.setMinimumHeight(200)
        self.setMinimumWidth(300)

    def set_data(self, data, labels=None):
        """Set chart data"""
        self.data = data
        self.labels = labels or [str(i+1) for i in range(len(data))]
        self.update()

    def set_color(self, color):
        """Set chart color"""
        self.color = QColor(color)
        self.update()

    def paintEvent(self, event):
        """Draw chart"""
        if not self.data:
            return

        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # Calculate dimensions
        margin = 40
        width = self.width() - 2 * margin
        height = self.height() - 2 * margin - 30  # Space for title

        # Draw title
        painter.setPen(Qt.white)
        font = QFont("Arial", 12, QFont.Bold)
        painter.setFont(font)
        painter.drawText(self.rect(), Qt.AlignTop | Qt.AlignHCenter, self.title)

        # Calculate max value
        max_value = max(self.data) if self.data else 1

        if self.chart_type == "bar":
            self._draw_bar_chart(painter, margin, margin + 30, width, height, max_value)
        else:
            self._draw_line_chart(painter, margin, margin + 30, width, height, max_value)

    def _draw_bar_chart(self, painter, x, y, width, height, max_value):
        """Draw bar chart"""
        bar_width = width / len(self.data)

        for i, value in enumerate(self.data):
            bar_height = (value / max_value) * height
            bar_x = x + i * bar_width + bar_width * 0.1
            bar_y = y + height - bar_height
            bar_w = bar_width * 0.8

            # Draw bar with gradient
            gradient = QLinearGradient(0, bar_y, 0, y + height)
            gradient.setColorAt(0, self.color)
            gradient.setColorAt(1, self.color.darker(150))

            painter.setBrush(gradient)
            painter.setPen(Qt.NoPen)
            painter.drawRoundedRect(int(bar_x), int(bar_y), int(bar_w), int(bar_height), 4, 4)

            # Draw value on top
            painter.setPen(Qt.white)
            font = QFont("Arial", 9)
            painter.setFont(font)
            painter.drawText(
                int(bar_x),
                int(bar_y - 5),
                int(bar_w),
                20,
                Qt.AlignCenter,
                str(int(value))
            )

    def _draw_line_chart(self, painter, x, y, width, height, max_value):
        """Draw line chart"""
        if len(self.data) < 2:
            return

        # Calculate points
        point_spacing = width / (len(self.data) - 1)
        points = []

        for i, value in enumerate(self.data):
            point_x = x + i * point_spacing
            point_y = y + height - (value / max_value) * height
            points.append(QPoint(int(point_x), int(point_y)))

        # Draw line
        pen = QPen(self.color, 3)
        painter.setPen(pen)

        for i in range(len(points) - 1):
            painter.drawLine(points[i], points[i + 1])

        # Draw points
        painter.setBrush(self.color)
        for point in points:
            painter.drawEllipse(point, 5, 5)


# ============================================================================
# END ADVANCED ANIMATED WIDGETS
# ============================================================================

class EmployeeForm(QWidget):
    """Placeholder for EmployeeForm - to be implemented"""
    pass

